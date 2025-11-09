# -*- coding: utf-8 -*-
"""
MainGUI.py - eCan.ai
"""

# ============================================================================
# 1. Standard Library Imports
# ============================================================================
import ast
import asyncio
import base64
import copy
import glob
import hashlib
import importlib
import importlib.util
import json
import math
import os
import platform
import requests
import socket
import random
import shutil
import re
import threading
import time
import traceback
import uuid
from datetime import datetime, timedelta, timezone
from csv import reader
from os.path import exists
from typing import List, Optional


# ============================================================================
# 2. Core Utility Imports
# ============================================================================
from utils.time_util import TimeUtil
from utils.logger_helper import logger_helper as logger
from utils.port_allocator import get_port_allocator
from bot.envi import getECBotDataHome

print(TimeUtil.formatted_now_with_ms() + " load MainGui start...")

# ============================================================================
# 3. Basic Model Imports (Required for startup)
# ============================================================================
from bot.ebbot import EBBOT
from bot.missions import EBMISSION
from bot.vehicles import VEHICLE
from common.models import BotModel, MissionModel, VehicleModel

print(TimeUtil.formatted_now_with_ms() + " load MainGui #0 finished...")

# ============================================================================
# 4. Network Library Imports (already included in standard library imports)
# ============================================================================

# ============================================================================
# 5. Cryptography Library Imports
# ============================================================================
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.fernet import Fernet

print(TimeUtil.formatted_now_with_ms() + " load MainGui #1 finished...")

# ============================================================================
# 6. Database Related Imports
# ============================================================================
from common.db_init import init_db, get_session
from common.services import MissionService, ProductService, SkillService, BotService, VehicleService

# ============================================================================
# 7. GUI Manager Imports
# ============================================================================
from gui.BotGUI import BotManager
from gui.MissionGUI import MissionManager
from gui.PlatoonGUI import PlatoonManager
from gui.ScheduleGUI import ScheduleManager
from gui.SkillManagerGUI import SkillManager
from gui.TrainGUI import TrainManager, ReminderManager
from gui.VehicleMonitorGUI import VehicleMonitorManager
from gui.ui_settings import SettingsManager

print(TimeUtil.formatted_now_with_ms() + " load MainGui #2 finished...")

# ============================================================================
# 8. Cloud Service Imports
# ============================================================================
from bot.Cloud import (send_dequeue_tasks_to_cloud, send_schedule_request_to_cloud,
                      send_update_missions_ex_status_to_cloud, set_up_cloud, upload_file,
                      send_add_missions_request_to_cloud, send_remove_missions_request_to_cloud,
                      send_update_missions_request_to_cloud, send_add_bots_request_to_cloud,
                      send_update_bots_request_to_cloud, send_remove_bots_request_to_cloud,
                      send_add_skills_request_to_cloud, send_get_bots_request_to_cloud,
                      send_query_chat_request_to_cloud, download_file, send_report_vehicles_to_cloud,
                      send_update_vehicles_request_to_cloud)

print(TimeUtil.formatted_now_with_ms() + " load MainGui #3 finished...")

# ============================================================================
# 9. Bot Module Imports
# ============================================================================
from bot.Logger import log3
from bot.WorkSkill import WORKSKILL
from bot.adsPowerSkill import formADSProfileBatchesFor1Vehicle, updateIndividualProfileFromBatchSavedTxt
from bot.basicSkill import processExternalHook, symTab, STEP_GAP, setMissionInput, getScreenSize
from bot.genSkills import genSkillCode, getWorkRunSettings, setWorkSettingsSkill, SkillGeneratorTable, ManagerTriggerTable
from bot.inventories import INVENTORY
from bot.wanChat import wanSendMessage, wanSendMessage8
from bot.network import myname, fieldLinks, commanderIP, commanderXport, runCommanderLAN, runPlatoonLAN
from gui.utils.system_info import get_system_info_manager, get_complete_system_info
from bot.readSkill import RAIS, ARAIS, first_step, get_printable_datetime, prepRunSkill, readPSkillFile, addNameSpaceToAddress, rpaRunAllSteps, running_step_index
from bot.labelSkill import handleExtLabelGenResults

print(TimeUtil.formatted_now_with_ms() + " load MainGui #4 finished...")

# ============================================================================
# 10. External Library Imports
# ============================================================================
import concurrent.futures
from qasync import QEventLoop

# ============================================================================
# 11. GUI Tool Imports
# ============================================================================
from gui.tool.MainGUITool import FileResource, StaticResource
from gui.encrypt import *
from gui.unified_browser_manager import get_unified_browser_manager
from auth.auth_manager import AuthManager

print(TimeUtil.formatted_now_with_ms() + " load MainGui #5 finished...")

# ============================================================================
# 12. Agent Module Imports (Most time-consuming, placed last)
# ============================================================================
from agent.db import initialize_ecan_database

print(TimeUtil.formatted_now_with_ms() + " load MainGui #6 finished...")

# ============================================================================
# 13. Global Variables and Configuration
# ============================================================================

START_TIME = 15      # 15 x 20 minute = 5 o'clock in the morning
Tzs = ["eastern", "central", "mountain", "pacific", "alaska", "hawaii"]
rpaConfig = None
ecb_data_homepath = getECBotDataHome()
in_data_string = ""

print(TimeUtil.formatted_now_with_ms() + " load MainGui finished...")


# class MainWindow(QWidget):
class MainWindow:
    def __init__(self, auth_manager: AuthManager, mainloop, ip,
                 user, homepath, machine_role, schedule_mode):
        """Initialize MainWindow with optimized non-blocking initialization"""
        self._init_start_time = time.time()
        logger.info("[MainWindow] 🚀 Starting optimized MainWindow initialization...")

        # Initialize status tracking first
        self._initialization_status = {
            'sync_init_complete': False,
            'async_init_complete': False,
            'fully_ready': False,
            'ui_ready': False,
            'critical_services_ready': False
        }
        
        # Initialize shutdown flag
        self._shutting_down = False

        # ============================================================================
        # PHASE 1: CRITICAL SYNCHRONOUS INITIALIZATION (UI-blocking, keep minimal)
        # ============================================================================
        logger.info("[MainWindow] 📋 Phase 1: Critical synchronous initialization...")

        # 1. Core system (essential for basic functionality)
        self._init_core_system(auth_manager, mainloop, ip, user, homepath, machine_role, schedule_mode)

        # 2. User & environment (lightweight)
        self._init_user_environment(user, machine_role)

        # 3. System information (lightweight)
        self._init_system_info()

        # 4. Directory & file system (essential paths)
        self._init_file_system()

        # 5. Configuration management (needed for other components)
        self._init_configuration_manager()

        # 6. Business objects initialization (lightweight data structures)
        self._init_business_objects()

        # 7. Network communication setup (lightweight)
        self._init_network_communication()

        # 8. Database initialization (essential for data persistence)
        self._init_database()

        # 9. Setup local web server
        def _ensure_no_proxy_entries(extra_hosts: list[str]):
            existing = (
                os.environ.get("NO_PROXY")
                or os.environ.get("no_proxy")
                or ""
            )
            entries = [part.strip() for part in existing.split(',') if part.strip()]
            for host in extra_hosts:
                if host not in entries:
                    entries.append(host)
            no_proxy_value = ",".join(entries)
            os.environ["NO_PROXY"] = no_proxy_value
            os.environ["no_proxy"] = no_proxy_value

        _ensure_no_proxy_entries(["localhost", "127.0.0.1"])

        from agent.mcp.server.server import set_server_main_win
        from gui.LocalServer import start_local_server_in_thread

        set_server_main_win(self)
        start_local_server_in_thread(self)

        # Mark UI as ready for display and sync init complete
        self._initialization_status['ui_ready'] = True
        self._initialization_status['sync_init_complete'] = True
        ui_ready_time = time.time() - self._init_start_time
        logger.info(f"[MainWindow] ✅ Phase 1 completed in {ui_ready_time:.2f}s - UI ready for display")
        logger.info("[MainWindow] 🎯 Synchronous initialization complete - UI can be displayed to user")

        # ============================================================================
        # PHASE 2: BACKGROUND INITIALIZATION (Non-blocking)
        # ============================================================================
        logger.info("[MainWindow] 🚀 Phase 2: Starting background initialization...")

        # Notify IPC Registry that system is ready, clear cache to ensure immediate effect
        self._initialization_status['fully_ready'] = True
        try:
            from gui.ipc.registry import IPCHandlerRegistry
            IPCHandlerRegistry.force_system_ready(True)
        except Exception as cache_e:
            logger.warning(f"[MainWindow] Failed to update IPC registry cache: {cache_e}")

        # Start background initialization immediately
        try:
            # Try to create task in existing event loop
            loop = asyncio.get_running_loop()
            loop.create_task(self._async_background_initialization())
            logger.info("[MainWindow] ✅ Background initialization task created successfully")
        except RuntimeError as e:
            logger.error(f"[MainWindow] ⚠️ No running event loop for background initialization: {e}")
            # Mark initialization complete directly to avoid frontend infinite waiting
            self._initialization_status['async_init_complete'] = True
            logger.info("[MainWindow] ✅ Marked async_init_complete=True due to no event loop")

        logger.info("[MainWindow] ✅ MainWindow basic initialization completed - background services starting")

    async def _update_vehicle_metrics_async(self, vehicle):
        """Asynchronously update vehicle performance metrics without blocking main thread"""
        try:
            logger.debug(f"[MainWindow] 📊 Starting async metrics update for vehicle: {vehicle.getName()}")
            
            # Delay briefly to let main thread complete initialization
            await asyncio.sleep(0.5)
            
            # Update performance metrics in background
            start_time = time.time()
            vehicle.updateSystemMetrics()
            elapsed_time = time.time() - start_time
            
            logger.info(f"[MainWindow] ✅ Updated performance metrics for vehicle: {vehicle.getName()} in {elapsed_time:.3f}s")
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to update vehicle metrics for {vehicle.getName()}: {e}")
            import traceback
            logger.debug(f"[MainWindow] Vehicle metrics error traceback: {traceback.format_exc()}")

    def _validate_and_fix_default_llm_model(self):
        """
        Validate that default_llm_model belongs to the current default_llm provider.
        If not, clear it and use the provider's default model.
        This prevents issues when switching providers.
        """
        try:
            default_llm = self.config_manager.general_settings.default_llm
            default_llm_model = self.config_manager.general_settings.default_llm_model
            
            # Use llm_manager to validate and fix
            corrected_model, was_fixed = self.config_manager.llm_manager.validate_and_fix_default_llm_model(
                default_llm, default_llm_model
            )
            
            if was_fixed:
                # Update and save the corrected model
                self.config_manager.general_settings.default_llm_model = corrected_model
                self.config_manager.general_settings.save()
        except Exception as e:
            logger.warning(f"[MainWindow] Error validating default_llm_model: {e}")

    def update_all_llms(self, reason="unknown"):
        """
        Update mainwin.llm and all agents' LLMs (skill_llm and browser_use LLM).
        
        This method should be called when:
        - LLM provider is changed
        - Proxy settings change
        - LLM configuration is updated
        
        Args:
            reason: Description of why the LLM is being updated (for logging)
        """
        try:
            logger.info(f"[MainWindow] 🔄 Updating all LLMs - Reason: {reason}")
            from agent.ec_skills.llm_utils.llm_utils import pick_llm
            
            # Recreate mainwin.llm
            # Important: Use allow_fallback=False to prevent overriding user's provider selection
            new_llm = pick_llm(
                self.config_manager.general_settings.default_llm,
                self.config_manager.llm_manager.get_all_providers(),
                self.config_manager,
                allow_fallback=False  # Don't auto-fallback when hot-updating
            )
            
            if not new_llm:
                logger.warning("[MainWindow] ⚠️ Failed to recreate LLM")
                return False
            
            old_llm_type = type(self.llm).__name__ if self.llm else "None"
            new_llm_type = type(new_llm).__name__
            self.llm = new_llm
            logger.info(f"[MainWindow] ✅ LLM recreated successfully - {old_llm_type} → {new_llm_type}")
            
            # Recreate browser_use LLM with new configuration
            new_browser_use_llm = None
            try:
                from agent.playwright import create_browser_use_llm
                new_browser_use_llm = create_browser_use_llm(mainwin=self)
                if new_browser_use_llm:
                    # Get detailed info for browser_use LLM
                    browser_llm_type = type(new_browser_use_llm).__name__
                    browser_llm_details = []
                    
                    # Extract model name
                    if hasattr(new_browser_use_llm, 'model_name'):
                        browser_llm_details.append(f"model={new_browser_use_llm.model_name}")
                    elif hasattr(new_browser_use_llm, 'model'):
                        browser_llm_details.append(f"model={new_browser_use_llm.model}")
                    
                    # Extract endpoint
                    if hasattr(new_browser_use_llm, 'openai_api_base') and new_browser_use_llm.openai_api_base:
                        browser_llm_details.append(f"endpoint={new_browser_use_llm.openai_api_base}")
                    elif hasattr(new_browser_use_llm, 'base_url') and new_browser_use_llm.base_url:
                        browser_llm_details.append(f"endpoint={new_browser_use_llm.base_url}")
                    
                    # Add provider info
                    default_llm = self.config_manager.general_settings.default_llm
                    if default_llm:
                        provider = self.config_manager.llm_manager.get_provider(default_llm)
                        if provider:
                            provider_display = provider.get('display_name', default_llm)
                            browser_llm_details.append(f"provider={provider_display}")
                    
                    detail_str = f" ({', '.join(browser_llm_details)})" if browser_llm_details else ""
                    logger.info(f"[MainWindow] ✅ Browser-use LLM recreated: {browser_llm_type}{detail_str}")
                else:
                    logger.warning("[MainWindow] ⚠️ Failed to recreate browser-use LLM")
            except Exception as e:
                logger.warning(f"[MainWindow] ⚠️ Error recreating browser-use LLM: {e}")
            
            # Update all agents' skill_llm and llm (browser_use)
            updated_agents = 0
            for agent in self.agents:
                # Update skill_llm
                if hasattr(agent, 'set_skill_llm'):
                    agent.set_skill_llm(self.llm)
                    updated_agents += 1
                    logger.debug(f"[MainWindow] Updated skill_llm for agent: {agent.card.name}")
                
                # Update agent.llm (browser_use LLM)
                if new_browser_use_llm and hasattr(agent, 'llm'):
                    agent.llm = new_browser_use_llm
                    logger.debug(f"[MainWindow] Updated browser-use LLM for agent: {agent.card.name}")
            
            logger.info(f"[MainWindow] ✅ Updated LLMs for {updated_agents} agents")
            return True
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Error updating LLMs: {e}")
            import traceback
            logger.debug(f"[MainWindow] Error traceback: {traceback.format_exc()}")
            return False
    
    def _register_proxy_change_callback(self):
        """
        Register callback with ProxyManager to recreate LLM instances when proxy state changes.
        
        This ensures that when proxy is enabled/disabled, all LLM instances will be recreated
        with the correct proxy configuration.
        """
        try:
            from agent.ec_skills.system_proxy import get_proxy_manager
            
            proxy_manager = get_proxy_manager()
            if not proxy_manager:
                logger.debug("[MainWindow] ProxyManager not available, skipping callback registration")
                return
            
            def on_proxy_change(proxies):
                """
                Callback fired when proxy state changes.
                Args:
                    proxies: None if proxy disabled, Dict if proxy enabled
                """
                if proxies:
                    proxy_info = f"HTTP: {proxies.get('http://', 'N/A')}, HTTPS: {proxies.get('https://', 'N/A')}"
                    logger.info(f"[MainWindow] 🌐 Proxy enabled - {proxy_info}")
                    reason = f"Proxy enabled - {proxy_info}"
                else:
                    logger.info("[MainWindow] 🌐 Proxy disabled - using direct connection")
                    reason = "Proxy disabled"
                
                # Use unified method to update all LLMs
                self.update_all_llms(reason=reason)
            
            # Register the callback
            unregister = proxy_manager.register_callback(on_proxy_change)
            logger.info("[MainWindow] ✅ Registered proxy change callback for LLM recreation")
            
            # Store unregister function for cleanup (if needed)
            self._proxy_callback_unregister = unregister
            
        except Exception as e:
            logger.warning(f"[MainWindow] Failed to register proxy change callback: {e}")

    def _schedule_delayed_metrics_update(self, vehicle):
        """Schedule delayed performance monitoring update, waiting for event loop availability"""
        try:
            logger.debug(f"[MainWindow] 🔄 Attempting delayed metrics update for vehicle: {vehicle.getName()}")

            # Try to get event loop again
            try:
                loop = asyncio.get_running_loop()
                loop.create_task(self._update_vehicle_metrics_async(vehicle))
                logger.debug(f"[MainWindow] ✅ Successfully scheduled delayed metrics update for vehicle: {vehicle.getName()}")
            except RuntimeError as e:
                # If event loop is still not available after delay, log as warning and continue with defaults
                logger.warning(f"[MainWindow] ⚠️ Event loop still not available for delayed metrics update: {e}")
                logger.info(f"[MainWindow] 📊 Vehicle {vehicle.getName()} will continue with default metrics")

        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to schedule delayed metrics update for {vehicle.getName()}: {e}")

    async def _async_background_initialization(self):
        """
        Perform heavy initialization operations in background to avoid blocking UI
        """
        try:
            logger.info("[MainWindow] 🔄 Starting background initialization phase...")

            # Phase 2A: Database and critical services (parallel where possible)
            logger.info("[MainWindow] 📊 Initializing database and critical services...")

            # Run database initialization in executor to avoid blocking
            await asyncio.get_event_loop().run_in_executor(
                None, self._init_database_services
            )

            # Now that database services are ready, check vehicles and load data
            logger.info("[MainWindow] 🚗 Checking vehicles after database services ready...")
            await asyncio.get_event_loop().run_in_executor(
                None, self._check_vehicles_with_database
            )

            # Phase 2B: Data loading (must be after database services are ready)
            logger.info("[MainWindow] 📂 Loading local data...")
            await asyncio.get_event_loop().run_in_executor(
                None, self._init_local_data_loading
            )

            # Phase 2C: Extensions and plugins (can run in parallel)
            logger.info("[MainWindow] 🔌 Loading extensions and plugins...")
            extensions_task = asyncio.get_event_loop().run_in_executor(
                None, self._init_extensions_and_plugins
            )

            # Phase 2D: Server and agent initialization (heavy operations)
            logger.info("[MainWindow] 🤖 Initializing servers and agents...")
            servers_task = asyncio.get_event_loop().run_in_executor(
                None, self._init_servers_and_agents
            )

            # Wait for remaining parallel services to complete
            await extensions_task
            await servers_task

            self._initialization_status['critical_services_ready'] = True
            logger.info("[MainWindow] ✅ Critical services ready")

            # Phase 2E: Task management and async tasks
            logger.info("[MainWindow] 📋 Initializing task management...")
            await asyncio.get_event_loop().run_in_executor(
                None, self._init_task_management
            )

            # Phase 2F: Start offline sync on startup (non-blocking)
            logger.info("[MainWindow] 🔄 Starting offline sync on startup (non-blocking)...")
            asyncio.get_event_loop().run_in_executor(
                None, self._startup_sync_offline_cloud_cache
            )
            
            # Initialize async tasks
            self._init_async_tasks()

            # Phase 2F: Final background services
            logger.info("[MainWindow] 🏁 Starting final background services...")
            await self._finalize_async_initialization()

            # Mark full initialization as complete
            self._initialization_status['async_init_complete'] = True
            
            total_time = time.time() - self._init_start_time
            logger.info(f"[MainWindow] ✅ Background initialization completed successfully in {total_time:.2f}s total")

        except Exception as e:
            logger.error(f"[MainWindow] ❌ Background initialization failed: {e}")
            import traceback
            logger.error(f"[MainWindow] Background initialization error traceback:\n{traceback.format_exc()}")
            # Even if background init fails, mark as complete to prevent hanging
            self._initialization_status['async_init_complete'] = True
            logger.info("[MainWindow] ✅ Marked async_init_complete=True after background initialization failure")

    async def _finalize_async_initialization(self):
        """Finalize async initialization and start final background services"""
        logger.info("[MainWindow] 🏁 Finalizing async initialization...")

        # Save current settings
        await asyncio.get_event_loop().run_in_executor(None, self.saveSettings)

        # Now that database services are available, save vehicles that were skipped earlier
        if hasattr(self, 'vehicles') and self.vehicles and hasattr(self, 'vehicle_service') and self.vehicle_service:
            logger.info("[MainWindow] 🚗 Saving vehicles to database (deferred from sync phase)...")
            # Run vehicle saving in executor to avoid blocking
            await asyncio.get_event_loop().run_in_executor(None, self._save_vehicles_to_database)

        # Log final vehicle status
        logger.info(f"[MainWindow] Final vehicle count: {len(getattr(self, 'vehicles', []))}")
        for v in getattr(self, 'vehicles', []):
            logger.debug(f"[MainWindow] Vehicle: {v.getName()}, Status: {v.getStatus()}")

        # Start cloud sync if needed (deferred from sync phase)
        if getattr(self, '_should_start_cloud_sync', False):
            logger.info("[MainWindow] 🌐 Starting deferred cloud data sync...")
            try:
                loop = asyncio.get_running_loop()
                loop.create_task(self._async_sync_cloud_data())
            except RuntimeError:
                logger.error("[MainWindow] No running event loop for cloud sync, skipping")

        # Start final background services - wait for agents to be ready before marking fully ready
        logger.info("[MainWindow] 🚀 Starting final background services...")
        try:
            loop = asyncio.get_running_loop()
            agents_task = loop.create_task(self.async_agents_init())
            loop.create_task(self._async_setup_browser_manager())
            loop.create_task(self._async_start_lightrag())
            self.wan_sub_task = loop.create_task(self._async_start_wan_chat())
            self.llm_sub_task = loop.create_task(self._async_start_llm_subscription())
            # self.cloud_show_sub_task = loop.create_task(self._async_start_cloud_show_subscription())
            logger.info("[MainWindow] ✅ All final background service tasks created successfully")
        except RuntimeError as e:
            logger.error(f"[MainWindow] ⚠️ No running event loop for final background services: {e}")
            logger.info("[MainWindow] 📋 Skipping background services - system will work with basic functionality")
            agents_task = None

        # Wait for agents initialization to complete before marking system fully ready
        try:
            if agents_task is not None:
                logger.info("[MainWindow] ⏳ Waiting for agents initialization to complete...")
                start_time = time.time()
                await agents_task
                elapsed_time = time.time() - start_time
                logger.info(f"[MainWindow] ✅ Agents initialization completed in {elapsed_time:.3f}s")
            else:
                logger.warning("[MainWindow] ⚠️ No agents task to wait for - skipping agent initialization")
                # Give system some time to stabilize
                await asyncio.sleep(1.0)
                logger.info("[MainWindow] 📋 System stabilization delay completed")
            
            # # Now mark system as fully ready since agents are loaded
            # self._initialization_status['fully_ready'] = True

            # # Notify IPC Registry that system is ready, clear cache to ensure immediate effect
            # try:
            #     from gui.ipc.registry import IPCHandlerRegistry
            #     IPCHandlerRegistry.force_system_ready(True)
            # except Exception as cache_e:
            #     logger.warning(f"[MainWindow] Failed to update IPC registry cache: {cache_e}")


            # Notify update home agents page
            from app_context import AppContext
            web_gui = AppContext.get_web_gui()
            web_gui.get_ipc_api().update_org_agents()

            logger.info("[MainWindow] 🎉 System is now fully ready with all data loaded!")
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Agents initialization failed: {e}")
            logger.warning("[MainWindow] ⚠️ System marked as ready despite agents initialization failure")

        # Start completely independent delayed task to copy example my_skills
        # This runs after all critical initialization is complete and won't block anything
        try:
            loop = asyncio.get_running_loop()
            loop.create_task(self._delayed_copy_example_my_skills())
            logger.debug("[MainWindow] 📚 Scheduled delayed copy of example my_skills (non-blocking)")
        except RuntimeError as e:
            logger.debug(f"[MainWindow] No event loop for delayed my_skills copy: {e}")

        logger.info("[MainWindow] ✅ Async initialization finalized")

        # Check LLM provider configuration and show onboarding guide if needed
        try:
            loop = asyncio.get_running_loop()
            loop.create_task(self.config_manager.llm_manager.check_and_show_onboarding())
            logger.debug("[MainWindow] 📋 Scheduled LLM provider onboarding check (non-blocking)")
        except RuntimeError as e:
            logger.debug(f"[MainWindow] No event loop for LLM provider onboarding check: {e}")

        # Register proxy state change callback to recreate LLM when proxy changes
        self._register_proxy_change_callback()

    def _save_vehicles_to_database(self):
        """Save vehicles to database (synchronous helper method for executor)"""
        try:
            for vehicle in self.vehicles:
                try:
                    self.saveVehicle(vehicle)
                    logger.debug(f"[MainWindow] Saved vehicle: {vehicle.getName()}")
                except Exception as e:
                    logger.error(f"[MainWindow] Failed to save vehicle {vehicle.getName()}: {e}")
        except Exception as e:
            logger.error(f"[MainWindow] Error in vehicle saving process: {e}")

    async def _delayed_copy_example_my_skills(self):
        """
        Delayed async task to copy example skills after system is fully initialized
        This runs completely independently and won't block any startup process
        """
        try:
            # Wait 5 seconds after system is fully ready to ensure no impact on startup
            logger.debug("[MainWindow] 📚 Waiting 5s before copying example my_skills...")
            await asyncio.sleep(5.0)
            
            # Run the actual copy operation in executor to avoid blocking event loop
            logger.debug("[MainWindow] 📚 Starting example my_skills copy in background thread...")
            await asyncio.get_event_loop().run_in_executor(
                None, self._copy_example_my_skills
            )
            logger.debug("[MainWindow] 📚 Example my_skills copy completed")
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Delayed my_skills copy failed: {e}")
            # Silently fail - this is a nice-to-have feature, not critical

    def _copy_example_my_skills(self):
        """
        Copy example skills from resource/my_skills to appdata/my_skills directory
        This method is synchronous and designed to be run in an executor
        """
        try:
            from config.app_info import app_info
            
            # Source directory: resource/my_skills
            source_dir = os.path.join(app_info.app_resources_path, "my_skills")
            
            # Target directory: appdata/my_skills  
            target_dir = os.path.join(app_info.appdata_path, "my_skills")
            
            # Check if source directory exists
            if not os.path.exists(source_dir):
                logger.debug(f"[MainWindow] 📂 Source my_skills directory not found: {source_dir}")
                return
            
            # Create target directory if it doesn't exist
            if not os.path.exists(target_dir):
                os.makedirs(target_dir, exist_ok=True)
                logger.debug(f"[MainWindow] 📁 Created target my_skills directory: {target_dir}")
            
            # Copy each skill directory from source to target
            copied_count = 0
            updated_count = 0
            skipped_count = 0
            
            for skill_name in os.listdir(source_dir):
                source_skill_path = os.path.join(source_dir, skill_name)
                target_skill_path = os.path.join(target_dir, skill_name)
                
                # Skip if not a directory
                if not os.path.isdir(source_skill_path):
                    continue
                
                try:
                    if os.path.exists(target_skill_path):
                        # Target exists: check if files need updating
                        # Only update files if source is newer or different
                        logger.debug(f"[MainWindow] 🔍 Skill already exists, checking for updates: {skill_name}")
                        
                        files_updated = 0
                        files_skipped = 0
                        
                        # Walk through source directory and compare files
                        for root, dirs, files in os.walk(source_skill_path):
                            # Calculate relative path from source
                            rel_path = os.path.relpath(root, source_skill_path)
                            target_root = os.path.join(target_skill_path, rel_path)
                            
                            # Create target directory if it doesn't exist
                            os.makedirs(target_root, exist_ok=True)
                            
                            # Check and copy files only if source is newer or different
                            for file in files:
                                source_file = os.path.join(root, file)
                                target_file = os.path.join(target_root, file)
                                
                                try:
                                    # Check if target file exists
                                    if os.path.exists(target_file):
                                        # Compare modification times
                                        source_mtime = os.path.getmtime(source_file)
                                        target_mtime = os.path.getmtime(target_file)
                                        
                                        # Only update if source is newer
                                        if source_mtime > target_mtime:
                                            shutil.copy2(source_file, target_file)
                                            files_updated += 1
                                            logger.debug(f"[MainWindow] 📝 Updated file: {os.path.join(rel_path, file)}")
                                        else:
                                            files_skipped += 1
                                    else:
                                        # Target file doesn't exist, copy it
                                        shutil.copy2(source_file, target_file)
                                        files_updated += 1
                                        logger.debug(f"[MainWindow] 📄 Copied new file: {os.path.join(rel_path, file)}")
                                except Exception as file_error:
                                    logger.debug(f"[MainWindow] ⚠️ Failed to process file {file}: {file_error}")
                        
                        if files_updated > 0:
                            updated_count += 1
                            logger.debug(f"[MainWindow] ✅ Updated {files_updated} file(s) in skill: {skill_name}")
                        else:
                            skipped_count += 1
                            logger.debug(f"[MainWindow] ⏭️ Skill up to date, skipped: {skill_name}")
                    else:
                        # Target doesn't exist: copy entire directory
                        shutil.copytree(source_skill_path, target_skill_path)
                        logger.debug(f"[MainWindow] ✅ Copied example skill: {skill_name}")
                        copied_count += 1
                except Exception as copy_error:
                    logger.debug(f"[MainWindow] ❌ Failed to copy/update skill {skill_name}: {copy_error}")
            
            if copied_count > 0:
                logger.info(f"[MainWindow] 🎉 Successfully copied {copied_count} example skill(s) to my_skills directory")
            if updated_count > 0:
                logger.info(f"[MainWindow] 🔄 Successfully updated {updated_count} example skill(s) in my_skills directory")
            if skipped_count > 0:
                logger.debug(f"[MainWindow] 📊 Skipped {skipped_count} existing skill(s)")
            if copied_count == 0 and updated_count == 0 and skipped_count == 0:
                logger.debug(f"[MainWindow] 📂 No skills found to copy from {source_dir}")
                
        except Exception as e:
            logger.debug(f"[MainWindow] ❌ Error copying example my_skills: {e}")
            # Silently continue - this is a nice-to-have feature

    def is_ui_ready(self) -> bool:
        """Check if UI is ready for display (minimal initialization complete)"""
        return self._initialization_status.get('ui_ready', False)

    def are_critical_services_ready(self) -> bool:
        """Check if critical services are ready"""
        return self._initialization_status.get('critical_services_ready', False)

    def get_initialization_progress(self) -> dict:
        """Get detailed initialization progress information"""
        return {
            'ui_ready': self._initialization_status.get('ui_ready', False),
            'critical_services_ready': self._initialization_status.get('critical_services_ready', False),
            'async_init_complete': self._initialization_status.get('async_init_complete', False),
            'fully_ready': self._initialization_status.get('fully_ready', False),
            'sync_init_complete': self._initialization_status.get('sync_init_complete', False)
        }



    def _init_core_system(self, auth_manager, mainloop, ip, user, homepath, machine_role, schedule_mode):
        """Initialize core system components"""
        logger.info("[MainWindow] 🔧 Initializing core system components...")

        # Core references
        self.auth_manager = auth_manager
        self.mainLoop: QEventLoop = mainloop
        self.ip = ip
        self.machine_role = machine_role
        self.schedule_mode_param = schedule_mode  # Store for potential config override

        # Path normalization
        self.homepath = homepath.rstrip('/')

        # Core queues for inter-component communication
        self.gui_net_msg_queue = asyncio.Queue()
        self.gui_rpa_msg_queue = asyncio.Queue()
        self.gui_manager_msg_queue = asyncio.Queue()
        self.virtual_cloud_task_queue = asyncio.Queue()
        self.gui_monitor_msg_queue = asyncio.Queue()
        self.gui_chat_msg_queue = asyncio.Queue()
        self.wan_chat_msg_queue = asyncio.Queue()

        # Core resources
        self.tz = self.obtainTZ()
        self.file_resource = FileResource(self.homepath)
        self.static_resource = StaticResource()
        self.session = set_up_cloud()
        self.threadPoolExecutor = concurrent.futures.ThreadPoolExecutor(max_workers=16)
        
        # Port allocation for thread-safe agent port management
        self._port_allocator = get_port_allocator()

        # Machine role configuration
        if "Platoon" in self.machine_role:
            self.functions = "buyer,seller"
        elif "Commander" in self.machine_role:
            self.functions = "manager,hr,it"
        else:
            self.functions = ""

        logger.info(f"[MainWindow] ✅ Core system initialized - Role: {machine_role}, Functions: {self.functions}")

    def _init_user_environment(self, user, machine_role):
        """Initialize user environment and identity"""
        logger.info("[MainWindow] 👤 Initializing user environment...")

        self.owner = user
        # Normalize user to a safe email-like value
        self.user = user if (user and isinstance(user, str) and "@" in user) else "unknown@local"

        # Build chat_id safely
        try:
            local_part, domain_part = self.user.split("@", 1)
        except ValueError:
            local_part, domain_part = self.user, "local"

        domain_part_sanitized = domain_part.replace(".", "_")
        self.chat_id = f"{local_part}_{domain_part_sanitized}"
        self.log_user = self.chat_id

        # User-specific paths
        self.my_ecb_data_homepath = f"{ecb_data_homepath}/{self.log_user}"
        self.ecb_data_homepath = ecb_data_homepath

        # Role-specific chat ID modification
        self.host_role = machine_role
        if "Only" in self.host_role:
            self.chat_id = self.chat_id + "_Commander"
        else:
            self.chat_id = self.chat_id + "_" + "".join(self.host_role.split())

        # User ID generation
        usrparts = self.user.split("@")
        usrdomainparts = usrparts[1].split(".")
        self.uid = usrparts[0] + "_" + usrdomainparts[0]

        logger.info(f"[MainWindow] ✅ User environment initialized - Chat ID: {self.chat_id}, UID: {self.uid}")

    def _init_system_info(self):
        """Initialize system information and hardware details"""
        logger.info("[MainWindow] 💻 Initializing system information...")

        # Use system information manager to get complete information
        self.system_info_manager = get_system_info_manager()
        system_info = self.system_info_manager.get_complete_system_info()
        
        # Basic system information
        self.os_info = system_info.get('os_info', 'Unknown OS')
        self.platform = system_info.get('platform', 'unk')
        self.system = system_info.get('system', 'Unknown')
        self.architecture = system_info.get('architecture', '64bit')

        # OS short name
        if self.system == "Windows":
            self.os_short = "win"
        elif self.system == "Linux":
            self.os_short = "linux"
        elif self.system == "Darwin":
            self.os_short = "mac"
        else:
            self.os_short = "other"
        
        # Processor information
        processor_info = system_info.get('processor', {})
        self.cpuinfo = processor_info
        self.processor = processor_info.get('brand_raw', 'Unknown Processor')
        self.cpu_cores = processor_info.get('count', 1)
        self.cpu_threads = processor_info.get('threads', 1)
        self.cpu_speed = processor_info.get('hz_advertised_friendly', 'Unknown Speed')

        # Memory information
        memory_info = system_info.get('memory', {})
        self.total_memory = memory_info.get('total_gb', 0.0)
        self.virtual_memory = type('VirtualMemory', (), {
            'total': memory_info.get('total', 0),
            'available': memory_info.get('available', 0),
            'percent': memory_info.get('percent', 0.0),
            'used': memory_info.get('used', 0)
        })()

        # Screen information
        self.screen_size = getScreenSize()

        # Machine identification information - using system information manager
        # self.machine_name = system_info.get('machine_name', 'Unknown-Computer')
        # Machine identification information - prefer OS hostname (avoid friendly names)
        try:
            candidates = []
            import socket, platform, os as _os
            if self.system == "Windows":
                candidates.extend([
                    _os.environ.get("COMPUTERNAME"),
                    socket.gethostname(),
                    platform.node(),
                ])
            else:
                candidates.extend([
                    _os.environ.get("HOSTNAME"),
                    socket.gethostname(),
                    platform.node(),
                ])
            candidates.append(system_info.get('machine_name'))
            print("candidates:", candidates)
            _mn = next((x for x in candidates if isinstance(x, str) and x.strip()), None)
            if isinstance(_mn, str):
                _mn = _mn.strip().strip('"').strip("'").replace("’", "'")
                _mn = " ".join(_mn.split())
                if "." in _mn:
                    _mn = _mn.split(".", 1)[0]
            self.machine_name = _mn or "Unknown-Computer"
        except Exception:
            self.machine_name = system_info.get('machine_name', 'Unknown-Computer')

        self.device_type = system_info.get('device_type', 'Computer')
        self.system_arch = system_info.get('system_arch', 'unknown')
        self.commander_name = ""

        logger.info(f"[MainWindow] ✅ System info initialized - OS: {self.os_info}, CPU: {self.processor}, Memory: {self.total_memory:.1f}GB")
        logger.info(f"[MainWindow] ✅ Device info - Name: {self.machine_name}, Type: {self.device_type}, Arch: {self.system_arch}")
        logger.info(f"[MainWindow] ✅ System info", system_info)


    def _init_file_system(self):
        """Initialize file system directories and paths"""
        logger.info("[MainWindow] 📁 Initializing file system...")

        # Create essential directories
        resource_data_dir = f"{self.my_ecb_data_homepath}/resource/data/"
        if not os.path.exists(resource_data_dir):
            os.makedirs(resource_data_dir)

        self.temp_dir = os.path.join(self.my_ecb_data_homepath, "temp")
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir, exist_ok=True)
            logger.info(f"Created temp directory: {self.temp_dir}")

        self.ads_profile_dir = self.my_ecb_data_homepath + "/ads_profiles/"
        if not os.path.exists(self.ads_profile_dir):
            os.makedirs(self.ads_profile_dir)

        # File paths
        self.VEHICLES_FILE = self.my_ecb_data_homepath + "/vehicles.json"
        self.dbfile = f"{self.my_ecb_data_homepath}/resource/data/myecb.db"
        self.product_catelog_file = f"{self.my_ecb_data_homepath}/resource/data/product_catelog.json"
        self.build_dom_tree_script_path = f"{self.homepath}/resource/build_dom_tree.js"

        logger.info(f"[MainWindow] ✅ File system initialized - Data path: {self.my_ecb_data_homepath}")

    def _init_configuration_manager(self):
        """Initialize configuration management system"""
        logger.info("[MainWindow] ⚙️ Initializing configuration manager...")

        from gui.manager import ConfigManager
        self.config_manager = ConfigManager(self.my_ecb_data_homepath)

        # Set display resolution after config_manager is initialized
        display_resolution = "D"+str(self.screen_size[0])+"X"+str(self.screen_size[1])
        self.config_manager.general_settings.display_resolution = display_resolution

        # Set default webdriver path if not already configured
        if not self.config_manager.general_settings.default_webdriver_path:
            self.config_manager.general_settings.default_webdriver_path = f"{self.homepath}/chromedriver-win64/chromedriver.exe"

        self.titles = ["Director", "Product Manager", "Engineer Manager", "Team Leader", "Engineer", "Sales", "Analyst", "Senior Analyst"]
        self.ranks = ["E6", "E7", "E8", "E9", "E10", "E11", "E12", "E13", "E14", "E15", "E16", "B1", "B2", "B3", "B4", "B5", "B6", "B7", "B8"]
        self.personalities = ["Introvert", "Extrovert"]

        logger.info(f"[MainWindow] ✅ Configuration manager initialized - Debug: {self.config_manager.general_settings.debug_mode}, Schedule: {self.config_manager.general_settings.schedule_mode}")

    def _init_database(self):
        """Initialize database and related services with parallel optimization"""
        logger.info("[MainWindow] 🗄️ Initializing database ...")

        # Initialize database manager and chat service
        start_time = time.time()
        
        # Initialize eCan database system
        self.ec_db_mgr = initialize_ecan_database(
            self.my_ecb_data_homepath, 
            auto_migrate=True
        )
        db_init_time = time.time() - start_time
        logger.info(f"[MainWindow] ✅ Database manager initialized with optimized pool in {db_init_time:.3f}s")
        
        # Load default template data for new users
        start_time = time.time()
        self._load_default_template_data()
        template_init_time = time.time() - start_time
        logger.info(f"[MainWindow] ✅ Default template data loaded in {template_init_time:.3f}s")

        # TODO need to remove this, use ec_db_mgr get_chat_service
        self.db_chat_service = self.ec_db_mgr.get_chat_service()

        if "Commander" in self.machine_role:
            # Initialize database for Commander role
            start_time = time.time()
            engine = init_db(self.dbfile)
            session = get_session(engine)
            db_init_time = time.time() - start_time
            logger.info(f"[MainWindow] 🗄️ ecbot Database engine created in {db_init_time:.3f}s")

            # Store engine and session for service initialization
            self._db_engine = engine
            self._db_session = session

    def _init_database_services(self):
        """Initialize database and related services with parallel optimization"""
        logger.info("[MainWindow] 🗄️ Initializing database services...")

        if "Commander" in self.machine_role:
            # Initialize database for Commander role

            # Initialize services in parallel using thread pool
            start_time = time.time()
            # Define services to initialize
            services_config = [
                ('bot_service', BotService),
                ('mission_service', MissionService),
                ('product_service', ProductService),
                ('skill_service', SkillService),
                ('vehicle_service', VehicleService)
            ]

            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                # Submit all service initialization tasks using the generic function
                future_to_service = {
                    executor.submit(self._init_service_threaded, service_class, service_name): service_name
                    for service_name, service_class in services_config
                }

                # Wait for all services to complete
                for future in concurrent.futures.as_completed(future_to_service):
                    service_name = future_to_service[future]
                    try:
                        service_instance = future.result()
                        setattr(self, service_name, service_instance)
                        logger.debug(f"[MainWindow] ✅ {service_name} initialized successfully")
                    except Exception as e:
                        logger.error(f"[MainWindow] ❌ Failed to initialize {service_name}: {e}")
                        setattr(self, service_name, None)

            services_init_time = time.time() - start_time
            logger.info(f"[MainWindow] ✅ Database services initialized in {services_init_time:.3f}s (parallel)")

        else:
            # Platoon role - no database services needed
            self.bot_service: BotService = None
            self.mission_service: MissionService = None
            self.product_service: ProductService = None
            self.skill_service: SkillService = None
            self.vehicle_service: VehicleService = None

            logger.info("[MainWindow] ✅ Database services skipped for Platoon role")

    def _init_service_threaded(self, service_class, service_name):
        """Generic function to initialize any service in a separate thread

        Args:
            service_class: The service class to instantiate (e.g., BotService)
            service_name: The name of the service for logging (e.g., 'bot_service')

        Returns:
            service_instance: The initialized service instance
        """
        try:
            logger.debug(f"[MainWindow] Initializing {service_name} in thread...")
            service_instance = service_class(self, self._db_session, self._db_engine)
            logger.debug(f"[MainWindow] {service_name} thread initialization completed")
            return service_instance
        except Exception as e:
            logger.error(f"[MainWindow] {service_name} initialization failed: {e}")
            raise

    def _check_vehicles_with_database(self):
        """Check vehicles after database services are ready"""
        logger.info("[MainWindow] 🚗 Checking vehicles with database services...")

        # Now we can safely call checkVehicles since vehicle_service is ready
        self.checkVehicles()
        logger.info(f"[MainWindow] Vehicles checked: {len(self.vehicles)} found")
        for v in self.vehicles:
            logger.debug(f"Vehicle: {v.getName()}, Status: {v.getStatus()}")

        # All vehicles created during checkVehicles() will be automatically saved
        # because vehicle_service is now guaranteed to be available

    def _load_default_template_data(self):
        """Load default template data for new users"""
        try:
            logger.info("[MainWindow] 📋 Loading default template data...")
            
            # Load organization template data
            self._load_organization_template()
            
            # TODO: Add other template data loading here (agents, skills, tasks, etc.)
            # self._load_agent_template()
            # self._load_skill_template()
            # self._load_task_template()
            
            logger.info("[MainWindow] ✅ Default template data loading completed")
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to load default template data: {e}")
            logger.error(traceback.format_exc())

    def _load_organization_template(self):
        """Load default organization structure from template"""
        try:
            from agent.ec_org_ctrl import get_ec_org_ctrl
            
            # Get org manager (use existing database manager to avoid conflicts)
            ec_org_ctrl = get_ec_org_ctrl(self.ec_db_mgr)
            
            # Load org template using the controller
            result = ec_org_ctrl.load_org_template()
            
            if result.get("success"):
                created_count = result.get("created_count", 0)
                if created_count > 0:
                    logger.info(f"[MainWindow] ✅ Successfully loaded organization template: {result.get('message')}")
                else:
                    logger.info(f"[MainWindow] ℹ️ {result.get('message')}")
            else:
                logger.error(f"[MainWindow] ❌ Failed to load organization template: {result.get('error')}")
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to load organization template: {e}")
            logger.error(traceback.format_exc())

    def _init_business_objects(self):
        """Initialize business objects and data structures"""
        logger.info("[MainWindow] 📊 Initializing business objects...")

        # Core business objects
        self.agent_skills = []
        self.agent_tasks = []
        self.agent_tools = []
        self.agent_knowledges = []

        # Bot and mission management
        self.bots = []
        self.missions = []
        self.skills = []
        self.vehicles = []
        self.platoons = []
        self.products = []
        self.inventories = []

        # Additional missing variables
        self.commanderName = ""

        # Reports and tracking
        self.todaysReport = []
        self.todaysReports = []
        self.todaysPlatoonReports = []

        # Mission and task tracking (missing variables)
        self.missionsToday = []
        self.todaysSchedule = {}
        self.todays_scheduled_task_groups = {}
        self.unassigned_scheduled_task_groups = {}
        self.unassigned_reactive_task_groups = {}

        # UI state management
        self.selected_bot_row = -1
        self.selected_mission_row = -1
        self.selected_bot_item = None
        self.selected_mission_item = None

        # Component managers
        self.bot_manager = None
        self.missionWin = None
        self.lightrag_server = None
        self.train_manager = None
        self.reminder_manager = None
        self.platoonWin = None
        self.unified_browser_manager = None

        # Bot states and profiles
        self.bot_states = ["active", "disabled", "banned", "deleted"]
        self.todays_bot_profiles = []
        self.bot_cookie_site_lists = {}

        # Working state
        self.botRank = "soldier"
        self.rpa_work_assigned_for_today = False
        self.running_mission = None
        self.botsFingerPrintsReady = False
        self.default_webdriver = None
        self.working_state = "running_idle"
        self.staff_officer_on_line = False
        self.DONE_WITH_TODAY = True

        # Utility objects
        from lzstring import LZString
        self.zipper = LZString()
        self.trMission = self.createTrialRunMission()

        # Initialize skill manager
        self.skill_manager = SkillManager(self)

        # Initialize settings manager
        self.settings_manager: SettingsManager = SettingsManager(self)

        # Data files
        self.sellerInventoryJsonData = None
        self.botJsonData = None
        self.fetch_schedule_counter = 1

        logger.info("[MainWindow] ✅ Business objects initialized")

    def _init_network_communication(self):
        """Initialize network communication and related services"""
        logger.info("[MainWindow] 🌐 Initializing network communication...")

        # Network state
        self.wan_connected = False
        self.wan_msg_subscribed = False
        self.websocket = None

        # Network configuration based on role
        if "Commander" in self.machine_role:
            self.tcpServer = None
            self.commanderXport = None
            self.commander_name = self.machine_name
            self.commanderIP = ""  # Initialize for Commander role
        elif self.machine_role == "Platoon":
            self.commanderXport = None
            self.commanderIP = commanderIP
            self.tcpServer = None

        # Vehicle monitoring
        self.vehicle_monitor = VehicleMonitorManager(self)

        # Start network services based on role
        if "Platoon" not in self.machine_role:
            logger.info("[MainWindow] Starting commander side networking...")
            self.lan_task = self.mainLoop.create_task(runCommanderLAN(self))
        else:
            logger.info("[MainWindow] Starting platoon side networking...")
            self.lan_task = self.mainLoop.create_task(runPlatoonLAN(self, self.mainLoop))

        logger.info(f"[MainWindow] ✅ Network communication initialized - Role: {self.machine_role}")

        # Note: Vehicle checking moved to background phase after database services are ready

    def _init_local_data_loading(self):
        """Initialize and load local data"""
        logger.info("[MainWindow] 📂 Loading local data initialization...")
        
        if "Commander" in self.machine_role:
            # Load vehicle configuration
            self.readVehicleJsonFile()
            logger.info(f"[MainWindow] Vehicle files loaded: {len(getattr(self, 'vehiclesJsonData', []))} vehicles")

            # Load local bot data for immediate UI responsiveness
            bots_data = self.bot_service.find_all_bots()
            logger.info(f"[MainWindow] Loading {len(bots_data)} bots from database")
            self.loadLocalBots(bots_data)

            # Create new bots from Excel if available
            self.createNewBotsFromBotsXlsx()

            # Load local mission data
            missions_data = self.mission_service.find_missions_by_createon()
            logger.info(f"[MainWindow] Loading {len(missions_data)} missions from database")
            self.loadLocalMissions(missions_data)

            # Update daily skillset
            self.dailySkillsetUpdate()

            # Mark that cloud sync should be started after async initialization
            if not self.config_manager.general_settings.is_debug_enabled() or self.config_manager.general_settings.is_auto_mode():
                logger.info("[MainWindow] Cloud sync will be started after async initialization...")
                self._should_start_cloud_sync = True
            else:
                logger.info("[MainWindow] Cloud sync skipped (debug mode or manual schedule)")
                self._should_start_cloud_sync = False
                
        logger.info("[MainWindow] ✅ Local data loading completed")

    def _init_extensions_and_plugins(self):
        """Initialize extensions and plugins"""
        logger.info("[MainWindow] 🔌 Initializing extensions and plugins...")

        # Load RAIS extensions
        rais_extensions_file = self.my_ecb_data_homepath + "/my_rais_extensions/my_rais_extensions.json"
        added_handlers = []

        if os.path.isfile(rais_extensions_file):
            try:
                with open(rais_extensions_file, 'r') as rais_extensions:
                    user_rais_modules = json.load(rais_extensions)
                    logger.info(f"Loading {len(user_rais_modules)} RAIS extensions")

                    for i, user_module in enumerate(user_rais_modules):
                        module_file = self.my_ecb_data_homepath + "/" + user_module["dir"] + "/" + user_module["file"]
                        added_ins = user_module['instructions']
                        module_name = os.path.splitext(user_module["file"])[0]

                        try:
                            spec = importlib.util.spec_from_file_location(module_name, module_file)
                            module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(module)

                            for ins in added_ins:
                                if hasattr(module, ins["handler"]):
                                    RAIS[ins["instruction name"]] = getattr(module, ins["handler"])
                                    ARAIS[ins["instruction name"]] = getattr(module, ins["handler"])
                                    added_handlers.append(ins["instruction name"])

                        except Exception as e:
                            logger.error(f"[MainWindow] Failed to load RAIS extension {module_name}: {e}")

            except Exception as e:
                logger.error(f"[MainWindow] Failed to load RAIS extensions file: {e}")

        # Load run experience file for icon matching optimization
        run_experience_file = self.my_ecb_data_homepath + "/run_experience.txt"
        icon_match_dict = {}

        if os.path.exists(run_experience_file):
            try:
                with open(run_experience_file, 'rb') as fileTBRead:
                    icon_match_dict = json.load(fileTBRead)
                logger.info(f"[MainWindow] Loaded {len(icon_match_dict)} icon matching experiences")
            except json.JSONDecodeError:
                logger.error("[MainWindow] Error: Invalid JSON format in run experience file")
            except Exception as e:
                logger.error(f"[MainWindow] Error loading run experience file: {e}")

        logger.info(f"✅ Extensions initialized - {len(added_handlers)} RAIS handlers, {len(icon_match_dict)} icon experiences")

    def _init_task_management(self):
        """Initialize task and work management"""
        logger.info("[MainWindow] 📋 Initializing task management...")

        # Initialize work queues
        self.todays_work = {"tbd": [], "allstat": "working"}
        self.reactive_work = {"tbd": [], "allstat": "working"}
        self.todays_completed = []
        self.reactive_completed = []
        self.num_todays_task_groups = 0
        self.num_reactive_task_groups = 0

        # Setup scheduled work fetching for Commander role
        if "Commander" in self.host_role:
            fetchCloudScheduledWork = {
                "name": "fetch schedule",
                "works": self.gen_default_fetch(),
                "status": "yet to start",
                "current widx": 0,
                "completed": [],
                "aborted": []
            }

            # Add to work queue if in auto mode and not debug
            if not self.config_manager.general_settings.is_debug_enabled() and self.config_manager.general_settings.is_auto_mode():
                logger.info("[MainWindow] Adding fetch schedule to work queue")
                self.todays_work["tbd"].append(fetchCloudScheduledWork)
            else:
                logger.info(f"[MainWindow] Skipping auto schedule - Debug: {self.config_manager.general_settings.debug_mode}, Mode: {self.config_manager.general_settings.schedule_mode}")

        logger.info("[MainWindow] ✅ Task management initialized")

    def _init_servers_and_agents(self):
        """Initialize servers and agent systems"""
        logger.info("[MainWindow] 🤖 Initializing servers and agents...")


        from agent.mcp.server.tool_schemas import build_agent_mcp_tools_schemas
        from agent.ec_skills.build_node import get_default_node_schemas


        # Validate and fix default_llm_model before initializing LLM
        self._validate_and_fix_default_llm_model()
        
        # Initialize LLM with proper error handling
        try:
            from agent.ec_skills.llm_utils.llm_utils import pick_llm
            self.llm = pick_llm(
                self.config_manager.general_settings.default_llm,
                self.config_manager.llm_manager.get_all_providers(),
                self.config_manager
            )
            if self.llm:
                logger.info(f"[MainWindow] ✅ LLM initialized successfully - Type: {type(self.llm).__name__}")
                # Try to get provider info from config
                default_llm = self.config_manager.general_settings.default_llm
                if default_llm:
                    provider = self.config_manager.llm_manager.get_provider(default_llm)
                    if provider:
                        model_name = provider.get('default_model', 'unknown')
                        provider_display = provider.get('display_name', default_llm)
                        logger.info(f"[MainWindow] 📋 Current LLM: Provider={provider_display}, Name={default_llm}, Model={model_name}, Class={type(self.llm).__name__}")
            else:
                logger.warning(f"[MainWindow] ⚠️ LLM initialization failed - LLM is None")
        except Exception as e:
            logger.error(f"[MainWindow] Failed to initialize LLM: {e}")
            self.llm = None

        # Initialize agent-related components
        self.agents = []
        self.mcp_tools_schemas = build_agent_mcp_tools_schemas()
        self.mcp_client = None
        self._sse_cm = None
        
        # Initialize browser file system path if needed
        if not self.config_manager.general_settings.browser_use_file_system_path:
            self.config_manager.general_settings.browser_use_file_system_path = os.path.join(
                self.my_ecb_data_homepath, 'browser_use_fs'
            )
            self.config_manager.general_settings.save()

        # Load GUI flowgram schema if configured
        gui_flowgram_schema = self.config_manager.general_settings.data.get("gui_flowgram_schema", "")
        if gui_flowgram_schema:
            node_schema_file = self.my_ecb_data_homepath + gui_flowgram_schema
            if os.path.exists(node_schema_file):
                try:
                    with open(node_schema_file, 'rb') as fileTBRead:
                        self.node_schemas = json.load(fileTBRead)
                    logger.info("[MainWindow] GUI flowgram schema loaded")
                except (json.JSONDecodeError, Exception) as e:
                    logger.error(f"[MainWindow] Failed to load GUI flowgram schema: {e}")
                    self.node_schemas = get_default_node_schemas()
            else:
                self.node_schemas = get_default_node_schemas()
        else:
            self.node_schemas = get_default_node_schemas()

        logger.info("[MainWindow] ✅ Servers and agents initialized")

    def _init_async_tasks(self):
        """Initialize async tasks and background services"""
        logger.info("[MainWindow] ⚡ Initializing async tasks...")

        # Setup peer communication tasks based on role
        if self.host_role != "Platoon":
            if self.host_role != "Staff Officer":
                self.peer_task = asyncio.create_task(self.servePlatoons(self.gui_net_msg_queue))
                logger.info("[MainWindow] Started platoon serving task")
            else:
                self.peer_task = asyncio.create_task(self.wait_forever())
                logger.info("[MainWindow] Started staff officer waiting task")
        else:
            self.peer_task = asyncio.create_task(self.serveCommander(self.gui_net_msg_queue))
            logger.info("[MainWindow] Started commander serving task")

        # Initialize core monitoring and communication tasks
        self.monitor_task = asyncio.create_task(self.runRPAMonitor(self.gui_monitor_msg_queue))
        self.chat_task = asyncio.create_task(self.connectChat(self.gui_chat_msg_queue))

        # Initialize core async tasks (non-blocking)
        loop = asyncio.get_event_loop()
        asyncio.run_coroutine_threadsafe(self.run_async_tasks(), loop)

        logger.info("[MainWindow] ✅ Async tasks initialized")

    def is_fully_initialized(self) -> bool:
        """Check if initialization is fully completed"""
        return self._initialization_status.get('fully_ready', False)
    
    def get_main_window_safely(self):
        """
        Safely get MainWindow instance with initialization check.
        
        Returns:
            bool: True if ready, False otherwise
        """
        try:
            return self.is_fully_initialized()
        except Exception as e:
            logger.error(f"[MainGUI] Error accessing MainWindow: {e}")
            return False

    def _start_lightrag_deferred(self):
        """Start LightRAG server in deferred mode."""
        try:
            from knowledge.lightrag_server import LightragServer
            from utils.env.secure_store import secure_store
            
            # Prepare environment variables for LightRAG server
            lightrag_env = {"APP_DATA_PATH": ecb_data_homepath + "/lightrag_data"}
            
            # Ensure OPENAI_API_KEY is passed to LightRAG server from secure store with user isolation
            from utils.env.secure_store import get_current_username
            username = get_current_username()
            openai_api_key = secure_store.get('OPENAI_API_KEY', username=username)
            if openai_api_key and openai_api_key.strip():
                lightrag_env['OPENAI_API_KEY'] = openai_api_key
                logger.info("[MainWindow] 🔑 OPENAI_API_KEY found in secure store and will be passed to LightRAG server (deferred)")
            else:
                logger.warning("[MainWindow] ⚠️ OPENAI_API_KEY not found in secure store (deferred)")

            self.lightrag_server = LightragServer(extra_env=lightrag_env)
            # Start server process but don't wait for it to be ready
            self.lightrag_server.start(wait_ready=False)
            logger.info("[MainWindow] LightRAG server started (deferred, non-blocking)")
        except Exception as e:
            logger.warning(f"[MainWindow] Deferred LightRAG start failed: {e}")


    def stop_lightrag_server(self):
        self.lightrag_server.stop()
        self.lightrag_server = None

    async def _async_sync_cloud_data(self):
        """
        Asynchronously sync cloud data in background without blocking UI
        """
        try:
            logger.info("[MainWindow] 🌐 Starting background cloud data synchronization...")

            # Sync bot data
            logger.info("📥 Syncing bot data from cloud...")
            await asyncio.get_event_loop().run_in_executor(
                None,
                self.bot_service.sync_cloud_bot_data,
                self.session,
                self.get_auth_token(),
                self
            )

            # Reload local bots after cloud sync
            logger.info("[MainWindow] 🔄 Reloading bot data after cloud sync...")
            bots_data = self.bot_service.find_all_bots()
            self.loadLocalBots(bots_data)
            logger.info(f"[MainWindow] ✅ Bot cloud sync completed - {len(bots_data)} bots loaded")

            # Sync mission data
            logger.info("[MainWindow] 📥 Syncing mission data from cloud...")
            await asyncio.get_event_loop().run_in_executor(
                None,
                self.mission_service.sync_cloud_mission_data,
                self.session,
                self.get_auth_token(),
                self
            )

            # Reload local missions after cloud sync
            logger.info("[MainWindow] 🔄 Reloading mission data after cloud sync...")
            missions_data = self.mission_service.find_missions_by_createon()
            self.loadLocalMissions(missions_data)
            logger.info(f"[MainWindow] ✅ Mission cloud sync completed - {len(missions_data)} missions loaded")

            # Update UI to reflect new data
            logger.info("[MainWindow] Cloud data sync completed successfully")
            logger.info("[MainWindow] 🎉 Background cloud data synchronization completed successfully!")

        except Exception as e:
            logger.error(f"[MainWindow] ❌ Background cloud sync failed: {e}")
            logger.error(f"[MainWindow] Cloud sync error details: {traceback.format_exc()}")
            logger.error(f"[MainWindow] Cloud sync failed: {str(e)}")

    def _startup_sync_offline_cloud_cache(self):
        """
        Startup sync: sync pending cache to cloud and start auto retry timer
        
        This is a BLOCKING synchronous function that:
        1. Checks if there are pending cached tasks
        2. If yes, syncs all cached tasks to cloud (blocking)
        3. Starts auto retry timer for periodic cache sync (every 5 minutes)
        
        Note: Loading data from cloud is handled by each module separately.
        """
        try:
            logger.info("[MainWindow] 🚀 Starting startup sync (blocking)...")
            
            # Get username
            username = self.user if hasattr(self, 'user') else None
            if not username:
                logger.warning("[MainWindow] No username available, skipping startup sync")
                return
            
            # Import sync manager and queue
            from agent.cloud_api.offline_sync_manager import get_sync_manager
            from agent.cloud_api.offline_sync_queue import get_sync_queue
            
            manager = get_sync_manager()
            queue = get_sync_queue()
            
            # Step 1: Check if there are pending tasks
            stats = queue.get_stats()
            pending_count = stats['pending_count']
            failed_count = stats['failed_count']
            total_count = pending_count + failed_count
            
            if total_count > 0:
                logger.info(f"[MainWindow] 📤 Found {total_count} tasks to sync ({pending_count} pending, {failed_count} failed)")
                logger.info(f"[MainWindow] Pending by type: {stats['pending_by_type']}")
                if failed_count > 0:
                    logger.info(f"[MainWindow] Failed by type: {stats['failed_by_type']}")
                
                # Use thread for sync, fully async, doesn't block any operations
                import threading
                import time
                
                def _sync_task():
                    """Background sync task"""
                    try:
                        start_time = time.time()
                        logger.info("[MainWindow] 🔄 Starting startup sync in background (timeout: 15s per task)...")
                        
                        # Limit to max 20 tasks, 15s timeout per task, include failed tasks
                        result = manager.sync_pending_queue(max_tasks=20, timeout_per_task=15.0, include_failed=True)
                        
                        elapsed = time.time() - start_time
                        logger.info(f"[MainWindow] ✅ Startup sync completed in {elapsed:.1f}s:")
                        logger.info(f"  - Total: {result['total']}")
                        logger.info(f"  - Synced: {result['synced']}")
                        logger.info(f"  - Failed: {result['failed']}")
                        
                        if result['failed'] > 0:
                            logger.warning(f"[MainWindow] ⚠️ {result['failed']} tasks failed (timeout or error), will retry later")
                        
                        if total_count > 20:
                            logger.info(f"[MainWindow] 💡 {total_count - 20} tasks will be synced by auto-retry timer")
                            
                    except Exception as e:
                        logger.error(f"[MainWindow] ❌ Startup sync error: {e}")
                        logger.error(f"[MainWindow] Traceback: {traceback.format_exc()}")
                
                # Start background thread (fully async, no waiting)
                sync_thread = threading.Thread(target=_sync_task, daemon=True, name="StartupSync")
                sync_thread.start()
                logger.info("[MainWindow] ✅ Startup sync started in background (non-blocking)")
            else:
                logger.info("[MainWindow] ✅ No pending tasks to sync")
            
            # Step 2: Start auto retry timer (background task for periodic sync)
            # This timer will automatically sync cached data every 5 minutes
            logger.info("[MainWindow] 🔄 Starting auto retry timer for periodic cache sync...")
            manager.start_auto_retry(interval=300)  # 300 seconds
            logger.info("[MainWindow] ✅ Auto retry timer started (interval: 300s)")
                
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Startup sync error: {e}")
            logger.error(f"[MainWindow] Startup sync error details: {traceback.format_exc()}")

    async def _async_setup_browser_manager(self):
        """
        Asynchronously setup browser manager and WebDriver in background
        """
        try:
            # Wait a bit to ensure main window is responsive first
            await asyncio.sleep(1.0)

            logger.info("[MainWindow] 🌐 Starting background browser manager initialization...")

            # Run browser manager setup in executor to avoid blocking
            await asyncio.get_event_loop().run_in_executor(
                None,
                self.setupUnifiedBrowserManager
            )

            # Start WebDriver initialization after browser manager is ready
            await self._start_webdriver_initialization()

            logger.info("[MainWindow] ✅ Browser manager and WebDriver initialization completed!")

        except Exception as e:
            logger.error(f"[MainWindow] ❌ Background browser setup failed: {e}")
            logger.error(f"[MainWindow] Browser setup error details: {traceback.format_exc()}")
            # Don't crash the app if browser setup fails
            # The app should continue to work without browser automation

    async def _async_start_lightrag(self):
        """
        Asynchronously start LightRAG server in background
        """
        try:
            # Wait a bit to ensure other services are ready
            await asyncio.sleep(0.5)

            logger.info("[MainWindow] 🧠 Starting LightRAG server initialization...")

            # Initialize LightRAG server in main thread to allow signal handlers
            # but run the actual server start in executor for non-blocking behavior
            from knowledge.lightrag_server import LightragServer
            from utils.env.secure_store import secure_store
            
            # Prepare environment variables for LightRAG server
            lightrag_env = {"APP_DATA_PATH": ecb_data_homepath + "/lightrag_data"}
            
            # Ensure OPENAI_API_KEY is passed to LightRAG server from secure store with user isolation
            from utils.env.secure_store import get_current_username
            username = get_current_username()
            openai_api_key = secure_store.get('OPENAI_API_KEY', username=username)
            if openai_api_key and openai_api_key.strip():
                lightrag_env['OPENAI_API_KEY'] = openai_api_key
                logger.info("[MainWindow] 🔑 OPENAI_API_KEY found in secure store and will be passed to LightRAG server")
            else:
                logger.warning("[MainWindow] ⚠️ OPENAI_API_KEY not found in secure store")
            
            self.lightrag_server = LightragServer(extra_env=lightrag_env)

            # Start the server process in executor (this is the blocking part)
            await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: self.lightrag_server.start(wait_ready=False)
            )

            logger.info("[MainWindow] ✅ LightRAG server initialization completed!")

        except Exception as e:
            logger.error(f"❌ LightRAG server initialization failed: {e}")
            logger.error(f"LightRAG server error details: {traceback.format_exc()}")



    async def _async_start_wan_chat(self):
        """
        Asynchronously start wan based chat service in background
        """
        try:
            # Wait a bit to ensure other services are ready
            await asyncio.sleep(0.5)

            logger.info("[MainWindow] 🧠 Starting websocket wan chat...")

            from bot.wanChat import subscribeToWanChat
            # Start the websocket subscribe coroutine as a background task
            token = self.get_auth_token()
            # Wait for token to become available if auth flow is still initializing
            wait_loops = 0
            while not token or not isinstance(token, str) or not token.strip():
                wait_loops += 1
                if wait_loops % 10 == 1:
                    logger.info("[MainWindow] Waiting for auth token before starting WAN chat...")
                await asyncio.sleep(0.5)
                token = self.get_auth_token()

            # Kick off WAN chat in background so this method can complete
            if getattr(self, 'wan_chat_task', None) and not self.wan_chat_task.done():
                # a previous task exists; cancel and replace
                try:
                    self.wan_chat_task.cancel()
                except Exception:
                    pass
            self.wan_chat_task = asyncio.create_task(subscribeToWanChat(self, token, self.chat_id))

            # Wait up to 15 seconds for subscription to be acknowledged
            for _ in range(30):
                if getattr(self, 'get_wan_msg_subscribed', None) and self.get_wan_msg_subscribed():
                    logger.info("[MainWindow] ✅ websocket wan chat initialization completed!")
                    break
                await asyncio.sleep(0.5)
            else:
                logger.warning("[MainWindow] ⚠️ WAN chat started but subscription not confirmed within timeout")

        except Exception as e:
            logger.error(f"[MainWindow] ❌ websocket wan chat initialization failed: {e}")
            logger.error(f"[MainWindow] websocket wan chat error details: {traceback.format_exc()}")



    async def _async_start_llm_subscription(self):
        """
        Asynchronously start eCan's own cloud side LLM service subscription
        """
        from agent.cloud_api.cloud_api import subscribe_cloud_llm_task
        try:
            # Check if shutting down before starting
            if hasattr(self, '_shutting_down') and self._shutting_down:
                logger.info("System is shutting down, skipping LLM subscription")
                return
                
            # Wait a bit to ensure other services are ready
            await asyncio.sleep(0.5)

            logger.info("🧠 Starting Cloud LLM Subscription...")

            # Initialize LightRAG server in main thread to allow signal handlers
            # but run the actual server start in executor for non-blocking behavior
            ws_host = self.getWSApiHost()
            ws_endpoint = self.getWSApiEndpoint()
            token = self.get_auth_token()
            masked_token = token[:15] + "..." + token[-4:] if len(token) > 20 else "***"
            logger.info("ws_host", ws_host, "token:", masked_token if token else "", "ws_endpoint:", ws_endpoint)
            acctSiteID = self.getAcctSiteID()
            print("acct site id:", acctSiteID)
            
            # Start the server process in executor and save references for cleanup
            ws, thread = await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: subscribe_cloud_llm_task(acctSiteID, token, ws_endpoint)
            )
            
            # Save references for cleanup
            self.cloud_llm_ws = ws
            self.cloud_llm_thread = thread

            logger.info("✅ Cloud LLM Subscription initialization completed!")

        except Exception as e:
            logger.error(f"❌ Cloud LLM Subscription initialization failed: {e}")
            logger.error(f"Cloud LLM Subscription error details: {traceback.format_exc()}")
            # Don't crash the app if Cloud LLM Subscription fails
            # The app should continue to work without Cloud LLM Subscription

    async def _async_start_cloud_show_subscription(self):
        """
        Asynchronously start eCan's own cloud side LLM service subscription
        """
        from agent.story.story_gen import subscribe_cloud_show
        try:
            # Check if shutting down before starting
            if hasattr(self, '_shutting_down') and self._shutting_down:
                logger.info("System is shutting down, skipping Cloud Show subscription")
                return

            # Wait a bit to ensure other services are ready
            await asyncio.sleep(0.5)

            logger.info("🧠 Starting Cloud LLM Subscription...")

            # Initialize LightRAG server in main thread to allow signal handlers
            # but run the actual server start in executor for non-blocking behavior
            ws_host = self.getWSApiHost()
            ws_endpoint = self.getWSApiEndpoint()
            token = self.get_auth_token()
            masked_token = token[:15] + "..." + token[-4:] if len(token) > 20 else "***"
            logger.info("ws_host", ws_host, "token:", masked_token if token else "", "ws_endpoint:", ws_endpoint)
            acctSiteID = self.getAcctSiteID()
            print("acct site id:", acctSiteID)

            # Start the server process in executor and save references for cleanup
            ws, thread = await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: subscribe_cloud_show(acctSiteID, token, ws_endpoint)
            )

            # Save references for cleanup
            self.cloud_show_ws = ws
            self.cloud_show_thread = thread

            logger.info("✅ Cloud Show Subscription initialization completed!")

        except Exception as e:
            err_msg = get_traceback(e, "ErrorCloudShowSubscription")
            logger.error(f"❌ {err_msg}")
            # Don't crash the app if Cloud LLM Subscription fails
            # The app should continue to work without Cloud LLM Subscription


    def get_auth_token(self):
        """Return a valid JWT for AppSync Authorization header.
        Prefer Cognito IdToken; fall back to AccessToken. Support multiple token shapes.
        """
        try:
            tokens = self.auth_manager.get_tokens()
            if not tokens or not isinstance(tokens, dict):
                return None
            # Common flat shapes from OAuth token exchange
            for k in ('IdToken', 'id_token'):
                if k in tokens and isinstance(tokens[k], str) and tokens[k]:
                    return tokens[k]
            for k in ('AccessToken', 'access_token'):
                if k in tokens and isinstance(tokens[k], str) and tokens[k]:
                    return tokens[k]
            # Nested shape from Cognito AWSSRP / refresh
            ar = tokens.get('AuthenticationResult') if isinstance(tokens, dict) else None
            if isinstance(ar, dict):
                for k in ('IdToken', 'AccessToken'):
                    if k in ar and isinstance(ar[k], str) and ar[k]:
                        return ar[k]
            return None
        except Exception as e:
            logger.error(f"Error getting auth token: {e}")
            return None



    async def async_agents_init(self):
        """
        Highly optimized asynchronous Agent initialization with UI non-blocking design
        """
        start_time = time.time()
        
        try:
            logger.info("[MainWindow] 🚀 Starting ultra-optimized async agents initialization...")
            local_server_port = self.get_local_server_port()
            
            # Phase 1: Instant basic setup (target: <50ms)
            logger.info("[MainWindow] ⚡ Phase 1: Instant setup...")
            phase1_start = time.time()
            
            # Initialize basic data structures instantly
            self.agent_skills = []
            self.agent_tasks = []
            self.agent_tools = []
            self.agent_knowledges = []
            
            # Non-blocking server readiness check
            await self._wait_for_server_ready(local_server_port)
            
            elapsed_phase1 = time.time() - phase1_start
            logger.info(f"[MainWindow] ✅ Phase 1 completed in {elapsed_phase1:.3f}s")
            
            # Phase 2: Aggressive parallel initialization (target: <2s)
            logger.info("[MainWindow] 🚀 Phase 2: Aggressive parallel initialization...")
            phase2_start = time.time()
            
            # Create all parallel tasks with timeout protection
            # Adjusted timeouts for slower machines
            tasks = []
            
            # 1. MCP tools with adaptive timeout based on system performance
            # Detect system performance and adjust timeout accordingly
            import psutil
            cpu_count = psutil.cpu_count()
            memory_gb = psutil.virtual_memory().total / (1024**3)
            
            # Adaptive timeout calculation
            base_timeout = 3.0
            if cpu_count >= 8 and memory_gb >= 16:
                # High-performance system
                mcp_timeout = base_timeout
            elif cpu_count >= 4 and memory_gb >= 8:
                # Medium-performance system  
                mcp_timeout = base_timeout * 1.5
            else:
                # Low-performance system
                mcp_timeout = base_timeout * 2.5
            
            logger.info(f"[MainWindow] 🎯 Adaptive MCP timeout: {mcp_timeout:.1f}s (CPU: {cpu_count}, RAM: {memory_gb:.1f}GB)")
            
            tasks.append(asyncio.wait_for(
                self._get_mcp_tools_async(), 
                timeout=mcp_timeout
            ))
            
            # 2. Placeholder for skills dependencies (simplified)
            async def prepare_skills_deps():
                await asyncio.sleep(0.01)
                return True
            tasks.append(asyncio.wait_for(prepare_skills_deps(), timeout=1.0))
            
            # 3. Placeholder for agent components (simplified)
            async def prepare_agent_components():
                await asyncio.sleep(0.01)
                return True
            tasks.append(asyncio.wait_for(prepare_agent_components(), timeout=1.0))
            
            # 4. Placeholder for background data (simplified)
            async def load_background_data():
                await asyncio.sleep(0.01)
                return True
            tasks.append(asyncio.wait_for(load_background_data(), timeout=1.0))
            
            # Execute all tasks in parallel with exception handling
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Process results with graceful degradation
            mcp_tools, skills_deps, agent_components, background_data = results
            
            # Handle MCP tools result
            if isinstance(mcp_tools, Exception):
                logger.warning(f"[MainWindow] ⚠️ MCP tools failed, using empty list: {mcp_tools}")
                self.mcp_tools = []
            else:
                self.mcp_tools = mcp_tools
                logger.info(f"[MainWindow] ✅ MCP tools ready: {len(self.mcp_tools)} tools")
            
            # Handle skills dependencies result
            if isinstance(skills_deps, Exception):
                logger.warning(f"[MainWindow] ⚠️ Skills dependencies failed: {skills_deps}")
            else:
                logger.info(f"[MainWindow] ✅ Skills dependencies prepared")
            
            # Handle agent components result
            if isinstance(agent_components, Exception):
                logger.warning(f"[MainWindow] ⚠️ Agent components failed: {agent_components}")
                # Continue with minimal components
            
            elapsed_phase2 = time.time() - phase2_start
            logger.info(f"[MainWindow] ✅ Phase 2 completed in {elapsed_phase2:.3f}s")
            
            # Phase 3: Streamlined agent assembly (target: <1s)
            logger.info("[MainWindow] 🎯 Phase 3: Streamlined agent assembly...")
            phase3_start = time.time()
            
            # Pipelined Agent construction and startup
            logger.info("[MainWindow] 🔧 Starting pipelined agent initialization...")
            
            # Check if required components are ready
            missing_components = []
            if not hasattr(self, 'llm') or self.llm is None:
                missing_components.append('LLM')
            if not hasattr(self, 'mcp_client'):
                missing_components.append('MCP Client')
            
            if missing_components:
                logger.warning(f"[MainWindow] ⚠️ Missing components: {missing_components}, skipping agent skills building")
                self.agent_skills = []
            else:
                logger.info(f"[MainWindow] ✅ All components ready - LLM: {type(self.llm)}, MCP Client: {self.mcp_client is not None}")
                
                # Start skill building task (asynchronous)
                agent_skills_task = asyncio.create_task(self._build_agent_skills_async())
                
                # Wait for skill building to complete
                try:
                    self.agent_skills = await agent_skills_task
                    logger.info(f"[MainWindow] ✅ Agent skills built: {len(self.agent_skills)} skills")
                except Exception as e:
                    logger.warning(f"[MainWindow] ⚠️ Agent skills building failed: {e}")
                    import traceback
                    logger.error(f"[MainWindow] Skills building traceback: {traceback.format_exc()}")
                    self.agent_skills = []

                # Start agent task building (asynchronous, parallel with skills)
                logger.info("[MainWindow] 📝 Building agent tasks...")
                agent_tasks_task = asyncio.create_task(self._build_agent_tasks_async())

                # Wait for agent task building to complete
                try:
                    self.agent_tasks = await agent_tasks_task
                    logger.info(f"[MainWindow] ✅ Agent tasks built: {len(self.agent_tasks)} agent tasks")
                except Exception as e:
                    logger.warning(f"[MainWindow] ⚠️ Agent tasks building failed: {e}")
                    import traceback
                    logger.error(f"[MainWindow] Agent tasks building traceback: {traceback.format_exc()}")
                    self.agent_tasks = []

            # Environment preparation (simplified handling)
            logger.info("[MainWindow] 🔧 Environment preparation completed")
            
            # Ultra-parallel Agent construction and startup
            try:
                logger.info("[MainWindow] 🚀 Building and launching agents with ultra-parallel optimization...")
                agents_built = await self._build_and_launch_agents_ultra_parallel()
                
                if agents_built:
                    logger.info(f"[MainWindow] ✅ Successfully built and launched {len(self.agents)} agents")

                    # TODO: Merge agent.tasks from built agents into mainwin.agent_tasks, This step will be deprecated in the future
                    self._merge_agent_tasks_to_memory()
                else:
                    logger.warning("[MainWindow] ⚠️ Agent ultra-parallel process completed with issues")

            except Exception as e:
                logger.error(f"[MainWindow] ❌ Agent ultra-parallel process failed: {e}")
                agents_built = False
            
            # Mark async initialization complete
            self._initialization_status['async_init_complete'] = True
            
            elapsed_phase3 = time.time() - phase3_start
            total_elapsed = time.time() - start_time
            
            logger.info(f"[MainWindow] ✅ Phase 3 completed in {elapsed_phase3:.3f}s")
            logger.info(f"[MainWindow] 🎉 Ultra-optimized initialization completed in {total_elapsed:.3f}s!")
            
            # Simple performance reporting
            logger.info("=" * 50)
            logger.info("📊 PERFORMANCE SUMMARY")
            logger.info("=" * 50)
            logger.info(f"⏱️  Total Time: {total_elapsed:.3f}s")
            logger.info(f"📋 Phase Breakdown:")
            logger.info(f"   Phase 1 (Setup): {elapsed_phase1:.3f}s")
            logger.info(f"   Phase 2 (Parallel): {elapsed_phase2:.3f}s") 
            logger.info(f"   Phase 3 (Assembly): {elapsed_phase3:.3f}s")
            logger.info(f"💻 System: CPU={cpu_count}, RAM={memory_gb:.1f}GB")
            logger.info("=" * 50)

        except Exception as e:
            total_elapsed = time.time() - start_time
            logger.error(f"[MainWindow] ❌ Optimized agents initialization failed after {total_elapsed:.3f}s: {e}")
            
            # Provide helpful error information
            error_msg = f"Agent initialization failed: {str(e)}\n\n" \
                       f"Optimization attempted but encountered issues.\n" \
                       f"Time taken: {total_elapsed:.3f}s\n\n" \
                       f"Possible causes:\n" \
                       f"• Server connection timeout\n" \
                       f"• Resource conflicts\n" \
                       f"• Configuration errors\n\n" \
                       f"Check logs for detailed information."
            
            # Show error message (simplified)
            self.showMsg(error_msg)
            raise

    async def _wait_for_server_ready(self, local_server_port: int):
        """Ultra-fast server readiness check with intelligent backoff"""
        
        logger.info(f"[MainWindow] 🔍 Fast server check on port {local_server_port}...")
        host = "127.0.0.1"
        
        # Quick port check first (usually succeeds immediately)
        for attempt in range(5):
            try:
                # Fast socket connection test
                sock = socket.create_connection((host, local_server_port), timeout=0.5)
                sock.close()
                logger.info(f"[MainWindow] ✅ Server ready on {host}:{local_server_port} (attempt {attempt + 1})")
                return True
            except (socket.error, OSError):
                if attempt < 4:
                    await asyncio.sleep(0.1)  # Very short wait
                continue
        
        # Fallback to HTTP check if socket check fails
        try:
            import httpx
            timeout = httpx.Timeout(connect=3.0, read=5.0, write=3.0, pool=3.0)  # Increased timeouts for slower machines
            async with httpx.AsyncClient(timeout=timeout) as client:
                response = await client.get(f"http://{host}:{local_server_port}/healthz")
                if response.status_code == 200:
                    logger.info(f"[MainWindow] ✅ Server HTTP ready on {host}:{local_server_port}")
                    return True
        except Exception as e:
            logger.warning(f"[MainWindow] ⚠️ Server check failed: {e}")
        
        return False



    async def _get_mcp_tools_async(self):
        """Enhanced cached MCP tools retrieval with memory caching"""
        
        # Memory cache strategy
        memory_cache_timeout = 300  # 5 minutes
        if hasattr(self, '_mcp_tools_cache'):
            cache_data, cache_time = self._mcp_tools_cache
            cache_age = time.time() - cache_time
            if cache_age < memory_cache_timeout:
                logger.info(f"[MainWindow] ⚡ Using memory cached MCP tools ({len(cache_data)} tools, age: {cache_age:.1f}s)")
                return cache_data
            else:
                logger.debug(f"[MainWindow] Memory cache expired (age: {cache_age:.1f}s > {memory_cache_timeout}s)")
        
        logger.info("[MainWindow] 📋 Getting MCP tools with timeout protection...")
        
        try:
            # Try to get MCP tools directly with timeout protection
            result = await asyncio.wait_for(self._get_mcp_tools_direct(), timeout=2.0)
            
            if result and len(result) > 0:
                # Update memory cache
                current_time = time.time()
                self._mcp_tools_cache = (result, current_time)
                
                logger.info(f"[MainWindow] ✅ Got {len(result)} MCP tools")
                return result
            else:
                logger.warning("[MainWindow] ⚠️ No MCP tools returned, using empty list")
                return []
            
        except asyncio.TimeoutError:
            logger.warning("[MainWindow] ⚠️ MCP tools request timed out (2s), using empty list")
            return []
        except Exception as e:
            logger.warning(f"[MainWindow] ⚠️ MCP tools failed: {e}, using empty list")
            return []

    async def _build_agent_skills_async(self):
        """Optimized parallel agent skills building"""
        logger.info("[MainWindow] 🔧 Building agent skills in parallel...")

        try:
            # Use the async build_agent_skills function directly
            from agent.ec_skills.build_agent_skills import build_agent_skills
            skills = await build_agent_skills(self)

            logger.info(f"[MainWindow] ✅ Built {len(skills)} agent skills")
            return skills or []

        except Exception as e:
            logger.warning(f"[MainWindow] ⚠️ Agent skills building failed: {e}")
            import traceback
            logger.debug(f"[MainWindow] Skills building traceback: {traceback.format_exc()}")
            return []

    async def _build_agent_tasks_async(self):
        """Optimized parallel agent tasks building"""
        logger.info("[MainWindow] 📝 Building agent tasks in parallel...")

        try:
            # Use the async build_agent_tasks function directly
            from agent.ec_agents.create_agent_tasks import build_agent_tasks
            agent_tasks = await build_agent_tasks(self)

            logger.info(f"[MainWindow] ✅ Built {len(agent_tasks)} agent tasks")
            return agent_tasks or []

        except Exception as e:
            logger.warning(f"[MainWindow] ⚠️ Agent tasks building failed: {e}")
            import traceback
            logger.debug(f"[MainWindow] Agent tasks building traceback: {traceback.format_exc()}")
            return []

    async def _obtain_agent_tools_async(self):
        """Asynchronously obtain Agent Tools"""
        try:
            from agent.ec_agents.obtain_agent_tools import obtain_agent_tools
            logger.debug("[MainWindow] 🛠️ Obtaining agent tools...")
            tools = obtain_agent_tools(self)
            return tools
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to obtain agent tools: {e}")
            raise

    async def _build_agent_knowledges_async(self):
        """Asynchronously build Agent Knowledges"""
        try:
            from agent.ec_agents.build_agent_knowledges import build_agent_knowledges
            logger.debug("[MainWindow] 📚 Building agent knowledges...")
            knowledges = build_agent_knowledges(self)
            return knowledges
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to build agent knowledges: {e}")
            raise

    async def wait_for_server_async(self, agent, timeout: float = 5.0):
        """Asynchronously wait for Agent server startup"""
        url = agent.get_card().url+'/ping'
        start = time.time()
        
        try:
            import aiohttp
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=2)) as session:
                while time.time() - start < timeout:
                    try:
                        async with session.get(url) as response:
                            if response.status == 200:
                                logger.info(f"✅ {agent.get_card().name} Server is up at {url}")
                                return True
                    except (aiohttp.ClientError, asyncio.TimeoutError):
                        pass
                    await asyncio.sleep(0.5)
        except Exception as e:
            logger.error(f"[MainWindow] Error checking server {url}: {e}")
        
        logger.warning(f"[MainWindow] ⚠️ Server did not start within {timeout} seconds, continuing anyway")
        return False

    def get_free_agent_ports(self, n):
        """
        Thread-safe port allocation for agents.
        Uses the global port allocator to prevent race conditions during parallel agent launch.
        """
        try:
            # Get currently used ports from existing agents
            used_ports = [ag.get_a2a_server_port() for ag in self.agents if ag is not None and hasattr(ag, 'get_a2a_server_port')]
            
            # Get port range from configuration
            local_agent_ports = self.config_manager.general_settings.local_agent_ports
                        
            # Use thread-safe allocator
            free_ports = self._port_allocator.get_free_ports(n, local_agent_ports, used_ports)
            
            logger.info(f"[MainWindow] Allocated ports: {free_ports}")
            return free_ports
            
        except Exception as e:
            logger.error(f"[MainWindow] Port allocation failed: {e}")
            # Fallback to original method if allocator fails
            used_ports = [ag.get_a2a_server_port() for ag in self.agents if ag is not None]
            local_agent_ports = self.config_manager.general_settings.local_agent_ports
            all_ports = range(local_agent_ports[0], local_agent_ports[1]+1)
            free_ports = [port for port in all_ports if port not in used_ports]
            
            if len(free_ports) < n:
                raise RuntimeError(f"[MainWindow] Only {len(free_ports)} free ports available, but {n} requested.")
            
            return free_ports[:n]


    def release_agent_port(self, port):
        """
        Release a port back to the allocator when an agent shuts down.
        """
        try:
            self._port_allocator.release_port(port)
            logger.info(f"[MainWindow] Released agent port: {port}")
        except Exception as e:
            logger.warning(f"[MainWindow] Failed to release port {port}: {e}")

    def release_agent_ports(self, ports):
        """
        Release multiple ports back to the allocator.
        """
        try:
            self._port_allocator.release_ports(ports)
            logger.info(f"[MainWindow] Released agent ports: {ports}")
        except Exception as e:
            logger.warning(f"[MainWindow] Failed to release ports {ports}: {e}")
    def save_agent_skill(self, skill):
        from agent.ec_skills.save_agent_skills import save_agent_skills
        return save_agent_skills(self, [skill])

    async def _get_mcp_tools_direct(self):
        """Direct MCP tools retrieval using original method"""
        try:
            from agent.mcp.local_client import mcp_client_manager
            from agent.mcp.config import mcp_http_base
            
            url = mcp_http_base()
            tl_result = await mcp_client_manager.list_tools(url)

            # Handle ListToolsResult object
            if hasattr(tl_result, 'tools'):
                tl = tl_result.tools
                logger.debug(f"✅ Successfully listed {len(tl)} MCP tools")
            elif isinstance(tl_result, list):
                tl = tl_result
                logger.debug(f"✅ Successfully listed {len(tl)} MCP tools")
            else:
                logger.warning(f"Unexpected tools result type: {type(tl_result)}")
                tl = []

            return tl or []
        except Exception as e:
            logger.debug(f"[MainWindow] Direct MCP tools failed: {e}")
            return []

    async def _build_and_launch_agents_ultra_parallel(self):
        """
        Ultra-parallel agent build and launch
        
        Integrates agents from three sources in parallel:
        1. Code-built agents (from builder functions)
        2. Local database agents (user-created via UI)
        3. Cloud agents (synced from cloud API)
        
        Returns:
            bool: True if agents were successfully built and initialized
        """
        try:
            logger.info("[MainWindow] 🚀 Starting ultra-parallel agent build and launch...")
            start_time = time.time()
            
            # Initialize agents list
            self.agents = []
            
            # Step 1: Parallel execution of THREE tasks
            logger.info("[MainWindow] 🔧 Starting 3-way parallel execution:")
            logger.info("[MainWindow]    1️⃣ Building code-built agents")
            logger.info("[MainWindow]    2️⃣ Loading local database agents")
            logger.info("[MainWindow]    3️⃣ Fetching cloud agents")
            
            built_agents, local_db_agents, cloud_agents = await asyncio.gather(
                self._build_code_agents_async(),
                self._load_local_db_agents(),
                self._fetch_cloud_agents(),
                return_exceptions=True
            )
            
            # Handle exceptions from parallel tasks
            if isinstance(built_agents, Exception):
                logger.error(f"[MainWindow] ❌ Failed to build code agents: {built_agents}")
                built_agents = []
            if isinstance(local_db_agents, Exception):
                logger.error(f"[MainWindow] ❌ Failed to load local DB agents: {local_db_agents}")
                local_db_agents = []
            if isinstance(cloud_agents, Exception):
                logger.error(f"[MainWindow] ❌ Failed to fetch cloud agents: {cloud_agents}")
                cloud_agents = []
            
            logger.info(f"[MainWindow] 📊 Parallel loading completed:")
            logger.info(f"[MainWindow]    - Code-built: {len(built_agents)}, DB: {len(local_db_agents)}, Cloud: {len(cloud_agents)}")
            
            # Step 2: Merge and deduplicate agents from all sources
            logger.info("[MainWindow] 🔀 Merging agents from all sources...")
            
            # Merge local DB and cloud agents (cloud overwrites local)
            merged_db_cloud = self._merge_agent_data(local_db_agents, cloud_agents)
            logger.info(f"[MainWindow] 📊 DB + Cloud merged: {len(merged_db_cloud)} agents")
            
            # Update local database with merged data
            await self._update_local_db_agents(merged_db_cloud)
            
            # Merge code-built agents with DB/Cloud agents (Priority: Code-built > Cloud > Local DB)
            self.agents = self._merge_all_agent_sources(
                code_built_agents=built_agents,
                db_cloud_agents=merged_db_cloud
            )
            
            logger.info(f"[MainWindow] ✅ Merged {len(self.agents)} unique agents")
            
            if len(self.agents) == 0:
                logger.error("[MainWindow] ❌ No agents available after merge")
                return False
            
            # Step 3: Verify all agents are EC_Agent objects (already converted in _merge_all_agent_sources)
            dict_count = sum(1 for a in self.agents if isinstance(a, dict))
            ec_agent_count = len(self.agents) - dict_count
            
            if dict_count > 0:
                logger.warning(f"[MainWindow] ⚠️ Found {dict_count} unconverted dict agents (expected 0)")
            
            logger.info(f"[MainWindow] ✅ Final agent list: {ec_agent_count} EC_Agent objects")
            
            # Log detailed agent and skill information before launch
            logger.info("[AGENT_INVENTORY] ========== Agent Skills Inventory ==========")
            for idx, agent in enumerate(self.agents):
                try:
                    agent_name = agent.get_card().name if hasattr(agent, 'get_card') and agent.get_card() else f"Agent_{idx}"
                    logger.info(f"[AGENT_INVENTORY] Agent {idx}: {agent_name}")
                    
                    if hasattr(agent, 'skills'):
                        skills_count = len(agent.skills) if agent.skills else 0
                        logger.info(f"[AGENT_INVENTORY]   - Skills count: {skills_count}")
                        
                        if skills_count > 0:
                            for skill_idx, skill in enumerate(agent.skills):
                                if skill is None:
                                    logger.error(f"[AGENT_INVENTORY]   - Skill[{skill_idx}]: None (MISSING!)")
                                else:
                                    skill_name = skill.name if hasattr(skill, 'name') else f"Skill_{skill_idx}"
                                    has_runnable = hasattr(skill, 'runnable') and skill.runnable is not None
                                    logger.info(f"[AGENT_INVENTORY]   - Skill[{skill_idx}]: {skill_name}, has_runnable: {has_runnable}")
                                    if not has_runnable:
                                        logger.error(f"[AGENT_INVENTORY]     ⚠️ Skill '{skill_name}' has no runnable!")
                        else:
                            logger.warning(f"[AGENT_INVENTORY]   - No skills assigned")
                    else:
                        logger.warning(f"[AGENT_INVENTORY]   - No 'skills' attribute")
                except Exception as e:
                    logger.error(f"[AGENT_INVENTORY] Error inspecting agent {idx}: {e}")
            logger.info("[AGENT_INVENTORY] =============================================")
            
            # Step 4: Launch agents in background (non-blocking)
            self._launch_agents_async(self.agents)
            
            total_time = time.time() - start_time
            logger.info(f"[MainWindow] 🎉 Ultra-parallel process completed in {total_time:.3f}s")
            logger.info(f"[MainWindow]    - {len(self.agents)} agents ready, launches running in background")
            
            return len(self.agents) > 0
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Ultra-parallel process failed: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    async def _build_code_agents_async(self):
        """Build all code-built agents in parallel
        
        Returns:
            List of successfully built (agent_name, agent) tuples
        """
        try:
            # Import agent builder functions
            from agent.ec_agents.my_twin_agent import set_up_my_twin_agent
            from agent.ec_agents.ec_helper_agent import set_up_ec_helper_agent
            from agent.ec_agents.ec_rpa_operator_agent import set_up_ec_rpa_operator_agent
            from agent.ec_agents.ec_tester_agent import set_up_ec_tester_agent
            from agent.ec_agents.ec_procurement_agent import set_up_ec_procurement_agent
            
            # Define agent configurations based on role
            agent_configs = []
            
            # Basic Agent (always needed)
            agent_configs.append({
                'name': 'My Twin',
                'builder': set_up_my_twin_agent
            })
            
            # Add other agents based on role
            if "Platoon" in self.machine_role:
                agent_configs.extend([
                    {'name': 'Helper', 'builder': set_up_ec_helper_agent},
                    {'name': 'RPA Operator', 'builder': set_up_ec_rpa_operator_agent},
                    {'name': 'Tester', 'builder': set_up_ec_tester_agent}
                ])
            else:
                agent_configs.append({
                    'name': 'Helper',
                    'builder': set_up_ec_helper_agent
                })
                
                if "ONLY" not in self.machine_role:
                    agent_configs.extend([
                        {'name': 'Procurement', 'builder': set_up_ec_procurement_agent},
                        {'name': 'Tester', 'builder': set_up_ec_tester_agent}
                    ])
            
            logger.info(f"[MainWindow] 📋 Building {len(agent_configs)} code-built agents in parallel")
            
            # Build all agents in parallel
            async def build_single(config):
                try:
                    loop = asyncio.get_event_loop()
                    agent = await loop.run_in_executor(None, config['builder'], self)
                    return (config['name'], agent) if agent else None
                except Exception as e:
                    logger.error(f"[MainWindow] ❌ Failed to build {config['name']}: {e}")
                    return None
            
            build_tasks = [build_single(config) for config in agent_configs]
            results = await asyncio.gather(*build_tasks, return_exceptions=True)
            
            # Filter successful builds
            built_agents = []
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    logger.error(f"[MainWindow] ❌ Failed to build {agent_configs[i]['name']}: {result}")
                elif result and result[1]:
                    built_agents.append(result)
                    logger.info(f"[MainWindow] ✅ Built {result[0]} agent ({len(built_agents)}/{len(agent_configs)})")
                else:
                    logger.warning(f"[MainWindow] ⚠️ {agent_configs[i]['name']} build returned None")
            
            logger.info(f"[MainWindow] 🎉 Code-built agents completed: {len(built_agents)}/{len(agent_configs)}")
            return built_agents
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to build code agents: {e}")
            return []

    def _launch_agents_async(self, agents_list):
        """Launch agents in background without blocking
        
        Creates a background task to launch all agents in parallel.
        This method returns immediately without waiting for launches to complete.
        
        Args:
            agents_list: List of EC_Agent objects to launch
        """
        async def launch_all_async():
            """Internal async function to launch all agents"""
            launchable_agents = [
                (f"Agent_{i}", agent) for i, agent in enumerate(agents_list)
                if hasattr(agent, 'launch') or hasattr(agent, 'start')
            ]
            
            if not launchable_agents:
                logger.info("[MainWindow] 🔍 No launchable agents found")
                return
            
            logger.info(f"[MainWindow] 🚀 Launching {len(launchable_agents)} agents in background (fire-and-forget)...")
            
            # Launch all agents in parallel using existing method
            launch_tasks = [
                self._launch_single_agent_with_name_async(name, agent)
                for name, agent in launchable_agents
            ]
            results = await asyncio.gather(*launch_tasks, return_exceptions=True)
            
            # Count results
            launched_count = sum(1 for r in results if r is True)
            logger.info(f"[MainWindow] 🎉 Background launch completed: {launched_count}/{len(launchable_agents)} agents launched")
        
        # Create task and run in background (fire-and-forget)
        asyncio.create_task(launch_all_async())
        logger.info(f"[MainWindow] 🔥 Agent launch task created (running in background)")
    
    async def _launch_single_agent_with_name_async(self, agent_name: str, agent):
        """Asynchronously launch a single Agent (returns launch result)"""
        try:
            # Log comprehensive agent and skill information
            agent_card_name = agent.get_card().name if hasattr(agent, 'get_card') and agent.get_card() else agent_name
            logger.info(f"[AGENT_LAUNCH] Starting launch for agent: {agent_card_name}")
            
            # Check and log agent skills
            if hasattr(agent, 'skills'):
                skills_count = len(agent.skills) if agent.skills else 0
                logger.info(f"[AGENT_SKILLS] Agent '{agent_card_name}' has {skills_count} skills")
                
                if skills_count > 0:
                    for idx, skill in enumerate(agent.skills):
                        if skill is None:
                            logger.error(f"[SKILL_MISSING] Agent '{agent_card_name}' skill[{idx}] is None!")
                        else:
                            skill_name = skill.name if hasattr(skill, 'name') else f"Skill_{idx}"
                            has_runnable = hasattr(skill, 'runnable') and skill.runnable is not None
                            runnable_type = type(skill.runnable).__name__ if has_runnable else "None"
                            logger.info(f"[SKILL_CHECK] Agent '{agent_card_name}' skill[{idx}]: {skill_name}, runnable: {runnable_type}")
                            
                            if not has_runnable:
                                logger.error(f"[SKILL_MISSING] Agent '{agent_card_name}' skill '{skill_name}' has runnable=None!")
                else:
                    logger.warning(f"[AGENT_SKILLS] Agent '{agent_card_name}' has no skills!")
            else:
                logger.warning(f"[AGENT_SKILLS] Agent '{agent_card_name}' has no 'skills' attribute")
            
            # Check if Agent has a launch method
            if hasattr(agent, 'launch') and callable(agent.launch):
                # Execute launch in thread pool (avoid blocking)
                loop = asyncio.get_event_loop()
                await loop.run_in_executor(None, agent.launch)
                # Wait for server to be ready
                await self.wait_for_server_async(agent, timeout=10.0)
                return True
            elif hasattr(agent, 'start') and callable(agent.start):
                loop = asyncio.get_event_loop()
                await loop.run_in_executor(None, agent.start)
                
                # 🔧 Critical fix: Wait for A2A Server to be fully ready before returning
                # Agent.start() launches Uvicorn in a thread but doesn't wait for it to be ready
                # We must explicitly wait for the server's /ping endpoint to respond
                logger.info(f"[MainWindow] 🔍 Waiting for {agent.get_card().name}'s A2A server to be ready...")
                server_ready = await self.wait_for_server_async(agent, timeout=15.0)
                
                if server_ready:
                    logger.info(f"[MainWindow] ✅ {agent.get_card().name}'s A2A server is ready")
                    return True
                else:
                    logger.error(f"[MainWindow] ❌ {agent.get_card().name}'s A2A server failed to start within 10s")
                    return False
            else:
                # If no specific launch method, try waiting for server readiness
                if hasattr(agent, 'get_card') and agent.get_card():
                    await self.wait_for_server_async(agent, timeout=3.0)
                    return True
                else:
                    logger.warning(f"[MainWindow] ⚠️ {agent.get_card().name} has no launch method or server card")
                    return False
                    
        except Exception as e:
            import traceback
            agent_name = "Unknown"
            try:
                agent_name = agent.get_card().name if hasattr(agent, 'get_card') else str(agent)
            except:
                pass
            logger.error(f"[MainWindow] ❌ Failed to launch {agent_name}: {e}")
            logger.error(f"[MainWindow] Traceback: {traceback.format_exc()}")
            return False

    async def _load_local_db_agents(self):
        """
        Load agents from local database using agent service
        Returns: List of agent dicts
        """
        try:
            logger.info("[MainWindow] 📂 Loading agents from local database...")
            logger.info(f"[MainWindow] 📂 Current user (original): {self.user}")
            logger.info(f"[MainWindow] 📂 Current user (sanitized): {self.log_user}")
            
            if not self.ec_db_mgr or not self.ec_db_mgr.agent_service:
                logger.warning("[MainWindow] ⚠️  Agent service not available")
                return []
            
            agent_service = self.ec_db_mgr.agent_service
            logger.info(f"[MainWindow] 📂 Agent service available: {agent_service}")
            
            # Use agent service method to load agents (run in executor to avoid blocking)
            # IMPORTANT: Use self.user (original email) not self.log_user (sanitized for filesystem)
            def load_from_service():
                logger.info(f"[MainWindow] 📂 Querying agents for owner: {self.user}")
                result = agent_service.get_agents_by_owner(self.user)
                logger.info(f"[MainWindow] 📂 Query result: success={result.get('success')}, data_count={len(result.get('data', []))}")
                if result.get("success"):
                    agents_data = result.get("data", [])
                    logger.info(f"[MainWindow] 📂 Found {len(agents_data)} agents in database")
                    return agents_data
                else:
                    logger.error(f"[MainWindow] ❌ Agent service query failed: {result.get('error')}")
                    return []
            
            agents = await asyncio.get_event_loop().run_in_executor(None, load_from_service)
            logger.info(f"[MainWindow] ✅ Loaded {len(agents)} agents from local database")
            return agents
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to load local DB agents: {e}")
            return []
    
    def _merge_agent_tasks_to_memory(self):
        """
        Merge agent.tasks from built agents into mainwin.agent_tasks
        
        This method consolidates tasks from all agents into the main window's task list,
        with deduplication by task ID and name.
        
        Returns:
            tuple: (merged_count, skipped_count, total_count)
        """
        try:
            logger.info("[MainWindow] 📝 Merging agent.tasks from built agents...")
            merged_count = 0
            skipped_count = 0

            # Build a set of existing task identifiers for fast lookup
            existing_task_ids = set()
            existing_task_names = set()
            for existing_task in self.agent_tasks:
                task_id = getattr(existing_task, 'id', None)
                task_name = getattr(existing_task, 'name', None)
                if task_id:
                    existing_task_ids.add(task_id)
                if task_name:
                    existing_task_names.add(task_name)

            logger.debug(f"[MainWindow] Existing tasks: {len(existing_task_ids)} with IDs, {len(existing_task_names)} with names")

            # Merge agent tasks with deduplication
            for agent in self.agents:
                if hasattr(agent, 'tasks') and agent.tasks:
                    agent_name = getattr(agent, 'name', 'Unknown')
                    logger.debug(f"[MainWindow] Processing {len(agent.tasks)} tasks from agent: {agent_name}")

                    for agent_task in agent.tasks:
                        task_id = getattr(agent_task, 'id', None)
                        task_name = getattr(agent_task, 'name', None)

                        # Check for duplicates by ID (primary) or name (fallback)
                        is_duplicate = False
                        if task_id and task_id in existing_task_ids:
                            is_duplicate = True
                            logger.debug(f"[MainWindow] Skipping duplicate task by ID: {task_id}")
                        elif task_name and task_name in existing_task_names:
                            is_duplicate = True
                            logger.debug(f"[MainWindow] Skipping duplicate task by name: {task_name}")

                        if not is_duplicate:
                            # Add to memory
                            self.agent_tasks.append(agent_task)

                            # Update tracking sets
                            if task_id:
                                existing_task_ids.add(task_id)
                            if task_name:
                                existing_task_names.add(task_name)

                            merged_count += 1
                            logger.debug(f"[MainWindow] Merged task: {task_name or task_id}")
                        else:
                            skipped_count += 1

            total_count = len(self.agent_tasks)
            logger.info(f"[MainWindow] ✅ Merged {merged_count} agent tasks from built agents")
            logger.info(f"[MainWindow] ⏭️  Skipped {skipped_count} duplicate agent tasks")
            logger.info(f"[MainWindow] 📊 Total agent tasks in memory: {total_count}")
            
            return merged_count, skipped_count, total_count
            
        except Exception as e:
            logger.warning(f"[MainWindow] ⚠️ Failed to merge agent.tasks: {e}")
            import traceback
            logger.debug(f"[MainWindow] Merge error traceback: {traceback.format_exc()}")
            return 0, 0, len(self.agent_tasks)
    
    async def _fetch_cloud_agents(self):
        """
        Fetch agents from cloud API
        Returns: List of agent dicts
        """
        try:
            logger.info("[MainWindow] 🌐 Fetching agents from cloud...")
            
            # Use load_agents_from_cloud to fetch agents from cloud
            from agent.ec_agents.agent_utils import load_agents_from_cloud
            
            # Run in executor to avoid blocking the event loop
            # Note: load_agents_from_cloud returns EC_Agent objects, not dicts
            cloud_agent_objects = await asyncio.get_event_loop().run_in_executor(
                None,
                load_agents_from_cloud,
                self
            )
            
            if not cloud_agent_objects:
                logger.info("[MainWindow] ℹ️  No agents returned from cloud")
                return []
            
            logger.info(f"[MainWindow] ✅ Fetched {len(cloud_agent_objects)} agents from cloud")
            
            # Convert EC_Agent objects to dicts for merging
            cloud_agent_dicts = []
            for agent_obj in cloud_agent_objects:
                try:
                    if hasattr(agent_obj, 'to_dict'):
                        agent_dict = agent_obj.to_dict(owner=self.user)
                    elif hasattr(agent_obj, 'card'):
                        # Extract basic info from agent card
                        agent_dict = {
                            'id': agent_obj.card.id if hasattr(agent_obj.card, 'id') else None,
                            'name': agent_obj.card.name if hasattr(agent_obj.card, 'name') else 'Unknown',
                            'description': agent_obj.card.description if hasattr(agent_obj.card, 'description') else '',
                            'owner': getattr(agent_obj, 'owner', 'unknown'),
                        }
                    else:
                        logger.warning(f"[MainWindow] Cannot convert agent object to dict: {agent_obj}")
                        continue
                    
                    cloud_agent_dicts.append(agent_dict)
                except Exception as e:
                    logger.warning(f"[MainWindow] Failed to convert agent to dict: {e}")
                    continue
            
            logger.info(f"[MainWindow] 🔄 Converted {len(cloud_agent_dicts)} cloud agents to dicts")
            return cloud_agent_dicts
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to fetch cloud agents: {e}")
            logger.error(f"[MainWindow] Traceback: {traceback.format_exc()}")
            return []
    
    def _merge_agent_data(self, local_agents, cloud_agents):
        """
        Merge local and cloud agent data
        Cloud data overwrites local data by agent ID
        
        Args:
            local_agents: List of agent dicts from local database
            cloud_agents: List of agent dicts from cloud API
        
        Returns:
            List of merged agent dicts
        """
        try:
            logger.info("[MainWindow] 🔀 Merging local and cloud agent data...")
            
            # Create a dict with local agents keyed by ID
            merged_dict = {agent.get('id'): agent for agent in local_agents}
            
            # Overwrite with cloud agents (cloud data takes precedence)
            for cloud_agent in cloud_agents:
                agent_id = cloud_agent.get('id')
                if agent_id:
                    merged_dict[agent_id] = cloud_agent
                    logger.debug(f"[MainWindow] 🔄 Cloud agent overwrites local: {agent_id}")
            
            merged_agents = list(merged_dict.values())
            logger.info(f"[MainWindow] ✅ Merged {len(merged_agents)} unique agents")
            
            return merged_agents
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to merge agent data: {e}")
            # Return local agents as fallback
            return local_agents
    
    def _merge_all_agent_sources(self, code_built_agents, db_cloud_agents):
        """
        Merge agents from all sources with deduplication into a unified list of EC_Agent objects
        
        Priority: Code-built > Cloud > Local DB
        
        Args:
            code_built_agents: List of (name, agent) tuples from code builders
            db_cloud_agents: List of agent dicts from DB/Cloud merge
        
        Returns:
            List of EC_Agent objects (unified type)
        """
        try:
            logger.info("[MainWindow] 🔀 Merging all agent sources into unified list...")
            
            # Import converter
            from agent.agent_converter import convert_agent_dict_to_ec_agent
            
            # Create a unified list of EC_Agent objects
            unified_agents = []
            
            # Track code-built agent names for deduplication
            code_built_names = set()
            
            # Step 1: Add all code-built agents (already EC_Agent objects)
            for name, agent in code_built_agents:
                if agent:
                    unified_agents.append(agent)
                    code_built_names.add(name)
                    logger.debug(f"[MainWindow] 🔧 Added code-built agent: {name}")
            
            # Step 2: Convert DB/Cloud agents (dicts) to EC_Agent objects
            converted_count = 0
            failed_count = 0
            
            for agent_data in db_cloud_agents:
                agent_name = agent_data.get('name', '')
                
                # Only keep DB/Cloud agent if no code-built agent with same name
                if agent_name not in code_built_names:
                    # Convert dict to EC_Agent object
                    ec_agent = convert_agent_dict_to_ec_agent(agent_data, self)
                    if ec_agent:
                        unified_agents.append(ec_agent)
                        converted_count += 1
                        logger.debug(f"[MainWindow] 📂 Converted DB/Cloud agent: {agent_name}")
                    else:
                        failed_count += 1
                        logger.warning(f"[MainWindow] ⚠️  Failed to convert DB/Cloud agent: {agent_name}")
                else:
                    logger.debug(f"[MainWindow] ⚠️  DB/Cloud agent '{agent_name}' overwritten by code-built agent")
            
            logger.info(f"[MainWindow] ✅ Merged all sources into unified list:")
            logger.info(f"[MainWindow]    - Code-built agents: {len(code_built_names)}")
            logger.info(f"[MainWindow]    - DB/Cloud agents converted: {converted_count}")
            if failed_count > 0:
                logger.warning(f"[MainWindow]    - Conversion failed: {failed_count}")
            logger.info(f"[MainWindow]    - Total EC_Agent objects: {len(unified_agents)}")
            
            return unified_agents
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to merge all agent sources: {e}")
            import traceback
            logger.error(traceback.format_exc())
            # Return code-built agents as fallback
            return [agent for _, agent in code_built_agents if agent]
    
    async def _update_local_db_agents(self, merged_agents):
        """
        Update local database with merged agent data
        
        Args:
            merged_agents: List of merged agent dicts
        """
        try:
            logger.info(f"[MainWindow] 💾 Updating local database with {len(merged_agents)} merged agents...")
            
            if not self.ec_db_mgr or not self.ec_db_mgr.agent_service:
                logger.warning("[MainWindow] ⚠️  Agent service not available, skipping DB update")
                return
            
            agent_service = self.ec_db_mgr.agent_service
            
            # Run database update in executor
            def update_db():
                try:
                    with agent_service.session_scope() as session:
                        from agent.db.models.agent_model import DBAgent
                        
                        updated_count = 0
                        for agent_data in merged_agents:
                            agent_id = agent_data.get('id')
                            if not agent_id:
                                continue
                            
                            # Check if agent exists
                            # IMPORTANT: Use self.user (original email) not self.log_user (sanitized)
                            existing = session.query(DBAgent).filter(
                                DBAgent.id == agent_id,
                                DBAgent.owner == self.user
                            ).first()
                            
                            if existing:
                                # Update existing agent
                                for key, value in agent_data.items():
                                    if hasattr(existing, key):
                                        # Convert timestamp to datetime for SQLite
                                        if key in ('created_at', 'updated_at') and isinstance(value, (int, float)):
                                            from datetime import datetime
                                            value = datetime.fromtimestamp(value / 1000.0)  # Convert ms to seconds
                                        setattr(existing, key, value)
                                updated_count += 1
                            else:
                                # Create new agent - convert timestamps
                                agent_dict = agent_data.copy()
                                for key in ('created_at', 'updated_at'):
                                    if key in agent_dict and isinstance(agent_dict[key], (int, float)):
                                        from datetime import datetime
                                        agent_dict[key] = datetime.fromtimestamp(agent_dict[key] / 1000.0)
                                new_agent = DBAgent(**agent_dict)
                                session.add(new_agent)
                                updated_count += 1
                        
                        session.commit()
                        logger.info(f"[MainWindow] ✅ Updated {updated_count} agents in database")
                        
                except Exception as e:
                    logger.error(f"[MainWindow] ❌ Database update failed: {e}")
                    session.rollback()
            
            await asyncio.get_event_loop().run_in_executor(None, update_db)
            
        except Exception as e:
            logger.error(f"[MainWindow] ❌ Failed to update local DB agents: {e}")


    def get_vehicle_ecbot_op_agent(self, v):
        # obtain agents on a vehicle.
        logger.debug(f"{len(self.agents)}")
        ecb_op_agent = next((ag for ag in self.agents if "ECBot RPA Operator Agent" in ag.card.name), None)
        logger.info("FOUND Operator......", ecb_op_agent.card.name)
        return ecb_op_agent

    # SC note - really need to have
    async def run_async_tasks(self):
        if self.host_role != "Staff Officer":
            self.rpa_task = asyncio.create_task(self.runbotworks(self.gui_rpa_msg_queue, self.gui_monitor_msg_queue))
            self.manager_task = asyncio.create_task(self.runmanagerworks(self.gui_manager_msg_queue, self.gui_rpa_msg_queue, self.gui_monitor_msg_queue))

        else:
            self.rpa_task = asyncio.create_task(self.wait_forever())

        # all_tasks = [self.peer_task, self.monitor_task, self.chat_task, self.rpa_task, self.wan_sub_task]
        all_tasks = [self.peer_task, self.monitor_task, self.chat_task, self.rpa_task]

        # atasks = []
        # for ag in self.agents:
        #     for task in ag.tasks:
        #         if task.trigger == "schedule":
        #             all_tasks.append(asyncio.create_task(ag.runner.launch_scheduled_run(task)))
        #         # await loop.run_in_executor(threading.Thread(), await self.runner.launch_scheduled_run(task), True)
        #         elif task.trigger == "message":
        #             all_tasks.append(asyncio.create_task(ag.runner.launch_reacted_run(task)))
        #         # await loop.run_in_executor(threading.Thread(), await self.runner.launch_reacted_run(task), True)
        #         elif task.trigger == "interaction":
        #             all_tasks.append(asyncio.create_task(ag.runner.launch_interacted_run(task)))
        #         # await loop.run_in_executor(threading.Thread(), await self.runner.launch_interacted_run(task), True)
        #         else:
        #             print("WARNING: UNRECOGNIZED task trigger type....")

        await asyncio.gather(*all_tasks)

    # 1) gather all skills (cloud + local public)
    # 2) analyze dependence and update data structure
    # 3) regenerate psk files for each skill
    # 4) build up skill_table (a look up table)
    def dailySkillsetUpdate(self):
        if not self.general_settings.is_test_mode():
            cloud_skills_results = self.skill_manager.fetch_my_skills()
            logger.trace("DAILY SKILL FETCH:", cloud_skills_results)
        else:
            cloud_skills_results = {"body": "{}"}
        existing_skids = [sk.getSkid() for sk in self.skills]
        logger.info("EXISTING SKIDS:", existing_skids)

        if 'body' in cloud_skills_results:
            # self.showMsg("db_skills_results:::::"+json.dumps(db_skills_results))
            cloud_skills = json.loads(cloud_skills_results["body"])
            logger.info("Cloud side skills fetched:" + str(len(cloud_skills)))

            # convert json to WORKSKILL object.
            for cloud_skill in cloud_skills:
                if cloud_skill["skid"] not in existing_skids:
                    logger.trace("db skill:" + json.dumps(cloud_skill))
                    cloud_work_skill = WORKSKILL(self, cloud_skill["name"])
                    cloud_work_skill.loadJson(cloud_skill)

                    # now read the cloud skill's local definition file to get
                    self.skills.append(cloud_work_skill)

            # this will handle all skill bundled into software itself.
            logger.info("load local private skills")
            self.loadLocalPrivateSkills()

            # read public skills from local json files and merge with what's just read from the cloud.
            # if there is any conlict will use the cloud data as the true data.
            self.loadPublicSkills()

            # Note: SkillManager no longer handles GUI display
            # Skills are now managed in the main skills list
            # You may need to implement a new display method or use existing skill display


            # for sanity immediately re-generate psk files... and gather dependencies info so that when user creates a new mission
            # when a skill is selected, its dependencies will added to mission's skills list.
            logger.info("SKIDS to be regenerated:", [sk.getSkid() for sk in self.skills])
            self.regenSkillPSKs()

        logger.trace("after daily sync SKIDS:", [sk.getSkid() for sk in self.skills])



    def addSkillRowsToSkillManager(self):
        # Note: SkillManager no longer handles GUI display
        # Skills are now managed in the main skills list
        # You may need to implement a new display method or use existing skill display
        pass

    def regenSkillPSKs(self):
        for ski, sk in enumerate(self.skills):
            # next_step is not used,
            sk_full_name = sk.getPlatform()+"_"+sk.getApp()+"_"+sk.getSiteName()+"_"+sk.getPage()+"_"+sk.getName()
            logger.trace("PSK FILE NAME::::::::::"+str(ski)+"::["+str(sk.getSkid())+"::"+sk.getPrivacy()+":::::"+sk_full_name, "fetchSchedule", self)
            if sk.getPrivacy() == "public":
                next_step, psk_file = genSkillCode(sk_full_name, sk.getPrivacy(), self.ecb_data_homepath, first_step, "light")
            else:
                self.showMsg("GEN PRIVATE SKILL PSK::::::" + sk_full_name)
                next_step, psk_file = genSkillCode(sk_full_name, sk.getPrivacy(), self.my_ecb_data_homepath, first_step, "light")
            logger.trace("PSK FILE:::::::::::::::::::::::::"+psk_file, "fetchSchedule", self)
            sk.setPskFileName(psk_file)
            # fill out each skill's depencies attribute
            sk.setDependencies(self.analyzeMainSkillDependencies(psk_file))
            logger.trace("RESULTING DEPENDENCIES:["+str(sk.getSkid())+"] ", sk.getDependencies())

    def get_helper_agent(self):
        return self.helper_agent



    def translateSiteName(self, site_text):
        if site_text in self.static_resource.SITES_SH_DICT.keys():
            return self.static_resource.SITES_SH_DICT[site_text]
        else:
            return site_text

    def translatePlatform(self, site_text):
        if site_text in self.static_resource.PLATFORMS_SH_DICT.keys():
            return self.static_resource.PLATFORMS_SH_DICT[site_text]
        else:
            return site_text

    def translateShortSiteName(self, site_text):
        if site_text in self.static_resource.SH_SITES_DICT.keys():
            return self.static_resource.SH_SITES_DICT[site_text]
        else:
            return site_text

    def translateShortPlatform(self, site_text):
        if site_text in self.static_resource.SH_PLATFORMS_DICT.keys():
            return self.static_resource.SH_PLATFORMS_DICT[site_text]
        else:
            return site_text



    def setCommanderName(self, cn):
        self.commander_name = cn

    def getCommanderName(self):
        return self.commander_name

    def _get_cpu_info_safely(self):
        """
        Safely get CPU information, avoiding multiprocessing issues

        Returns:
            dict: CPU information dictionary containing fields like brand_raw and hz_advertised_friendly
        """
        from utils.cpu_info_helper import get_cpu_info_safely
        return get_cpu_info_safely()

    def getWifis(self):
        return self.config_manager.general_settings.get_wifi_networks()

    def getWebDriverPath(self):
        return self.config_manager.general_settings.default_webdriver_path



    def getWebDriver(self):
        return self.default_webdriver

    def setWebDriver(self, driver):
        self.default_webdriver = driver

    def getWebCrawler(self):
        return self.async_crawler

    def setupUnifiedBrowserManager(self):
        """Setup unified browser manager"""
        try:
            self.unified_browser_manager = get_unified_browser_manager()

            # Pass file system path
            file_system_path = self.config_manager.general_settings.browser_use_file_system_path

            if self.unified_browser_manager.initialize(file_system_path=file_system_path):
                logger.info("✅ Browser manager initialized successfully")
            else:
                logger.error("❌ Browser manager initialization failed")
                self.unified_browser_manager = None

        except Exception as e:
            logger.error(f"Browser manager setup failed: {e}")
            self.unified_browser_manager = None

    async def _start_webdriver_initialization(self):
        """Initializes the WebDriver using the simplified WebDriverManager."""
        try:
            from gui.webdriver.manager import get_webdriver_manager

            logger.info("🚀 Starting WebDriver initialization...")

            # Get the manager instance
            manager = await get_webdriver_manager()

            # Initialize the manager. This is now a self-contained async process.
            success = await manager.initialize()

            if success:
                new_path = await manager.get_webdriver_path()
                # If we found a new, valid path that is different from the one in settings, update and save.
                if new_path and new_path != self.config_manager.general_settings.default_webdriver_path:
                    logger.info(f"WebDriver path updated to: {new_path}. Saving settings...")
                    self.config_manager.general_settings.default_webdriver_path = new_path
                    self.config_manager.general_settings.save()
                else:
                    # Ensure it's set even if not saved
                    self.config_manager.general_settings.default_webdriver_path = new_path

                logger.info(f"✅ WebDriver initialization successful. Path: {self.config_manager.general_settings.default_webdriver_path}")
            else:
                logger.error("❌ WebDriver initialization failed.")

        except Exception as e:
            logger.error(f"An exception occurred during WebDriver initialization: {e}")

    @property
    def async_crawler(self):
        try:
            manager = self.unified_browser_manager
            if manager and manager.is_ready():
                return manager.get_async_crawler()
        except Exception as e:
            logger.error(f"Error accessing async_crawler: {e}")
        return None

    @property
    def browser_session(self):
        try:
            manager = self.unified_browser_manager
            if manager and manager.is_ready():
                return manager.get_browser_session()
        except Exception as e:
            logger.error(f"Error accessing browser_session: {e}")
        return None

    @property
    def browser_use_controller(self):
        try:
            manager = getattr(self, 'unified_browser_manager', None)
            if manager and manager.is_ready():
                return manager.get_browser_use_controller()
        except Exception as e:
            logger.error(f"Error accessing browser_use_controller: {e}")
        return None

    @property
    def browser_use_file_system(self):
        try:
            manager = getattr(self, 'unified_browser_manager', None)
            if manager and manager.is_ready():
                return manager.get_browser_use_file_system()
        except Exception as e:
            logger.error(f"Error accessing browser_use_file_system: {e}")
        return None

    def getBrowserSession(self):
        return self.browser_session

    def getBrowserUseController(self):
        return self.browser_use_controller

    def load_build_dom_tree_script(self):
        script = ""
        try:
            logger.debug("Loading build dom tree script...", self.build_dom_tree_script_path)
            with open(self.build_dom_tree_script_path, 'r', encoding='utf-8') as file:
                script = file.read()
            return script
        except FileNotFoundError:
            logger.error(f"Error: The file {self.build_dom_tree_script_path} was not found.")
            return ""
        except IOError as e:
            logger.error(f"Error reading {self.build_dom_tree_script_path}: {e}")
            return ""

    #async def networking(self, platoonCallBack):
    def set_host_role(self, role):
        self.host_role = role

    def set_default_wifi(self, default_wifi):
        self.config_manager.general_settings.default_wifi = default_wifi
        self.config_manager.general_settings.save()

    def get_default_wifi(self):
        return self.config_manager.general_settings.default_wifi

    def set_default_printer(self, default_printer):
        self.config_manager.general_settings.default_printer = default_printer
        self.config_manager.general_settings.save()

    def get_default_printer(self):
        return self.config_manager.general_settings.default_printer


    def saveSettings(self):
        try:
            self.showMsg("saving all settings...")
            return self.config_manager.save_all_settings()
        except Exception as e:
            logger.error(f"Error saving settings: {e}")
            return False

    @property
    def general_settings(self):
        """
        Unified access to general settings.
        Use this property to access any general settings field directly:
        
        Examples:
            self.general_settings.debug_mode
            self.general_settings.schedule_mode
            self.general_settings.default_wifi
            self.general_settings.lan_api_endpoint
        """
        return self.config_manager.general_settings
    
    def get_server_base_url(self) -> str:
        """
        Get the base URL of the local server.
        
        This is the unified method for getting server URL across the application.
        It directly uses the server_manager_instance to get the actual running server URL.
        
        Returns:
            str: Base URL like "http://localhost:4668"
        """
        try:
            from gui.LocalServer import server_manager_instance
            if server_manager_instance:
                return server_manager_instance.get_server_url()
        except Exception as e:
            from utils.logger_helper import logger_helper as logger
            logger.warning(f"Failed to get server URL from server_manager: {e}")
        
        # Fallback: construct from config port
        port = self.general_settings.local_server_port
        return f"http://localhost:{port}"

    def get_host_role(self):
        return self.host_role


    def setCommanderXPort(self, xport):
        self.commanderXport = xport

    def getGuiMsgQueue(self):
        return self.gui_net_msg_queue

    def setIP(self, ip):
        self.ip = ip

    def getUser(self):
        return self.user

    def getNetworkApiEngine(self):
        return self.config_manager.general_settings.network_api_engine

    def getLanOCREndpoint(self):
        return self.config_manager.general_settings.ocr_api_endpoint

    def getLanOCRApiKey(self):
        return self.config_manager.general_settings.ocr_api_key

    def getLanApiEndpoint(self):
        return self.config_manager.general_settings.lan_api_endpoint

    def get_local_server_port(self):
        return self.config_manager.general_settings.local_server_port

    def getWanApiEndpoint(self):
        return self.config_manager.general_settings.wan_api_endpoint

    def getWanApiKey(self):
        return self.config_manager.general_settings.wan_api_key

    def getWSApiEndpoint(self):
        return self.config_manager.general_settings.ws_api_endpoint

    def getWSApiHost(self):
        return self.config_manager.general_settings.ws_api_host

    def getAcctSiteID(self):
        site = self.machine_name
        user = self.user.replace("@", "_").replace(".", "_")
        return f"{user}_{site}"

    def setMILANServer(self, ip, port="8848"):
        self.config_manager.general_settings.lan_api_endpoint = f"http://{ip}:{port}"
        self.config_manager.general_settings.save()
        logger.info(f"lan_api_endpoint: {self.config_manager.general_settings.lan_api_endpoint}")

    def setLanDBServer(self, ip, port="5080"):
        self.config_manager.general_settings.local_user_db_host = ip
        self.config_manager.general_settings.local_user_db_port = port
        self.config_manager.general_settings.save()

    def getDisplayResolution(self):
        return self.config_manager.general_settings.display_resolution

    async def runTodaysLocalWork(self):
        # send a request to commander for today's scheduled work.
        workReq = {"type": "reqResendWorkReq", "ip": self.ip, "content": "now"}
        await self.send_json_to_commander(self.commanderXport, workReq)

    # 1) prepre ads profile cookies
    # 2) group them by vehicle
    # 3) assign them. (move the troop to the vehicle(host computer where they belong， Bots, Missions, Skills, ADS related data and files.)
    def handleCloudScheduledWorks(self, bodyobj):
        if bodyobj:
            log3("handleCloudScheduledWorks...."+str(len(bodyobj))+" "+str(type(bodyobj)), "fetchSchedule", self)
            # print("bodyobj:", bodyobj)
            for nm in bodyobj["added_missions"]:
                today = datetime.today()
                formatted_today = today.strftime('%Y-%m-%d')
                bd_parts = nm["createon"].split()
                nm["createon"] = formatted_today + " " + bd_parts[1]

            # log3("cloud schedule works:" + json.dumps(bodyobj), "fetchSchedule", self)
            log3("BEGIN ASSIGN INCOMING MISSION....", "fetchSchedule", self)
            self.build_cookie_site_lists()
            # convert new added mission json to MISSIONs object
            newlyAdded = self.addNewlyAddedMissions(bodyobj)
            # now that todays' newly added missions are in place, generate the cookie site list for the run.
            self.num_todays_task_groups = self.num_todays_task_groups + len(bodyobj["task_groups"])
            print("num_todays_task_groups:", self.num_todays_task_groups)
            # self.todays_scheduled_task_groups = self.groupTaskGroupsByOS(bodyobj["task_groups"])
            #  turn this into a per-vehicle flattend list of tasks (vehicle name based dictionary).
            self.todays_scheduled_task_groups = self.reGroupByBotVehicles(bodyobj["task_groups"])
            self.unassigned_scheduled_task_groups = self.todays_scheduled_task_groups
            print("current unassigned task groups:", list(self.unassigned_scheduled_task_groups.keys()))
            for vn in self.unassigned_scheduled_task_groups:
                print(f"unassigned task groups:{vn} {len(self.unassigned_scheduled_task_groups[vn])}")
            # print("current work to do:", self.todays_work)
            # for works on this host, add to the list of todos, otherwise send to the designated vehicle.
            # self.assignWork()

            # log3("current unassigned scheduled task groups after assignwork:"+json.dumps(self.unassigned_scheduled_task_groups), "fetchSchedule", self)
            # log3("current work to do after assignwork:"+json.dumps(self.todays_work), "fetchSchedule", self)

            self.logDailySchedule(json.dumps(bodyobj))
        else:
            log3("WARN: empty obj", "fetchSchedule", self)
            self.warn("Warning: NO schedule generated.")

    # this is more for the convinence of isolated testing ....
    def reGenWorksForVehicle(self, vehicle):

        if len(self.todaysSchedule) > 0:
            log3("reGenWorksForVehicle...." + str(len(self.todaysSchedule)) + " " + str(type(self.todaysSchedule)),
                 "fetchSchedule", self)
            # print("todaysSchedule:", self.todaysSchedule)

            vname = vehicle.getName()
            if vname in self.todaysSchedule["task_groups"]:
                vbids = []
                tzs = self.todaysSchedule["task_groups"][vname].keys()
                for tz in tzs:
                    vbids = vbids + [vw["bid"] for vw in self.todaysSchedule["task_groups"][vname][tz]]
                vadded_ms = [m for m in self.todaysSchedule["added_missions"] if m["botid"] in vbids]
                print("vbids:", vbids)
                print("vadded mids:", [m["mid"] for m in vadded_ms])
                vtg = {"task_groups": {vname: self.todaysSchedule["task_groups"][vname]},
                       "added_missions": vadded_ms}

                print("vtg:", vtg)

            # now that todays' newly added missions are in place, generate the cookie site list for the run.
            self.num_todays_task_groups = self.num_todays_task_groups + len(vtg["task_groups"])
            print("regen num_todays_task_groups:", self.num_todays_task_groups)

            self.todays_scheduled_task_groups = self.reGroupByBotVehicles(vtg["task_groups"])
            self.unassigned_scheduled_task_groups = self.todays_scheduled_task_groups
            # print("current unassigned task groups:", self.unassigned_scheduled_task_groups)
            # assignWork() will take care of the rest, it will check any unassigned work and assign them.

            log3("current unassigned scheduled task groups after assignwork:"+json.dumps(self.unassigned_scheduled_task_groups), "fetchSchedule", self)
            log3("current work to do after assignwork:"+json.dumps(self.todays_work), "fetchSchedule", self)

        else:
            log3("WARN: empty obj", "fetchSchedule", self)
            self.warn("Warning: NO schedule generated.")


    def addTestMissions(self, bodyobj):
        for m in bodyobj["added_missions"]:
            new_mission = EBMISSION(self)
            self.fill_mission(new_mission, m, bodyobj["task_groups"])
            self.setPrivateAttributesBasedOnPast(new_mission)
            new_mission.updateDisplay()
            self.missions.append(new_mission)
            # missionModel removed - UI components no longer needed

    # this function fetches schedule and assign work based on fetch schedule results...
    def fetchSchedule(self, ts_name, settings, forceful=False):
        log3("time stamp " + datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] + " start fetching schedule...", "fetchSchedule", self)
        ex_stat = "Completed:0"
        try:
            # before even actual fetch schedule, automatically all new customer buy orders from the designated directory.
            # self.newBuyMissionFromFiles()
            # self.createNewBotsFromBotsXlsx()
            # self.createNewMissionsFromOrdersXlsx()
            today = datetime.now()
            # Format the date as yyyymmdd
            yyyymmdd = today.strftime("%Y%m%d")
            sf_name = "schedule" + yyyymmdd+".json"
            schedule_file = os.path.join(self.my_ecb_data_homepath + "/runlogs", sf_name)
            todaysScheduleExists = os.path.exists(schedule_file)
            log3("Done handling today's new Buy orders...", "fetchSchedule", self)
            bodyobj = {}
            # next line commented out for testing purpose....
            if not self.config_manager.general_settings.is_debug_enabled() and self.config_manager.general_settings.is_auto_mode():
                log3("schedule setting:"+json.dumps(settings), "fetchSchedule", self)

                log3(f"schedule file {schedule_file} exists: {todaysScheduleExists}", "fetchSchedule", self)
                if not todaysScheduleExists or forceful:
                    jresp = send_schedule_request_to_cloud(self.session, self.get_auth_token(), ts_name, settings, self.getWanApiEndpoint())
                    log3(f"schedule JRESP: {len(jresp['body'])} bytes", "fetchSchedule", self)
                else:
                    with open(schedule_file, "r") as sf:
                        jresp = json.load(sf)
            else:
                log3("debug mode, skipping cloud fetch schedule", "fetchSchedule", self)
                jresp = {}

            if "errorType" in jresp:
                screen_error = True
                log3("ERROR Type: "+json.dumps(jresp["errorType"])+"ERROR Info: "+json.dumps(jresp["errorInfo"]), "fetchSchedule", self)
            else:
                # first, need to decompress the body.
                # very important to use compress and decompress on Base64
                if not self.config_manager.general_settings.is_debug_enabled() and self.config_manager.general_settings.is_auto_mode():
                    if not todaysScheduleExists or forceful:
                        uncompressed = self.zipper.decompressFromBase64(jresp["body"])   # commented out for testing
                    else:
                        uncompressed = "{}"
                else:
                    uncompressed = "{}"
                print("unzip schedule done....")
                # for testing purpose, short circuit the cloud fetch schedule and load a tests schedule from a tests
                # json file instead.

                # uncompressed = jresp["body"]
                if uncompressed != "":
                    self.showMsg("body string:!"+str(len(uncompressed))+"::")

                    bodyobj = {"task_groups": {}, "added_missions": []}

                    if not self.config_manager.general_settings.is_debug_enabled() and self.config_manager.general_settings.is_auto_mode():
                        if not todaysScheduleExists or forceful:
                            bodyobj = json.loads(uncompressed)                      # for test purpose, comment out, put it back when test is done....
                        else:
                            bodyobj = jresp
                    else:
                        log3("debug mode, using test vector....", "fetchSchedule", self)
                        file = 'C:/temp/scheduleResultTest1.json'
                        if exists(file):
                            with open(file) as test_schedule_file:
                                bodyobj = json.load(test_schedule_file)
                                self.addTestMissions(bodyobj)

                    # self.handleCloudScheduledWorks(bodyobj)
                else:
                    self.warn("Warning: Empty Network Response.")

            if ((not todaysScheduleExists) or forceful) and (not self.config_manager.general_settings.is_debug_enabled()) and (self.config_manager.general_settings.is_auto_mode()):
                log3(f"saving schedule file {schedule_file}", "fetchSchedule", self)

                with open(schedule_file, 'w') as sf:
                    json.dump(bodyobj, sf, indent=4)
                sf.close()

            print("fetch schedule time stamp " + datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] + " done with fetch schedule....", list(bodyobj.keys()), len(bodyobj["added_missions"]))
            self.todaysSchedule = bodyobj
            return bodyobj
        # ni is already incremented by processExtract(), so simply return it.
        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorFetchSchedule:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorFetchSchedule: traceback information not available:" + str(e)
            self.showMsg(ex_stat)
            return {}

    def warn(self, msg, level="info"):
        logger.warning(msg)

    def showMsg(self, msg, level="info"):
        logger.info(msg)

    def logDailySchedule(self, netSched):
        now = datetime.now()  # current date and time
        year = now.strftime("%Y")
        month = now.strftime("%m")
        day = now.strftime("%d")
        time = now.strftime("%H:%M:%S - ")
        dailyScheduleLogFile = self.my_ecb_data_homepath + "/runlogs/{}/{}/schedule{}{}{}.txt".format(self.log_user, year, month, day, year)
        if os.path.isfile(dailyScheduleLogFile):
            log3("append to daily schedule file:" + dailyScheduleLogFile, "fetchSchedule", self)
            file1 = open(dailyScheduleLogFile, "a")  # append mode
            file1.write(json.dumps(time+netSched) + "\n=====================================================================\n")
            file1.close()
        else:
            log3("daily schedule file not exist:"+dailyScheduleLogFile, "fetchSchedule", self)
            file1 = open(dailyScheduleLogFile, "w")  # write mode
            file1.write(json.dumps(time+netSched) + "\n=====================================================================\n")
            file1.close()

    def saveDailyRunReport(self, runStat):
        now = datetime.now()  # current date and time
        year = now.strftime("%Y")
        month = now.strftime("%m")
        day = now.strftime("%d")
        time = now.strftime("%H:%M:%S - ")
        dailyRunReportFile = self.my_ecb_data_homepath + "/runlogs/{}/{}/runreport{}{}{}.txt".format(self.log_user, year, month, day, year)

        if os.path.isfile(dailyRunReportFile):
            with open(dailyRunReportFile, 'a') as f:

                f.write(time+json.dumps(runStat) + "\n")

                f.close()
        else:
            with open(dailyRunReportFile, 'w') as f:

                f.write(time+json.dumps(runStat) + "\n")

                f.close()


    def fill_mission(self, blank_m, m, tgs):
        # print("BLANK:", m)
        blank_m.loadNetRespJson(m)
        # self.showMsg("after fill mission paramter:"+str(blank_m.getRetry()))
        mconfig = None
        for v in tgs:
            tz_group = tgs[v]
            for tz in tz_group:
                if len(tz_group[tz]) > 0:
                    for bot_works in tz_group[tz]:
                        for bw in bot_works["bw_works"]:
                            if m["mid"] == bw["mid"]:
                                # now add this mission to the list.
                                self.showMsg("found a bw mission matching mid.... "+str(bw["mid"]))
                                mconfig = bw["config"]
                                break
                        if mconfig:
                            break

                        for ow in bot_works["other_works"]:
                            if m["mid"] == ow["mid"]:
                                # now add this mission to the list.
                                self.showMsg("found a other mission matching mid.... "+str(ow["mid"]))
                                mconfig = ow["config"]
                                break
                        if mconfig:
                            break
                if mconfig:
                    break
            if mconfig:
                break
        print("SETTING CONFIG:", mconfig)
        blank_m.setConfig(mconfig)

    # after fetching today's schedule, update missions data structure since some walk/buy routine will be created.
    # as well as some daily routines.... will be generated as well....
    # one of the key thing to do here is the fill out the private attribute from the most recent past similar missions.
    def addNewlyAddedMissions(self, resp_data):
        # for each received work mission, check whether they're in the self.missions already, if not, create them and
        # add to the missions list.
        mb_words = ""
        task_groups = resp_data["task_groups"]
        for v in task_groups:
            tg = task_groups[v]
            for tz in tg:
                for wg in tg[tz]:
                    for w in wg["bw_works"]:
                        mb_words = mb_words + "M"+str(w["mid"])+"B"+str(wg["bid"]) + ", "

                    for w in wg["other_works"]:
                        mb_words = mb_words + "M"+str(w["mid"])+"B"+str(wg["bid"]) + ", "

        log3(mb_words, "fetchSchedule", self)

        newAdded = []
        newly_added_missions = resp_data["added_missions"]
        true_newly_added = []       # newly_added_missions includes some previous incompleted missions, they're not really NEW.
        log3("Added MS:"+json.dumps(["M"+str(m["mid"])+"B"+str(m["botid"]) for m in newly_added_missions]), "fetchSchedule", self)
        loadedMids = [m.getMid() for m in self.missions]
        for m in newly_added_missions:
            if m["mid"] not in loadedMids:
                new_mission = EBMISSION(self)
                self.fill_mission(new_mission, m, task_groups)
                self.setPrivateAttributesBasedOnPast(new_mission)
                new_mission.updateDisplay()
                self.missions.append(new_mission)
                # missionModel removed - UI components no longer needed
                log3("adding mission.... "+str(new_mission.getRetry()), "fetchSchedule", self)
                true_newly_added.append(new_mission)
                newAdded.append(new_mission)
            else:
                log3("this mission already exists:"+str(m["mid"]), "fetchSchedule", self)
                # in such a case, simply sync up the data
                existingMission = self.getMissionByID(m["mid"])
                # now, update data from cloud...
                existingMission.loadNetRespJson(m)
                newAdded.append(existingMission)

        if not self.config_manager.general_settings.debug_mode:
            self.addMissionsToLocalDB(true_newly_added)

        return(newAdded)

    # this is really about setting up fingerprint profile automatically for a new Mission
    # if it's not already set up, basically using bot's email + site in cuspas to create
    # the relevant profile xlsx from the superset .txt version of the bot's profile (based on email only)
    def setPrivateAttributesBasedOnPast(self, newMission):
        print("new mission type and cuspas:", newMission.getType(), newMission.getCusPAS())

        foundBot = self.getBotByID(newMission.getBid())
        if foundBot:
            botEmail = foundBot.getEmail()

        similar = [m for m in self.missions if m.getType() == newMission.getType() and m.getCusPAS() == newMission.getCusPAS()]
        similarWithFingerPrintProfile = [m for m in similar if m.getFingerPrintProfile()]

        print("similar w fpp: ", [m.getFingerPrintProfile() for m in similarWithFingerPrintProfile])
        if similarWithFingerPrintProfile:
            mostRecent = similarWithFingerPrintProfile[-1]

            newMission.setFingerPrintProfile(mostRecent.getFingerPrintProfile())
            print("newy set fpp:", newMission.getFingerPrintProfile())

    def getBotByID(self, bid):
        found_bot = next((bot for i, bot in enumerate(self.bots) if bot.getBid() == bid), None)
        return found_bot

    def getMissionByID(self, mid):
        found_mission = next((mission for i, mission in enumerate(self.missions) if mission.getMid() == mid), None)
        return found_mission

    def getSkillByID(self, skid):
        found_skill = next((skill for i, skill in enumerate(self.skills) if skill.getSkid() == skid), None)
        return found_skill

    def formBotsJsons(self, botids):
        result = []
        for bid in botids:
            # result = result + json.dumps(self.getBotByID(bid).genJson()).replace('"', '\\"')
            found_bot = next((bot for i, bot in enumerate(self.bots) if bot.getBid() == bid), None)
            if found_bot:
                result.append(found_bot.genJson())

        return result


    def formMissionsJsons(self, mids):
        result = []
        for mid in mids:
            # result = result + json.dumps(self.getMissionByID(mid).genJson()).replace('"', '\\"')
            found_mission = next((mission for i, mission in enumerate(self.missions) if mission.getMid() == mid), None)
            if found_mission:
                mj = found_mission.genJson()
                result.append(mj)

        return result

    def formSkillsJsons(self, skids):
        result = []
        all_skids = [sk.getSkid() for sk in self.skills]
        self.showMsg("all known skids:"+json.dumps(all_skids))
        for skid in skids:
            # result = result + json.dumps(self.getMissionByID(mid).genJson()).replace('"', '\\"')
            found_skill = next((sk for i, sk in enumerate(self.skills) if sk.getSkid() == skid), None)
            if found_skill:
                print("found skill")
                result.append(found_skill.genJson())
            else:
                self.showMsg("ERROR: skill id not found [" + str(skid)+"]")
        return result

    def formBotsMissionsSkillsString(self, botids, mids, skids):
        # result = "{\"bots\": " + self.formBotsString(botids) + ",\"missions\": " + self.formMissionsString(mids) + "}"
        BMS_Json = {"bots": self.formBotsJsons(botids), "missions": self.formMissionsJsons(mids), "skills": self.formSkillsJsons(skids)}

        return json.dumps(BMS_Json)

    def formBotsMissionsSkillsJsonData(self, botids, mids, skids):
        return self.formBotsJsons(botids),self.formMissionsJsons(mids),self.formSkillsJsons(skids)

    def getAllBotidsMidsSkidsFromTaskGroup(self, task_group):
        # bids = []
        # mids = []
        # for key, value in task_group.items():
        #     if isinstance(value, list) and len(value) > 0:
        #         for assignment in value:
        #             bids.append(assignment["bid"])
        #             for work in assignment["bw_works"]:
        #                 mids.append(work["mid"])
        #             for work in assignment["other_works"]:
        #                 mids.append(work["mid"])

        bids = [task["bid"] for task in task_group]
        mids = [task["mid"] for task in task_group]

        # Convert the set back to a list
        bids_set = set(bids)
        bids = list(bids_set)

        mids_set = set(mids)
        mids = list(mids_set)

        # at this points all skills should have been fetched, dependencies analyzed and skills regenerated, so just gather them....
        needed_skills = []
        print("check m skills: " + json.dumps(mids))
        print("all mids: " + json.dumps([m.getMid() for m in self.missions]))
        for mid in mids:
            m = next((mission for i, mission in enumerate(self.missions) if mission.getMid() == mid), None)

            if m:
                print("m skillls: ", mid, m.getMid(), m.getSkills(), type(m.getSkills()))
                if isinstance(m.getSkills(), list):
                    m_skids = m.getSkills()
                else:
                    m_skids = [int(skstring.strip()) for skstring in m.getSkills().strip().split(",")]
                print("m_skids: ", m_skids)
                if m_skids:
                    needed_skills = needed_skills + m_skids
                    m_main_skid = m_skids[0]

                    m_main_skill = next((sk for i, sk in enumerate(self.skills) if sk.getSkid() == m_main_skid), None)
                    if m_main_skill:
                        print("found skill")
                        needed_skills = needed_skills + m_main_skill.getDependencies()
                        print("needed skills add dependencies", m_main_skill.getDependencies())
                    else:
                        self.showMsg("ERROR: skill id not found - " + str(m_main_skid))
                else:
                    self.showMsg("ERROR: mission has no skill "+str(mid))
            else:
                self.showMsg("ERROR: mission ID not found " + str(mid))

        if len(needed_skills) > 0:
            skill_set = set(needed_skills)
            skids = list(skill_set)
        else:
            skids = []
        self.showMsg("bids in the task group:: "+json.dumps(bids))
        self.showMsg("mids in the task group:: "+json.dumps(mids))
        self.showMsg("skids in the task group:: " + json.dumps(skids))
        return bids, mids, skids

    # assumption, tg will not be empty.
    def getTaskGroupOS(self, tg):
        # get the 1st mission, and get its cuspas and extract platform part of the cuspas.
        for tz in tg.keys():
            if len(tg[tz]) > 0:
                # if len(tg[tz][0]["bw_works"]) > 0:
                #     mission_id = tg[tz][0]["bw_works"][0]["mid"]
                # else:
                #     mission_id = tg[tz][0]["other_works"][0]["mid"]
                #
                # midx = next((i for i, mission in enumerate(self.missions) if str(mission.getMid()) == mission_id), -1)
                # platform = self.missions[midx].getPlatform()
                platform = tg[tz][0]["cuspas"]
                break
        self.showMsg("Platform of the group:: "+platform)
        return platform

    # flatten all tasks associated with a vehicle.
    def flattenTaskGroup(self, vTasks):
        try:
            tgbs = []

            # flatten across time zone
            for tz in vTasks.keys():
                tgbs = tgbs + vTasks[tz]

            all_works = []

            for tgb in tgbs:
                bid = tgb["bid"]

                for bw in tgb["bw_works"]:
                    bw["bid"] = bid
                    all_works.append(bw)

                for other in tgb["other_works"]:
                    other["bid"] = bid
                    all_works.append(other)

            self.showMsg("after flatten and aggregation, total of "+str(len(all_works))+"tasks in this group!")
            time_ordered_works = sorted(all_works, key=lambda x: x["start_time"], reverse=False)
        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorFlattenTasks:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorFlattenTasks: traceback information not available:" + str(e)

        return time_ordered_works


    def groupTaskGroupsByOS(self, tgs):
        result = {
            "win": [tg for tg in tgs if "win" in self.getTaskGroupOS(tg)],
            "mac": [tg for tg in tgs if "mac" in self.getTaskGroupOS(tg)],
            "linux": [tg for tg in tgs if "linux" in self.getTaskGroupOS(tg)]
        }
        return result

    # note there could be schedule conflict here, because on cloud side, the schedule are assigned sequentially without knowing
    # which bot on which vehicle, if 2 bots are on the same vehicle, cloud doesn't know it and could assign them to two
    # vehicle group and cause them to be assigned with the same time slot. Q: should cloud side change assignment algorithm?
    # or should time assignment be done locally anyways? should bot cloud DB includes vehicle info? if cloud side includes
    # vehicle info, what algorithm should it be?
    def reGroupByBotVehicles(self, tgs):
        vtgs = {}
        for vehicle in tgs.keys():
            vtasks = self.flattenTaskGroup(tgs[vehicle])
            # for vtask in vtasks:
            #     found_bot = next((b for i, b in enumerate(self.bots) if b.getBid() == vtask["bid"]), None)
            #     bot_v = found_bot.getVehicle()
            #     print("bot_v:", bot_v, "task v:", vehicle)
            vtgs[vehicle] = vtasks
        return vtgs


    def getUnassignedVehiclesByOS(self):
        self.showMsg("N vehicles " + str(len(self.vehicles)))
        result = {
            "win": [v for v in self.vehicles if v.getOS().lower() in "Windows".lower() and len(v.getBotIds()) == 0],
            "mac": [v for v in self.vehicles if v.getOS().lower() in "Mac".lower() and len(v.getBotIds()) == 0],
            "linux": [v for v in self.vehicles if v.getOS().lower() in "Linux".lower() and len(v.getBotIds()) == 0]
        }
        self.showMsg("N vehicles win " + str(len(result["win"]))+" " + str(len(result["mac"]))+" " + str(len(result["linux"])))
        if self.host_role == "Commander" and not self.rpa_work_assigned_for_today:
            print("checking commander", self.todays_work["tbd"])
            if len([wk for wk in self.todays_work["tbd"] if wk["name"] == "automation"]) == 0:
                self.showMsg("myself unassigned "+self.getIP())
                # put in a dummy V
                self_v = VEHICLE(self)
                self_v.setIP(self.getIP())
                ipfields = self.getIP().split(".")
                ip = ipfields[len(ipfields) - 1]
                self_v.setVid(ip)
                if self.platform == "win":
                    self.showMsg("add myself to win based v list")
                    result["win"].insert(0, self_v)
                elif self.platform == "mac":
                    self.showMsg("add myself to mac based v list")
                    result["mac"].insert(0, self_v)
                else:
                    self.showMsg("add myself to linux based v list")
                    result["linux"].insert(0, self_v)

        return result


    def groupVehiclesByOS(self):
        self.showMsg("groupVehiclesByOS>>>>>>>>>>>> "+self.host_role)
        result = {
            "win": [v for v in self.vehicles if v.getOS() == "Windows"],
            "mac": [v for v in self.vehicles if v.getOS() == "Mac"],
            "linux": [v for v in self.vehicles if v.getOS() == "Linux"]
        }
        self.showMsg("all vehicles>>>>>>>>>>>> " + json.dumps(result))
        self.showMsg("now take care of commander machine itself in case of being a dual role commander")
        if self.host_role == "Commander":
            self.showMsg("checking commander>>>>>>>>>>>>>>>>>>>>>>>>> "+self.getIP())
            # put in a dummy V
            self_v = VEHICLE(self)
            self_v.setIP(self.getIP())
            ipfields = self.getIP().split(".")
            ip = ipfields[len(ipfields) - 1]
            self_v.setVid(ip)
            if self.platform == "win":
                result["win"].insert(0, self_v)
            elif self.platform == "mac":
                result["mac"].insert(0, self_v)
            else:
                result["linux"].insert(0, self_v)

        return result


    # generate a buy associated browse-search configuration
    # on the cloud side, a search config should have been attached to the buy_*** mission,
    # in case we just started execute a buy mission (i.e. the first steps are addcart or
    # addcart and pay, then:
    # 0) randomly select a search to swap the actual buy-related search.
    # 1) expand search result pages to 5 pages (we'll search up to 5 pages for the designated product.
    # 2) on each result pages of this selected search, make a to-be-opened product,(sel_type "cus" and add purchase)
    #    to the to-be-opened products list.
    # in case we are onto the later stage of buy (such as check shipping status, feedback etc.)
    # we would simply
    #  0) add "purchase" to first product on the first porduct list page of the selected search.
    # this would trigger the skill to go directly to the account's orders list and perform the buy
    # step.
    def gen_new_buy_search(self, work, mission):
        # simply modify mission's search configuration to fit our need.
        # we'll randomely pick one of the searches and modify its parameter.
        log3(f"gen buy related search...", "buyConfig", self)
        nth_search = random.randrange(0, len(work["config"]["searches"]))
        n_pages = len(work["config"]["searches"][nth_search]["prodlist_pages"])
        log3(f"nth_search-{nth_search}", "buyConfig", self)

        work["config"]["searches"][nth_search]["entry_paths"]["type"] = "Search"
        work["config"]["searches"][nth_search]["entry_paths"]["words"] = [mission.getSearchKW()]

        # simply duplate the last prodlist_pages enough times to satisfy up to 5 pages requreiment
        if work["name"].split("_")[1] in ["addCart", "addCartPay"]:
            last_page = work["config"]["searches"][nth_search]["prodlist_pages"][n_pages-1]
            if n_pages < 5:  # we will browse up to 5 pages for a product purchase.
                for i in range(5-n_pages):
                    work["config"]["searches"][nth_search]["prodlist_pages"].append(copy.deepcopy(last_page))

            # on each pages, add the target buy product onto the list.
            for page in work["config"]["searches"][nth_search]["prodlist_pages"]:
                if work["name"].split("_")[1] in ["addCart", "addCartPay"]:
                    target_buy = {
                        "selType": "cus",   # this is key, the skill itself will do the swapping of search terms once it see "cus" here.
                        "detailLvl": 3,
                        "purchase": [{
                                    "action": work["name"].split("_")[1],
                                    "asin": mission.getASIN(),
                                    "seller": mission.getStore(),
                                    "brand": mission.getBrand(),
                                    "img": mission.getImagePath(),
                                    "title": mission.getTitle(),
                                    "variations": mission.getVariations(),
                                    "rating": mission.getRating(),
                                    "feedbacks": mission.getFeedbacks(),
                                    "price": mission.getPrice(),
                                    "follow_seller": mission.getFollowSeller(),
                                    "follow_price": mission.getFollowPrice()
                                }]
                    }
                log3(f"added target buy cart pay: {target_buy}", "buyConfig", self)
                page["products"].append(target_buy)

        elif work["name"].split("_")[1] in ["pay", "checkShipping", "rate", "feedback", "checkFB"]:
            # in all other case, simply replace last st product of the 1st page.
            first_page = work["config"]["searches"][nth_search]["prodlist_pages"][0]
            first_page["products"][0] = {
                        "selType": "cus",  # this is key,
                        "detailLvl": 0,
                        "purchase": [
                            {
                                "action": work["name"].split("_")[1],
                                "asin": mission.getASIN(),
                                "seller": mission.getStore(),
                                "brand": mission.getBrand(),
                                "img": mission.getImagePath(),
                                "title": mission.getTitle(),
                                "variations": mission.getVariations(),
                                "rating": mission.getRating(),
                                "feedbacks": mission.getFeedbacks(),
                                "price": mission.getPrice(),
                                "order_id": mission.getOrderID(),
                                "feedback_rating": mission.getFeedbackRating(),
                                "feedback_title": mission.getFeedbackTitle(),
                                "feedback_text": mission.getFeedbackText(),
                                "feedback_image": mission.getFeedbackImgLink(),
                                "feedback_video": mission.getFeedbackVideoLink(),
                                "feedback_instructions": mission.getFeedbackInstructions(),
                                "follow_seller": mission.getFollowSeller(),
                                "follow_price": mission.getFollowPrice()
                            }
                        ]
                    }
            log3(f"set up run time swap with buy related search to replace cloud search {first_page['products'][0]}", "buyConfig", self)
        log3("Modified Buy Search Work:"+json.dumps(work), "buyConfig", self)


    def findWorkFromMission(self, mission):
        found = False
        foundWork = None
        if self.todaysSchedule():
            for vname in self.todaysSchedule["task_groups"]:
                tzs = self.todaysSchedule["task_groups"][vname].keys()
                for tz in tzs:
                    for w in self.todaysSchedule["task_groups"][vname][tz]["bw_works"]:
                        if w["mid"] == mission.getMid():
                            found = True
                            foundWork = w
                            break

                    if not found:
                        for w in self.todaysSchedule["task_groups"][vname][tz]["other_works"]:
                            if w["mid"] == mission.getMid():
                                found = True
                                foundWork = w
                                break

                    else:
                        break
                if found:
                    break

        return foundWork

    def gen_random_search_term(self, mission):
        search_settings = self.config_manager.search_settings.data
        main_cats = list(search_settings["search_terms"]["amz"].keys())
        main_cat_idx = random.randint(0, len(main_cats))
        main_cat = main_cats[main_cat_idx]
        sub1_cats = list(search_settings["search_terms"]["amz"][main_cat].keys())
        sub1_cat_idx = random.randint(0, len(sub1_cats))
        sub1_cat = sub1_cats[sub1_cat_idx]
        terms = search_settings["search_terms"]["amz"][main_cat][sub1_cat]
        terms_idx = random.randint(0, len(terms))
        search_term = terms[terms_idx]
        return search_term

    def gen_random_product_params(self, mission):
        search_settings = self.config_manager.search_settings.data
        random_st_idx = random.randint(0, len(search_settings["selType_selections"]))
        random_dl_idx = random.randint(0, len(search_settings["detailLvl_selections"]))
        product_params = {
            "selType": search_settings["selType_selections"][random_st_idx],
            "detailLvl": search_settings["selType_selections"][random_dl_idx],
            "purchase": []
        }

        return product_params

    def gen_random_page_params(self, mission):
        search_settings = self.config_manager.search_settings.data
        random_flow_idx = random.randint(0, len(search_settings["flow_selections"]))
        pg_params = {
            "flow_type": search_settings["flow_selections"][random_flow_idx],
            "products": []
        }
        nProducts = random.randint(1, search_settings["max_browse_products_per_page"]+1)
        for n in range(nProducts):
            productConfig = self.gen_random_product_params(mission)
            pg_params["products"].append(productConfig)

        return pg_params

    def gen_random_search_params(self, mission):
        search = {
            "type": "browse_routine",
            "site": "amz",
            "os": "win",
            "app": "ads",
            "entry_paths": {
                "type": "Search",
                "words": [self.gen_random_search_term(mission)]
            },
            "top_menu_item": "",
            "prodlist_pages": [],
            "buy_cfg": None
        }
        search_settings = self.config_manager.search_settings.data
        nPages = random.randint(1, search_settings["max_browse_pages"]+1)
        for n in range(nPages):
            pageConfig = self.gen_random_page_params(mission)
            search["prodlist_pages"].append(pageConfig)

        return search

    def gen_random_search_config(self, mission):
        config = {"estRunTime": 1, "searches": []}
        search_settings = self.config_manager.search_settings.data
        nSearches = random.randint(1, search_settings["max_searches"]+1)
        log3(f"gen nsearches:{nSearches}", "buyConfig", self)
        for n in range(nSearches):
            search = self.gen_random_search_params(mission)
            config["searches"].append(search)
        return config

    def gen_buy_search_config(self, mission):
        log3(f"cofigure buy search {mission.getMid()}", "buyConfig", self)
        work = self.findWorkFromMission(mission)
        work["config"] = self.gen_random_search_config(mission)

        # simply modify mission's search configuration to fit our need.
        # we'll randomely pick one of the searches and modify its parameter.
        nth_search = random.randrange(0, len(work["config"]["searches"]))
        n_pages = len(work["config"]["searches"][nth_search]["prodlist_pages"])

        work["config"]["searches"][nth_search]["entry_paths"]["type"] = "Search"
        work["config"]["searches"][nth_search]["entry_paths"]["words"] = [mission.getSearchKW()]

        # simply duplate the last prodlist_pages enough times to satisfy up to 5 pages requirement
        if work["name"].split("_")[1] in ["addCart", "addCartPay"]:
            last_page = work["config"]["searches"][nth_search]["prodlist_pages"][n_pages-1]
            if n_pages < 5:  # we will browse up to 5 pages for a product purchase.
                for i in range(5-n_pages):
                    work["config"]["searches"][nth_search]["prodlist_pages"].append(copy.deepcopy(last_page))

            # on each pages, add the target buy product onto the list.
            for page in work["config"]["searches"][nth_search]["prodlist_pages"]:
                if work["name"].split("_")[1] in ["addCart", "addCartPay"]:
                    target_buy = {
                        "selType": "cus",   # this is key, the skill itself will do the swapping of search terms once it see "cus" here.
                        "detailLvl": 3,
                        "purchase": [{
                                    "action": work["name"].split("_")[1],
                                    "asin": mission.getASIN(),
                                    "seller": mission.getStore(),
                                    "brand": mission.getBrand(),
                                    "img": mission.getImagePath(),
                                    "title": mission.getTitle(),
                                    "variations": mission.getVariations(),
                                    "rating": mission.getRating(),
                                    "feedbacks": mission.getFeedbacks(),
                                    "price": mission.getPrice(),
                                    "follow_seller": mission.getFollowSeller(),
                                    "follow_price": mission.getFollowPrice()
                                }]
                    }
                log3(f"added target buy: {target_buy}", "buyConfig", self)
                page["products"].append(target_buy)

        mission.setConfig(work["config"])
        log3("Modified Buy Work:"+json.dumps(work), "buyConfig", self)


    def gen_prod_sel(self):
        idx = math.floor(random.random() * (len(self.static_resource.PRODUCT_SEL_TYPES.length) - 1))
        return self.static_resource.PRODUCT_SEL_TYPES[idx]


    # given a derived buy mission, find out the original buy mission that was put in order by the users.
    # this is done thru searching ticket number. since this is likely to be a mission created 2 wks ago,
    # might not be loaded from memory, so directly search DB.
    def find_original_buy(self, buy_mission):
        # Construct the SQL query with a parameterized IN clause
        if buy_mission.getTicket() == 0:
            print("original buy mission ticket")
            # this is test mode special ticket, so provide some test vector.
            original_buy_mission = EBMISSION(self)
            original_buy_mission.setMid(0)
            original_buy_mission.setASIN("B0D1BY5VTM")
            original_buy_mission.setStore("Tikom")
            original_buy_mission.setFollowSeller("")
            original_buy_mission.setBrand("Tikom")
            original_buy_mission.setImagePath("")
            original_buy_mission.setSearchKW("dumb bells")
            original_buy_mission.setTitle("Tikom Robot Vacuum and Mop Combo with LiDAR Navigation, L9000 Robotic Vacuum Cleaner with 4000Pa Suction,150Min Max, 14 No-Go Zones, Smart Mapping, Good for Pet Hair, Carpet, Hard Floor")
            original_buy_mission.setVariations("")
            original_buy_mission.setRating("5.0")
            original_buy_mission.setFeedbacks("23")
            original_buy_mission.setPrice(229.99)
            original_buy_mission.setFollowPrice(0.0)
            original_buy_mission.setCustomerID("")
        else:
            db_data = self.mission_service.find_missions_by_ticket(buy_mission.getTicket())
            print("buy mission ticket:", buy_mission.getTicket())
            self.showMsg("same ticket missions: " + json.dumps(db_data.to_dict()))
            if len(db_data) != 0:
                original_buy_mission = EBMISSION(self)
                original_buy_mission.loadDBData(db_data)
                self.missions.append(original_buy_mission)
                # missionModel removed - UI components no longer needed
            else:
                original_buy_mission = None

        return original_buy_mission


    # if function will add buy task related search if there is any 1st stage buy type of missions. (Note a buy mission will always have a same CUSPUS browse action go along with it.
    # will go into the configuration of the browse mission, if there is a keyword search run, go the last one, and swap out the auto assigned
    # search phrase and replace with the buy search phrase. If there is no keyword search run, then simply create one and replace whatever the last
    # search with the buy related prodcut search flow, when we complete the mission and report the status, we'll do it just as the original browse
    # mission is done. This way, the cloud side will have no idea what's being processed.
    # in case  there is no same CUSPAS browse mission go along with a buy type, create one anyways, but this could affect capacity.
    # so really, we need cloud side to coordinate the buy-bowse coupling which I think it's there...
    # task name will be mainType_subType for example buy_addCart or goodFB_pay
    # main types will be: "buy", "goodFB", "badFB", "goodRating", "badRating"
    # sub types will be: 'addCart', 'pay', "checkShipping", 'rate', 'feedback', "checkFB"
    # 06-07-2024 actually not add, but again replace the configuration, otherwise, time will be wasted...
    def add_buy_searchs(self, p_task_groups):
        print("add buy to taskgroup:", p_task_groups)

        #1st find all 1st stage buy missions.
        self.showMsg("task name:" + json.dumps([tsk["name"]  for tsk in p_task_groups]))
        buys = [tsk for tsk in p_task_groups if (tsk["name"].split("_")[0] in self.static_resource.BUY_TYPES)]
        initial_buys = []
        later_buys = []
        for buy in buys:
            buy_parts = buy["name"].split("_")
            if len(buy_parts) > 1:
                if buy_parts[1] in ['addCart', 'pay', 'addCartPay']:
                    initial_buys.append(buy)
                else:
                    later_buys.append(buy)

        print(f"# buys:{len(buys)}, {len(initial_buys)}, {len(later_buys)}")
        for buytask in buys:
            # make sure we do search before buy
            midx = next( (i for i, mission in enumerate(self.missions) if str(mission.getMid()) == str(buytask["mid"])), -1)
            if midx >= 0:
                task_mission = self.missions[midx]
                original_buy = self.find_original_buy(task_mission)
                # first, fill the mission with original buy's private attributes for convenience.
                if original_buy:
                    task_mission.setASIN(original_buy.getASIN())
                    task_mission.setTitle(original_buy.getTitle())
                    task_mission.setVariations(original_buy.getVariations())
                    task_mission.setFollowSeller(original_buy.getFollowSeller())
                    task_mission.setStore(original_buy.getStore())
                    task_mission.setBrand(original_buy.getBrand())
                    task_mission.setImagePath(original_buy.getImagePath())
                    task_mission.setRating(original_buy.getRating())
                    task_mission.setFeedbacks(original_buy.getFeedbacks())
                    task_mission.setPrice(original_buy.getPrice())
                    task_mission.setFollowPrice(original_buy.getFollowPrice())
                    task_mission.setResult(original_buy.getResult())
                    task_mission.setSearchKW(original_buy.getSearchKW())

                    self.gen_new_buy_search(buytask, task_mission)
                else:
                    log3("ERROR: could NOT find original buy mission!")
            else:
                log3(f"buy mission not found {midx} {buytask['mid']}")
    # 1) group vehicle based on OS
    # 2) matche unassigned task group to vehicle based on OS.
    # 3) generate ADS profile xls for bots on that vehicle.
    # 4) modify task in case of buy related task....
    # 5) empower that vehicle with bots(including profiles), missions, tasks, skills
    # SC-06/27/2024 this algorithm asssumes, any bots can run on any vehicle as long as role, skill platform matches.
    # but this could be aggressive, bots and vehicles relationship could be fixed. in that case, we'll need a different
    # algorithm. which leas to assignWork2() where a vehicle-bot relationship will be read out at the beginning and
    # maintained in constant. when work group is scheduled. we will regroups bots' vehicle, and then generated associated
    # ads profiles, bots, missions, tasks, skills file.....
    # otherwise, send works to platoons to execute.
    async def assignWork(self):
        # tasks should already be sorted by botid,
        try:
            nsites = 0

            v_groups = self.getUnassignedVehiclesByOS()                      #result will {"win": win_vs, "mac": mac_vs, "linux": linux_vs}
            # print some debug info.
            for key in v_groups:
                log3("num vehicles in "+key+" :"+str(len(v_groups[key])), "assignWork", self)
                if len(v_groups[key]) > 0:
                    for k, v in enumerate(v_groups[key]):
                        log3("Vehicle OS:"+key+"["+str(k)+"]"+json.dumps(v.genJson())+"\n", "assignWork", self)
            print("assigning work....")

            # log3("unassigned_scheduled_task_groups: "+json.dumps(self.unassigned_scheduled_task_groups), "assignWork", self)
            tbd_unassigned = []
            for vname in self.unassigned_scheduled_task_groups:
                log3("assignwork scheduled checking vehicle: "+vname, "assignWork", self)
                p_task_groups = self.unassigned_scheduled_task_groups[vname]      # flattend per vehicle tasks.
                # log3("p_task_groups: "+json.dumps(p_task_groups), "assignWork", self)
                # print("p_task_groups:", p_task_groups)
                if len(p_task_groups) > 0:
                    print("some work to assign...", self.machine_name)
                    if self.machine_name in vname:
                        vehicle = self.getVehicleByName(vname)

                        if vehicle:
                            # if commander participate work, give the first(0th) work to self.

                            # in case this is an e-commerce work that requires finger print browser, then prepare here.
                            # all_works = [work for tg in p_task_groups for work in tg.get("works", [])]
                            # SC - at this point, p_task_groups should already be a flattened list of tasks
                            batched_tasks, ads_profiles = formADSProfileBatchesFor1Vehicle(p_task_groups, vehicle, self)
                            # batched_tasks now contains the flattened tasks in a vehicle, sorted by start_time, so no longer need complicated structure.
                            log3("arranged for today on this machine...."+vname, "assignWork", self)

                            # handle any buy-side tasks. - no long needs this, will let skill itself take care of it.
                            # self.add_buy_searchs(batched_tasks)

                            # current_tz, current_group = self.setTaskGroupInitialState(p_task_groups[0])
                            self.todays_work["tbd"].append({"name": "automation", "works": batched_tasks, "status": "yet to start", "current widx": 0, "vname": vname, "completed": [], "aborted": []})
                            vidx = 0
                            self.rpa_work_assigned_for_today = True
                            self.updateUnassigned("scheduled", vname, p_task_groups, tbd_unassigned)
                    else:
                        # vidx = i
                        vehicle = self.getVehicleByName(vname)
                        print("VVV:",vehicle)
                        log3("assign work for vehicle:"+vname, "assignWork", self)
                        if vehicle:
                            print("assign for other machine...", vname, vehicle.getVid(), vehicle.getStatus())

                            if not vehicle.getTestDisabled():
                                print("set up schedule for vehicle", vname)
                                await self.vehicleSetupWorkSchedule(vehicle, p_task_groups)
                                if "running" in vehicle.getStatus():
                                    self.updateUnassigned("scheduled", vname, p_task_groups, tbd_unassigned)
            if tbd_unassigned:
                log3("deleting alread assigned schedule task groups", "assignWork", self)
                for vname in tbd_unassigned:
                    del self.unassigned_scheduled_task_groups[vname]

                tbd_unassigned = []

            for vname in self.unassigned_reactive_task_groups:
                log3("assignwork reactive checking vehicle: "+vname, "assignWork", self)
                p_task_groups = self.unassigned_reactive_task_groups[vname]      # flattend per vehicle tasks.
                log3("p_task_groups: "+json.dumps(p_task_groups), "assignWork", self)
                if len(p_task_groups) > 0:

                    if self.machine_name in vname:
                        vehicle = self.getVehicleByName(vname)

                        if vehicle:
                            # if commander participate work, give the first(0th) work to self.

                            # in case this is an e-commerce work that requires finger print browser, then prepare here.
                            # all_works = [work for tg in p_task_groups for work in tg.get("works", [])]
                            # SC - at this point, p_task_groups should already be a flattened list of tasks
                            batched_tasks, ads_profiles = formADSProfileBatchesFor1Vehicle(p_task_groups, vehicle, self)
                            # batched_tasks now contains the flattened tasks in a vehicle, sorted by start_time, so no longer need complicated structure.
                            log3("arranged for today on this machine...."+vname, "assignWork", self)

                            #need to do some prep work here if the work needs to download certain files......
                            for tsk in batched_tasks:
                                if tsk["name"] == "sellFullfill_genECBLabels":
                                    tskMission = next((m for i, m in enumerate(self.missions) if m.getMid() == tsk["mid"]), None)
                                    self.downloadForFullfillGenECBLabels(tskMission.getConfig()[1], tsk['config'][1][0])

                            # current_tz, current_group = self.setTaskGroupInitialState(p_task_groups[0])
                            self.reactive_work["tbd"].append({"name": "automation", "works": batched_tasks, "status": "yet to start", "current widx": 0, "vname": vname, "completed": [], "aborted": []})
                            vidx = 0
                            self.rpa_work_assigned_for_today = True
                            self.updateUnassigned("reactive", vname, p_task_groups, tbd_unassigned)
                    else:

                        # vidx = i
                        vehicle = self.getVehicleByName(vname)
                        log3("assign reactive work for vehicle:"+vname, "assignWork", self)
                        if not vehicle.getTestDisabled():
                            await self.vehicleSetupWorkSchedule(vehicle, p_task_groups, False)
                            if "running" in vehicle.getStatus():
                                self.updateUnassigned("reactive", vname, p_task_groups, tbd_unassigned)

            if tbd_unassigned:
                log3("deleting alread assigned reactive task groups", "assignWork", self)
                for vname in tbd_unassigned:
                    del self.unassigned_reactive_task_groups[vname]

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorAssignWork:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorAssignWork: traceback information not available:" + str(e)
            log3(ex_stat, "assignWork", self)


    def getVehicleByName(self, vname):
        found_vehicle = next((v for i, v in enumerate(self.vehicles) if v.getName() == vname), None)
        return found_vehicle


    async def empower_platoon_with_skills(self, platoon_link, skill_ids):
        # at this point skilll PSK files should be ready to use, send these files to the platton so that can use them.
        for skid in skill_ids:
            found_skill = next((sk for i, sk in enumerate(self.skills) if sk.getSkid() == skid), None)
            if found_skill:
                if found_skill.getPrivacy() == "public":
                    psk_file = self.homepath + found_skill.getPskFileName()
                else:
                    psk_file = self.my_ecb_data_homepath + found_skill.getPskFileName()
                log3("Empowering platoon with skill PSK"+psk_file, "empower_platoon_with_skills", self)
                await self.send_file_to_platoon(platoon_link, "skill psk", psk_file)
            else:
                log3("ERROR: skid NOT FOUND [" + str(skid) + "]", "empower_platoon_with_skills", self)


    def setTaskGroupInitialState(self, tg):
        initial_tz = ""
        initial_group = ""
        for tz_key in tg:
            if len(tg[tz_key]) > 0:
                initial_tz = tz_key
                if len(tg[tz_key][0]['bw_works']) > 0:
                    initial_group = 'bw_works'
                else:
                    initial_group = 'other_works'
            break
        return initial_tz, initial_group


    # find to todos.,
    # 1) check whether need to fetch schedules, this is highest priority, it occurs at 2am, there should be no requested work anyways...
    # 2) checking whether need to do RPA
    #    2a) there are two queues: reactive and scheduled, reactive take higher priority since they are almost always
    #        customer request driven work.
    # the key data structure is self.todays_work["tbd"] which should be an array of either 1 or 2 elements.
    # either 1 or 2 elements depends on the role, if commander_only or platoon, will be 1 element,
    # if commander (which means commander can do tasks too) then there will be 2 elements.
    # in case of 1 element, it will be the actuall bot tasks to be done for platton or the fetch schedule task for Comander Only.
    # in case of 2 elements, the 0th element will be the fetch schedule, the 1st element will be the bot tasks(as a whole)
    # self.todays_work = {"tbd": [], "allstat": "working"}
    def checkNextToRun(self):
        log3("checkNextToRun: todays tbd... "+json.dumps(self.todays_work["tbd"]), "checkNextToRun", self)
        log3("checkNextToRun: reactive tbd... " + json.dumps(self.reactive_work["tbd"]), "checkNextToRun", self)
        nextrun = None
        runType = "scheduled"
        # go thru tasks and check the 1st task whose designated start_time has passed.
        pt = datetime.now()
        ten_hours = timedelta(hours=10)

        # Add 10 hours to the present date and time, some temp hack here to prevent something to run....
        pt = pt + ten_hours

        if len(self.todays_work["tbd"]) > 0:
            if ("Completed" not in self.todays_work["tbd"][0]["status"]) and (self.todays_work["tbd"][0]["name"] == "fetch schedule"):
                # in case the 1st todos is fetch schedule
                log3("set up first todo", "checkNextToRun", self)
                if self.ts2time(int(self.todays_work["tbd"][0]["works"][0]["start_time"]/1)) < pt:
                    log3("set up next run"+json.dumps(self.todays_work["tbd"][0]), "checkNextToRun", self)
                    nextrun = self.todays_work["tbd"][0]
            elif "Completed" not in self.todays_work["tbd"][0]["status"]:
                # in case the 1st todos is an automation task.
                log3("self.todays_work[\"tbd\"][0] : "+json.dumps(self.todays_work["tbd"][0]), "checkNextToRun", self)
                log3("time right now is: "+str(self.time2ts(pt)), "checkNextToRun", self)

                # determin next task group:

                if self.reactive_work["tbd"]:
                    nextrun = self.reactive_work["tbd"][0]
                else:
                    # check schedule work queue only when there is nothing in the reactive work queue
                    current_work_idx = self.todays_work["tbd"][0]["current widx"]

                    # if time is up to run the next work group,
                    if self.todays_work["tbd"][0]["works"]:
                        if self.ts2time(int(self.todays_work["tbd"][0]["works"][current_work_idx]["start_time"])) < pt:
                            log3("next run is now set up......", "checkNextToRun", self)
                            nextrun = self.todays_work["tbd"][0]
                        else:
                            nextrun = {}
                    else:
                        nextrun = {}
                log3("nextRUN>>>>>: "+json.dumps(nextrun), "checkNextToRun", self)
            else:
                log3("now today's schedule work are all finished, only serve the reactive work...", "checkNextToRun", self)
                if self.reactive_work["tbd"]:
                    nextrun = self.reactive_work["tbd"][0]
                log3("nextRUN reactive>>>>>: " + json.dumps(nextrun), "checkNextToRun", self)
        else:
            # if there is no schedule task to run, check whether there is reactive tasks to do, if so, do it asap.
            if self.reactive_work["tbd"]:
                log3("run contracted work netxt", "checkNextToRun", self)
                nextrun = self.reactive_work["tbd"][0]
                runType = "reactive"
            else:
                log3("no contract work to run", "checkNextToRun", self)
        return nextrun, runType

    def getNumUnassignedWork(self):
        num = 0
        for key in self.unassigned_scheduled_task_groups:
            num = num + len(self.unassigned_scheduled_task_groups[key])
        for key in self.unassigned_reactive_task_groups:
            num = num + len(self.unassigned_reactive_task_groups[key])
        return num


    def checkToDos(self):
        log3("checking todos...... "+json.dumps(self.todays_work["tbd"]), "checkToDos", self)
        nextrun = None
        # go thru tasks and check the 1st task whose designated start_time has passed.
        pt = datetime.now()
        ten_hours = timedelta(hours=10)

        # Add 10 hours to the current date and time
        pt = pt + ten_hours
        if len(self.todays_work["tbd"]) > 0:
            if ("Completed" not in self.todays_work["tbd"][0]["status"]) and (self.todays_work["tbd"][0]["name"] == "fetch schedule"):
                # in case the 1st todos is fetch schedule
                log3("checking fetch time", "checkToDos", self)
                if self.ts2time(int(self.todays_work["tbd"][0]["works"]["eastern"][0]["other_works"][0]["start_time"]/1)) < pt:
                    nextrun = self.todays_work["tbd"][0]
            elif "Completed" not in self.todays_work["tbd"][0]["status"]:
                # in case the 1st todos is an automation task.
                log3("eastern:"+json.dumps(self.todays_work["tbd"][0]["works"]["eastern"]), "checkToDos", self)
                log3("self.todays_work[\"tbd\"][0] : "+json.dumps(self.todays_work["tbd"][0]), "checkToDos", self)
                tz = self.todays_work["tbd"][0]["current tz"]

                bith = self.todays_work["tbd"][0]["current bidx"]

                # determin next task group:
                current_bw_idx = self.todays_work["tbd"][0]["current widx"]
                current_other_idx = self.todays_work["tbd"][0]["current oidx"]
                log3("time right now is: "+self.time2ts(pt)+"("+str(pt)+")"+datetime.now().strftime('%Y-%m-%d %H:%M:%S')+" tz:"+tz+" bith:"+str(bith)+" bw idx:"+str(current_bw_idx)+"other idx:"+str(current_other_idx), "checkToDos", self)

                if current_bw_idx < len(self.todays_work["tbd"][0]["works"][tz][bith]["bw_works"]):
                    current_bw_start_time = self.todays_work["tbd"][0]["works"][tz][bith]["bw_works"][current_bw_idx]["start_time"]
                    log3("current bw start time: " + str(current_bw_start_time), "checkToDos", self)
                else:
                    # just give it a huge number so that, this group won't get run
                    current_bw_start_time = 1000
                log3("current_bw_start_time: "+str(current_bw_start_time), "checkToDos", self)

                if current_other_idx < len(self.todays_work["tbd"][0]["works"][tz][bith]["other_works"]):
                    current_other_start_time = self.todays_work["tbd"][0]["works"][tz][bith]["other_works"][current_other_idx]["start_time"]
                    log3("current bw start time: " + str(current_other_start_time), "checkToDos", self)
                else:
                    # in case, all just give it a huge number so that, this group won't get run
                    current_other_start_time = 1000
                log3("current_other_start_time: "+current_other_start_time, "checkToDos", self)

                # if a buy-walk task is scheduled earlier than other tasks, arrange the buy-walk task, otherwise arrange other works.
                if current_bw_start_time < current_other_start_time:
                    grp = "bw_works"
                    wjth = current_bw_idx
                elif current_bw_start_time > current_other_start_time:
                    grp = "other_works"
                    wjth = current_other_idx
                else:
                    # if both gets 1000 value, that means there is nothing to run.
                    grp = ""
                    wjth = -1

                self.todays_work["tbd"][0]["current grp"] = grp
                log3("tz: "+tz+" bith: "+str(bith)+" grp: "+grp+" wjth: "+str(wjth), "checkToDos", self)

                if wjth >= 0:
                    if self.ts2time(int(self.todays_work["tbd"][0]["works"][tz][bith][grp][wjth]["start_time"]/3)) < pt:
                        self.showMsg("next run is now set up......")
                        nextrun = self.todays_work["tbd"][0]
                log3("nextRUN>>>>>: "+json.dumps(nextrun), "checkToDos", self)
        return nextrun


    def findWorksToBeRetried(self, todos):
        retries = copy.deepcopy(todos)
        log3("MISSIONS needs retry: "+str(retries), "findWorksToBeRetried", self)
        return retries

    def findMissonsToBeRetried(self, todos):
        retries = copy.deepcopy(todos)
        for key1, value1 in todos.items():
            # regions
            if isinstance(value1, dict):
                for key2, value2 in value1.items():
                    # botids
                    if isinstance(value2, dict):
                        for key3, value3 in value2.items():
                            # groups
                            if isinstance(value3, dict):
                                for key4, value4 in value3.items():
                                    # missions
                                    if isinstance(value4, dict):
                                        if "Completed" in value4["status"]:
                                            junk = retries[key1][key2][key3].pop(key4)

        #now point to the 1st item in this todo list

        log3("MISSIONS needs retry: "+str(retries), "findMissonsToBeRetried", self)
        return retries

    def flatten_todos(self, todos):
        all_missions = {}
        for key1, value1 in todos.items():
            # regions
            if isinstance(value1, dict):
                for key2, value2 in value1.items():
                    # botids
                    if isinstance(value2, dict):
                        for key3, value3 in value2.items():
                            # groups
                            if isinstance(value3, dict):
                                all_missions.update(value3)
        return all_missions


    def loadSkillFile(self, skname, pub):
        #slap on a file path prefix, then read in the file.
        skillsubnames = skname.split("_")
        actionname = ''.join(skillsubnames[2:len(skillsubnames)])
        if pub:
            skill_file = self.homepath + "resource/skills/public/" + skname + "/scripts/" + skname + ".psk"
        else:
            skill_file = self.my_ecb_data_homepath + "/my_skills/" + skname + "/scripts/" + skname + ".psk"

        log3("loadSKILLFILE: "+skill_file, "loadSkillFile", self)
        stepKeys = readPSkillFile(skname, skill_file, lvl=0)

        return stepKeys


    # fill in real address to some placeholders
    def reAddrAndUpdateSteps(self, pskJson, init_step_idx, work_settings):
        # self.showMsg("PSK JSON::::: "+json.dumps(pskJson))
        newPskJson = {}
        log3("New Index:"+str(init_step_idx), "reAddrAndUpdateSteps", self)
        new_idx = init_step_idx
        old_keys = list(pskJson.keys())
        for key in old_keys:
            if "step" in key:
                new_key = "step "+str(new_idx)
                newPskJson[new_key] = pskJson[key]
                new_idx = new_idx + STEP_GAP
                # print("old/new key:", key, new_key, pskJson[key])
                if "Create Data" in newPskJson[new_key]['type']:
                    if newPskJson[new_key]['data_name'] == "sk_work_settings":
                        newPskJson[new_key]["key_value"] = work_settings
                        # newPskJson[new_key]["key_value"] = copy.deepcopy(work_settings)
                        # newPskJson[new_key]["key_value"]["commander_link"] = ""
                        log3("REPLACED WORKSETTINGS HERE: "+new_key+" :::: "+json.dumps(newPskJson[new_key]), "reAddrAndUpdateSteps", self)

                pskJson.pop(key)

        # self.showMsg("PSK JSON after address and update step::::: "+json.dumps(newPskJson))
        return new_idx, newPskJson

    # run one bot one time slot at a time，for 1 bot and 1 time slot, there should be only 1 mission running
    async def runRPA(self, worksTBD, rpa_msg_queue, monitor_msg_queue):
        global rpaConfig
        global skill_code

        all_done = False
        try:
            worksettings = getWorkRunSettings(self, worksTBD)
            mid2br = worksettings["mid"]

            if (not self.checkMissionAlreadyRun(worksettings)) or mid2br in self.general_settings.mids_forced_to_run:
                log3("worksettings: bid, mid "+str(worksettings["botid"])+" "+str(worksettings["mid"])+" "+str(worksettings["midx"])+" "+json.dumps([m.getFingerPrintProfile() for m in self.missions]), "runRPA", self)

                bot_idx = next((i for i, b in enumerate(self.bots) if str(b.getBid()) == str(worksettings["botid"])), -1)
                if bot_idx >= 0:
                    log3("found BOT to be run......"+str(self.bots[bot_idx].getEmail()), "runRPA", self)
                    running_bot = self.bots[bot_idx]

                rpaScripts = []

                # generate walk skills on the fly.
                self.running_mission = self.missions[worksettings["midx"]]

                # no finger print profile, no run for ads.
                if 'ads' in self.running_mission.getCusPAS() and self.running_mission.getFingerPrintProfile() == "":
                    log3("ERROR ADS mission has no profile: " + str(self.running_mission.getMid()) + " " + self.running_mission.getCusPAS() + " " + self.running_mission.getFingerPrintProfile(), "runRPA", self)
                    runResult = "ErrorRPA ADS mission has no profile " + str(self.running_mission.getMid())
                    self.update1MStat(worksettings, runResult)
                    self.update1WorkRunStatus(worksTBD, worksettings["midx"])
                else:
                    log3("current RUNNING MISSION: "+json.dumps(self.running_mission.genJson()), "runRPA", self)
                    log3("RPA all skill ids:"+json.dumps([sk.getSkid() for sk in self.skills]), "runRPA", self)
                    if self.running_mission.getSkills() != "":
                        rpaSkillIdWords = self.running_mission.getSkills().split(",")
                        log3("current RUNNING MISSION SKILL: "+json.dumps(self.running_mission.getSkills()), "runRPA", self)
                        rpaSkillIds = [int(skidword.strip()) for skidword in rpaSkillIdWords]

                        log3("rpaSkillIds: "+json.dumps(rpaSkillIds)+" "+str(type(rpaSkillIds[0]))+" "+" running mission id: "+str(self.running_mission.getMid()), "runRPA", self)

                        # get skills data structure by IDs
                        all_skids = [sk.getSkid() for sk in self.skills]
                        log3("all skills ids:"+json.dumps([sk.getSkid() for sk in self.skills]), "runRPA", self)
                        rpaSkillIds = list(dict.fromkeys(rpaSkillIds))
                        log3("rpaSkillIds:"+json.dumps(rpaSkillIds), "runRPA", self)

                        relevant_skills = [self.skills[all_skids.index(skid)] for skid in rpaSkillIds]

                        log3("N relevant skills:"+str(len(relevant_skills))+json.dumps([sk.getSkid() for sk in relevant_skills]), "runRPA", self)
                        relevant_skill_ids = [sk.getSkid() for sk in self.skills if sk.getSkid() in rpaSkillIds]
                        relevant_skill_ids = list(set(relevant_skill_ids))
                        log3("relevant skills ids: "+json.dumps(relevant_skill_ids), "runRPA", self)
                        dependent_skids=[]
                        for sk in relevant_skills:
                            log3("add dependency: " + json.dumps(sk.getDependencies()) + "for skill#" + str(sk.getSkid()), "runRPA", self)
                            dependent_skids = dependent_skids + sk.getDependencies()

                        dependent_skids = list(set(dependent_skids))
                        dependent_skids = [skid for skid in dependent_skids if skid not in relevant_skill_ids]
                        log3("all dependencies: "+json.dumps(dependent_skids), "runRPA", self)

                        dependent_skills = [sk for sk in self.skills if sk.getSkid() in dependent_skids]
                        relevant_skills = relevant_skills + dependent_skills
                        relevant_skill_ids = relevant_skill_ids + dependent_skids

                        if len(relevant_skill_ids) < len(rpaSkillIds):
                            s = set(relevant_skill_ids)
                            missing = [x for x in rpaSkillIds if x not in s]
                            log3("ERROR: Required Skills not found:"+json.dumps(missing), "runRPA", self)


                        log3("all skids involved in this skill: "+json.dumps([sk.getSkid() for sk in relevant_skills]), "runRPA", self)
                        all_skill_codes = []
                        step_idx = 0
                        for sk in relevant_skills:
                            log3("settingSKKKKKKKK: "+str(sk.getSkid())+" "+sk.getName()+" "+str(worksettings["b_email"]), "runRPA", self)
                            setWorkSettingsSkill(worksettings, sk)
                            # self.showMsg("settingSKKKKKKKK: "+json.dumps(worksettings, indent=4))

                            # readPSkillFile will remove comments. from the file
                            if sk.getPrivacy() == "public":
                                sk_dir = self.homepath
                            else:
                                sk_dir = self.my_ecb_data_homepath
                            pskJson = readPSkillFile(worksettings["name_space"], sk_dir+sk.getPskFileName(), lvl=0)
                            # self.showMsg("RAW PSK JSON::::"+json.dumps(pskJson))

                            # now regen address and update settings, after running, pskJson will be updated.
                            step_idx, pskJson = self.reAddrAndUpdateSteps(pskJson, step_idx, worksettings)
                            # self.showMsg("AFTER READDRESS AND UPDATE PSK JSON::::" + json.dumps(pskJson))

                            addNameSpaceToAddress(pskJson, worksettings["name_space"], lvl=0)

                            # self.showMsg("RUNNABLE PSK JSON::::"+json.dumps(pskJson))

                            # save the file to a .rsk file (runnable skill) which contains json only with comments stripped off from .psk file by the readSkillFile function.
                            rskFileName = sk_dir + sk.getPskFileName().split(".")[0] + ".rsk"
                            rskFileDir = os.path.dirname(rskFileName)
                            if not os.path.exists(rskFileDir):
                                os.makedirs(rskFileDir)
                            log3("rskFileName: "+rskFileName+" step_idx: "+str(step_idx), "runRPA", self)
                            with open(rskFileName, "w") as outfile:
                                json.dump(pskJson, outfile, indent=4)
                            outfile.close()

                            all_skill_codes.append({"ns": worksettings["name_space"], "skfile": rskFileName})

                        log3("all_skill_codes: "+json.dumps(all_skill_codes), "runRPA", self)

                        rpa_script = prepRunSkill(all_skill_codes)
                        # log3("generated ready2run: "+json.dumps(rpa_script), "runRPA", self)
                        # self.showMsg("generated psk: " + str(len(rpa_script.keys())))

                        # doing this just so that the code below can run multiple codes if needed. but in reality
                        # prepRunSkill put code in a global var "skill_code", even if there are multiple scripts,
                        # this has to be corrected because, the following append would just have multiple same
                        # skill_code...... SC, but for now this is OK, there is no multiple script scenario in
                        # forseaable future.
                        rpaScripts.append(rpa_script)
                        # self.showMsg("rpaScripts:["+str(len(rpaScripts))+"] "+json.dumps(rpaScripts))
                        log3("rpaScripts:["+str(len(rpaScripts))+"] "+str(len(relevant_skills))+" "+str(worksettings["midx"])+" "+str(len(self.missions)), "runRPA", self)

                        # Before running do the needed prep to get "fin" input parameters ready.
                        # this is the case when this mission is run as an independent server, the input
                        # of the mission will come from the another computer, and there might even be
                        # files to be downloaded first as the input to the mission.
                        if worksettings["as_server"]:
                            log3("SETTING MISSSION INPUT:"+json.dumps(self.running_mission.getConfig()), "runRPA", self)
                            setMissionInput(self.running_mission.getConfig())


                        # (steps, mission, skill, mode="normal"):
                        # it_items = (item for i, item in enumerate(self.skills) if item.getSkid() == rpaSkillIds[0])
                        # self.showMsg("it_items: "+json.dumps(it_items))
                        # for it in it_items:
                        #     self.showMsg("item: "+str(it.getSkid()))
                        # running_skill = next((item for i, item in enumerate(self.skills) if item.getSkid() == int(rpaSkillIds[0])), -1)
                        # self.showMsg("running skid:"+str(rpaSkillIds[0])+"len(self.skills): "+str(len(self.skills))+"skill 0 skid: "+str(self.skills[0].getSkid()))
                        # self.showMsg("running skill: "+json.dumps(running_skill))
                        # runStepsTask = asyncio.create_task(rpaRunAllSteps(rpa_script, self.missions[worksettings["midx"]], relevant_skills[0], rpa_msg_queue, monitor_msg_queue))
                        # runResult = await runStepsTask

                        log3("BEFORE RUN: " + worksettings["b_email"], "runRPA", self)
                        runResult = await rpaRunAllSteps(rpa_script, self.missions[worksettings["midx"]], relevant_skills[0], rpa_msg_queue, monitor_msg_queue)

                        # for retry test purpose:
                        # runResult = "Incomplete Error"

                        # finished 1 mission, update status and update pointer to the next one on the list.... and be done.
                        # the timer tick will trigger the run of the next mission on the list....
                        log3("UPDATEing completed mission status:: "+str(worksettings["midx"])+"RUN result:"+runResult, "runRPA", self)
                        self.update1MStat(worksettings, runResult)
                        self.update1WorkRunStatus(worksTBD, worksettings["midx"])
                    else:
                        log3("UPDATEing ERROR mmission status:: " + str(worksettings["midx"]) + "RUN result: " + "Incomplete: ERRORRunRPA:-1", "runRPA", self)
                        self.update1MStat(worksettings, "Incomplete: ERRORRunRPA:No Skill To Run")
                        self.update1WorkRunStatus(worksTBD, worksettings["midx"])
                        raise Exception('ERROR: NO SKILL TO RUN!')
            else:
                log3("mission already ran " + str(worksettings["mid"]), "runRPA", self)
                log3("mission ALREADY Completed today: " + str(worksettings["mid"]), "runRPA", self)
                runResult = "Completed:0 Skip Rerun"
                self.update1WorkRunStatus(worksTBD, worksettings["midx"])

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorRanRPA:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorRanRPA: traceback information not available:" + str(e)
            log3(ex_stat, "runRPA", self)
            runResult = "Incomplete: ERRORRunRPA:-1"

        log3("botid, mid:"+str(worksettings["botid"]) + " "+str(worksettings["mid"]), "runRPA", self)
        return worksettings["botid"], worksettings["mid"], runResult


    # run one bot one time slot at a time，for 1 bot and 1 time slot, there should be only 1 mission running
    async def run1ManagerMission(self, mission, self_in_queue, rpa_msg_queue, monitor_msg_queue):
        global rpaConfig
        global skill_code

        all_done = False
        try:
            worksettings = getWorkRunSettings(self, mission)
            log3("manager worksettings: bid, mid "+str(worksettings["botid"])+" "+str(worksettings["mid"])+" "+str(worksettings["midx"])+" "+json.dumps([m.getFingerPrintProfile() for m in self.missions]), "runRPA", self)

            print("manager work settings:", worksettings)
            rpaScripts = []

            # generate walk skills on the fly.
            self.running_manager_mission = mission

            # no finger print profile, no run for ads.
            if 'ads' in self.running_manager_mission.getCusPAS() and self.running_manager_mission.getFingerPrintProfile() == "":
                log3("ERROR ADS mission has no profile: " + str(self.running_manager_mission.getMid()) + " " + self.running_mission.getCusPAS() + " " + self.running_mission.getFingerPrintProfile(), "runRPA", self)
                runResult = "ErrorRPA ADS mission has no profile " + str(self.running_manager_mission.getMid())
            else:
                log3("current RUNNING MISSION: "+json.dumps(self.running_manager_mission.genJson()), "runRPA", self)
                log3("RPA all skill ids:"+json.dumps([sk.getSkid() for sk in self.skills]), "runRPA", self)
                if self.running_manager_mission.getSkills() != "":
                    rpaSkillIdWords = self.running_manager_mission.getSkills().split(",")
                    log3("current RUNNING MISSION SKILL: "+json.dumps(self.running_manager_mission.getSkills()), "runRPA", self)
                    rpaSkillIds = [int(skidword.strip()) for skidword in rpaSkillIdWords]

                    log3("rpaSkillIds: "+json.dumps(rpaSkillIds)+" "+str(type(rpaSkillIds[0]))+" "+" running mission id: "+str(self.running_manager_mission.getMid()), "runRPA", self)

                    # get skills data structure by IDs
                    all_skids = [sk.getSkid() for sk in self.skills]
                    log3("all skills ids:"+json.dumps([sk.getSkid() for sk in self.skills]), "runRPA", self)
                    rpaSkillIds = list(dict.fromkeys(rpaSkillIds))
                    log3("rpaSkillIds:"+json.dumps(rpaSkillIds), "runRPA", self)

                    relevant_skills = [self.skills[all_skids.index(skid)] for skid in rpaSkillIds]

                    log3("N relevant skills:"+str(len(relevant_skills))+json.dumps([sk.getSkid() for sk in relevant_skills]), "runRPA", self)
                    relevant_skill_ids = [sk.getSkid() for sk in self.skills if sk.getSkid() in rpaSkillIds]
                    relevant_skill_ids = list(set(relevant_skill_ids))
                    log3("relevant skills ids: "+json.dumps(relevant_skill_ids), "runRPA", self)
                    dependent_skids=[]
                    for sk in relevant_skills:
                        log3("add dependency: " + json.dumps(sk.getDependencies()) + "for skill#" + str(sk.getSkid()), "runRPA", self)
                        dependent_skids = dependent_skids + sk.getDependencies()

                    dependent_skids = list(set(dependent_skids))
                    dependent_skids = [skid for skid in dependent_skids if skid not in relevant_skill_ids]
                    log3("all dependencies: "+json.dumps(dependent_skids), "runRPA", self)

                    dependent_skills = [sk for sk in self.skills if sk.getSkid() in dependent_skids]
                    relevant_skills = relevant_skills + dependent_skills
                    relevant_skill_ids = relevant_skill_ids + dependent_skids

                    if len(relevant_skill_ids) < len(rpaSkillIds):
                        s = set(relevant_skill_ids)
                        missing = [x for x in rpaSkillIds if x not in s]
                        log3("ERROR: Required Skills not found:"+json.dumps(missing), "runRPA", self)


                    log3("all skids involved in this skill: "+json.dumps([sk.getSkid() for sk in relevant_skills]), "runRPA", self)
                    all_skill_codes = []
                    step_idx = 0
                    for sk in relevant_skills:
                        log3("settingSKKKKKKKK: "+str(sk.getSkid())+" "+sk.getName()+" "+str(worksettings["b_email"]), "runRPA", self)
                        setWorkSettingsSkill(worksettings, sk)
                        # self.showMsg("settingSKKKKKKKK: "+json.dumps(worksettings, indent=4))

                        # readPSkillFile will remove comments. from the file
                        if sk.getPrivacy() == "public":
                            sk_dir = self.homepath
                        else:
                            sk_dir = self.my_ecb_data_homepath
                        pskJson = readPSkillFile(worksettings["name_space"], sk_dir+sk.getPskFileName(), lvl=0)
                        # self.showMsg("RAW PSK JSON::::"+json.dumps(pskJson))

                        # now regen address and update settings, after running, pskJson will be updated.
                        step_idx, pskJson = self.reAddrAndUpdateSteps(pskJson, step_idx, worksettings)
                        # self.showMsg("AFTER READDRESS AND UPDATE PSK JSON::::" + json.dumps(pskJson))

                        addNameSpaceToAddress(pskJson, worksettings["name_space"], lvl=0)

                        # self.showMsg("RUNNABLE PSK JSON::::"+json.dumps(pskJson))

                        # save the file to a .rsk file (runnable skill) which contains json only with comments stripped off from .psk file by the readSkillFile function.
                        rskFileName = sk_dir + sk.getPskFileName().split(".")[0] + ".rsk"
                        rskFileDir = os.path.dirname(rskFileName)
                        if not os.path.exists(rskFileDir):
                            os.makedirs(rskFileDir)
                        log3("rskFileName: "+rskFileName+" step_idx: "+str(step_idx), "runRPA", self)
                        with open(rskFileName, "w") as outfile:
                            json.dump(pskJson, outfile, indent=4)
                        outfile.close()

                        all_skill_codes.append({"ns": worksettings["name_space"], "skfile": rskFileName})

                    log3("all_skill_codes: "+json.dumps(all_skill_codes), "runRPA", self)

                    rpa_script = prepRunSkill(all_skill_codes)
                    # log3("generated ready2run: "+json.dumps(rpa_script), "runRPA", self)
                    # self.showMsg("generated psk: " + str(len(rpa_script.keys())))

                    # doing this just so that the code below can run multiple codes if needed. but in reality
                    # prepRunSkill put code in a global var "skill_code", even if there are multiple scripts,
                    # this has to be corrected because, the following append would just have multiple same
                    # skill_code...... SC, but for now this is OK, there is no multiple script scenario in
                    # forseaable future.
                    rpaScripts.append(rpa_script)
                    # self.showMsg("rpaScripts:["+str(len(rpaScripts))+"] "+json.dumps(rpaScripts))
                    log3("rpaScripts:["+str(len(rpaScripts))+"] "+str(len(relevant_skills))+" "+str(worksettings["midx"])+" "+str(len(self.missions)), "runRPA", self)

                    # Before running do the needed prep to get "fin" input parameters ready.
                    # this is the case when this mission is run as an independent server, the input
                    # of the mission will come from the another computer, and there might even be
                    # files to be downloaded first as the input to the mission.
                    if worksettings["as_server"]:
                        log3("SETTING MISSSION INPUT:"+json.dumps(self.running_manager_mission.getConfig()), "runRPA", self)
                        setMissionInput(self.running_manager_mission.getConfig())


                    log3("MANAGER BEFORE RUN: " + worksettings["b_email"], "runRPA", self)
                    runResult = await rpaRunAllSteps(rpa_script, self.running_manager_mission, relevant_skills[0], rpa_msg_queue, monitor_msg_queue)


                    # finished 1 mission, update status and update pointer to the next one on the list.... and be done.
                    # the timer tick will trigger the run of the next mission on the list....
                    log3("UPDATEing 1 completed mmission status:: "+str(worksettings["midx"])+"RUN result:"+runResult, "runRPA", self)
                    mission.setResult(runResult)
                else:
                    log3("UPDATEing ERROR mmission status:: " + str(worksettings["midx"]) + "RUN result: " + "Incomplete: ERRORRunRPA:-1", "runRPA", self)
                    mission.setResult("Incomplete: ERRORRunRPA:-1")
                    raise Exception('ERROR: NO SKILL TO RUN!')


        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorRun1ManagerMission:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorRun1ManagerMission: traceback information not available:" + str(e)

            print(ex_stat)
            log3(ex_stat, "run1managerMission", self)
            runResult = "Incomplete: ERRORRunRPA:-1"

        log3("manager mission run result:"+json.dumps(runResult), "runRPA", self)
        return runResult

    def checkMissionAlreadyRun(self, worksettings):
        alreadyRun = False
        mid = str(worksettings["mid"])
        missionReportFile = worksettings["log_path_prefix"] + "run_result.json"
        print("check mission already run missionReportFile:", missionReportFile)
        if os.path.exists(missionReportFile):
            with open(missionReportFile, "r", encoding="utf-8") as mrf:
                m_report_json = json.load(mrf)
                if "Completed" in m_report_json[mid]:
                    alreadyRun = True
        return alreadyRun

    def save1MStatToFile(self, worksettings, result):
        # save mission run status to a local file. so that if re-run and we realized the mission
        # has already being completed, we don't run it again. of course we'd have a force run
        # setting in general settings, such that if set, that would overide it.
        mid = worksettings["mid"]
        mission = self.missions[worksettings["midx"]]
        missionReportFile = worksettings["log_path_prefix"]+"run_result.json"
        print("saving missionReportFile:", missionReportFile)
        # read-modify-write
        if os.path.exists(missionReportFile):
            with open(missionReportFile, "r", encoding="utf-8") as mrf:
                m_report_json = json.load(mrf)
            m_report_json[mid] = result
            with open(missionReportFile, "w", encoding="utf-8") as mrf:
                json.dump(m_report_json, mrf, indent=4)
        else:
            # no file yet, just write it.
            os.makedirs(worksettings["log_path_prefix"], exist_ok=True)
            with open(missionReportFile, "w", encoding="utf-8") as mrf:
                m_report_json = {mid: result}
                json.dump(m_report_json, mrf, indent=4)


    def update1MStat(self, worksettings, result):
        midx = worksettings["midx"]
        log3("1 mission run completed."+str(midx)+" "+str(self.missions[midx].getMid())+" "+str(self.missions[midx].getRetry())+" "+str(self.missions[midx].getNRetries())+"status:"+result, "update1MStat", self)
        self.missions[midx].setStatus(result)
        self.save1MStatToFile(worksettings, result)
        retry_count = self.missions[midx].getNRetries()
        self.missions[midx].setNRetries(retry_count + 1)
        log3("update1MStat:"+str(midx)+":"+str(self.missions[midx].getMid())+":"+str(self.missions[midx].getNRetries()), "update1MStat", self)
        bid = self.missions[midx].getBid()

        # if platoon send this updated info to commander boss
        if "Platoon" in self.host_role:
            self.sendCommanderMissionsStatMsg([self.missions[midx].getMid()])
            missionResultFiles = self.getMissionResultFileNames(self.missions[midx])
            if missionResultFiles:
                self.send_mission_result_files_to_commander(self.commanderXport, self.missions[midx].getMid(), "zip", missionResultFiles)
        elif "Commander" in self.host_role:
            self.updateMissionsStatToCloud([self.missions[midx]])

    # update mission status to the cloud db, and to local data structure, local DB, and to chat？
    # "mid": mid,
    # "botid": self.missions[mid].getBid(),
    # "sst": self.missions[mid].getEstimatedStartTime(),
    # "sd": self.missions[mid].getEstimatedRunTime(),
    # "ast": self.missions[mid].getActualStartTime(),
    # "aet": self.missions[mid].getActualEndTime(),
    # "status": m_stat,
    # "error": m_err
    def updateMStats(self, mStats):
        inMids = [m["mid"] for m in mStats]
        foundMissions = []
        for mstat in mStats:
            foundMission = next((m for i, m in enumerate(self.missions) if m.getMid() == mstat["mid"]), None)
            if foundMission:
                foundMission.setStatus= mstat.get("status", foundMission.getStatus())
                if "ast" in mstat:
                    mStartTime = mstat.get("ast", foundMission.getActualStartTime())
                    mStartDate = mstat.get("ast", foundMission.getAsd())
                    foundMission.setActualStartTime = mStartTime
                    foundMission.setAsd = mStartDate

                if "aet" in mstat:
                    mEndTime = mstat.get("aet", foundMission.getActualEndTime())
                    mEndDate = mstat.get("act", foundMission.getAcd())

                    foundMission.setActualEndTime = mEndTime
                    foundMission.setAcd = mEndDate

                foundMission.setError = mstat.get("error", foundMission.getError())
                foundMissions.append(foundMission)

        # update missions to cloud DB and local DB
        if foundMissions:
            self.updateMissions(foundMissions)

    # check where a mission is supposed to return any resulting files, if so, return the list of full path file names.
    def getMissionResultFileNames(self, mission):
        # this is really a case by case thing, the scheme is really in the mission.config
        # if there are any expected outfiles, they ought to be in mission.config.
        # and mission.config json format depends on the mission itself.
        mConfig = mission.getConfig()
        if "out_files" in mConfig:
            return mConfig["out_files"]
        else:
            return []


    def updateMissionsStatToCloud(self, missions):
        mstats = [{"mid": m.getMid(), "status": m.getStatus()} for m in missions]
        send_update_missions_ex_status_to_cloud(self.session, mstats, self.get_auth_token(), self.getWanApiEndPoint())


    def updateUnassigned(self, tg_type, vname, task_group, tbd):
        tg_mids = [tsk["mid"] for tsk in task_group]
        if tg_type == "scheduled":
            if vname in self.unassigned_scheduled_task_groups:
                # maybe an expensive way of remove a task from the group.
                self.unassigned_scheduled_task_groups[vname] = [tsk for tsk in self.unassigned_scheduled_task_groups[vname] if tsk['mid'] not in tg_mids]

                # if this vehcle no longer has unassigned work, then remove this vehicle from unassigned_task_group
                if not self.unassigned_scheduled_task_groups[vname]:
                    tbd.append(vname)
                    log3("Remove alredy assigned mission from unassigned scheduled list", "updateUnassigned", self)
            else:
                log3(vname+" NOT FOUND in unassigned scheduled work group", "updateUnassigned", self)
        elif tg_type == "reactive":
            if vname in self.unassigned_reactive_task_groups:
                # maybe an expensive way of remove a task from the group.
                self.unassigned_reactive_task_groups[vname] = [tsk for tsk in self.unassigned_reactive_task_groups[vname] if tsk['mid'] not in tg_mids]

                # if this vehcle no longer has unassigned work, then remove this vehicle from unassigned_task_group
                if not self.unassigned_reactive_task_groups[vname]:
                    tbd.append(vname)
                    log3("Remove already assigned mission from unassigned reactive list", "updateUnassigned", self)
            else:
                log3(vname+" NOT FOUND in unassigned reactive work group", "updateUnassigned", self)

        # find and delete mission from the work group.

    #update next mission pointer, return -1 if exceed the end of it.
    def update1WorkRunStatus(self, worksTBD, midx):

        this_stat = self.missions[midx].getStatus()
        worksTBD["current widx"] = worksTBD["current widx"] + 1

        log3("updatin 1 work run status:"+this_stat+" "+str(worksTBD["current widx"])+" "+str(len(worksTBD["works"])), "update1WorkRunStatus", self)

        if worksTBD["current widx"] >= len(worksTBD["works"]):
            worksTBD["current widx"] = self.checkTaskGroupCompleteness(worksTBD)
            log3("current widx pointer after checking retries:"+str(worksTBD["current widx"])+" "+str(len(worksTBD["works"])), "update1WorkRunStatus", self)
            if worksTBD["current widx"] >= len(worksTBD["works"]):
                log3("current work group is COMPLETED.", "update1WorkRunStatus", self)
                worksTBD["status"] = "Completed"


        log3("current widx pointer now at:"+str(worksTBD["current widx"])+" worksTBD status: "+worksTBD["status"], "update1WorkRunStatus", self)


    def checkTaskGroupCompleteness(self, worksTBD):
        mids = [w["mid"] for w in worksTBD["works"]]
        next_run_index = len(mids)
        for j, mid in enumerate(mids):
            midx = next((i for i, m in enumerate(self.missions) if m.getMid() == mid), -1)
            if midx != -1:
                this_stat = self.missions[midx].getStatus()
                n_2b_retried = self.missions[midx].getRetry()
                retry_count = self.missions[midx].getNRetries()
                log3("check retries: "+str(mid)+" "+str(self.missions[midx].getMid())+" n2b retries: "+str(n_2b_retried)+" n retried: "+str(retry_count), "checkTaskGroupCompleteness", self)
                if "Complete" not in this_stat and retry_count < n_2b_retried:
                    log3("scheduing retry#:"+str(j)+" MID: "+str(mid), "checkTaskGroupCompleteness", self)
                    next_run_index = j
                    break
        return next_run_index

    def updateRunStatus(self, worksTBD, midx):
        works = worksTBD["works"]

        tz = worksTBD["current tz"]
        grp = worksTBD["current grp"]
        bidx = worksTBD["current bidx"]
        widx = worksTBD["current widx"]
        oidx = worksTBD["current oidx"]
        switch_tz = False
        switch_grp = False
        switch_bot = False
        if grp == "other_works":
            idx = oidx
        else:
            idx = widx

        this_stat = self.missions[midx].getStatus()

        log3("TZ:"+tz+" GRP:"+grp+" BIDX:"+str(bidx)+" WIDX:"+str(widx)+" OIDX:"+str(oidx)+" THIS STATUS:"+this_stat, "updateRunStatus", self)

        if "Completed" in this_stat:
            # check whether need to switch group?
            if grp == None:
                # just the begining....
                tzi = 0
                switch_tz = True
            else:
                # update after already started
                if len(works[tz]) > 0:
                    if grp == "other_works":
                        if len(works[tz][bidx][grp])-1 > oidx:
                            oidx = oidx + 1
                        else:
                            # all other_works are done. simply go to the next bw_works if there are more
                            # simply switch group
                            grp = "bw_works"
                            # but if no more work after switching grp, switch timezone.
                            if len(works[tz][bidx][grp]) > 0:
                                if widx > len(works[tz][bidx][grp])-1:
                                    if bidx < len(works[tz]) - 1:
                                        bidx = bidx + 1
                                        switch_bot = True
                                    else:
                                        # in case this is the last bot, then switch timezone.
                                        switch_tz = True
                                else:
                                    switch_grp = True
                                    widx = widx + 1
                            else:
                                # all other_works and bw_works of this region(timezone) are done, check to see whether to switch bot.
                                if bidx < len(works[tz])-1:
                                    bidx = bidx + 1
                                    switch_bot = True
                                else:
                                    # in case this is the last bot, then switch timezone.
                                    switch_tz = True
                    else:
                        # bw works
                        if len(works[tz][bidx][grp])-1 > widx:
                            widx = widx + 1
                        else:
                            # all walk-buy works are done. simply go to the next other_works  if there are more
                            grp = "other_works"
                            if len(works[tz][bidx][grp]) > 0:
                                if oidx > len(works[tz][bidx][grp])-1:
                                    if bidx < len(works[tz]) - 1:
                                        bidx = bidx + 1
                                        switch_bot = True
                                    else:
                                        # switch tz.
                                        switch_tz = True
                                else:
                                    switch_grp = True
                                    oidx = oidx + 1
                            else:
                                if bidx < len(works[tz])-1:
                                    bidx = bidx + 1
                                    switch_bot = True
                                else:
                                    # switch tz.
                                    switch_tz = True
                    # now compare time.
                    if switch_tz == False:
                        if switch_bot == False:
                            if switch_grp == False:
                                if works[tz][bidx]["other_works"][oidx]["start_time"] < works[tz][bidx]["bw_works"][widx]["start_time"]:
                                    worksTBD["current grp"] = "other_works"
                                else:
                                    worksTBD["current grp"] = "bw_works"
                            else:
                                worksTBD["current grp"] = grp
                        else:
                            # if bot is changed, oidx and widx restart from 0.
                            oidx = 0
                            widx = 0
                            log3("SWITCHED BOT:"+str(bidx), "updateRunStatus", self)
                            if len(works[tz][bidx]["other_works"]) > 0 and len(works[tz][bidx]["bw_works"]) > 0:
                                if works[tz][bidx]["other_works"][oidx]["start_time"] < works[tz][bidx]["bw_works"][widx]["start_time"]:
                                    worksTBD["current grp"] = "other_works"
                                else:
                                    worksTBD["current grp"] = "bw_works"
                            elif len(works[tz][bidx]["other_works"]) > 0:
                                worksTBD["current grp"] = "other_works"
                            else:
                                worksTBD["current grp"] = "bw_works"

                        worksTBD["current bidx"] = bidx
                        worksTBD["current widx"] = widx
                        worksTBD["current oidx"] = oidx
                        worksTBD["current tz"] = tz
                else:
                    switch_tz = True

            # check whether need to switch region?
            if switch_tz:
                tzi = Tzs.index(tz)
                while tzi < len(Tzs) and len(works[tz]) == 0:
                    tzi = tzi + 1

                if tzi < len(Tzs):
                    tz = Tzs[tzi]
                    log3("SWITCHED TZ: "+tz, "updateRunStatus", self)
                    if len(works[tz][bidx]["other_works"]) > 0 and len(works[tz][bidx]["bw_works"]) > 0:
                        # see which one's start time is earlier
                        if works[tz][bidx]["other_works"][0]["start_time"] < works[tz][bidx]["bw_works"][0]["start_time"]:
                            worksTBD["current grp"] = "other_works"
                            worksTBD["current bidx"] = 0
                            worksTBD["current widx"] = -1
                            worksTBD["current oidx"] = 0
                        else:
                            worksTBD["current grp"] = "bw_works"
                            worksTBD["current bidx"] = 0
                            worksTBD["current widx"] = 0
                            worksTBD["current oidx"] = -1
                    elif len(works[tz][bidx]["other_works"]) > 0:
                        worksTBD["current grp"] = "other_works"
                        worksTBD["current bidx"] = 0
                        worksTBD["current widx"] = -1
                        worksTBD["current oidx"] = 0
                    elif len(works[tz][bidx]["bw_works"]) > 0:
                        worksTBD["current grp"] = "bw_works"
                        worksTBD["current bidx"] = 0
                        worksTBD["current widx"] = 0
                        worksTBD["current oidx"] = -1

                else:
                    # already reached the last region in this todo group, consider this group done.
                    # now check whether there is any failed missions, if there is, now it's time to set
                    # up to re-run it, simply by set the pointers to it.
                    log3("all workdsTBD exhausted...", "updateRunStatus", self)
                    rt_tz, rt_bid, rt_grp, rt_mid = self.findNextMissonsToBeRetried(worksTBD)
                    if rt_tz == "":
                        # if nothing is found, we're done with this todo list...
                        worksTBD["status"] == "Completed"
                    else:
                        # now set the pointer to the next mission that needs to be retried....
                        tz = rt_tz
                        worksTBD["current grp"] = rt_grp
                        worksTBD["current bidx"] = int(rt_bid)
                        if rt_grp == "bw_works":
                            worksTBD["current widx"] = int(rt_mid)
                            worksTBD["current oidx"] = 0
                        else:
                            worksTBD["current oidx"] = int(rt_mid)
                            worksTBD["current widx"] = 0

        worksTBD["current tz"] = tz


    def findNextMissonsToBeRetried(self, workgroup):
        found = False
        works = workgroup["works"]
        while not found:
            tz = workgroup["current tz"]
            grp = workgroup["current grp"]
            bidx = workgroup["current bidx"]
            widx = workgroup["current widx"]
            oidx = workgroup["current oidx"]

            switch_tz = False
            switch_grp = False
            switch_bot = False

            # check whether need to switch group?
            if grp == None:
                # just the begining....
                tzi = 0
                switch_tz = True
            else:
                # update after already started
                if len(works[tz]) > 0:
                    if grp == "other_works":
                        if len(works[tz][bidx][grp]) - 1 > oidx:
                            oidx = oidx + 1
                        else:
                            # all other_works are done. simply go to the next bw_works if there are more
                            # simply switch group
                            grp = "bw_works"
                            # but if no more work after switching grp, switch timezone.
                            if len(works[tz][bidx][grp]) > 0:
                                if widx > len(works[tz][bidx][grp]) - 1:
                                    if bidx < len(works[tz]) - 1:
                                        bidx = bidx + 1
                                        switch_bot = True
                                    else:
                                        # in case this is the last bot, then switch timezone.
                                        switch_tz = True
                                else:
                                    switch_grp = True
                                    widx = widx + 1
                            else:
                                # all other_works and bw_works of this region(timezone) are done, check to see whether to switch bot.
                                if bidx < len(works[tz]) - 1:
                                    bidx = bidx + 1
                                    switch_bot = True
                                else:
                                    # in case this is the last bot, then switch timezone.
                                    switch_tz = True
                    else:
                        # bw works
                        if len(works[tz][bidx][grp]) - 1 > widx:
                            widx = widx + 1
                        else:
                            # all walk-buy works are done. simply go to the next other_works  if there are more
                            grp = "other_works"
                            if len(works[tz][bidx][grp]) > 0:
                                if oidx > len(works[tz][bidx][grp]) - 1:
                                    if bidx < len(works[tz]) - 1:
                                        bidx = bidx + 1
                                        switch_bot = True
                                    else:
                                        # switch tz.
                                        switch_tz = True
                                else:
                                    switch_grp = True
                                    oidx = oidx + 1
                            else:
                                if bidx < len(works[tz]) - 1:
                                    bidx = bidx + 1
                                    switch_bot = True
                                else:
                                    # switch tz.
                                    switch_tz = True
                    # now compare time.
                    if switch_tz == False:
                        if switch_bot == False:
                            if switch_grp == False:
                                if works[tz][bidx]["other_works"][oidx]["start_time"] < \
                                        works[tz][bidx]["bw_works"][widx]["start_time"]:
                                    workgroup["current grp"] = "other_works"
                                else:
                                    workgroup["current grp"] = "bw_works"
                            else:
                                workgroup["current grp"] = grp
                        else:
                            # if bot is changed, oidx and widx restart from 0.
                            oidx = 0
                            widx = 0
                            if works[tz][bidx]["other_works"][oidx]["start_time"] < \
                                    works[tz][bidx]["bw_works"][widx][
                                        "start_time"]:
                                workgroup["current grp"] = "other_works"
                            else:
                                workgroup["current grp"] = "bw_works"

                        workgroup["current bidx"] = bidx
                        workgroup["current widx"] = widx
                        workgroup["current oidx"] = oidx
                        workgroup["current tz"] = tz
                else:
                    switch_tz = True

            # check whether need to switch region?
            if switch_tz:
                tzi = Tzs.index(tz)
                while tzi < len(Tzs) and len(works[tz]) == 0:
                    tzi = tzi + 1

                if tzi < len(Tzs):
                    tz = Tzs[tzi]
                    if len(works[tz][bidx]["other_works"]) > 0 and len(works[tz][bidx]["bw_works"]) > 0:
                        # see which one's start time is earlier
                        if works[tz][bidx]["other_works"][0]["start_time"] < works[tz][bidx]["bw_works"][0][
                            "start_time"]:
                            workgroup["current grp"] = "other_works"
                            workgroup["current bidx"] = 0
                            workgroup["current widx"] = -1
                            workgroup["current oidx"] = 0
                        else:
                            workgroup["current grp"] = "bw_works"
                            workgroup["current bidx"] = 0
                            workgroup["current widx"] = 0
                            workgroup["current oidx"] = -1
                    elif len(works[tz][bidx]["other_works"]) > 0:
                        workgroup["current grp"] = "other_works"
                        workgroup["current bidx"] = 0
                        workgroup["current widx"] = -1
                        workgroup["current oidx"] = 0
                    elif len(works[tz][bidx]["bw_works"]) > 0:
                        workgroup["current grp"] = "bw_works"
                        workgroup["current bidx"] = 0
                        workgroup["current widx"] = 0
                        workgroup["current oidx"] = -1
                else:
                    # this is the case we have reach the last mission of the todo list...
                    tz, bid, grp, mid  = self.findFirstMissonsToBeRetried(works)
                    if tz == "":
                        # in such a case there is nothing to retry. consider it done....
                        found = True
                        workgroup["status"] = "Completed"
                    else:
                        workgroup["current tz"] = tz
                        workgroup["current bidx"] = bid
                        workgroup["current grp"] = grp
                        if grp == "bw_works":
                            workgroup["current widx"] = mid
                        else:
                            workgroup["current oidx"] = mid

            workgroup["current tz"] = tz

            if grp == "other_works":
                idx = oidx
            else:
                idx = widx

            mission_id = works[tz][bidx][grp][idx]["mid"]
            midx = next((i for i, mission in enumerate(self.missions) if str(mission.getMid()) == mission_id), -1)
            this_stat = self.missions[midx].getStatus()
            n_retries = self.missions[midx].getNRetries()
            if "Completed" not in this_stat and n_retries > 0:
                found = True


    # go thru all todos to find the first mission that's incomplete and retry count is not down to 0 yet.
    def findFirstMissonsToBeRetried(self, todos):
        found = False
        mid = grp = bid = tz = ""
        for key1, value1 in todos.items():
            # regions
            if isinstance(value1, dict):
                for key2, value2 in value1.items():
                    # botids
                    if isinstance(value2, dict):
                        for key3, value3 in value2.items():
                            # groups
                            if isinstance(value3, dict):
                                mid = 0
                                for item in value3.items():
                                    # missions
                                    mission_id = item["mid"]
                                    midx = next((i for i, mission in enumerate(self.missions) if str(mission.getMid()) == mission_id), -1)
                                    this_stat = self.missions[midx].getStatus()
                                    n_retry = self.missions[midx].getRetry()
                                    if "Completed" not in this_stat and n_retry > 0:
                                        found = True
                                        grp = key3
                                        bid = key2
                                        tz = key1
                                        break
                                    else:
                                        mid = mid + 1
                            if found:
                                break
                    if found:
                        break
            if found:
                break
        #now point to the 1st item in this todo list

        log3("MISSIONS needs retry: "+tz+" "+str(bid)+" "+grp+" "+str(mid), "findFirstMissonsToBeRetried", self)
        return tz, bid, grp, mid




    #convert time zone, time slot to datetime
    # the time slot is defined as following:
    # time slot is defined as a 20 minute interval, an entire day has 72 slots indexed 0~71
    # counting timezone, and starts from eastern standard time, the timezone will extend to
    # cover hawaii, which is 5 timezone away from eastern, so total time zone slots are
    # 72+15=87 or index 0~86.
    def ts2time(self, ts):
        thistime = datetime.now()
        zerotime = datetime(thistime.date().year, thistime.date().month, thistime.date().day, 0, 0, 0)
        if ts < 0:      # in case of timeslot is -1 it means run it asap, so make it 0 zero time of today.
            ts = 0
        time_change = timedelta(minutes=20*ts)
        runtime = zerotime + time_change
        return runtime

    def time2ts(self, pdt):
        thistime = datetime.now()
        zerotime = datetime(thistime.date().year, thistime.date().month, thistime.date().day, 0, 0, 0)
        # Get the time difference in seconds
        ts = int((pdt - zerotime).total_seconds()/1200)         # computer time slot in 20minuts chunk

        return ts


    def runBotTask(self, task):
        self.working_state = "running_working"
        task_mission = self.missions[task.mid]
        # run all the todo steps
        # (steps, mission, skill, mode="normal"):
        runResult = rpaRunAllSteps(task.todos, task_mission.parent_settings)

    def newBotView(self):
        # Logic for creating a new bot:
        # pop out a new windows for user to set parameters for a new bot.
        # at the moment, just add an icon.
        #new_bot = EBBOT(self)
        # Icon removed - no longer needed for UI
        #self.centralWidget.setText("<b>File > New</b> clicked")
        if self.bot_manager == None:
            self.bot_manager = BotManager(self)
            self.bot_manager.set_mode("new")
            self.bot_manager.load_bot(EBBOT(self))
            # Note: BotManager is now a data handler, not a GUI window
            # You may need to implement a new GUI or use existing bot display methods


    def trainNewSkill(self):
        if self.train_manager == None:
            self.train_manager = TrainManager(self)
            self.reminder_manager = ReminderManager(self)
        self.showMsg("train new skill....")
        # Note: TrainManager is now a data handler, not a GUI window
        # You may need to implement a new GUI or use existing skill display methods
        self.train_manager.set_cloud_credentials(self.session, self.get_auth_token())

    def findAllBot(self):
        self.bot_service.find_all_bots()

    def logout(self):
        """Initiate graceful logout and cleanup of background tasks/servers."""
        if getattr(self, "_cleanup_in_progress", False):
            return
        self._cleanup_in_progress = True
        self.showMsg("logging out (graceful shutdown)........")
        try:
            # Run async cleanup without blocking the UI thread
            self.mainLoop.create_task(self._async_cleanup_and_logout())
        except Exception as e:
            logger.warning(f"Failed to schedule async cleanup: {e}")

    async def _async_cleanup_and_logout(self):
        """Asynchronously cleanup background tasks, servers, and resources, then logout."""
        logger.info("[MainWindow] 🧹 Starting comprehensive cleanup for logout...")
        
        # Set shutdown flag to prevent WAN Chat reconnections
        self._shutting_down = True
        
        # Unregister proxy change callback
        try:
            if hasattr(self, '_proxy_callback_unregister') and self._proxy_callback_unregister:
                self._proxy_callback_unregister()
                self._proxy_callback_unregister = None
                logger.info("[MainWindow] ✅ Proxy change callback unregistered")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error unregistering proxy callback: {e}")
        
        # Stop sync manager auto retry timer
        try:
            from agent.cloud_api.offline_sync_manager import get_sync_manager
            manager = get_sync_manager()
            manager.stop_auto_retry()
            logger.info("[MainWindow] ✅ Sync manager auto retry timer stopped")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error stopping sync manager: {e}")
        
        # Stop LightRAG server
        try:
            if hasattr(self, 'lightrag_server') and self.lightrag_server:
                self.stop_lightrag_server()
                logger.info("[MainWindow] ✅ LightRAG server stopped")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error stopping LightRAG server: {e}")

        # Stop local Starlette server (uvicorn) and join thread
        try:
            from gui.LocalServer import stop_local_server, MCPHandler
            
            # First, clean up MCP session manager BEFORE stopping the server
            # This prevents the TaskGroup context issues
            try:
                logger.info("[MainWindow] 🧹 Pre-cleaning MCP session manager...")
                await MCPHandler.cleanup()
                logger.info("[MainWindow] ✅ MCP session manager pre-cleaned")
            except Exception as e:
                logger.warning(f"[MainWindow] ❌ Error pre-cleaning MCP session manager: {e}")
            
            # Then stop the server
            stop_local_server()
            logger.info("[MainWindow] ✅ Local Starlette server stopped")
            
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error stopping local server: {e}")

        # Close MCP client manager session
        try:
            from agent.mcp.local_client import mcp_client_manager
            await mcp_client_manager.close()
            logger.info("[MainWindow] ✅ MCP client manager closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing MCP client manager: {e}")

        # Close MCP client if present
        try:
            if hasattr(self, 'mcp_client') and self.mcp_client:
                if hasattr(self.mcp_client, 'close'):
                    await self.mcp_client.close()
                self.mcp_client = None
                logger.info("[MainWindow] ✅ MCP client closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing MCP client: {e}")

        # Close SSE connection manager if present
        try:
            if hasattr(self, '_sse_cm') and self._sse_cm:
                if hasattr(self._sse_cm, 'close'):
                    await self._sse_cm.close()
                self._sse_cm = None
                logger.info("[MainWindow] ✅ SSE connection manager closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing SSE connection manager: {e}")

        # Close websocket if present
        try:
            if getattr(self, 'websocket', None):
                ws = self.websocket
                close_fn = getattr(ws, 'close', None)
                if close_fn:
                    if asyncio.iscoroutinefunction(close_fn):
                        await close_fn()
                    else:
                        close_fn()
                self.websocket = None
                logger.info("[MainWindow] ✅ Websocket closed")
        except Exception as e:
            logger.debug(f"[MainWindow] ❌ Error closing websocket: {e}")

        # Close unified browser manager
        try:
            if hasattr(self, 'unified_browser_manager') and self.unified_browser_manager:
                if hasattr(self.unified_browser_manager, 'cleanup'):
                    cleanup_fn = self.unified_browser_manager.cleanup
                    if asyncio.iscoroutinefunction(cleanup_fn):
                        await cleanup_fn()
                    else:
                        cleanup_fn()
                elif hasattr(self.unified_browser_manager, 'close'):
                    close_fn = self.unified_browser_manager.close
                    if asyncio.iscoroutinefunction(close_fn):
                        await close_fn()
                    else:
                        close_fn()
                self.unified_browser_manager = None
                logger.info("[MainWindow] ✅ Unified browser manager closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing unified browser manager: {e}")

        # Close default webdriver if present
        try:
            if hasattr(self, 'default_webdriver') and self.default_webdriver:
                self.default_webdriver.quit()
                self.default_webdriver = None
                logger.info("[MainWindow] ✅ Default webdriver closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing default webdriver: {e}")

        # Close TCP server if present
        try:
            if hasattr(self, 'tcpServer') and self.tcpServer:
                if hasattr(self.tcpServer, 'close'):
                    self.tcpServer.close()
                self.tcpServer = None
                logger.info("[MainWindow] ✅ TCP server closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing TCP server: {e}")

        # Close commander transport if present
        try:
            if hasattr(self, 'commanderXport') and self.commanderXport:
                if hasattr(self.commanderXport, 'close'):
                    self.commanderXport.close()
                self.commanderXport = None
                logger.info("[MainWindow] ✅ Commander transport closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing commander transport: {e}")

        # Close session (cloud connection)
        try:
            if hasattr(self, 'session') and self.session:
                if hasattr(self.session, 'close'):
                    close_fn = self.session.close
                    if asyncio.iscoroutinefunction(close_fn):
                        await close_fn()
                    else:
                        close_fn()
                self.session = None
                logger.info("[MainWindow] ✅ Cloud session closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing cloud session: {e}")

        # Signal all agent TaskRunner loops to stop (while-loops in threads)
        try:
            from agent.tasks import TaskRunnerRegistry
            TaskRunnerRegistry.stop_all()
            logger.info("[MainWindow] ✅ TaskRunners stopped")
        except Exception as e:
            logger.debug(f"[MainWindow] ❌ Error stopping TaskRunners: {e}")

        # Cancel asyncio tasks we manage
        to_cancel = []
        for name in (
            'lan_task', 'peer_task', 'monitor_task', 'chat_task', 'wan_sub_task',
            'rpa_task', 'manager_task', 'wan_chat_task', 'llm_sub_task', 'cloud_show_sub_task'
        ):
            try:
                t = getattr(self, name, None)
                if t and not t.done():
                    t.cancel()
                    to_cancel.append(t)
            except Exception:
                pass
        if to_cancel:
            try:
                await asyncio.gather(*to_cancel, return_exceptions=True)
                logger.info(f"[MainWindow] ✅ Cancelled {len(to_cancel)} asyncio tasks")
            except Exception:
                pass

        # Close Cloud LLM WebSocket and thread
        try:
            if hasattr(self, 'cloud_llm_ws') and self.cloud_llm_ws:
                self.cloud_llm_ws.close()
                logger.info("[MainWindow] ✅ Cloud LLM WebSocket closed")
            if hasattr(self, 'cloud_llm_thread') and self.cloud_llm_thread:
                # Thread is daemon, so it will be terminated when main process exits
                logger.info("[MainWindow] ✅ Cloud LLM thread will terminate with main process")
        except Exception as e:
            logger.debug(f"[MainWindow] ❌ Error closing Cloud LLM WebSocket: {e}")

        # Shut down ThreadPoolExecutor
        try:
            if hasattr(self, 'threadPoolExecutor') and self.threadPoolExecutor:
                self.threadPoolExecutor.shutdown(wait=False, cancel_futures=True)
                logger.info("[MainWindow] ✅ ThreadPoolExecutor shutdown")
        except Exception as e:
            logger.debug(f"[MainWindow] ❌ Error shutting down ThreadPoolExecutor: {e}")

        # Close database services and connections
        try:
            # Close chat service
            if hasattr(self, 'db_chat_service') and self.db_chat_service:
                if hasattr(self.db_chat_service, 'close'):
                    self.db_chat_service.close()
                self.db_chat_service = None
                logger.info("[MainWindow] ✅ Database chat service closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing database chat service: {e}")

        try:
            # Close database engine and session
            if hasattr(self, '_db_engine') and self._db_engine:
                self._db_engine.dispose()
                self._db_engine = None
                logger.info("[MainWindow] ✅ Database engine disposed")
            
            if hasattr(self, '_db_session') and self._db_session:
                self._db_session.close()
                self._db_session = None
                logger.info("[MainWindow] ✅ Database session closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing database engine/session: {e}")

        # Clear IPC registry system ready cache
        try:
            from gui.ipc.registry import IPCHandlerRegistry
            IPCHandlerRegistry.clear_system_ready_cache()
            logger.debug("[MainWindow] ✅ IPC registry system ready cache cleared")
        except Exception as e:
            logger.debug(f"[MainWindow] ❌ Error clearing IPC registry cache: {e}")

        # Close database manager and clean up database connections
        try:
            if hasattr(self, 'ec_db_mgr') and self.ec_db_mgr:
                self.ec_db_mgr.close()
                self.ec_db_mgr = None
                logger.info("[MainWindow] ✅ Database manager closed")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error closing database manager: {e}")

        # Clear component managers
        try:
            manager_attrs = [
                'bot_manager', 'missionWin', 'train_manager', 'reminder_manager',
                'platoonWin', 'skill_manager', 'settings_manager', 'vehicle_monitor',
                'config_manager'
            ]
            for attr in manager_attrs:
                if hasattr(self, attr):
                    manager = getattr(self, attr)
                    if manager and hasattr(manager, 'cleanup'):
                        try:
                            if asyncio.iscoroutinefunction(manager.cleanup):
                                await manager.cleanup()
                            else:
                                manager.cleanup()
                        except Exception as e:
                            logger.debug(f"[MainWindow] Error cleaning up {attr}: {e}")
                    setattr(self, attr, None)
            logger.info("[MainWindow] ✅ Component managers cleared")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error clearing component managers: {e}")

        # Clear resource objects
        try:
            resource_attrs = ['file_resource', 'static_resource', 'zipper']
            for attr in resource_attrs:
                if hasattr(self, attr):
                    setattr(self, attr, None)
            logger.info("[MainWindow] ✅ Resource objects cleared")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error clearing resource objects: {e}")

        # Clear data collections
        try:
            data_attrs = [
                'agents', 'botJsonData', 'sellerInventoryJsonData', 'mcp_tools_schemas',
                'todays_work', 'reactive_work'
            ]
            for attr in data_attrs:
                if hasattr(self, attr):
                    data = getattr(self, attr)
                    if isinstance(data, (list, dict)):
                        data.clear()
                    else:
                        setattr(self, attr, None)
            logger.info("[MainWindow] ✅ Data collections cleared")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error clearing data collections: {e}")

        # Clear AppContext references
        try:
            from app_context import AppContext
            if AppContext.cleanup_instance():
                logger.info("[MainWindow] ✅ AppContext cleared")
            else:
                logger.warning("[MainWindow] ⚠️  AppContext cleanup had issues")
        except Exception as e:
            logger.warning(f"[MainWindow] ❌ Error clearing AppContext: {e}")

        # Finally, call auth logout and close window
        try:
            if hasattr(self, 'auth_manager') and self.auth_manager:
                self.auth_manager.logout()
                logger.info("[MainWindow] ✅ Auth manager logout completed")
        except Exception as e:
            logger.debug(f"[MainWindow] ❌ Auth logout error: {e}")

        logger.info("[MainWindow] 🎉 Comprehensive cleanup completed - logged out successfully")


    def addNewBots(self, new_bots):
        # Logic for creating a new bot:
        api_bots = []
        self.showMsg("adding new bots....")
        for new_bot in new_bots:
            api_bots.append({
                # "bid": new_bot.getBid(),
                "owner": self.owner,
                "roles": new_bot.getRoles(),
                "org": new_bot.getOrg(),
                "pubbirthday": new_bot.getPubBirthday(),
                "gender": new_bot.getGender(),
                "location": new_bot.getLocation(),
                "levels": new_bot.getLevels(),
                "birthday": new_bot.getBirthdayTxt(),
                "interests": new_bot.getInterests(),
                "status": new_bot.getStatus(),
                "delDate": new_bot.getInterests(),
                "name": new_bot.getName(),
                "pseudoname": new_bot.getPseudoName(),
                "nickname": new_bot.getNickName(),
                "addr": new_bot.getInterests(),
                "shipaddr": new_bot.getInterests(),
                "phone": new_bot.getPhone(),
                "email": new_bot.getEmail(),
                "epw": new_bot.getEmPW(),
                "backemail": new_bot.getBackEm(),
                "backemailpw": new_bot.getBackEmPW(),
                "ebpw": new_bot.getAcctPw(),
                "backemail_site": new_bot.getBackEmSite(),
                "createon": new_bot.getCreateOn(),
                "vehicle": new_bot.getVehicle()
            })
        jresp = send_add_bots_request_to_cloud(self.session, new_bots, self.get_auth_token(), self.getWanApiEndpoint())

        if "errorType" in jresp:
            screen_error = True
            self.showMsg("ERROR Type: "+json.dumps(jresp["errorType"])+"ERROR Info: "+json.dumps(jresp["errorInfo"]))
        else:
            self.showMsg("jresp:"+json.dumps(jresp))
            jbody = jresp["body"]
            #now that add is successfull, update local file as well.
            # first, update bot ID both in data structure and in GUI display.
            for i, resp_rec in enumerate(jresp["body"]):
                new_bots[i].setBid(resp_rec["bid"])
                new_bots[i].setInterests(resp_rec["interests"])
                self.bots.append(new_bots[i])
                # botModel removed - UI components no longer needed
                self.updateBotRelatedVehicles(new_bots[i])
            # botModel removed - UI components no longer needed
            # now add bots to local DB.
            if not self.config_manager.general_settings.debug_mode:
                self.bot_service.insert_bots_batch(jbody, api_bots)

    def updateBots(self, bots, localOnly=False):
        # potential optimization here, only if cloud side related attributes changed, then we do update on the cloud side.
        # otherwise, only update locally.
        api_bots = []
        for abot in bots:
            api_bots.append({
                "bid": abot.getBid(),
                "owner": self.owner,
                "roles": abot.getRoles(),
                "org": abot.getOrg(),
                "pubbirthday": abot.getPubBirthday(),
                "gender": abot.getGender(),
                "location": abot.getLocation(),
                "levels": abot.getLevels(),
                "birthday": abot.getBirthdayTxt(),
                "interests": abot.getInterests(),
                "status": abot.getStatus(),
                "delDate": abot.getInterests(),
                "createon": abot.getCreateOn(),
                "vehicle": abot.getVehicle(),
                "name": abot.getName(),
                "pseudoname": abot.getPseudoName(),
                "nickname": abot.getNickName(),
                "addr": abot.getAddr(),
                "shipaddr": abot.getShippingAddr(),
                "phone": abot.getPhone(),
                "email": abot.getEmail(),
                "epw": abot.getEmPW(),
                "backemail": abot.getBackEm(),
                "backemailpw": abot.getBackEmPW(),
                "ebpw": abot.getAcctPw(),
                "backemail_site": abot.getAcctPw()
            })
            # self.updateBotRelatedVehicles(abot)
        if not localOnly:
            jresp = send_update_bots_request_to_cloud(self.session, bots, self.get_auth_token(), self.getWanApiEndpoint())
            if "errorType" in jresp:
                screen_error = True
                self.showMsg("ERROR Type: "+json.dumps(jresp["errorType"]), "ERROR Info: "+json.dumps(jresp["errorInfo"]))
            else:
                print("update bot jresp:", jresp)
                # jbody = jresp["body"]
                jbody = jresp
                if jbody['numberOfRecordsUpdated'] == len(bots):
                    for i, abot in enumerate(bots):
                        api_bots[i]["vehicle"] = abot.getVehicle()
                    self.bot_service.update_bots_batch(api_bots)
                else:
                    self.showMsg("WARNING: bot NOT updated in Cloud!")
        else:
            print("updating local only....")
            self.bot_service.update_bots_batch(api_bots)

    # update in cloud, local DB, and local Memory
    def updateBotsWithJsData(self, bjs, localOnly=False):
        try:
            api_bots = []
            for abot in bjs:
                api_bots.append({
                    "bid": abot["pubAttributes"]["bid"],
                    "owner": self.owner,
                    "roles": abot["pubAttributes"]["roles"],
                    "org": abot["pubAttributes"]["org"],
                    "pubbirthday": abot["pubAttributes"]["pubbirthday"],
                    "gender": abot["pubAttributes"]["gender"],
                    "location": abot["pubAttributes"]["location"],
                    "levels": abot["pubAttributes"]["levels"],
                    "birthday": abot["privateProfile"]["birthday"],
                    "interests": abot["pubAttributes"]["interests"],
                    "status": abot["pubAttributes"]["status"],
                    "delDate": abot["pubAttributes"]["delDate"],
                    "createon": abot["privateProfile"]["createon"],
                    "vehicle": abot["pubAttributes"]["vehicle"],
                    "name": abot["privateProfile"]["bid"],
                    "pseudoname": abot["pubAttributes"]["pseudo_name"],
                    "nickname": abot["pubAttributes"]["pseudo_nick_name"],
                    "addr": abot["privateProfile"]["addr"],
                    "shipaddr": abot["privateProfile"]["shipping_addr"],
                    "phone": abot["privateProfile"]["phone"],
                    "email": abot["privateProfile"]["email"],
                    "epw": abot["privateProfile"]["email_pw"],
                    "backemail": abot["privateProfile"]["backup_email"],
                    "backemailpw": abot["privateProfile"]["backup_email_pw"],
                    "ebpw": abot["privateProfile"]["acct_pw"],
                    "backemail_site": abot["privateProfile"]["backup_email_site"],
                })
                self.updateBotRelatedVehicles(abot)

            if not localOnly:
                jresp = send_update_bots_request_to_cloud(self.session, bjs, self.get_auth_token(), self.getWanApiEndpoint())
                if "errorType" in jresp:
                    screen_error = True
                    self.showMsg("ERROR Type: " + json.dumps(jresp["errorType"]),
                                 "ERROR Info: " + json.dumps(jresp["errorInfo"]))
                else:
                    jbody = jresp["body"]
                    if jbody['numberOfRecordsUpdated'] == len(bjs):
                        self.bot_service.update_bots_batch(api_bots)

                        # finally update into in-memory data structure.

                    else:
                        self.showMsg("WARNING: bot NOT updated in Cloud!")

            else:
                self.bot_service.update_bots_batch(api_bots)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorUpdateBotsWithJsData:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorUpdateBotsWithJsData: traceback information not available:" + str(e)

            self.showMsg(ex_stat)



    def updateBotRelatedVehicles(self, bot):
        if bot.getVehicle() is not None and bot.getVehicle() != "" and bot.getVehicle() != "NA":
            # if last assigned vehicle is changed. remove botid from last assigned vehicle and botid to the new vehicle.
            vname = bot.getVehicle().split(":")[0]

            # update local DB
            currentVehicleInLocalDB = self.vehicle_service.find_vehicle_by_name(vname)
            previousVehicleInLocalDB = self.vehicle_service.find_vehicle_by_botid(str(bot.getBid()))

            if currentVehicleInLocalDB is not None:
                if previousVehicleInLocalDB is not None:
                    if currentVehicleInLocalDB.name != previousVehicleInLocalDB.name:
                        # update the current vehicle in local DB
                        bot_ids = ast.literal_eval(currentVehicleInLocalDB.bot_ids)
                        if bot.getBid() not in bot_ids:
                            bot_ids.append(bot.getBid())
                            currentVehicleInLocalDB.bot_ids = str(bot_ids)
                            self.vehicle_service.update_vehicle(currentVehicleInLocalDB)

                        # update the previous vehicle in local DB
                        self.vehicle_service.remove_bot_from_current_vehicle(str(bot.getBid()), previousVehicleInLocalDB)
                else:
                    bot_ids = ast.literal_eval(currentVehicleInLocalDB.bot_ids)
                    if bot.getBid() not in bot_ids:
                        bot_ids.append(bot.getBid())
                        currentVehicleInLocalDB.bot_ids = str(bot_ids)
                        self.vehicle_service.update_vehicle(currentVehicleInLocalDB)
            else:
                log3("ERROR: bot's vehicle non-exists in local DB...")

            # update local data structure.
            currentVehicle = next((v for i, v in enumerate(self.vehicles) if vname in v.getName()), None)
            previousVehicle = next((v for i, v in enumerate(self.vehicles) if bot.getBid() in v.getBotIds()), None)

            # update vehicle data structure
            if previousVehicle:
                if currentVehicle:
                    if previousVehicle.getName() != currentVehicle.getName():
                        previousVehicle.removeBot(bot.getBid())
                        currentVehicle.addBot(bot.getBid())
                else:
                    log3("ERROR: bot's vehicle non-exists...")
            else:
                if currentVehicle:
                    currentVehicle.addBot(bot.getBid())

    def addNewMissions(self, new_missions):
        # Logic for creating a new mission:
        try:
            self.showMsg("adding a .... new... mission")
            addedNewMissions = []
            jresp = send_add_missions_request_to_cloud(self.session, new_missions,
                                                       self.get_auth_token(), self.getWanApiEndpoint())
            if "errorType" in jresp:
                print("jresp:", jresp)
                self.showMsg("Error Add New Mission From File: "+json.dumps(jresp))
            else:
                jbody = jresp["body"]
                # now that delete is successfull, update local file as well.
                self.showMsg("JUST ADDED mission: "+str(len(jbody))+json.dumps(jbody))

                # Note not all mission will be added, if the cloud scheduling algorithm could NOT
                # find a bot for the mission, it will not be added.....

                for i, added in enumerate(jbody):
                    new_mission = next((m for i, m in enumerate(self.new_mission) if m.getTicket() == added["ticket"]), None)
                    if new_mission:
                        new_mission.setMid(jbody[i]["mid"])
                        new_mission.setTicket(jbody[i]["ticket"])
                        new_mission.setEstimatedStartTime(jbody[i]["esttime"])
                        new_mission.setEstimatedRunTime(jbody[i]["runtime"])
                        new_mission.setEsd(jbody[i]["esd"])
                        self.missions.append(new_mission)
                        # missionModel removed - UI components no longer needed
                        addedNewMissions.append(new_mission)
                if not self.config_manager.general_settings.debug_mode:
                    api_missions = []
                    for new_mission in addedNewMissions:
                        api_missions.append({
                            # "mid": new_mission.getMid(),
                            "ticket": new_mission.getMid(),
                            "botid": new_mission.getBid(),
                            "status": new_mission.getStatus(),
                            "createon": new_mission.getBD(),
                            "esd": new_mission.getEsd(),
                            "ecd": new_mission.getEcd(),
                            "asd": new_mission.getAsd(),
                            "abd": new_mission.getAbd(),
                            "aad": new_mission.getAad(),
                            "afd": new_mission.getAfd(),
                            "acd": new_mission.getAcd(),
                            "actual_start_time": new_mission.getActualStartTime(),
                            "est_start_time": new_mission.getEstimatedStartTime(),
                            "actual_run_time": new_mission.getActualRunTime(),
                            "est_run_time": new_mission.getEstimatedRunTime(),
                            "n_retries": new_mission.getNRetries(),
                            "cuspas": new_mission.getCusPAS(),
                            "search_cat": new_mission.getSearchCat(),
                            "search_kw": new_mission.getSearchKW(),
                            "pseudo_store": new_mission.getPseudoStore(),
                            "pseudo_brand": new_mission.getPseudoBrand(),
                            "pseudo_asin": new_mission.getPseudoASIN(),
                            "repeat": new_mission.getRetry(),
                            "mtype": new_mission.getMtype(),
                            "mconfig": new_mission.getConfig(),
                            "skills": new_mission.getSkills(),
                            "delDate": new_mission.getDelDate(),
                            "asin": new_mission.getASIN(),
                            "stores": new_mission.getStore(),
                            "follow_seller": new_mission.getFollowSeller(),
                            "brand": new_mission.getBrand(),
                            "image": new_mission.getImagePath(),
                            "title": new_mission.getTitle(),
                            "variations": new_mission.getVariations(),
                            "rating": new_mission.getRating(),
                            "feedbacks": new_mission.getFeedbacks(),
                            "price": new_mission.getPrice(),
                            "follow_price": new_mission.getFollowPrice(),
                            "customer": new_mission.getCustomerID(),
                            "platoon": new_mission.getPlatoonID(),
                            "fingerprint_profile": new_mission.getFingerPrintProfile(),
                            "original_req_file": new_mission.getReqFile(),
                            "as_server": new_mission.getAsServer(),
                            "result": ""
                        })
                    self.mission_service.insert_missions_batch(jbody, api_missions)

                mid_list = [mission.getMid() for mission in addedNewMissions]
                local_mission_rows = self.mission_service.find_missions_by_mids(mid_list)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorAddNewMissions:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorAddNewMissions: traceback information not available:" + str(e)

            print(ex_stat)



    def updateMissions(self, missions):
        # potential optimization here, only if cloud side related attributes changed, then we do update on the cloud side.
        # otherwise, only update locally.
        api_missions = []
        for amission in missions:
            api_missions.append({
                "mid": amission.getMid(),
                "ticket": amission.getMid(),
                "botid": amission.getBid(),
                "status": amission.getStatus(),
                "createon": amission.getBD(),
                "esd": amission.getEsd(),
                "ecd": amission.getEcd(),
                "asd": amission.getAsd(),
                "abd": amission.getAbd(),
                "aad": amission.getAad(),
                "afd": amission.getAfd(),
                "acd": amission.getAcd(),
                "actual_start_time": amission.getActualStartTime(),
                "est_start_time": amission.getEstimatedStartTime(),
                "actual_run_time": amission.getActualRunTime(),
                "est_run_time": amission.getEstimatedRunTime(),
                "n_retries": amission.getNRetries(),
                "cuspas": amission.getCusPAS(),
                "search_cat": amission.getSearchCat(),
                "search_kw": amission.getSearchKW(),
                "pseudo_store": amission.getPseudoStore(),
                "pseudo_brand": amission.getPseudoBrand(),
                "pseudo_asin": amission.getPseudoASIN(),
                "repeat": amission.getRetry(),
                "type": amission.getMtype(),
                "config": amission.getConfig(),
                "skills": amission.getSkills(),
                "delDate": amission.getDelDate(),
                "asin": amission.getASIN(),
                "stores": amission.getStore(),
                "follow_seller": amission.getFollowSeller(),
                "brand": amission.getBrand(),
                "image": amission.getImagePath(),
                "title": amission.getTitle(),
                "variations": amission.getVariations(),
                "rating": amission.getRating(),
                "feedbacks": amission.getFeedbacks(),
                "price": amission.getPrice(),
                "follow_price": amission.getFollowPrice(),
                "customer": amission.getCustomerID(),
                "platoon": amission.getPlatoonID(),
                "result": amission.getResult(),
                "as_server": amission.getAsServer(),
                "fingerprint_profile": amission.getFingerPrintProfile(),
                "original_req_file": amission.getReqFile()
            })

        jresp = send_update_missions_request_to_cloud(self.session, missions, self.get_auth_token(), self.getWanApiEndpoint())
        if "errorType" in jresp:
            screen_error = True
            self.showMsg("ERROR Type: "+json.dumps(jresp["errorType"])+"ERROR Info: "+json.dumps(jresp["errorInfo"]))
        else:
            jbody = jresp["body"]
            self.showMsg("Update Cloud side result:"+json.dumps(jbody))
            if jbody['numberOfRecordsUpdated'] == len(missions):
                self.mission_service.update_missions_by_id(api_missions)
                mid_list = [mission.getMid() for mission in missions]
                self.mission_service.find_missions_by_mids(mid_list)
            else:
                self.showMsg("WARNIN: cloud NOT updated.", "warn")

    def addBotsMissionsSkillsFromCommander(self, botsJson, missionsJson, skillsJson):
        existinBids = [b.getBid() for b in self.bots]
        existinMids = [m.getMid() for m in self.missions]
        existinSkids = [sk.getSkid() for sk in self.skills]
        print("existinBids:", existinBids)
        # print("botsJson:", botsJson)

        usefulBotsJson = [bj for bj in botsJson if bj['privateProfile']["email"]]
        # self.showMsg("BOTS String:"+str(type(botsJson))+json.dumps(botsJson))
        # self.showMsg("Missions String:"+str(type(missionsJson))+json.dumps(missionsJson))
        # self.showMsg("Skills String:" + str(type(skillsJson)) + json.dumps(skillsJson))
        for bjs in usefulBotsJson:
            if int(bjs["pubProfile"]["bid"]) not in existinBids:
                self.newBot = EBBOT(self)
                self.newBot.loadJson(bjs)
                self.newBot.updateIcon()
                self.bots.append(self.newBot)
                # botModel removed - UI components no longer needed
                bot_profile_name = self.my_ecb_data_homepath + "/ads_profiles/"+bjs["privateProfile"]["email"].split("@")[0]+".txt"
                if bot_profile_name not in self.todays_bot_profiles:
                    self.todays_bot_profiles.append(bot_profile_name)

        for mjs in missionsJson:
            if int(mjs["pubAttributes"]["missionId"]) not in existinMids:
                self.newMission = EBMISSION(self)
                self.newMission.loadJson(mjs)
                self.newMission.updateIcon()
                self.missions.append(self.newMission)
                # missionModel removed - UI components no longer needed

        for skjs in skillsJson:
            if int(skjs["skid"]) not in existinSkids:
                self.newSkill = WORKSKILL(self, skjs["name"])
                self.newSkill.loadJson(skjs)
                self.newSkill.updateIcon()
                self.skills.append(self.newSkill)
                # self.skillModel.appendRow(self.newSkill)

        print("done setting bots, missions, skills from commander")

    def addConnectingVehicle(self, vname, vip):
        try:
            # ipfields = vinfo.peername[0].split(".")

            if len(self.vehicles) > 0:
                v_host_names = [v.getName().split(":")[0] for v in self.vehicles]
                print("existing vehicle "+json.dumps(v_host_names))
            else:
                vids = []

            found_fl = next((fl for i, fl in enumerate(fieldLinks) if vname in fl["name"]), None)

            if vname not in v_host_names:
                self.showMsg("adding a new vehicle..... "+vname+" "+vip)
                newVehicle = VEHICLE(self)
                newVehicle.setIP(vip)
                newVehicle.setVid(vip.split(".")[3])
                newVehicle.setName(vname+":")
                if found_fl:
                    print("found_fl IP:", found_fl["ip"])
                    newVehicle.setFieldLink(found_fl)
                    newVehicle.setStatus("connecting")
                self.saveVehicle(newVehicle)
                self.vehicles.append(newVehicle)
                # Vehicle GUI model removed - vehicles managed through PlatoonManager
                if hasattr(self, 'platoon_manager') and self.platoon_manager:
                    self.platoon_manager.add_vehicle(newVehicle)

                resultV = newVehicle
            else:
                self.showMsg("Reconnected: "+vip)
                foundV = next((v for i, v in enumerate(self.vehicles) if vname in v.getName()), None)
                foundV.setIP(vip)
                foundV.setStatus("connecting")
                if found_fl:
                    print("found_fl IP:", found_fl["ip"])
                    foundV.setFieldLink(found_fl)

                resultV = foundV

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorAddConnectingVehicle:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorAddConnectingVehicle: traceback information not available:" + str(e)

            self.showMsg(ex_stat)
        print("added connecting vehicle:", resultV.getName(), resultV.getStatus())
        return resultV



    def markVehicleOffline(self, vip, vname):
        global fieldLinks
        lostName = vname.split(":")[0]
        self.showMsg("marking vehicle offline: "+vip+" "+json.dumps([v.getIP()+":"+v.getName() for v in self.vehicles]))

        found_v_idx = next((i for i, v in enumerate(self.vehicles) if lostName in v.getName()), -1)
        print("found_v_idx", found_v_idx)
        if found_v_idx > 0:
            print("markingoff......")
            found_v = self.vehicles[found_v_idx]
            found_v.setStatus("offline")

        return found_v

    # add vehicles based on fieldlinks.
    def checkVehicles(self):
        import time
        start_time = time.time()
        
        self.showMsg("adding self as a vehicle if is Commander.....")
        existing_names = [v.getName().split(":")[0] for v in self.vehicles]
        print("existing v names:", existing_names)
        if self.machine_role == "Commander":
            # Check if this machine is already in the vehicle list
            if self.machine_name not in existing_names:
                # should add this machine to vehicle list.
                newVehicle = VEHICLE(self, self.machine_name+":"+self.os_short, self.ip)
                newVehicle.setStatus("running_idle")
                
                # Set complete system information
                if hasattr(self, 'architecture'):
                    # Get architecture from platform.architecture()[0] and convert to common format
                    arch_mapping = {
                        '64bit': 'x86_64' if 'intel' in self.processor.lower() or 'amd' in self.processor.lower() else 'arm64',
                        '32bit': 'x86'
                    }
                    newVehicle.setArch(arch_mapping.get(self.architecture, self.architecture))
                else:
                    # Fallback: detect architecture
                    import platform
                    machine = platform.machine().lower()
                    if machine in ['x86_64', 'amd64']:
                        newVehicle.setArch('x86_64')
                    elif machine in ['arm64', 'aarch64']:
                        newVehicle.setArch('arm64')
                    else:
                        newVehicle.setArch(machine)
                
                # Set functions based on role
                if hasattr(self, 'functions'):
                    newVehicle.setFunctions(self.functions)
                
                # Set unique vehicle ID based on IP
                ip_parts = self.ip.split('.')
                if len(ip_parts) >= 4:
                    newVehicle.setVid(ip_parts[-1])  # Use last octet as ID
                else:
                    newVehicle.setVid(str(hash(self.machine_name) % 1000))  # Fallback ID
                
                # Set device information and performance metrics using system info manager
                newVehicle.setType(self.device_type)
                newVehicle.setLocation(f"Local - {self.machine_name}")
                newVehicle.setBattery(100)  # Desktop/laptop default to 100%
                newVehicle.setCurrentTask("System Management")
                
                # Set maintenance schedule (example: next maintenance in 30 days)
                from datetime import datetime, timedelta
                next_maintenance = datetime.now() + timedelta(days=30)
                newVehicle.setNextMaintenance(next_maintenance.strftime("%Y-%m-%d"))
                
                # Schedule async system metrics update (non-blocking)
                try:
                    # Try to create task in existing event loop
                    loop = asyncio.get_running_loop()
                    loop.create_task(self._update_vehicle_metrics_async(newVehicle))
                    logger.debug(f"[MainWindow] ✅ Scheduled async metrics update for vehicle: {newVehicle.getName()}")
                except RuntimeError as e:
                    # This is expected during startup when event loop is not running yet
                    logger.debug(f"[MainWindow] ⏳ Event loop not running yet for vehicle metrics update: {e}")
                    # Use QTimer for delayed execution, waiting for event loop to be available
                    from PySide6.QtCore import QTimer
                    timer = QTimer()
                    timer.timeout.connect(lambda v=newVehicle: self._schedule_delayed_metrics_update(v))
                    timer.setSingleShot(True)
                    timer.start(2000)  # Retry after 2 seconds
                    logger.info(f"[MainWindow] 📊 Scheduled delayed metrics update for vehicle: {newVehicle.getName()}")
                
                self.saveVehicle(newVehicle)
                self.vehicles.append(newVehicle)
                # Vehicle GUI model removed - vehicles managed through PlatoonManager
                if hasattr(self, 'platoon_manager') and self.platoon_manager:
                    self.platoon_manager.add_vehicle(newVehicle)
                logger.info(f"[MainWindow] Added local machine as vehicle: {self.machine_name} (arch: {newVehicle.getArch()}, functions: {newVehicle.getFunctions()})")
            else:
                logger.info(f"[MainWindow] Local machine already exists in vehicle list: {self.machine_name}")

        # Record performance statistics
        end_time = time.time()
        logger.info(f"[MainWindow] checkVehicles completed in {end_time - start_time:.3f}s")

        self.showMsg("adding already linked vehicles.....")
        for i in range(len(fieldLinks)):
            if fieldLinks[i]["name"].split(":")[0] not in existing_names:
                self.showMsg("a fieldlink....."+json.dumps(fieldLinks[i]["ip"]))
                newVehicle = VEHICLE(self, fieldLinks[i]["name"], fieldLinks[i]["ip"])
                newVehicle.setFieldLink(fieldLinks[i])
                newVehicle.setStatus("running_idle")        # if already has a fieldlink, that means it's tcp connected.
                ipfields = fieldLinks[i]["ip"].split(".")
                ip = ipfields[len(ipfields)-1]
                newVehicle.setVid(ip)
                self.saveVehicle(newVehicle)
                self.vehicles.append(newVehicle)
                # Vehicle GUI model removed - vehicles managed through PlatoonManager
                if hasattr(self, 'platoon_manager') and self.platoon_manager:
                    self.platoon_manager.add_vehicle(newVehicle)

    def saveVehicle(self, vehicle: VEHICLE):
        """Save vehicle to database - only called when database services are ready"""
        # Check if this is a Commander role (only Commander has database services)
        if "Commander" not in self.machine_role:
            logger.debug(f"[MainWindow] Platoon role - no database save needed for vehicle: {vehicle.getName()}")
            return

        # At this point, vehicle_service MUST be available because saveVehicle should only be called
        # after database services are initialized. If it's not available, it's a programming error.
        if not hasattr(self, 'vehicle_service') or self.vehicle_service is None:
            logger.error(f"[MainWindow] PROGRAMMING ERROR: saveVehicle called before vehicle_service is ready!")
            logger.error(f"[MainWindow] This indicates a timing issue in the initialization sequence.")
            raise RuntimeError(f"saveVehicle called before vehicle_service initialization - fix the calling sequence!")

        try:
            v = self.vehicle_service.find_vehicle_by_ip(vehicle.ip)
            if v is None:
                # Create new vehicle record
                vehicle_model = VehicleModel()
                vehicle_model.arch = vehicle.arch
                vehicle_model.bot_ids = str(vehicle.bot_ids)
                vehicle_model.daily_mids = str(vehicle.daily_mids)
                vehicle_model.ip = vehicle.ip
                vehicle_model.mstats = str(vehicle.mstats)
                vehicle_model.name = vehicle.name
                vehicle_model.os = vehicle.os
                vehicle_model.cap = vehicle.CAP
                vehicle_model.status = vehicle.status
                self.vehicle_service.insert_vehicle(vehicle_model)
                vehicle.id = vehicle_model.id
                logger.debug(f"[MainWindow] Created new vehicle record: {vehicle.getName()} (ID: {vehicle.id})")
            else:
                # Update existing vehicle record and sync data
                vehicle.setVid(v.id)
                vehicle.setBotIds(ast.literal_eval(v.bot_ids))
                vehicle.setMStats(ast.literal_eval(v.mstats))
                vehicle.setMids(ast.literal_eval(v.daily_mids))
                v.cap = vehicle.CAP
                v.status = vehicle.status  # Update status in database
                self.vehicle_service.update_vehicle(v)
                logger.debug(f"[MainWindow] Updated vehicle record: {vehicle.getName()} (ID: {vehicle.id})")

        except Exception as e:
            logger.error(f"[MainWindow] Failed to save vehicle {vehicle.getName()}: {e}")
            raise  # Re-raise the exception to make the error visible

    def fetchVehicleStatus(self, rows):
        cmd = '{\"cmd\":\"reqStatusUpdate\", \"missions\":\"all\"}'
        effective_rows = list(filter(lambda r: r >= 0, rows))
        if len(effective_rows) > 0:
            self.sendToPlatoonsByRowIdxs(effective_rows, cmd)

    def sendPlatoonCommand(self, command, rows, mids):
        self.showMsg("hello???")
        if command == "refresh":
            # cmd = '{\"cmd\":\"reqStatusUpdate\", \"missions\":\"all\"}'
            cmd = {"cmd":"reqStatusUpdate", "missions":"all"}
        elif command == "halt":
            # cmd = '{\"cmd\":\"reqHaltMissions\", \"missions\":\"all\"}'
            cmd = {"cmd":"reqHaltMissions", "missions":"all"}
        elif command == "resume":
            # cmd = '{\"cmd\":\"reqResumeMissions\", \"missions\":\"all\"}'
            cmd = {"cmd":"reqResumeMissions", "missions":"all"}
        elif command == "cancel this":
            mission_list_string = ','.join(str(x) for x in mids)
            # cmd = '{\"cmd\":\"reqCancelMissions\", \"missions\":\"'+mission_list_string+'\"}'
            cmd = {"cmd":"reqCancelMissions", "missions": mission_list_string}
        elif command == "cancel all":
            # cmd = '{\"cmd\":\"reqCancelAllMissions\", \"missions\":\"all\"}'
            cmd = {"cmd":"reqCancelAllMissions", "missions":"all"}
        else:
            # cmd = '{\"cmd\":\"ping\", \"missions\":\"all\"}'
            cmd = {"cmd":"ping", "missions":"all"}

        self.showMsg("cmd is: "+cmd)
        if len(rows) > 0:
            effective_rows = list(filter(lambda r: r >= 0, rows))
        else:
            effective_rows = []

        self.showMsg("effective_rows:"+json.dumps(effective_rows))
        self.sendToPlatoonsByRowIdxs(effective_rows, cmd)


    def cancelVehicleMission(self, rows):
        # cmd = '{\"cmd\":\"reqCancelMission\", \"missions\":\"all\"}'
        cmd = {"cmd": "reqCancelMission", "missions": "all"}
        effective_rows = list(filter(lambda r: r >= 0, rows))
        if len(effective_rows) > 0:
            self.sendToPlatoonsByRowIdxs(effective_rows, cmd)

    # this function sends commands to platoon(s)
    def sendToPlatoonsByRowIdxs(self, idxs, cmd={"cmd": "ping"}):
        # this shall bring up a windows, but for now, simply send something to a platoon for network testing purpose...
        #if self.platoonWin == None:
        #    self.platoonWin = PlatoonWindow(self)
        #self.BotNewWin.resize(400, 200)
        #self.platoonWin.show()
        self.showMsg("sending commands.....")
        self.showMsg("tcp connections....."+json.dumps([flk["ip"] for flk in fieldLinks]))

        if len(idxs) == 0:
            idxs = range(len(fieldLinks))  # Use fieldLinks count instead of GUI model

        # if not self.tcpServer == None:
        if len(fieldLinks) > 0:
            self.showMsg("Currently, there are ("+str(len(fieldLinks))+") connection to this server.....")
            for i in range(len(fieldLinks)):
                if i in idxs:
                    self.send_json_to_platoon(fieldLinks[i], cmd)
                    self.showMsg("cmd sent on link:"+str(i)+":"+json.dumps(cmd))
        else:
            self.showMsg("Warning..... TCP server not up and running yet...")

    # this function sends commands to platoon(s)
    def sendToVehicleByVip(self, vip, cmd={"cmd": "ping"}):
        self.showMsg("sending commands to vehicle by vip.....")
        self.showMsg("tcp connections....." + vip + " " + json.dumps([flk["ip"] for flk in fieldLinks]))

        link = next((x for i, x in enumerate(fieldLinks) if x["ip"] == vip), None)

        # if not self.tcpServer == None:
        if link:
            self.send_json_to_platoon(link, cmd)
            self.showMsg("cmd sent on link:" + str(vip) + ":" + json.dumps(cmd))
        else:
            self.showMsg("Warning..... TCP server not up and running yet...")


    # This function translate bots data structure matching ebbot.py to Json format for file storage.
    def genBotsJson(self):
        bjs = []
        for bot in self.bots:
            self.showMsg("bot gen json0...." + str(len(self.bots)))
            bjs.append(bot.genJson())
        #self.showMsg(json.dumps(bjs))
        return bjs


    # This function translate bots data from Json format to the data structure matching ebbot.py
    def translateBotsJson(self):
        for bj in self.botJsonData:
            new_bot = EBBOT(self)
            new_bot.setJsonData(bj)
            self.bots.append(new_bot)


    def readBotJsonFile(self):
        if exists(self.file_resource.BOTS_FILE):
            with open(self.file_resource.BOTS_FILE, 'r') as file:
                self.botJsonData = json.load(file)
                self.translateBotsJson(self.botJsonData)

            file.close()


    def readVehicleJsonFile(self):
        self.showMsg("Reading Vehicle Json File: "+self.VEHICLES_FILE)
        self.vehiclesJsonData = self.config_manager.get_vehicles()
        if self.vehiclesJsonData:
            self.translateVehiclesJson(self.vehiclesJsonData)
        else:
            self.showMsg("WARNING: Vehicle Json File NOT FOUND: " + self.VEHICLES_FILE)

    def translateVehiclesJson(self, vjds):
        all_vnames = [v.getName() for v in self.vehicles]
        print("vehicles names in the vehicle json file:", all_vnames)
        for vjd in vjds:
            if vjd["name"] not in all_vnames:
                print("add new vehicle to local vehicle data structure but no yet added to GUI", vjd["name"])
                new_v = VEHICLE(self)
                new_v.loadJson(vjd)
                new_v.setStatus("offline")      # always set to offline when load from file. will self correct as we update it later....
                self.saveVehicle(new_v)
                self.vehicles.append(new_v)
                # Vehicle GUI model removed - vehicles managed through PlatoonManager
                if hasattr(self, 'platoon_manager') and self.platoon_manager:
                    self.platoon_manager.add_vehicle(new_v)
            else:
                if "test_disabled" in vjd:
                    foundV = self.getVehicleByName(vjd["name"])
                    foundV.setTestDisabled(vjd["test_disabled"])

    def saveVehiclesJsonFile(self):
        if self.VEHICLES_FILE:
            try:
                vehiclesdata = []
                for v in self.vehicles:
                    vehiclesdata.append(v.genJson())

                self.showMsg("WRITE TO VEHICLES_FILE: " + self.VEHICLES_FILE)
                self.vehiclesJsonData = vehiclesdata
                return self.config_manager.save_vehicles(vehiclesdata)
            except Exception as e:
                logger.error(f"Unable to save file: {self.VEHICLES_FILE}, error: {e}")
                return False
        else:
            self.showMsg("Vehicles json file does NOT exist.")


    def translateInventoryJson(self):
        #self.showMsg("Translating JSON to data......."+str(len(self.sellerInventoryJsonData)))
        for bj in self.sellerInventoryJsonData:
            new_inventory = INVENTORY()
            new_inventory.setJsonData(bj)
            self.inventories.append(new_inventory)


    def readSellerInventoryJsonFile(self, inv_file):
        if inv_file == "":
            inv_file_name = self.product_catelog_file
        else:
            inv_file_name = inv_file

        self.showMsg("product catelog file: "+inv_file_name)
        if exists(inv_file_name):
            self.showMsg("Reading inventory file: "+inv_file_name)
            with open(inv_file_name, 'r', encoding='utf-8') as file:
                self.sellerInventoryJsonData = json.load(file)
                self.translateInventoryJson()
        else:
            self.showMsg("NO inventory file found!")

    def getSellerProductCatelog(self):
        return self.sellerInventoryJsonData

    # This function translate bots data structure matching ebbot.py to Json format for file storage.
    def genMissionsJson(self):
        mjs = []
        for mission in self.missions:
            self.showMsg("mission gen json0...." + str(len(self.missions)))
            mjs.append(mission.genJson())
        #self.showMsg(json.dumps(bjs))
        return mjs


    # This function translate bots data from Json format to the data structure matching ebbot.py
    def translateMissionsJson(self):
        for mj in self.missionJsonData:
            new_mission = EBMISSION()
            new_mission.setJsonData(mj)
            self.missions.append(new_mission)


    def readMissionJsonFile(self):
        if exists(self.file_resource.MISSIONS_FILE):
            with open(self.file_resource.MISSIONS_FILE, 'r') as file:
                self.missionJsonData = json.load(file)
                self.translateMissionsJson(self.missionJsonData)

    def readCSVFiles(self):
        # read files from the local disk and. bot file in csv file format.
        # text, icon, ebtype, email, empw, phone, backemail, acctpw, fn, ln,
        # pfn, pln, pnn, loc, age, mf, interests, role,
        # platform, os, machine, browser, past_schedule, next_schedule
        # state(green/mature), state_start_date, last_walk_date, last_rv_date,
        names_path = 'C:/CrawlerData/names'
        nfiles = os.listdir(names_path)
        nfiles = list(filter(lambda f: f.endswith('.csv'), nfiles))
        nfiles = list(filter(lambda f: f.startswith('bot_'), nfiles))
        for bf in nfiles:
            botjson = {}
            with open((names_path + bf), 'r') as read_obj:
                # pass the file object to reader() to get the reader object
                csv_reader = reader(read_obj)
                rows = list(csv_reader)
                i=0
                # Iterate over each row in the csv using reader object
                botjson["text"] = rows[1][i]
                i = i + 1
                botjson["icon"] = rows[1][i]
                i = i + 1
                botjson["ebtype"] = rows[1][i]
                i = i + 1
                botjson["private_profile"] = {}
                botjson["private_profile"]["fn"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["ln"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["email"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["emailpw"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["acctpw"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["phone"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["backemail"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["prox0"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["prox1"] = rows[1][i]
                i = i + 1
                botjson["private_profile"]["prox2"] = rows[1][i]
                i = i + 1
                botjson["public_profile"] = {}
                botjson["public_profile"]["pfn"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["fln"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["fnn"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["loc"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["age"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["mf"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["interests"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["role"] = rows[1][i]
                i = i + 1
                botjson["public_profile"]["org"] = rows[1][i]
                i = i + 1
                botjson["settings"] = {}
                botjson["settings"]["platform"] = rows[1][i]
                i = i + 1
                botjson["settings"]["browser"] = rows[1][i]
                i = i + 1
                botjson["settings"]["machine"] = rows[1][i]
                i = i + 1
                botjson["settings"]["os"] = rows[1][i]
                i = i + 1
                botjson["status"] = {}
                botjson["status"]["state"] = rows[1][i]
                i = i + 1
                botjson["status"]["levels"] = rows[1][i]
                i = i + 1
                botjson["status"]["lvl_start_date"] = rows[1][i]
                i = i + 1
                botjson["status"]["last_walk_date"] = rows[1][i]
                i = i + 1
                botjson["status"]["last_fb_date"] = rows[1][i]
                i = i + 1
                botjson["status"]["last_mi_date"] = rows[1][i]
                i = i + 1
                botjson["status"]["walks_in1m"] = rows[1][i]
                i = i + 1
                botjson["status"]["mi_in1m"] = rows[1][i]
                i = i + 1
                botjson["status"]["fb_in1m"] = rows[1][i]
                i = i + 1

    def scheduleCalendarView(self):
        # Logic for the bot-mission-scheduler
        # pop out a new windows for user to view and schedule the missions.
        # at the moment, just add an icon.
        #new_bot = EBBOT(self)
        # Icon removed - no longer needed for UI
        #self.centralWidget.setText("<b>File > New</b> clicked")
        self.schedule_manager = ScheduleManager()
        #self.BotNewWin.resize(400, 200)
        # Note: ScheduleManager is now a data handler, not a GUI window
        # You may need to implement a new GUI or use existing schedule display methods

    def newMissionView(self):
        if self.missionWin == None:
            self.missionWin = MissionManager(self)
            self.missionWin.setOwner(self.owner)
        else:
            self.missionWin.setMode("new")

        # Note: MissionManager is now a data handler, not a GUI window
        # You may need to implement a new GUI or use existing mission display methods
        self.showMsg("Mission manager created for new mission")

    def newVehiclesView(self):
        if self.platoon_manager == None:
            self.showMsg("creating platoon manager....")
            self.platoon_manager = PlatoonManager(self, "init")
        else:
            self.showMsg("Shows existing platoon manager...")
        # Note: PlatoonManager is now a data handler, not a GUI window
        # You may need to implement a new GUI or use existing platoon display methods

    def eventFilter(self, source, event):
        return super().eventFilter(source, event)

    def editCusMission(self):
        # File actions
        if self.missionWin:
            self.showMsg("populating mission data............")
            self.missionWin.setMission(self.selected_cus_mission_item)
        else:
            self.showMsg("creating a new mission manager............")
            self.missionWin = MissionManager(self)
            self.showMsg("done create mission manager............"+str(self.selected_cus_mission_item.getMid())+" skills:"+self.selected_cus_mission_item.getSkills())
            self.missionWin.setMission(self.selected_cus_mission_item)

        self.missionWin.setMode("update")
        # Note: MissionManager is now a data handler, not a GUI window
        self.showMsg("edit mission" + str(self.selected_mission_row))


    def cloneCusMission(self):
        # File actions
        new_mission = self.selected_cus_mission_item
        # new_bot.setText()
        self.addNewMissions([new_mission])
        self.searchLocalMissions()

    def deleteCusMission(self):
        # File actions - confirmation dialog removed, proceeding with deletion
        logger.info("Deleting mission - confirmation dialog removed")
        api_removes = []

        # QListView and missionModel removed - UI components no longer needed
        # Mission deletion now handled through data models only
        if len(self.missions):
            for mission in self.missions:
                api_removes.append({"id": mission.getMid(), "owner": "", "reason": ""})

            # remove on the cloud side, local DB side, and MainGUI side
            self.deleteMissionsWithJsons(False, api_removes)

    # note: the mjs is in this format [{"id": mid, "owner": "", "reason": ""} .... ]
    def deleteMissionsWithJsons(self, del_gui, mjs):
        try:
            # remove on the cloud side
            if del_gui:
                print("delete GUI missions")

            # remove on the cloud side
            jresp = send_remove_missions_request_to_cloud(self.session, mjs,
                                                          self.get_auth_token(), self.getWanApiEndpoint())
            self.showMsg("DONE WITH CLOUD SIDE REMOVE MISSION REQUEST.....")
            if "errorType" in jresp:
                screen_error = True
                self.showMsg(
                    "Delete Missions ERROR Type: " + json.dumps(jresp["errorType"]) + "ERROR Info: " + json.dumps(
                        jresp["errorInfo"]))
            else:
                self.showMsg("JRESP:" + json.dumps(jresp) + "<>" + json.dumps(jresp['body']) + "<>" + json.dumps(
                    jresp['body']['$metadata']) + "<>" + json.dumps(jresp['body']['numberOfRecordsUpdated']))
                meta_data = jresp['body']['$metadata']
                if jresp['body']['numberOfRecordsUpdated'] == 0:
                    self.showMsg("WARNING: CLOUD SIDE MISSION DELETE NOT EXECUTED.")

                for m in mjs:
                    # missionTBDId = next((x for x in self.missions if x.getMid() == m["id"]), None)
                    self.mission_service.delete_missions_by_mid(m["id"])

                for m in mjs:
                    midx = next((i for i, x in enumerate(self.missions) if x.getMid() == m["id"]), -1)
                    self.showMsg("removeing MID:" + str(midx))
                    # If the element was found, remove it using pop()
                    if midx != -1:
                        self.missions.pop(midx)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorDeleteMissionsWithJsons:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorCreateMissionsWithJsons: traceback information not available:" + str(e)
            log3(ex_stat)


    def updateCusMissionStatus(self, amission):
        # send this mission's status to Cloud
        api_missions = [amission]
        # jresp = send_update_missions_request_to_cloud(self.session, api_missions, self.get_auth_token(), self.getWanApiEndpoint())
        # if "errorType" in jresp:
        #     screen_error = True
        #     self.showMsg("Delete Bots ERROR Type: "+json.dumps(jresp["errorType"])+"ERROR Info: "+json.dumps(jresp["errorInfo"]))
        # else:
        #     jbody = json.loads(jresp["body"])
        #     # now that delete is successfull, update local file as well.
        #     self.writeMissionJsonFile()


    def updateMissionsWithJsData(self, mjs):
        try:
            api_missions = []
            for amission in mjs:
                api_missions.append({
                    'mid': amission["pubAttributes"]["missionId"],
                    'ticket': amission["pubAttributes"]["ticket"],
                    'botid': amission["pubAttributes"]["bot_id"],
                    'status': amission["pubAttributes"]["status"],
                    'createon': amission["pubAttributes"]["createon"],
                    'esd': amission["pubAttributes"]["esd"],
                    'ecd': amission["pubAttributes"]["ecd"],
                    'asd': amission["pubAttributes"]["asd"],
                    'abd': amission["pubAttributes"]["abd"],
                    'aad': amission["pubAttributes"]["aad"],
                    'afd': amission["pubAttributes"]["afd"],
                    'acd': amission["pubAttributes"]["acd"],
                    'actual_start_time': amission["pubAttributes"]["actual_start_time"],
                    'est_start_time': amission["pubAttributes"]["est_start_time"],
                    'actual_runtime': amission["pubAttributes"]["actual_run_time"],
                    'est_runtime': amission["pubAttributes"]["est_run_time"],
                    'n_retries': amission["pubAttributes"]["repeat"],
                    'cuspas': amission["pubAttributes"]["cuspas"],
                    'category': amission["pubAttributes"]["category"],
                    'phrase': amission["pubAttributes"]["phrase"],
                    'pseudoStore': amission["pubAttributes"]["pseudo_store"],
                    'pseudoBrand': amission["pubAttributes"]["pseudo_brand"],
                    'pseudoASIN': amission["pubAttributes"]["pseudo_asin"],
                    'type': amission["pubAttributes"]["ms_type"],
                    'config': amission["pubAttributes"]["config"],
                    'skills': amission["pubAttributes"]["skills"],
                    'delDate': amission["pubAttributes"]["del_date"],
                    'asin': amission["privateProfile"]["item_number"],
                    'stores': amission["privateProfile"]["seller"],
                    'follow_seller': amission["privateProfile"]["follow_seller"],
                    'brand': amission["privateProfile"]["brand"],
                    'img': amission["privateProfile"]["imglink"],
                    'title': amission["privateProfile"]["title"],
                    'variations': amission["privateProfile"]["variations"],
                    'rating': amission["privateProfile"]["rating"],
                    'feedbacks': amission["privateProfile"]["feedbacks"],
                    'price': amission["privateProfile"]["price"],
                    'follow_price': amission["privateProfile"]["follow_price"],
                    'fingerprint_profile': amission["privateProfile"]["fingerprint_profile"],
                    'customer': amission["privateProfile"]["customer_id"],
                    'platoon': amission["pubAttributes"]["platoon_id"],
                    'result': amission["privateProfile"]["result"],
                    'as_server': amission["pubAttributes"]["as_server"],
                    'original_req_file': amission["privateProfile"]["original_req_file"]
                })

            jresp = send_update_bots_request_to_cloud(self.session, mjs, self.get_auth_token(), self.getWanApiEndpoint())
            if "errorType" in jresp:
                screen_error = True
                self.showMsg("ERROR Type: " + json.dumps(jresp["errorType"]),
                             "ERROR Info: " + json.dumps(jresp["errorInfo"]))
            else:
                jbody = jresp["body"]
                if jbody['numberOfRecordsUpdated'] == len(mjs):
                    self.bot_service.update_bots_batch(api_missions)

                    # finally update into in-memory data structure.

                else:
                    self.showMsg("WARNING: bot NOT updated in Cloud!")

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorUpdateMissionsWithJsData:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorUpdateMissionsWithJsData: traceback information not available:" + str(e)

            self.showMsg(ex_stat)


    def runCusMissionNowSync(self):
        print("force mission to run now")
        asyncio.create_task(self.runCusMissionNow(self.selected_cus_mission_item, self.gui_rpa_msg_queue, self.gui_monitor_msg_queue))

    async def runCusMissionNow(self, amission, gui_rpa_queue, gui_monitor_queue):
        # check if psk is already there, if not generate psk, then run it.
        self.showMsg("run mission now...."+str(amission.getBid()))
        executor = self.getBotByID(amission.getBid())

        tempMissionTasks = [{
            "name": amission.getType(),
            "mid": amission.getMid(),
            "ticket": amission.getTicket(),
            "cuspas": amission.getCusPAS(),
            "bid": amission.getBid(),
            "skills": amission.getSkills(),
            "config": amission.getConfig(),
            "trepeat": 1,
            "fingerprint_profile": amission.getFingerPrintProfile(),
            "start_time": 1            # make this task due 00:20 am, which should have been passed by now, so to catch up, the schedule will run this at the first possible chance.
        }]

        # ads_profile_batches_fnames = genAdsProfileBatchs(self, self.ip, tempMissionTasks)
        print("updated tempMissionTasks:", tempMissionTasks)
        widx = len(self.todays_work["tbd"])
        self.todays_work["tbd"].append({"name": "automation", "works": tempMissionTasks, "status": "Assigned", "current widx": widx, "completed": [], "aborted": []})


    async def markCusMissionCompleted(self, amission, gui_rpa_queue, gui_monitor_queue):
        # check if psk is already there, if not generate psk, then run it.
        self.showMsg("run mission now...."+str(amission.getBid()))
        amission.setStatus("Completed:0")
        if "Commander" in self.host_role:
            self.updateMissionsStatToCloud([amission])

    def editBot(self):
        if self.bot_manager == None:
            self.bot_manager = BotManager(self)
        self.bot_manager.load_bot(self.selected_bot_item)

        # Note: BotManager is now a data handler, not a GUI window
        # You may need to implement a new GUI or use existing bot display methods
        self.bot_manager.set_mode("update")
        self.showMsg("edit bot" + str(self.selected_bot_row))

    def cloneBot(self):
        # File actions
        new_bot = self.selected_bot_item
        # new_bot.setText()
        self.addNewBots([new_bot])
        self.searchLocalBots()

    def deleteBot(self):
        # File actions - confirmation dialog removed, proceeding with deletion
        logger.info("Deleting bot - confirmation dialog removed")
        api_removes = []
        # items = [self.selected_bot_item]
        # QListView and botModel removed - UI components no longer needed
        # Bot deletion now handled through data models only
        if len(self.bots):
            for bot in self.bots:
                api_removes.append({"id": bot.getBid(), "owner": "", "reason": ""})

            # remove on the cloud side, local side, MainGUI side.
            self.deleteBotsWithJsons(False, api_removes)

    # note: the bjs is in this format [{"id": bid, "owner": "", "reason": ""} .... ]
    def deleteBotsWithJsons(self, del_gui, bjs):
        try:
            # remove on the cloud side
            if del_gui:
                print("delete GUI bots")

            # now the common part.
            jresp = send_remove_bots_request_to_cloud(self.session, bjs,
                                                      self.get_auth_token(), self.getWanApiEndpoint())
            self.showMsg("DONE WITH CLOUD SIDE REMOVE BOT REQUEST.....")
            if "errorType" in jresp:
                screen_error = True
                self.showMsg("Delete Bots ERROR Type: " + json.dumps(jresp["errorType"]) + "ERROR Info: " + json.dumps(
                    jresp["errorInfo"]))
            else:
                self.showMsg("JRESP:" + json.dumps(jresp) + "<>" + json.dumps(jresp['body']))
                if jresp['body']['numberOfRecordsUpdated'] == 0:
                    self.showMsg("WARNING: CLOUD SIDE DELETE NOT EXECUTED.")

                for b in bjs:
                    botTBDId = next((x for x in self.bots if x.getBid() == b["id"]), None)
                    self.bot_service.delete_bots_by_botid(b["id"])

                for b in bjs:
                    bidx = next((i for i, x in enumerate(self.bots) if x.getBid() == b["id"]), -1)

                    # If the element was found, remove it using pop()
                    if bidx != -1:
                        self.bots.pop(bidx)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorDeleteBotsWithJsons:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorCreateBotsWithJsons: traceback information not available:" + str(e)
            log3(ex_stat)


    # data format conversion. nb is in EBBOT data structure format., nbdata is json
    def fillNewBotPubInfo(self, nbjson, nb):
        self.showMsg("filling bot public data for bot-" + str(nbjson["pubProfile"]["bid"]))
        nb.setNetRespJsonData(nbjson)

    def fillNewBotFullInfo(self, nbjson, nb):
        self.showMsg("filling bot data for bot-" + str(nbjson["pubProfile"]["bid"]))
        nb.loadJson(nbjson)

    # this function can only be called by a manager or HR head.
    def syncBotAccounts(self):
        try:
            # run a hook function to bring in external accounts.
            acctRows = self.runGetBotAccountsHook()
            print("ACCT ROWS:", acctRows)
            # then from there, figure out newly added accounts
            # from newly added accounts, screen the ones ready to be converted to a Bot/Agent
            # rows are updated....
            qualified, rowsNeedUpdate, botsNeedUpdate, vehiclesNeedUpdate = self.screenBuyerBotCandidates(acctRows, self.bots)
            print("qualified:", len(qualified), qualified)
            print("rowsNeedUpdate:", len(rowsNeedUpdate), rowsNeedUpdate)
            print("botsNeedUpdate:", len(botsNeedUpdate), [b.getAddr() for b in botsNeedUpdate])
            # turn qualified acct into bots/agents
            self.hireBuyerBotCandidates(qualified)
            # create new ads power profile for the newly added accounts.

            # genInitialADSProfiles(qualified)

            # call another hook function update the rowsNeedUpdate
            rowsNeedUpdate = rowsNeedUpdate + qualified
            results = self.runUpdateBotAccountsHook(rowsNeedUpdate)

            self.updateBots(botsNeedUpdate)

            if vehiclesNeedUpdate:
                self.updateVehicles(vehiclesNeedUpdate)

        except Exception as e:
            # Log and skip errors gracefully
            ex_stat = f"Error in SyncBotAccounts: {traceback.format_exc()} {str(e)}"
            print(f"{ex_stat}")


    def runGetBotAccountsHook(self):
        try:
            params = {"all": True}      # this will get all rows in accounts table.
            runStat = self.runExternalHook("hr_recruit_get_candidates_hook", params)
            # runStat = self.runExternalHook("get_accounts_hook", params)
            print("runStat:", runStat)
            if "Complete" in runStat:
                acctRows = symTab["hook_result"]["candidates"]

        except Exception as e:
            # Log and skip errors gracefully
            ex_stat = f"Error in runGetBotAccountsHook: {traceback.format_exc()} {str(e)}"
            print(f"{ex_stat}")

        return acctRows

    def runTeamPrepHook(self):
        try:
            params = {"all": True}      # this will get all rows in accounts table.
            runStat = self.runExternalHook("team_prep_hook", params)
            # runStat = self.runExternalHook("get_accounts_hook", params)
            print("runStat:", runStat)
            if "Complete" in runStat:
                runnable_work = symTab["hook_result"]["candidates"]

        except Exception as e:
            # Log and skip errors gracefully
            ex_stat = f"Error in runTeamPrepHook: {traceback.format_exc()} {str(e)}"
            print(f"{ex_stat}")

        return runnable_work

    def runUpdateBotAccountsHook(self, rows):
        try:
            params = {"rows": rows}
            runStat = self.runExternalHook("update_accounts_hook", params)

        except Exception as e:
            # Log and skip errors gracefully
            ex_stat = f"Error in runUpdateBotAccountsHook: {traceback.format_exc()} {str(e)}"
            print(f"{ex_stat}")

        return runStat

    def createBotsFromFilesOrJsData(self, bfiles):
        try:
            bots_from_file = []
            botsJson = []
            for filename in bfiles:
                if filename:
                    if isinstance(filename, str):
                        if "json" in filename:
                            try:
                                api_bots = []
                                with open(filename, 'r', encoding='utf-8') as uncompressed:
                                    filebbots = json.load(uncompressed)
                                    if filebbots:
                                        # Add bots to the relevant data structure and add these bots to the cloud and local DB.
                                        for fb in filebbots:
                                            new_bot = EBBOT(self)
                                            self.fillNewBotFullInfo(fb, new_bot)
                                            bots_from_file.append(new_bot)
                                    else:
                                        self.warn("Warning: NO bots found in file.")
                            except (FileNotFoundError, json.JSONDecodeError) as e:
                                self.warn(f"Error opening or decoding JSON file: {filename} - {e}")

                        elif "xlsx" in filename:
                            try:
                                log3("working on file:" + str(filename))
                                import openpyxl
                                xls = openpyxl.load_workbook(filename, data_only=True)
                                botsJson = []
                                title_cells = []

                                # Process each sheet in the Excel file
                                for idx, sheet in enumerate(xls.sheetnames):
                                    ws = xls[sheet]

                                    for ri, row in enumerate(ws.iter_rows(values_only=True)):
                                        # Capture header titles from the first row of the first sheet
                                        if idx == 0 and ri == 0:
                                            title_cells = [cell for cell in row]
                                        elif ri > 0 and len(row) == len(title_cells):  # Ensure row length matches headers
                                            botJson = {}
                                            for ci, cell in enumerate(title_cells):
                                                # Format dates if necessary
                                                if cell == "DoB" and row[ci]:
                                                    botJson[cell] = row[ci].strftime('%Y-%m-%d')
                                                else:
                                                    botJson[cell] = row[ci]
                                            botsJson.append(botJson)

                                log3("total # of bot rows read:" + str(len(botsJson)))
                                log3("all jsons from bot xlsx file:" + json.dumps(botsJson, ensure_ascii=False))
                                for bjson in botsJson:
                                    new_bot = EBBOT(self)
                                    new_bot.loadXlsxData(bjson)
                                    bots_from_file.append(new_bot)
                                    print(new_bot.genJson())

                            except FileNotFoundError as e:
                                self.warn(f"Excel file not found: {filename} - {e}")
                            except Exception as e:
                                self.warn(f"Error processing Excel file: {filename} - {e}")

                        else:
                            self.showMsg("ERROR: bot files must either be in .json format or .xlsx format!")
                    else:
                        # this is the case where input is already in json format, so just directly use them.
                        jsData = filename

                        new_bot = EBBOT(self)
                        self.fillNewBotFullInfo(jsData, new_bot)
                        self.assignBotVehicle(new_bot)
                        bots_from_file.append(new_bot)

                else:
                    self.warn("No file provided.")

            if len(bots_from_file) > 0:
                print("adding new bots to both cloud and local DB... update BID and Interests along the way since they're cloud generated.")
                self.addNewBots(bots_from_file)
                firstAddedBotId = bots_from_file[0].getBid()
                return firstAddedBotId, bots_from_file

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorCreateBotsFromFilesOrJsData:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorCreateBotsFromFilesOrJsData: traceback information not available:" + str(e)
            log3(ex_stat)
            return 0, []

    # data format conversion. nb is in EBMISSION data structure format., nbdata is json
    def fillNewMissionFromCloud(self, nmjson, nm):
        self.showMsg("filling mission data")
        nm.setNetRespJsonData(nmjson)

    def addMissionsToLocalDB(self, missions: List[EBMISSION]):
        local_missions: List[MissionModel] = []

        # Extract all mids from the new missions
        new_mids = [mission.getMid() for mission in missions]

        # Query existing mids in the local database
        existing_missions = self.mission_service.find_missions_by_mids(new_mids)
        existing_mids = {mission.mid for mission in existing_missions}

        for new_mission in missions:
            if new_mission.getMid() in existing_mids:
                log3(f"Mission with mid {new_mission.getMid()} already exists. Skipping.", "debug", self)
                continue

            local_mission = MissionModel()
            local_mission.mid = new_mission.getMid()
            local_mission.ticket = new_mission.getTicket()
            local_mission.botid = new_mission.getBid()
            local_mission.status = new_mission.getStatus()
            local_mission.createon = new_mission.getBD()
            local_mission.owner = self.owner
            local_mission.esd = new_mission.getEsd()
            local_mission.ecd = new_mission.getEcd()
            local_mission.asd = new_mission.getAsd()
            local_mission.abd = new_mission.getAbd()
            local_mission.aad = new_mission.getAad()
            local_mission.afd = new_mission.getAfd()
            local_mission.acd = new_mission.getAcd()
            local_mission.actual_start_time = new_mission.getActualStartTime()
            local_mission.est_start_time = new_mission.getEstimatedStartTime()
            local_mission.actual_runtime = new_mission.getActualRunTime()
            local_mission.est_runtime = new_mission.getEstimatedRunTime()
            local_mission.n_retries = new_mission.getNRetries()
            local_mission.cuspas = new_mission.getCusPAS()
            local_mission.category = new_mission.getSearchCat()
            local_mission.phrase = new_mission.getSearchKW()
            local_mission.pseudoStore = new_mission.getPseudoStore()
            local_mission.pseudoBrand = new_mission.getPseudoBrand()
            local_mission.pseudoASIN = new_mission.getPseudoASIN()
            local_mission.type = new_mission.getType()
            local_mission.config = json.dumps(new_mission.getConfig())
            local_mission.skills = new_mission.getSkills()
            local_mission.delDate = new_mission.getDelDate()
            local_mission.asin = new_mission.getASIN()
            local_mission.store = new_mission.getStore()
            local_mission.follow_seller = new_mission.getFollowSeller()
            local_mission.brand = new_mission.getBrand()
            local_mission.img = new_mission.getImagePath()
            local_mission.title = new_mission.getTitle()
            local_mission.rating = new_mission.getRating()
            local_mission.feedbacks = new_mission.getFeedbacks()
            local_mission.price = new_mission.getPrice()
            local_mission.follow_price = new_mission.getFollowPrice()
            local_mission.fingerprint_profile = new_mission.getFingerPrintProfile()
            local_mission.original_req_file = new_mission.getReqFile()
            local_mission.customer = new_mission.getCustomerID()
            local_mission.platoon = new_mission.getPlatoonID()
            local_mission.result = new_mission.getResult()
            local_mission.variations = new_mission.getVariations()
            local_mission.as_server = new_mission.getAsServer()
            local_missions.append(local_mission)

        if local_missions:
            self.mission_service.insert_missions_batch_(local_missions)

    def createMissionsFromFilesOrJsData(self, mfiles):
        missionsJson = []
        mTypeTable = {
            "browse": "browse",
            "free_review": "buy",
            "direct_review": "directbuy",
            "product_rating": "goodRating",
            "product_feedback": "goodFB",
            "store_rating": "storeRating",
            "store_feedback": "storeFB",
            "add_to_cart": "addCart",
        }
        for filename in mfiles:
            if filename != "":
                if isinstance(filename, str):
                    dataType = "file"
                    if "json" in filename:
                        api_missions = []
                        # self.showMsg("body string:"+uncompressed+"!"+str(len(uncompressed))+"::")
                        dataType = "missionJSFile"
                        with open(filename, 'r', encoding='utf-8') as f:
                            filebmissions = json.load(f)
                        if len(filebmissions) > 0:
                            #add bots to the relavant data structure and add these bots to the cloud and local DB.

                            jresp = send_add_missions_request_to_cloud(self.session, filebmissions,
                                                                   self.get_auth_token(), self.getWanApiEndpoint())

                            if "errorType" in jresp:
                                screen_error = True
                                self.showMsg("ERROR Type: "+json.dumps(jresp["errorType"])+"ERROR Info: "+json.dumps(jresp["errorInfo"]))
                            else:
                                self.showMsg("jresp type: "+str(type(jresp))+" "+str(len(jresp["body"])))
                                jbody = jresp["body"]
                                # now that add is successfull, update local file as well.

                                # now add missions to local DB.
                                new_missions: List[EBMISSION] = []
                                for i in range(len(jbody)):
                                    self.showMsg(str(i))
                                    new_mission = EBMISSION(self)
                                    # move json file based mission into MISSION data structure.
                                    new_mission.loadJson(filebmissions)
                                    self.fillNewMissionFromCloud(jbody[i], new_mission)
                                    self.missions.append(new_mission)
                                    # missionModel removed - UI components no longer needed
                                    new_missions.append(new_mission)

                                if not self.config_manager.general_settings.debug_mode:
                                    self.addMissionsToLocalDB(new_missions)

                        else:
                            self.warn("NO missions found in file.")

                    elif "xlsx" in filename:
                        dataType = "businessXlsxFile"
                        # if getting missions from xlsx file it's automatically assumed that the
                        # the mission will be for amz buy.
                        log3("working on order file:"+filename)
                        mJsons = self.convert_orders_xlsx_to_json(filename)
                        log3("mJsons from xlsx:" + json.dumps(mJsons))

                        # now if quantity is N, there will be N missions created.
                        # and add other required missions parameters....
                        for mJson in mJsons:
                            if "email" not in mJson:
                                pkString = "songc@yahoo.com"
                            elif not mJson["email"]:
                                pkString = "songc@yahoo.com"
                            else:
                                pkString = mJson["email"]
                            mJson["pseudoStore"] = self.generateShortHash(pkString+":"+mJson.get("stores", "NoneStore"))
                            mJson["pseudoBrand"] = self.generateShortHash(pkString+":"+mJson.get("brand", "NoneBrand"))
                            mJson["pseudoASIN"] = self.generateShortHash(pkString+":"+mJson["asin"])

                            ft = mJson.get("feedback_type")
                            if ft:
                                mJson["type"] = mTypeTable.get(ft, "buy")
                            else:
                                mJson["type"] = "buy"

                            # each buy should be a separate mission.
                            n_orders = int(mJson["quantity"])
                            missionsJson = missionsJson + [copy.deepcopy(mJson) for _ in range(n_orders)]

                        log3("total # of orders rows read: "+str(len(mJsons)))
                        log3("mJsons after conversion:"+json.dumps(mJsons))
                        m = sum(int(item["quantity"]) for item in mJsons)
                        log3("total # of missions to be generated: " + str(m))
                else:
                    log3("add missions from direct list of jsons, no data manipulation here.")
                    dataType = "businessJSData"
                    missionsJson = mfiles

        missions_from_file = []
        for mjson in missionsJson:
            new_mission = EBMISSION(self)
            if dataType == "businessJSData":
                new_mission.loadBusinessesDBData(mjson)
            else:
                new_mission.loadXlsxData(mjson)
            missions_from_file.append(new_mission)
            # new_mission.genJson()

        print("about to really add these missions to cloud and local DB...")
        if missions_from_file:
            # during the process of this the cloud generated mid should be updated to JSON.
            self.addNewMissions(missions_from_file)


        return missionsJson


    def process_original_xlsx_file(self, file_path):
        # Read the Excel file, skipping the first two rows
        import pandas as pd
        df = pd.read_excel(file_path, skiprows=2)

        # Drop rows where all elements are NaN
        df.dropna(how='all', inplace=True)

        # Convert each row to a JSON object and append to a list
        json_list = df.to_dict(orient='records')

        #add a reverse link back to
        for jl in json_list:
            jl["file_link"] = file_path

        print("READ XLSX::", json_list)
        return json_list

    def update_original_xlsx_file(self, file_path, mission_data):
        # Read the Excel file, skipping the first two rows
        dir_path = os.path.dirname(file_path)

        import pandas as pd
        df = pd.read_excel(file_path, skiprows=2)

        # Drop rows where all elements are NaN
        df.dropna(how='all', inplace=True)

        # Convert each row to a JSON object and append to a list
        json_list = df.to_dict(orient='records')

        mission_ids = [mission["mission ID"] for mission in mission_data]
        completion_dates = [mission["completion date"] for mission in mission_data]

        # Add new columns with default or empty values
        df['mission ID'] = mission_ids[:len(df)]
        df['completion date'] = completion_dates[:len(df)]

        # Get the new file name using the first row of the "mission ID" column
        new_mission_id = df.loc[0, 'mission ID']
        base_name = os.path.basename(file_path)
        new_file_name = f"{os.path.splitext(base_name)[0]}_{new_mission_id}.xlsx"
        new_file_path = os.path.join(dir_path, new_file_name)

        # Save the updated DataFrame to a new file
        df.to_excel(new_file_path, index=False)

        print(f"File saved as {new_file_name}")


    def newMissionFromNewReq(self, reqJson, reqFile):
        new_mission = EBMISSION(self)
        new_mission.loadAMZReqData(reqJson, reqFile)
        return new_mission



    # sc - 10/11/2024 - add new missions are they in todays_work?
    def newBuyMissionFromFiles(self):
        dtnow = datetime.now()

        recent = dtnow - timedelta(days=3)
        date_word = dtnow.strftime("%Y%m%d")
        year = dtnow.strftime("%Y")
        month = f"m{dtnow.month}"
        day = f"m{dtnow.day}"

        new_orders_dir = self.my_ecb_data_homepath + "/new_orders/ORDER" + date_word + "/"
        new_orders_dir = os.path.join(self.config_manager.general_settings.new_orders_dir, year, month, day)
        self.showMsg("working on new orders:" + new_orders_dir)

        new_buy_missions = []
        if os.path.isdir(new_orders_dir):
            files = os.listdir(new_orders_dir)
            xlsx_files = [os.path.join(new_orders_dir, file) for file in files if os.path.isfile(os.path.join(new_orders_dir, file)) and file.endswith('.xlsx')]

            #each row of each xlsx file becomes a new mission
            for xlsx_file in xlsx_files:
                # stores, brand, execution time, quantity, asin, search term, title, page number, price, variation, product image, fb type, fb title, fb contents, notes
                buy_mission_reqs = self.process_original_xlsx_file(xlsx_file)

                for buy_req in buy_mission_reqs:
                    n_buys = int(buy_req["quantity"])
                    if len(n_buys) > 0:
                        for n in range(n_buys):
                            print("creating new buy mission:", n)
                            new_buy_missions.append(self.newMissionFromNewReq(buy_req, xlsx_file))

        # now that we have created all the new missions,
        # create the mission in the cloud and local DB.
        # cloud side first

        if len(new_buy_missions) > 0:
            jresp = send_add_missions_request_to_cloud(self.session, new_buy_missions, self.get_auth_token(), self.getWanApiEndpoint())

            if "errorType" in jresp:
                screen_error = True
                self.showMsg( "ERROR Type: " + json.dumps(jresp["errorType"]) + "ERROR Info: " + json.dumps(jresp["errorInfo"]))
            else:
                self.showMsg("jresp type: " + str(type(jresp)) + " " + str(len(jresp["body"])))
                jbody = jresp["body"]
                # now that add is successfull, update local file as well.

                # now update mission ID
                for i in range(len(jbody)):
                    new_buy_missions.setMid(jbody[i]["mid"])

                #now add to local DB.
                if not self.config_manager.general_settings.debug_mode:
                    self.addMissionsToLocalDB(new_buy_missions)

                #add to local data structure
                self.missions = self.missions + new_buy_missions
                # missionModel removed - UI components no longer needed

        return new_buy_missions

    def fillNewSkill(self, nskjson, nsk):
        self.showMsg("filling skill data")
        nsk.setNetRespJsonData(nskjson)

    def showSkillManager(self):
        # Note: SkillManager is now a data handler, not a GUI window
        # You may need to implement a new GUI or use existing skill display methods
        self.showMsg("Skill Manager is now a data handler. Use skill_manager methods directly.")

    # def uploadSkill(self):
    #     filename, _ = QFileDialog.getOpenFileName(
    #         self,
    #         "Upload Skill File",
    #         '',
    #         "Skill Json Files (*.json)"
    #     )
    #     if filename != "":
    #         # ("body string:", uncompressed, "!", len(uncompressed), "::")
    #         sk_dir = os.path.abspath(filename)
    #         anchor_dir = sk_dir + "/" + os.path.basename(filename).split(".")[0] + "/images"
    #         scripts_dir = sk_dir + "/" + os.path.basename(filename).split(".")[0] + "/scripts"
    #         anchor_files = os.listdir(anchor_dir)
    #         for af in anchor_files:
    #             full_af_name = anchor_dir + "/" + af
    #             jresp = upload_file(self.session, full_af_name, self.get_auth_token(),  self.getWanApiEndpoint(), "anchor")

    #         csk_file = scripts_dir + "/" + os.path.basename(filename).split(".")[0] + ".csk"
    #         jresp = upload_file(self.session, csk_file, self.get_auth_token(),  self.getWanApiEndpoint(), "csk")


    # def newSkillFromFile(self):
    #     filename, _ = QFileDialog.getOpenFileName(
    #         self,
    #         "Open Skill File",
    #         '',
    #         "Skill Json Files (*.json)"
    #     )
    #     self.showMsg("loading skill from a file..."+filename)
    #     if filename != "":
    #         api_skills = []
    #         try:
    #             with open(filename, 'r') as new_skill_file:
    #                 # self.showMsg("body string:"+uncompressed+"!"+str(len(uncompressed))+"::")
    #                 skill_json = json.load(new_skill_file)
    #                 if skill_json:
    #                     #add skills to the relavant data structure and add these bots to the cloud and local DB.
    #                     # send_add_skills_to_cloud
    #                     jresp = send_add_skills_request_to_cloud(self.session, [skill_json], self.get_auth_token(), self.getWanApiEndpoint())

    #                     if "errorType" in jresp:
    #                         screen_error = True
    #                         self.showMsg("ERROR Type: "+json.dumps(jresp["errorType"])+"ERROR Info: "+json.dumps(jresp["errorInfo"]))
    #                     else:
    #                         self.showMsg("jresp type: "+str(type(jresp))+" "+str(len(jresp["body"])))
    #                         jbody = jresp["body"]
    #                         # now that add is successfull, update local file as well.

    #                         # now add bot to local DB.

    #                         for i in range(len(jbody)):
    #                             self.showMsg(str(i))
    #                             new_skill = WORKSKILL(self, jbody[i]["name"])
    #                             self.fillNewSkill(jbody[i], new_skill)
    #                             self.skills.append(new_skill)
    #                             # self.skillModel.appendRow(new_skill)
    #                             api_skills.append({
    #                                 "skid": new_skill.getSkid(),
    #                                 "owner": new_skill.getOwner(),
    #                                 "platform": new_skill.getPlatform(),
    #                                 "app": new_skill.getApp(),
    #                                 "applink": new_skill.getAppLink(),
    #                                 "appargs": new_skill.getAppArgs(),
    #                                 "site": new_skill.getSiteName(),
    #                                 "sitelink": new_skill.getSite(),
    #                                 "name": new_skill.getName(),
    #                                 "path": new_skill.getPath(),
    #                                 "main": new_skill.getMain(),
    #                                 "createdon": new_skill.getCreatedOn(),
    #                                 "extensions": "",
    #                                 "runtime": new_skill.getRunTime(),
    #                                 "price_model": new_skill.getPriceModel(),
    #                                 "price": new_skill.getPrice(),
    #                                 "privacy": new_skill.getPrivacy(),
    #                             })
    #                             self.skill_service.insert_skill(api_skills[i])
    #                 else:
    #                     self.warn("NO skills in the file.")
    #         except Exception as e:
    #             traceback_info = traceback.extract_tb(e.__traceback__)
    #             # Extract the file name and line number from the last entry in the traceback
    #             if traceback_info:
    #                 ex_stat = "ErrorLoadSkillFile:" + traceback.format_exc() + " " + str(e)
    #             else:
    #                 ex_stat = "ErrorLoadSkillFile: traceback information not available:" + str(e)
    #             logger.debug(ex_stat)
    #             logger.debug("load skill file error."))

    def find_dependencies(self, main_file, visited, dependencies):
        if main_file in visited:
            return

        visited.add(main_file)

        # "type": "Use Skill",
        # "skill_name": "update_tracking",
        # "skill_path": "public/win_chrome_etsy_orders",
        # "skill_args": "gs_input",
        # "output": "total_label_cost"
        logger.trace("TRYING...."+main_file, "fetchSchedule", self)
        if os.path.exists(main_file):
            logger.trace("OPENING...."+main_file, "fetchSchedule", self)
            try:
                with open(main_file, 'r') as psk_file:
                    code_jsons = json.load(psk_file)
            except json.JSONDecodeError as e:
                logger.error(f"JSON parsing error in file {main_file}: Line {e.lineno}, Column {e.colno}, Position {e.pos}: {e}")
                logger.warning(f"Skipping malformed JSON file: {main_file}")
                return dependencies  # Return current dependencies and continue
            except Exception as e:
                logger.error(f"Error reading file {main_file}: {e}")
                return dependencies

            # go thru all steps.
            for key in code_jsons.keys():
                if "type" in code_jsons[key]:
                    if code_jsons[key]["type"] == "Use Skill":
                        if "public" in code_jsons[key]["skill_path"]:
                            # For public skills, prioritize reading from user data directory, fallback to install directory if not exists
                            user_skill_file = self.ecb_data_homepath + "/resource/skills/" + code_jsons[key]["skill_path"] + "/" + code_jsons[key]["skill_name"] + ".psk"
                            install_skill_file = self.homepath + "/resource/skills/" + code_jsons[key]["skill_path"] + "/" + code_jsons[key]["skill_name"] + ".psk"

                            if os.path.exists(user_skill_file):
                                dependency_file = user_skill_file
                            else:
                                dependency_file = install_skill_file
                        else:
                            dependency_file = self.my_ecb_data_homepath + code_jsons[key]["skill_path"] + "/" + code_jsons[key]["skill_name"] + ".psk"

                        if dependency_file not in dependencies:
                            dependencies.add(dependency_file)
                            self.find_dependencies(dependency_file, visited, dependencies)
        # self.platform+"_"+self.App()+"_"+self.site_name+"_"+self.page+"_"+self.name is the output string format

    def analyzeMainSkillDependencies(self, main_psk):
        dependencies = set()
        visited = set()
        if os.path.exists(main_psk):
            self.find_dependencies(main_psk, visited, dependencies)
            if len(dependencies) > 0:
                dep_list = list(dependencies)
            else:
                dep_list = []
            logger.info("found dependency:"+json.dumps(dep_list))

            dep_ids = []
            for dep in dep_list:
                skid = self.findSkillIDWithSkillFileName(dep)
                dep_ids.append((skid, dep))

            existing_skill_ids = []
            for dp in dep_ids:
                if dp[0] == -1:
                    logger.error("ERROR: missing skill dependent skills file:"+str(dp[1]))
                else:
                    existing_skill_ids.append(dp[0])
            # existing_skill_ids = filter(lambda x: x == -1, dep_ids)
            logger.info("existing_skill_ids:"+json.dumps(existing_skill_ids))
        else:
            existing_skill_ids = []

        return existing_skill_ids


    def findSkillIDWithSkillFileName(self, skill_file_name):
        skidx = next((i for i, x in enumerate(self.skills) if x.matchPskFileName(skill_file_name)), -1)
        if skidx >= 0:
            return self.skills[skidx].getSkid()
        else:
            return -1

    def loadPublicSkills(self):
        skill_def_files = []
        skid_files = []
        psk_files = []
        csk_files = []
        json_files = []

        skdir = self.homepath + "/resource/skills/public/"
        logger.info("LISTING pub skills:", skdir, os.walk(skdir))
        # Iterate over all files in the directory
        # Walk through the directory tree recursively
        for root, dirs, files in os.walk(skdir):
            for file in files:
                if file.endswith(".json"):
                    file_path = os.path.join(root, file)
                    skill_def_files.append(file_path)
                    logger.trace("load all public skill definition json file:" + file + "::" + file_path)

        # self.showMsg("local skill files: "+json.dumps(skill_def_files))

        # if json exists, use json to guide what to do
        existing_skids = [sk.getSkid() for sk in self.skills]
        logger.info("existing public skids:", existing_skids)
        for file_path in skill_def_files:
            logger.trace("working on:", file_path)
            with open(file_path) as json_file:
                sk_data = json.load(json_file)
                json_file.close()
                logger.trace("loading public skill f: " + str(sk_data["skid"]) + " " + file_path)
                if sk_data["skid"] not in existing_skids:
                    new_skill = WORKSKILL(self, sk_data["name"], sk_data["path"])
                    new_skill.loadJson(sk_data)
                    self.skills.append(new_skill)
                    logger.debug("added public new skill:", sk_data["skid"], new_skill.getSkid(), new_skill.getPskFileName(),
                          new_skill.getPath())
                else:
                    existingSkill = next((x for i, x in enumerate(self.skills) if x.getSkid() == sk_data["skid"]), None)
                    if existingSkill:
                        # these are the only attributes that could be local only.
                        existingSkill.setAppLink(sk_data['app_link'])
                        existingSkill.setAppArgs(sk_data['app_args'])
                        existingSkill.add_procedural_skill(sk_data['procedural_skill'])
                        existingSkill.add_cloud_skill(sk_data['cloud_skill'])


        self.showMsg("Added Local public Skills:" + str(len(self.skills)))


    # load locally stored skills
    def loadLocalPrivateSkills(self):
        try:
            skill_def_files = []
            skid_files = []
            psk_files = []
            csk_files = []
            json_files = []

            skdir = self.my_ecb_data_homepath + "/my_skills/"
            logger.info("LISTING myskills:", skdir, os.walk(skdir))
            # Iterate over all files in the directory
            # Walk through the directory tree recursively
            for root, dirs, files in os.walk(skdir):
                for file in files:
                    if file.endswith(".json"):
                        file_path = os.path.join(root, file)
                        skill_def_files.append(file_path)
                        logger.debug("load private skill definition json file:" + file+"::"+file_path)

            # self.showMsg("local skill files: "+json.dumps(skill_def_files))

            # if json exists, use json to guide what to do
            existing_skids = [sk.getSkid() for sk in self.skills]
            for file_path in skill_def_files:
                logger.trace("working on:", file_path)
                with open(file_path) as json_file:
                    sk_data = json.load(json_file)
                    json_file.close()
                    logger.debug("sk_data::", sk_data)
                    self.showMsg("loading private skill f: "+str(sk_data["skid"])+" "+file_path)
                    if sk_data["skid"] in existing_skids:
                        new_skill = WORKSKILL(self, sk_data["name"], sk_data["path"])
                        new_skill.loadJson(sk_data)
                        self.skills.append(new_skill)
                        logger.debug("added private new skill:", new_skill.getSkid(), new_skill.getPskFileName(), new_skill.getPath())
                    else:
                        #update the existing skill or no even needed?
                        found_skill = next((x for x in self.skills if x.getSkid()==sk_data["skid"]), None)
                        # if found_skill:


                    this_skill_dir = skdir+sk_data["platform"]+"_"+sk_data["app"]+"_"+sk_data["site_name"]+"_"+sk_data["page"]+"/"
                    gen_string = sk_data["platform"]+"_"+sk_data["app"]+"_"+sk_data["site_name"]+"_"+sk_data["page"]+"_"+sk_data["name"]
                    self.showMsg("total skill files loaded: "+str(len(self.skills)))
                    self.load_external_functions(this_skill_dir, sk_data["name"], gen_string, sk_data["generator"])
                    # no need to run genSkillCode, since once in table, will be generated later....
                    # genSkillCode(sk_full_name, privacy, root_path, start_step, theme)

            self.showMsg("Added Local Private Skills:"+str(len(self.skills)))

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorLoadLocalPrivateSkills:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorLoadLocalPrivateSkills: traceback information not available:" + str(e)
            self.showMsg(ex_stat)


    #  in case private skill use certain external functions, load them
    def load_external_functions(self, sk_dir, sk_name, gen_string, generator):
        try:
            generator_script = sk_dir+sk_name+".py"
            generator_diagram = sk_dir + sk_name + ".skd"
            added_handlers = []
            self.showMsg("Generator:"+" "+sk_dir+" "+sk_name+" "+gen_string+" "+generator+" "+generator_script+" "+generator_diagram)
            if os.path.isfile(generator_script):
                spec = importlib.util.spec_from_file_location(sk_name, generator_script)
                # Create a module object from the spec
                module = importlib.util.module_from_spec(spec)
                # Load the module
                spec.loader.exec_module(module)

                if hasattr(module, generator):
                    self.showMsg("add key-val pair: "+gen_string+" "+generator)
                    SkillGeneratorTable[gen_string+"_my"] = lambda w, x, y, z: getattr(module, generator)(w, x, y, z)

            elif os.path.isfile(generator_diagram):
                self.showMsg("gen psk from diagram.")

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorLoadExternalFunction:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorLoadExternalFunction: traceback information not available:" + str(e)
            self.showMsg(ex_stat)


    def matchSkill(self, sk_long_name, sk):
        sk_words = sk_long_name.split("_")
        sk_name = "_".join(sk_words[4:])
        if sk.getPlatform() == sk_words[0] and sk.getApp() == sk_words[1] and sk.getSiteName() == sk_words[2] and sk.getName() == sk_name:
            return True
        else:
            return False


    def checkIsMain(self, sk_long_name):
        is_main = False
        # first find out the skill based on sk_long_name.
        sk = next((x for x in self.skills if self.matchSkill(sk_long_name, x)), None)
        # then check whether this is a main skill
        if sk:
            if sk.getIsMain():
                is_main = True

        return is_main

    def newProductsFromFile(self):

        self.showMsg("loading products from a local file or DB...")
        api_products = []
        uncompressed = open(self.my_ecb_data_homepath + "/resource/testdata/newproducts.json")
        if uncompressed != None:
            # self.showMsg("body string:"+uncompressed+"!"+str(len(uncompressed))+"::")
            fileproducts = json.load(uncompressed)
            if len(fileproducts) > 0:
                self.product_service.find_all_products()

            else:
                self.warn("NO products found in file.")
        else:
            self.warn("No tests products file")

    # try load bots from local database, if nothing in th local DB, then
    # try to fetch bots from local json files (this is mostly for testing).
    def loadLocalBots(self, db_data: List[BotModel]):
        try:
            dict_results = [result.to_dict() for result in db_data]
            self.showMsg("get local bots from DB::" + json.dumps(dict_results))
            if len(db_data) != 0:
                self.bots = []
                # botModel removed - UI components no longer needed
                for row in db_data:
                    self.showMsg("loading a bot: "+json.dumps(row.to_dict()))
                    new_bot = EBBOT(self)
                    new_bot.loadDBData(row)
                    print("hello????")
                    # new_bot.updateDisplay()
                    self.bots.append(new_bot)

                    self.addBotToVehicle(new_bot)
            else:
                self.showMsg("WARNING: local bots DB empty!")
        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorloadLocalBots:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorloadLocalBots: traceback information not available:" + str(e)
            log3(ex_stat)

    def addBotToVehicle(self, new_bot):

        if new_bot.getVehicle() != "" and new_bot.getVehicle() != "NA":
            found_v = next((x for x in self.vehicles if x.getName() == new_bot.getVehicle()), None)

            if found_v:
                nadded = found_v.addBot(new_bot.getBid())
                if nadded == 0:
                    self.showMsg("WARNING: vehicle reached full capacity!")
            else:
                self.showMsg("WARNING: bot vehicle NOT FOUND!")
        else:
            self.showMsg("WARNING: bot vehicle NOT ASSIGNED!")

    # load locally stored mission, but only for the past 7 days, otherwise, there would be too much......
    def loadLocalMissions(self, db_data: List[MissionModel]):
        dict_results = [result.to_dict() for result in db_data]
        # self.showMsg("get local missions from db::" + json.dumps(dict_results))
        if len(db_data) != 0:
            self.missions = []
            # missionModel removed - UI components no longer needed
            for row in db_data:
                # self.showMsg("loading a mission: "+json.dumps(row.to_dict()))
                new_mission = EBMISSION(self)
                new_mission.loadDBData(row)
                # new_mission.setData(row)
                self.cuspas_to_diaplayable(new_mission)
                new_mission.updateDisplay()
                self.missions.append(new_mission)
        else:
            self.showMsg("WARNING: local mission DB empty!")

    def cuspas_to_diaplayable(self, a_mission):
        cuspas_parts = a_mission.getCusPAS().split(",")
        a_mission.setPlatform(self.translateShortPlatform(cuspas_parts[0]))
        a_mission.setApp(cuspas_parts[1])
        a_mission.setSite(self.translateShortSiteName(cuspas_parts[2]))


    # fetch all bots stored in the cloud.
    def getAllBotsFromCloud(self):
        # File actions
        #resp = send_get_bots_request_to_cloud(self.session, self.cog.access_token, self.getWanApiEndpoint())
        jresp = send_get_bots_request_to_cloud(self.session, self.get_auth_token(), self.getWanApiEndpoint())
        if "errorType" in jresp:
            screen_error = True
            self.showMsg("Gat All Bots ERROR Type: "+json.dumps(jresp["errorType"])+"ERROR Info: "+json.dumps(jresp["errorInfo"]))
        else:
            self.showMsg("resp body"+json.dumps(jresp["body"]))
            #jbody = json.loads(jresp["body"])
            # now that fetch all bots from the cloud side is successfull, need to compare with local data and merge:

    def setOwner(self, owner):
        self.owner = owner

    def get_vehicle_settings(self, forceful="false"):
        import tzlocal
        vsettings = {
            "vwins": len([v for v in self.vehicles if v.getOS() == "win"]),
            "vmacs": len([v for v in self.vehicles if v.getOS() == "mac"]),
            "vlnxs": len([v for v in self.vehicles if v.getOS() == "linux"]),
            "forceful": forceful,
            "tz": str(tzlocal.get_localzone())
        }

        print("v timezone:", tzlocal.get_localzone())
        # add self to the compute resource pool
        if self.host_role == "Commander":
            if self.platform == "win":
                vsettings["vwins"] = vsettings["vwins"] + 1
            elif self.platform == "mac":
                vsettings["vmacs"] = vsettings["vmacs"] + 1
            else:
                vsettings["vlnxs"] = vsettings["vlnxs"] + 1
        return vsettings

    # the message queue is for messsage from tcpip task to the GUI task. OPENAI's fix
    async def servePlatoons(self, msgQueue):
        self.showMsg("starting servePlatoons")

        while True:
            logger.trace("listening to platoons" + datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])

            if not msgQueue.empty():
                try:
                    # Process all available messages in the queue
                    while not msgQueue.empty():
                        net_message = await msgQueue.get()
                        print("received net message:", type(net_message), net_message)
                        if isinstance(net_message, str):
                            if len(net_message) > 256:
                                mlen = 256
                            else:
                                mlen = len(net_message)
                            self.showMsg(
                                "received queued msg from platoon..... [" + str(msgQueue.qsize()) + "]" + net_message[:mlen])

                            print("platoon server received message from queu...")
                            # Parse the message into parts
                            msg_parts = net_message.split("!")
                            if len(msg_parts) >= 3:  # Check for valid message structure
                                if msg_parts[1] == "net data":
                                    await self.processPlatoonMsgs(msg_parts[2], msg_parts[0])
                                elif msg_parts[1] == "connection":
                                    print("received connection message: " + msg_parts[0] + " " + msg_parts[2])
                                    if self.platoon_manager is None:
                                        self.platoon_manager = PlatoonManager(self, "conn")
                                    addedV = self.addConnectingVehicle(msg_parts[2], msg_parts[0])
                                    # await asyncio.sleep(8)
                                    # if len(self.vehicles) > 0:
                                    #     print("pinging platoon: " + str(len(self.vehicles) - 1) + msg_parts[0])
                                    #     self.sendToVehicleByVip(msg_parts[0])
                                elif msg_parts[1] == "net loss":
                                    print("received net loss")
                                    found_vehicle = self.markVehicleOffline(msg_parts[0], msg_parts[2])
                                    vehicle_report = self.prepVehicleReportData(found_vehicle)
                                    resp = send_report_vehicles_to_cloud(
                                        self.session,
                                        self.get_auth_token(),
                                        vehicle_report,
                                        self.getWanApiEndpoint()
                                    )
                                    self.saveVehiclesJsonFile()
                        elif isinstance(net_message, dict):
                            print("process json from queue:")

                        msgQueue.task_done()

                except asyncio.QueueEmpty:
                    print("Queue unexpectedly empty when trying to get message.")
                except Exception as e:
                    print(f"Error processing Commander message: {e}")

            else:
                # if nothing on queue, do a quick check if any vehicle needs a ping-pong check
                for v in self.vehicles:
                    if "connecting" in v.getStatus():
                        print("pinging platoon: " + v.getIP())
                        self.sendToVehicleByVip(v.getIP())
            await asyncio.sleep(1)  # Short sleep to avoid busy-waiting



    # this is be run as an async task.
    async def runbotworks(self, gui_rpa_queue, gui_monitor_queue):
        # run all the work
        try:
            running = True
            wan_pre_time = datetime.now()
            lan_pre_time = datetime.now()
            while running:
                log3("runbotwork Task.....", "runbotworks", self)
                logger.trace("runbotworks................")
                current_time = datetime.now()

                # check whether there is vehicle for hire, if so, check any contract work in the queue
                # if so grab it.
                contractWorks = await self.checkCloudWorkQueue()

                # if there is actual work, 1) deque from virutal cloud queue, 2) put it into local unassigned work list.
                # and the rest will be taken care of by the work dispatcher...
                self.arrangeContractWorks(contractWorks)

                #print only first 3 and last 3 items.
                log3("real work starts here...."+json.dumps([m.getFingerPrintProfile() for i, m in enumerate(self.missions) if i<3 or i > len(self.missions)-4]), "runbotworks", self)
                botTodos = None
                if self.working_state == "running_idle":
                    log3("idle checking.....", "runbotworks", self)
                    if self.getNumUnassignedWork() > 0:
                        log3(get_printable_datetime() + " - Found unassigned work: "+str(self.getNumUnassignedWork())+"<>"+datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "runbotworks", self)
                        await self.assignWork()

                    log3("check next to run"+str(len(self.todays_work["tbd"]))+" "+str(len(self.reactive_work["tbd"]))+" "+str(self.getNumUnassignedWork()), "runbotworks", self)
                    botTodos, runType = self.checkNextToRun()
                    log3("fp profiles of mission: "+json.dumps([m.getFingerPrintProfile() for i, m in enumerate(self.missions) if i < 3 or i > len(self.missions)-4]), "runbotworks", self)
                    if botTodos:
                        log3("working on..... "+botTodos["name"], "runbotworks", self)
                        self.working_state = "running_working"

                        if botTodos["name"] == "automation":
                            # run 1 bot's work
                            log3("running RPA.............."+json.dumps([m.getFingerPrintProfile() for m in self.missions]), "runbotworks", self)
                            if "Completed" not in botTodos["status"]:
                                log3("time to run RPA........"+json.dumps(botTodos), "runbotworks", self)
                                last_start = int(datetime.now().timestamp()*1)

                                current_bid, current_mid, run_result = await self.runRPA(botTodos, gui_rpa_queue, gui_monitor_queue)
                                last_end = int(datetime.now().timestamp()*1)

                            # else:
                                # now need to chop off the 0th todo since that's done by now....
                                #
                                log3("total # of works:"+str(botTodos["current widx"])+":"+str(len(botTodos["works"])), "runbotworks", self)
                                if current_mid >= 0:
                                    current_run_report = self.genRunReport(runType, last_start, last_end, current_mid, current_bid, run_result)

                                # if all tasks in the task group are done, we're done with this group.
                                if botTodos["current widx"] >= len(botTodos["works"]):
                                    log3("POP a finished task from queue after runRPA", "runbotworks", self)
                                    # update GUI display to move missions in this task group to the completed missions list.
                                    if self.todays_work["tbd"][0]:
                                        log3("None empt first WORK GROUP", "runbotworks", self)
                                        just_finished = copy.deepcopy(self.todays_work["tbd"][0])
                                        self.updateCompletedMissions(just_finished)
                                        self.todays_completed.append(just_finished)

                                        finished = self.todays_work["tbd"].pop(0)
                                        log3("JUST FINISHED A WORK GROUP:"+json.dumps(finished), "runbotworks", self)
                                    else:
                                        log3("empty first WORK GROUP", "runbotworks", self)


                                if len(self.todays_work["tbd"]) == 0:
                                    if self.host_role == "Platoon":
                                        log3("Platoon Done with today!!!!!!!!!", "runbotworks", self)
                                        await self.doneWithToday()
                                    else:
                                        # check whether we have collected all reports so far, there is 1 count difference between,
                                        # at this point the local report on this machine has not been added to toddaysReports yet.
                                        # this will be done in doneWithToday....
                                        log3("n todaysPlatoonReports: "+str(len(self.todaysPlatoonReports))+" n todays_completed: "+str(len(self.todays_completed)), "runbotworks", self)
                                        log3("todaysPlatoonReports"+json.dumps(self.todaysPlatoonReports), "runbotworks", self)
                                        log3("todays_completed"+json.dumps(self.todays_completed), "runbotworks", self)
                                        if len(self.todaysPlatoonReports) == self.num_todays_task_groups:
                                            log3("Commander Done with today!!!!!!!!!", "runbotworks", self)
                                            await self.doneWithToday()
                        else:
                            log3("Unrecogizable todo...."+botTodos["name"], "runbotworks", self)
                            log3("POP a unrecognized task from queue", "runbotworks", self)
                            self.todays_work["tbd"].pop(0)

                    else:
                        # nothing to do right now. check if all of today's work are done.
                        # if my own works are done and all platoon's reports are collected.
                        logger.trace("empty to do...")
                        if self.host_role == "Platoon":
                            if len(self.todays_work["tbd"]) == 0:
                                await self.doneWithToday()
                            else:
                                self.todays_work["tbd"].pop(0)

                if self.working_state != "running_idle":
                    # clear to make next round ready to work
                    self.working_state = "running_idle"

                log3("running bot works whenever there is some to run....", "runbotworks", self)
                await asyncio.sleep(3)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "Errorwanrunbotworks:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "Errorwanrunbotworks traceback information not available:" + str(e)
            log3(ex_stat, "runbotworks", self)

    def checkManagerToRuns(self, managerMissions):
        """
        Determine which missions are ready to run based on their schedule.

        Args:
            managerMissions (list): A list of mission data structures.

        Returns:
            list: Missions ready to run.
        """
        try:
            missions_to_run = []
            current_time = datetime.now()

            for mission in managerMissions:
                print("checking next to run mission:", mission.getMid())
                # Parse repeat_last and repeat_until as datetime
                repeat_last = datetime.strptime(mission.getRepeatLast(), "%Y-%m-%d %H:%M:%S")
                repeat_until = datetime.strptime(mission.getRepeatUntil(), "%Y-%m-%d")
                print("repeat_last, repeat_until:", mission.getRepeatLast(), mission.getRepeatUntil())
                # Get the time slot as hours and minutes
                esttime_index = int(mission.getEstimatedStartTime())  # Index of the 15-min time slot (0–95)
                print("esttime_index", esttime_index)
                hours, minutes = divmod(esttime_index * 15, 60)
                print("hours, minutes:", hours, minutes)

                # Determine the baseline date for repetition
                if mission.getRepeatOn() == "now":
                    if mission.getRepeatType() == "by day":
                        repeat_on_date = datetime.strptime(mission.getEsd(), "%Y-%m-%d")
                    else:
                        repeat_on_date = current_time.date()
                elif mission.getRepeatOn() in self.static_resource.WEEK_DAY_TYPES:
                    repeat_on_date = self._get_next_weekday_date(mission.getRepeatOn())
                else:
                    repeat_on_date = datetime.strptime(mission.getRepeatOn(), "%Y-%m-%d").date()

                print("repeat on date::", repeat_on_date)
                # Combine the baseline date with the time slot
                repeat_on_time = datetime.combine(repeat_on_date, datetime.min.time()).replace(hour=hours, minute=minutes)
                print("repeat on time::", repeat_on_date)

                # Check for non-repeating missions
                if mission.getRepeatType() == "none":
                    if current_time >= repeat_on_time:
                        missions_to_run.append(mission)
                        continue

                # Check for repeating missions
                elif mission.getRepeatType() in self.static_resource.REPEAT_TYPES:
                    # Calculate the repeat interval
                    repeat_interval = self._compute_repeat_interval(mission.getRepeatUnit(), mission.getRepeatNumber(),
                                                                    repeat_on_time)

                    # Determine the supposed last scheduled repetition time
                    elapsed_time = (current_time - repeat_on_time).total_seconds()
                    elapsed_intervals = max(0, int(elapsed_time // repeat_interval.total_seconds())) if isinstance(
                        repeat_interval, timedelta) else self._calculate_elapsed_intervals_manual(repeat_on_time,
                                                                                                  current_time,
                                                                                                  repeat_interval)
                    supposed_last_run = repeat_on_time + elapsed_intervals * repeat_interval
                    print("supposed last run:", supposed_last_run)
                    # Calculate the next scheduled run
                    next_scheduled_run = supposed_last_run + repeat_interval
                    print("next scheduled run:", next_scheduled_run, "repeat_last::", repeat_last)

                    # If the current time is past the supposed last run, schedule the mission
                    if current_time <= repeat_until:
                        if repeat_last < (supposed_last_run - repeat_interval*0.5) or current_time >= next_scheduled_run:
                            print("time to run now....")
                            missions_to_run.append(mission)
                        elif self.config_manager.general_settings.debug_mode:
                            if self.fetch_schedule_counter:
                                missions_to_run.append(mission)
                                self.fetch_schedule_counter = self.fetch_schedule_counter -1


        except Exception as e:
            # Log and skip errors gracefully
            ex_stat = f"Error in check manager to runs: {traceback.format_exc()} {str(e)}"
            missions_to_run = []
            print(ex_stat)

        return missions_to_run

    def _compute_repeat_interval(self, repeat_unit, repeat_number, start_time):
        """
        Calculate the interval for repetition using timedelta.

        Args:
            repeat_unit (str): Unit of repetition ("second", "minute", "hour", etc.).
            repeat_number (int): Number of units for the interval.

        Returns:
            timedelta: The repeat interval.
        """
        if repeat_unit == "second":
            interval = timedelta(seconds=repeat_number)
        elif repeat_unit == "minute":
            interval = timedelta(minutes=repeat_number)
        elif repeat_unit == "hour":
            interval = timedelta(hours=repeat_number)
        elif repeat_unit == "day":
            interval = timedelta(days=repeat_number)
        elif repeat_unit == "week":
            interval = timedelta(weeks=repeat_number)
        elif repeat_unit == "month":
            interval = self._add_months(start_time, repeat_number)  # Custom month logic
        elif repeat_unit == "year":
            interval = self._add_years(start_time, repeat_number)  # Custom year logic
        else:
            print("invalid repeat unit")
            raise ValueError(f"Invalid repeat_unit: {repeat_unit}")

        print("interval:", interval)
        return interval

    def _add_months(self, start_time, months):
        """
        Manually add months to a datetime, adjusting for month overflow.

        Args:
            start_time (datetime): The starting date.
            months (int): Number of months to add.

        Returns:
            datetime: The resulting datetime.
        """
        new_month = (start_time.month - 1 + months) % 12 + 1
        year_increment = (start_time.month - 1 + months) // 12
        new_year = start_time.year + year_increment

        # Handle day overflow (e.g., adding 1 month to Jan 31 should result in Feb 28/29)
        try:
            updated_start_time = start_time.replace(year=new_year, month=new_month)
            print("month updated start time:", updated_start_time)
            return updated_start_time
        except ValueError:
            # For invalid days (e.g., Feb 30), use the last day of the month
            updated_start_time = start_time.replace(year=new_year, month=new_month, day=28) + timedelta(days=1) - timedelta(days=1)
            print("error month updated start time:", updated_start_time)
            return updated_start_time

    def _add_years(self, start_time, years):
        """
        Manually add years to a datetime, adjusting for leap years.

        Args:
            start_time (datetime): The starting date.
            years (int): Number of years to add.

        Returns:
            datetime: The resulting datetime.
        """
        try:
            updated_start_time = start_time.replace(year=start_time.year + years)
            print("year updated start time:", updated_start_time)
            return updated_start_time
        except ValueError:
            # For Feb 29 on non-leap years, fallback to Feb 28
            updated_start_time = start_time.replace(year=start_time.year + years, day=28)
            print("error year updated start time:", updated_start_time)
            return updated_start_time

    def _calculate_elapsed_intervals_manual(self, start_time, current_time, interval):
        """
        Calculate the number of elapsed intervals for manual month/year intervals.

        Args:
            start_time (datetime): The baseline start time.
            current_time (datetime): The current time.
            interval: Function to calculate the next interval (e.g., _add_months).

        Returns:
            int: Number of elapsed intervals.
        """
        intervals = 0
        next_time = start_time

        while next_time <= current_time:
            next_time = interval(next_time)
            intervals += 1

        print("intervals:", intervals)
        return intervals - 1  # Subtract 1 because the last addition exceeds current_time


    def _get_next_weekday_date(self, target_weekday):
        """
        Calculate the date of the next occurrence of the target weekday.

        Args:
            target_weekday (str): Target weekday ("M", "Tu", "W", etc.).

        Returns:
            date: The date of the next occurrence of the target weekday.
        """
        weekday_map = {"M": 0, "Tu": 1, "W": 2, "Th": 3, "F": 4, "Sa": 5, "Su": 6}
        current_date = datetime.now()
        current_weekday = current_date.weekday()
        target_weekday_num = weekday_map[target_weekday]
        print("target_weekday_num", target_weekday_num)
        days_ahead = (target_weekday_num - current_weekday) % 7
        if days_ahead == 0:  # If today is the target weekday, schedule for the next week
            days_ahead = 7

        print("days_ahead", days_ahead)
        next_week_day = (current_date + timedelta(days=days_ahead)).date()
        print("next week day:", next_week_day)
        return next_week_day



    async def runManagerMissions(self, missions, in_queue, out_team_queue, out_gui_queue):
        for mission in missions:
            #update the mission's last repeat time.
            mission.updateRepeatLast()
            await self.run1ManagerMission(mission, in_queue, out_team_queue, out_gui_queue)


    def genOneTimeMissionWithSkill(self, skid, mtype, botid):
        # simply search the past mission and check whether there are
        # already mission running this skill, if there is simply copy it and run.
        # if nothing found, then create a brand new mission on the fly.
        foundMission = next((x for i, x in enumerate(self.missions) if x.getSkills().startswith(str(skid)+',')), None)
        if foundMission:
            log3(f"duplicate the found mission {foundMission.getMid()}", "runmanagerworks", self)
            # newMisssion = copy.deepcopy(foundMission)
            newMisssion = foundMission
        else:
            log3(f"create a new mission based on skill {skid}...", "runmanagerworks", self)
            today = datetime.now()
            formatted_date = today.strftime("%Y-%m-%d")
            future_date = today + timedelta(days=1)
            formatted_future = future_date.strftime("%Y-%m-%d")
            far_future_date = today + timedelta(days=1000)
            formatted_far_future = far_future_date.strftime("%Y-%m-%d")
            mdbd = MissionModel()
            mdbd.mid = 0
            mdbd.ticket = 0
            mdbd.botid = botid
            mdbd.status = "Assiggned"
            mdbd.createon = formatted_date
            mdbd.owner = self.owner
            mdbd.esd = formatted_date
            mdbd.ecd = formatted_date
            mdbd.asd = formatted_future
            mdbd.abd = formatted_future
            mdbd.aad = formatted_future
            mdbd.afd = formatted_future
            mdbd.acd = formatted_future
            mdbd.actual_start_time = 0
            mdbd.est_start_time = 0
            mdbd.actual_runtime = 0
            mdbd.est_runtime = 30
            mdbd.n_retries = 3
            mdbd.cuspas = "win,chrome,amz"
            mdbd.category = ""
            mdbd.phrase = ""
            mdbd.pseudoStore = ""
            mdbd.pseudoBrand = ""
            mdbd.pseudoASIN = ""
            mdbd.type = mtype
            mdbd.config = "{}"
            mdbd.skills = str(skid)
            mdbd.delDate = formatted_far_future
            mdbd.asin = ""
            mdbd.store = ""
            mdbd.follow_seller = ""
            mdbd.brand = ""
            mdbd.img = ""
            mdbd.title = ""
            mdbd.rating = ""
            mdbd.feedbacks = ""
            mdbd.price = 0
            mdbd.follow_price = 0
            mdbd.fingerprint_profile = ""
            mdbd.original_req_file = ""
            mdbd.customer = ""
            mdbd.platoon = ""
            mdbd.result = ""
            mdbd.variations = ""
            mdbd.as_server = False
            newMisssion = EBMISSION(self)
            newMisssion.loadDBData(mdbd)

        return newMisssion


    # for now this is mainly used for after team run, a result to trigger some housekeeping work.
    # like process new orders, turn them into new missions, and so on....
    # the message will likely,
    async def processManagerNetMessage(self, msg, managers, in_queue, out_team_queue, out_gui_queue):
        print(f"recevied manager msg type: {msg['type']}")
        if msg["type"] in ManagerTriggerTable:
            otm = self.genOneTimeMissionWithSkill(ManagerTriggerTable[msg["type"]][0], ManagerTriggerTable[msg["type"]][1], managers[0].getBid())
            print("ready to run manager 1 mission....")
            result = await self.run1ManagerMission(otm, in_queue, out_team_queue, out_gui_queue)


    async def runmanagerworks(self, gui_manager_queue, manager_rpa_queue, gui_monitor_queue):
        # run all the work
        try:
            running = True
            while running:
                log3("runmanagerwork Task.....", "runmanagerworks", self)
                current_time = datetime.now()

                # check mission queue, how to make this flexible? (just run the mission)
                # check msg queue, (msg source: flask server, there needs to be a
                #                      api msg <-> handler skill table, there needs to be a
                #                       generic function to create a mission given the skill and run
                #                       it. and the skill can be overwritten with custom skill).
                # check time. @certain time, time based, read out all manager missions, user can
                #                  create missions and let them use certain skill and run at certain time.
                managerBots, managerMissions = self.findManagerMissionsOfThisVehicle()
                logger.trace("# manager missions:", len(managerMissions))
                managerToRun = self.checkManagerToRuns(managerMissions)

                if managerToRun:
                    print("there is some repeat type mission to run....")
                    await self.runManagerMissions(managerToRun, gui_manager_queue, manager_rpa_queue, gui_monitor_queue)

                if not gui_manager_queue.empty():
                    # Process all available messages in the queue
                    logger.trace("recevied manager queued msg...")
                    while not gui_manager_queue.empty():
                        net_message = await gui_manager_queue.get()
                        await self.processManagerNetMessage(net_message, managerBots, gui_manager_queue, manager_rpa_queue, gui_monitor_queue)
                else:
                    # always run some clean up after night
                    logger.trace("manager msg queue empty...")
                    if current_time.hour == 0 and current_time.minute < 10:
                        # do some data structure and state cleaning and get rid   the
                        # next day
                        log3("clear work related data structure", "runmanagerworks", self)
                        self.todays_scheduled_task_groups = {}
                        self.unassigned_scheduled_task_groups = {}  # per vehicle, flatten task list
                        self.unassigned_reactive_task_groups = {}
                        self.rpa_work_assigned_for_today = False

                    # start work after 5:30am
                    target_time = current_time.replace(hour=5, minute=30, second=0, microsecond=0)
                    # if manually start platoon after 5:30am, and if todays_scheduled_task_groups
                    # still empty, then check todo file.....
                    if current_time > target_time and "Platoon" in self.machine_role:
                        if not self.todays_scheduled_task_groups:
                            # check the local schedule file. if there, load it.
                            yyyymmdd = current_time.strftime("%Y%m%d")
                            sf_name = "todos" + yyyymmdd + ".json"
                            todays_todo_file = os.path.join(self.my_ecb_data_homepath + "/runlogs", sf_name)
                            # with todays_todo_file name this won't run on commander.
                            if os.path.exists(todays_todo_file):
                                if os.path.getsize(todays_todo_file) > 128:
                                    with open(todays_todo_file, "r") as tdf:
                                        msg = json.load(tdf)
                                        tdf.close()

                                        self.setupScheduledTodos(msg)
                                else:
                                    print("WARNING: invalid todo file")



                await asyncio.sleep(3)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "Errorwanrunmanagerworks:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "Errorwanrunmanagerworks traceback information not available:" + str(e)
            log3(ex_stat, "runmanagerworks", self)


    def setupScheduledTodos(self, msg):
        if msg:
            localworks = msg["todos"]
            self.addBotsMissionsSkillsFromCommander(msg["bots"], msg["missions"], msg["skills"])

            # this is the time to rebuild skills to make them up to date....
            self.dailySkillsetUpdate()

            log3("received work request:"+json.dumps(localworks), "serveCommander", self)
            # send work into work Queue which is the self.todays_work["tbd"] data structure.

            self.todays_work["tbd"].append({"name": "automation", "works": localworks, "status": "yet to start", "current widx": 0, "vname": self.machine_name+":"+self.os_short, "completed": [], "aborted": []})
            log3("after assigned work, "+str(len(self.todays_work["tbd"]))+" todos exists in the queue. "+json.dumps(self.todays_work["tbd"]), "serveCommander", self)

            platform_os = self.platform            # win, mac or linux
            vname = self.machine_name + ":" + self.os_short
            self.todays_scheduled_task_groups[vname] = localworks
            self.unassigned_scheduled_task_groups[vname] = localworks

            # generate ADS loadable batch profiles ((vTasks, vehicle, commander):)
            batched_tasks, ads_profiles = formADSProfileBatchesFor1Vehicle(localworks, self, self)
            # clean up the reports on this vehicle....
            self.todaysReports = []
            self.DONE_WITH_TODAY = False
        else:
            print("nothing to arrange to do....")

    #update a vehicle's missions status
    # rx_data is a list of mission status for each mission that belongs to the vehicle.
    def updateVMStats(self, rx_data):
        foundV = None
        for v in self.vehicles:
            if v.getIP() == rx_data["ip"]:
                log3("found vehicle by IP", "runbotworks", self)
                foundV = v
                break

        if foundV:
            log3("updating vehicle Mission status...", "runbotworks", self)
            foundV.setMStats(rx_data)



    # msg in json format
    # { sender: "ip addr", type: "intro/status/report", content : "another json" }
    # content format varies according to type.
    async def processPlatoonMsgs(self, msgString, ip):
        try:
            global running_step_index, fieldLinks
            fl_ips = [x["ip"] for x in fieldLinks]
            if len(msgString) < 128:
                log3("Platoon Msg Received:"+msgString+" from::"+ip+"  "+str(len(fieldLinks)) + json.dumps(fl_ips))
            else:
                log3("Platoon Msg Received: ..." + msgString[-127:0] + " from::" + ip + "  " + str(len(fieldLinks)) + json.dumps(
                    fl_ips))
            msg = json.loads(msgString)

            found = next((x for x in fieldLinks if x["ip"] == ip), None)
            found_vehicle = next((x for x in self.vehicles if x.getIP() == msg["ip"]), None)

            # first, check ip and make sure this from a know vehicle.
            if msg["type"] == "intro" or msg["type"] == "pong":
                if found:
                    self.showMsg("recevied a vehicle introduction/pong:" + msg["content"]["name"] + ":" + msg["content"]["os"] + ":"+ msg["content"]["machine"])

                    if found_vehicle:
                        print("found a vehicle to set.... "+found_vehicle.getOS())
                        if "connecting" in found_vehicle.getStatus():
                            found_vehicle.setStatus("running_idle")

                        if "Windows" in msg["content"]["os"]:
                            found_vehicle.setOS("Windows")
                            found_vehicle.setName(msg["content"]["name"]+":win")
                        elif "Mac" in msg["content"]["os"]:
                            found_vehicle.setOS("Mac")
                            found_vehicle.setName(msg["content"]["name"] + ":mac")
                        elif "Lin" in msg["content"]["os"]:
                            found_vehicle.setOS("Linux")
                            found_vehicle.setName(msg["content"]["name"] + ":linux")

                        print("now found vehicle" + found_vehicle.getName() + " " + found_vehicle.getOS())
                        # this is a good juncture to update vehicle status on cloud and local DB and JSON file.
                        #  now
                        vehicle_report = self.prepVehicleReportData(found_vehicle)
                        resp = send_report_vehicles_to_cloud(self.session,
                                                             self.get_auth_token(),
                                                             vehicle_report,
                                                             self.getWanApiEndpoint())
                        self.saveVehiclesJsonFile()

                        # sync finger print profiles from that vehicle.
                        if  msg["type"] == "pong":
                            # Launch async function without blocking
                            asyncio.ensure_future(self.syncFingerPrintOnConnectedVehicle(found_vehicle))

            elif msg["type"] == "status":
                # update vehicle status display.
                self.showMsg(msg["content"])
                log3("msg type:" + "status", "servePlatoons", self)
                self.showMsg("recevied a status update message:"+msg["content"])
                if self.platoonWin:
                    self.showMsg("updating platoon WIN")
                    self.platoonWin.updatePlatoonStatAndShow(msg, fieldLinks)
                    self.platoonWin.show()
                else:
                    self.showMsg("ERROR: platoon win not yet exists.......")

                self.updateVMStats(msg)

                # update mission status to the cloud and to local data structure and to chat？
                # "mid": mid,
                # "botid": self.missions[mid].getBid(),
                # "sst": self.missions[mid].getEstimatedStartTime(),
                # "sd": self.missions[mid].getEstimatedRunTime(),
                # "ast": self.missions[mid].getActualStartTime(),
                # "aet": self.missions[mid].getActualEndTime(),
                # "status": m_stat,
                # "error": m_err
                if msg["content"]:
                    mStats = json.loads(msg["content"])

                    self.updateMStats(mStats)
                else:
                    print("WARN: status contents empty.")

            elif msg["type"] == "report":
                # collect report, the report should be already organized in json format and ready to submit to the network.
                log3("msg type:"+"report", "servePlatoons", self)
                #msg should be in the following json format {"ip": self.ip, "type": "report", "content": []]}
                self.todaysPlatoonReports.append(msg)

                # now using ip to find the item added to self.self.todays_work["tbd"]
                task_idx = 0
                found = False
                for item in self.todays_work["tbd"]:
                    if "ip" in item:
                        if item["ip"] == msg["ip"]:
                            found = True
                            break
                    task_idx = task_idx + 1

                if found:
                    self.showMsg("pop a finising a remotely executed task...."+str(task_idx))
                    finished = self.todays_work["tbd"].pop(task_idx)
                    self.todays_completed.append(finished)

                    # Here need to update completed mission display subwindows.
                    self.updateCompletedMissions(finished)

                self.showMsg("len todays's reports: "+str(len(self.todaysPlatoonReports))+" len todays's completed:"+str(len(self.todays_completed)))
                self.showMsg("completd: "+json.dumps(self.todays_completed))

                # update vehicle status, now becomes idle again.
                self.updateVehicleStatusToRunningIdle(msg["ip"])

                # keep statistics on all platoon runs.
                if len(self.todaysPlatoonReports) == self.num_todays_task_groups:
                    # check = all(item in List1 for item in List2)
                    # this means all reports are collected, this is the last missing piece, ready to send to cloud.
                    await self.doneWithToday()
                    self.num_todays_task_groups = 0
            elif msg["type"] == "botsADSProfilesUpdate":
                log3("received botsADSProfilesUpdate message", "servePlatoons", self)
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text
                self.receivePlatoonBotsADSProfileUpdateMessage(msg)
            elif msg["type"] == "botsADSProfilesBatchUpdate":
                log3("received botsADSProfilesBatchUpdate message", "servePlatoons", self)
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text
                remote_outdated = self.receiveBotsADSProfilesBatchUpdateMessage(msg)
                self.expected_vehicle_responses[found_vehicle.getName()] = "Yes"

                if self.allResponded():
                    log3("all ads profiles updated...", "servePlatoons", self)
                    self.botsFingerPrintsReady = True

                if remote_outdated:
                    log3("remote outdated...", "servePlatoons", self)
                    self.batchSendFingerPrintProfilesToCommander(remote_outdated)

                # now the profiles are updated. send this vehicle's schedule to it.
                vname = found_vehicle.getName()
                log3("setup vehicle to do some work..."+vname, "servePlatoons", self)

                if self.unassigned_scheduled_task_groups:
                    if vname in self.unassigned_scheduled_task_groups:
                        p_task_groups = self.unassigned_scheduled_task_groups[vname]
                    else:
                        print(f"{vname} not found in unassigned_scheduled_task_groups empty")
                        print("keys:", list(self.unassigned_scheduled_task_groups.keys()))
                        p_task_groups = []
                else:
                    if self.todays_scheduled_task_groups:
                        if vname in self.todays_scheduled_task_groups:
                            p_task_groups = self.todays_scheduled_task_groups[vname]
                        else:
                            print(f"{vname} not found in todays_scheduled_task_groups empty")
                            print("keys:", list(self.todays_scheduled_task_groups.keys()))
                            p_task_groups = []
                    else:
                        print("time stamp "+datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]+" todays_scheduled_task_groups empty")
                        p_task_groups = []
                await self.vehicleSetupWorkSchedule(found_vehicle, p_task_groups)

            elif msg["type"] == "missionResultFile":
                self.showMsg("received missionResultFile message")
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text
                self.receivePlatoonMissionResultFilesMessage(msg)

            elif msg["type"] == "reqResendWorkReq":
                log3("received reqResendWorkReq message")
                # get work for this vehicle and send setWork
                self.reGenWorksForVehicle(found_vehicle)
                # self.vehicleSetupWorkSchedule(found_vehicle, self.todays_scheduled_task_groups)

            elif msg["type"] == "chat":
                log3("received chat message")
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text
                self.receiveBotChatMessage(msg["content"])

            elif msg["type"] == "exlog":
                self.showMsg("received exlog message")
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text
                self.receiveBotLogMessage(msg["content"])
            elif msg["type"] == "heartbeat":
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text

                if found_vehicle:
                    # this will set status as well as the last_update_time parameter
                    if found_vehicle.getStatus() != msg["content"]["vstatus"]:
                        found_vehicle.setStatus(msg["content"]["vstatus"])

                log3("Heartbeat From Vehicle: "+msg["ip"], "servePlatoons", self)
            else:
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text
                self.showMsg("unknown type:"+msg["contents"])

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorprocessPlatoonMsgs:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorprocessPlatoonMsgs: traceback information not available:" + str(e)
            log3(ex_stat, "servePlatoons", self)

            self.showMsg(ex_stat)

    def allResponded(self):
        alldone = False

        alldone = all([self.expected_vehicle_responses[v] for v in self.expected_vehicle_responses])

        return alldone


    # what's received here is a ADS profile for one individual bot, for safety, save the existing
    # file to file.old so that we at least always have two copies and in case something is wrong
    # we can at least go back to the previous copy.
    def receivePlatoonBotsADSProfileUpdateMessage(self, pMsg):
        file_name = self.my_ecb_data_homepath + pMsg["file_name"]           # msg["file_name"] should start with "/"
        file_name_wo_extension = os.path.basename(file_name).split(".")[0]
        file_name_dir = os.path.dirname(file_name)
        new_filename = file_name_dir + "/" + file_name_wo_extension + "_old.txt"
        os.rename(file_name, new_filename)

        file_type = pMsg["file_type"]
        file_contents = pMsg["file_contents"].encode('latin1')  # Encode string to binary data
        with open(file_name, 'wb') as file:
            file.write(file_contents)
            file.close()

    def receiveBotsADSProfilesBatchUpdateMessage(self, pMsg):
        """
            Receive multiple fingerprint profiles sent from the sender side.
            Args:
                pMsg: A dictionary containing:
                      - "profiles": A list of dictionaries, each containing:
                          - "file_name": The name of the file to be saved
                          - "file_type": The type of the file (e.g., txt)
                          - "timestamp": The timestamp of the incoming file
                          - "file_contents": The base64-encoded content of the file
            """
        try:
            remote_outdated = []
            profiles = pMsg.get("profiles", [])
            if not profiles:
                log3("ErrorReceiveBatchProfiles: No profiles received.")
                return

            for profile in profiles:
                # Resolve full file path
                file_name = os.path.basename(profile["file_name"])
                incoming_file_name = os.path.join(self.ads_profile_dir, file_name)
                incoming_file_timestamp = profile.get("timestamp")
                file_contents = base64.b64decode(profile["file_contents"])  # Decode base64-encoded binary data

                # Check if the file already exists
                if os.path.exists(incoming_file_name):
                    # Compare timestamps
                    existing_file_timestamp = os.path.getmtime(incoming_file_name)

                    if incoming_file_timestamp > existing_file_timestamp:
                        # Incoming file is newer, replace the existing file
                        with open(incoming_file_name, "wb") as file:
                            file.write(file_contents)

                        os.utime(incoming_file_name, (incoming_file_timestamp, incoming_file_timestamp))
                        log3(f"Updated profile: {incoming_file_name} (newer timestamp)")
                    else:
                        # Incoming file is older, skip saving
                        if incoming_file_timestamp < existing_file_timestamp:
                            remote_outdated.append(incoming_file_name)
                        log3(f"Skipped profile: {incoming_file_name} (existing file is newer or the same)")
                else:
                    # File doesn't exist, save it
                    with open(incoming_file_name, "wb") as file:
                        file.write(file_contents)
                    # Optionally, set the timestamp to the incoming file's timestamp (if desired)
                    os.utime(incoming_file_name, (incoming_file_timestamp, incoming_file_timestamp))
                    log3(f"Saved new profile: {incoming_file_name}")

                log3(f"Successfully updated profile: {incoming_file_name}")
            return remote_outdated

        except Exception as e:
            # Handle and log errors
            traceback_info = traceback.extract_tb(e.__traceback__)
            if traceback_info:
                ex_stat = "ErrorReceiveBatchFPProfiles:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorReceiveBatchFPProfiles: traceback information not available:" + str(e)
            self.showMsg(ex_stat)


    def receivePlatoonMissionResultFilesMessage(self, pMsg):
        file_name = pMsg["file_name"]           # msg["file_name"] should start with "/"

        file_contents = pMsg["file_contents"].encode('latin1')  # Encode string to binary data
        with open(file_name, 'wb') as file:
            file.write(file_contents)
            file.close()


    def updateCompletedMissions(self, finished):
        finished_works = finished["works"]
        finished_mids = []
        finished_midxs = []
        finished_missions = []

        # Log all current mission IDs
        self.showMsg("All mission ids: " + json.dumps([m.getMid() for m in self.missions]))

        # Collect all finished mission IDs
        if len(finished_works) > 0:
            for bi in range(len(finished_works)):
                finished_mids.append(finished_works[bi]["mid"])
        self.showMsg("Finished MIDS: " + json.dumps(finished_mids))

        # Find the indexes of the finished missions in the missions list
        for mid in finished_mids:
            found_i = next((i for i, mission in enumerate(self.missions) if mission.getMid() == mid), -1)
            self.showMsg("Found midx: " + str(found_i))
            if found_i >= 0:
                finished_midxs.append(found_i)

        # Sort the finished mission indexes
        sorted_finished_midxs = sorted(finished_midxs, key=lambda midx: midx, reverse=True)
        self.showMsg("Finished MID INDEXES: " + json.dumps(sorted_finished_midxs))

        # Iterate through the sorted mission indexes
        for midx in sorted_finished_midxs:
            found_mission = self.missions[midx]

            # Log the mission status
            self.showMsg(f"Just finished mission [{found_mission.getMid()}] status: {found_mission.getStatus()}")

            # Ensure the mission is still valid and not deleted
            if found_mission is None or not found_mission:
                self.showMsg("Mission object is invalid or already deleted.")
                continue  # Skip to the next mission if this one is invalid

            # Try to update the mission icon safely
            try:
                # Mission icons removed - no longer needed for UI
                if "Completed" in found_mission.getStatus():
                    logger.info("Mission completed successfully")
                else:
                    logger.info("Mission failed or incomplete")
            except RuntimeError as e:
                self.showMsg(f"Error processing mission status: {str(e)}")
                continue  # Skip to the next mission if there's an error

            # missionModel and completedMissionModel removed - UI components no longer needed
            # Mission completion now handled through data models only


    def genMissionStatusReport(self, mids, test_mode=True):
        # assumptions: mids should have already been error checked.
        self.showMsg("mids: "+json.dumps(mids))
        results = []
        if test_mode:
            # just to tests commander GUI can handle the result
            result = {"mid": 1, "botid": 0, "sst": "2023-11-09 01:12:02", "ast": "2023-11-09 01:12:02", "sd": "600", "aet": "2023-11-09 01:22:12", "status": "Scheduled", "error": ""}
            results.append(result)
            result = {"mid": 1, "botid": 0, "sst": "2023-11-09 01:12:02", "ast": "2023-11-09 01:12:02", "sd": "600", "aet": "2023-11-09 01:22:12", "status": "Completed", "error": ""}
            results.append(result)
            result = {"mid": 1, "botid": 0, "sst": "2023-11-09 01:12:02", "ast": "2023-11-09 01:12:02", "sd": "600", "aet": "2023-11-09 01:22:12", "status": "Running", "error": ""}
            results.append(result)
            result = {"mid": 1, "botid": 0, "sst": "2023-11-09 01:12:02", "ast": "2023-11-09 01:12:02", "sd": "500", "aet": "2023-11-09 01:22:12", "status": "Warned", "error": "505"}
            results.append(result)
            result = {"mid": 1, "botid": 0, "sst": "2023-11-09 01:12:02", "ast": "2023-11-09 01:12:02", "sd": "300", "aet": "2023-11-09 01:22:12", "status": "Aborted", "error": "5"}
            results.append(result)
        else:
            for mid in mids:
                if mid > 0 and mid < len(self.missions):
                    self.showMsg("working on MID:"+str(mid))
                    m_stat_parts = self.missions[mid].getStatus().split(":")
                    m_stat = m_stat_parts[0]
                    if len(m_stat_parts) > 1:
                        m_err = m_stat_parts[1]
                    else:
                        m_err = ""
                    result = {
                        "mid": mid,
                        "botid": self.missions[mid].getBid(),
                        "sst": self.missions[mid].getEstimatedStartTime(),
                        "sd": self.missions[mid].getEstimatedRunTime(),
                        "ast": self.missions[mid].getActualStartTime(),
                        "aet": self.missions[mid].getActualEndTime(),
                        "status": m_stat,
                        "error": m_err
                    }
                    results.append(result)


        self.showMsg("mission status result:"+json.dumps(results))
        return results

    def platoonHasNoneTodo(self):
        none2do = True
        if self.machine_role == "Platoon":
            platform_os = self.platform
            vname = self.machine_name + ":" + self.os_short
            # check either some RPA is being run right now, or today's rpa has being all done.
            if self.working_state == "running_working" or \
                self.todays_scheduled_task_groups[vname] or \
                self.unassigned_scheduled_task_groups[vname] or \
                self.DONE_WITH_TODAY:
                none2do = False

        return none2do


    async def todo_wait_in_line(self, request):
        try:
            print("task waiting in line.....", request)
            await self.gui_net_msg_queue.put(request)
            print("todo now in line....", datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])
            return("rpa tasks queued")
        except Exception as e:
            ex_stat = "ErrorPlatoonWaitInLine:" + traceback.format_exc() + " " + str(e)
            print(f"{ex_stat}")
            return (f"Error: {ex_stat}")


    async def rpa_wait_in_line(self, request):
        try:
            print("task waiting in line.....")
            await self.gui_rpa_msg_queue.put(request)
            print("rpa tasks now in line....", datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])
        except Exception as e:
            ex_stat = "ErrorRPAWaitInLine:" + traceback.format_exc() + " " + str(e)
            print(f"{ex_stat}")


    async def serveCommander(self, msgQueue):
        log3("starting serve Commanders", "serveCommander", self)
        heartbeat = 0
        while True:
            try:
                heartbeat = heartbeat + 1
                if heartbeat > 255:
                    heartbeat = 0

                if heartbeat%16 == 0:
                    # sends a heart beat to commander

                    hbJson = {
                        "ip": self.ip,
                        "type": "heartbeat",
                        "content" : {
                            "vstatus": self.working_state,
                            "running_mid": self.running_mission.getMid() if self.running_mission else 0,
                            "running_instruction": running_step_index
                        }
                    }
                    msg = json.dumps(hbJson)
                    # send to commander
                    msg_with_delimiter = msg + "!ENDMSG!"
                    logger.debug("platoon heartbeat", "wan_log", self, self.running_mission, running_step_index, "~^v^v~")
                    if self.commanderXport:
                        log3("sending heartbeat", "serveCommander", self)
                        if self.commanderXport and not self.commanderXport.is_closing():
                            self.commanderXport.write(msg_with_delimiter.encode('utf8'))
                        # self.commanderXport.get_loop().call_soon(lambda: print("HB MSG SENT2COMMANDER..."))
                elif heartbeat%19 == 0:
                    # no need to do this, just make sure commander always send set schedule command
                    # after a ping-pong sequence...... and after syncing fingerprint profiles....
                    if False and self.platoonHasNoneTodo():
                        workReq = {"type": "reqResendWorkReq", "ip": self.ip, "content": "now"}
                        await self.send_json_to_commander(self.commanderXport, workReq)
            except (json.JSONDecodeError, AttributeError) as e:
                # Handle JSON encoding or missing attributes issues
                log3(f"Error encoding heartbeat JSON or missing attribute: {e}", "serveCommander", self)
            except OSError as e:
                # Handle network-related errors
                log3(f"Error sending heartbeat to Commander: {e}", "serveCommander", self)

            logger.info("serving commander, checking queue...")
            if not msgQueue.empty():
                try:
                    net_message = await msgQueue.get()
                    # log3("From Commander, recevied queued net message: "+net_message, "serveCommander", self)
                    self.processCommanderMsgs(net_message)
                    msgQueue.task_done()
                except asyncio.QueueEmpty:
                    # If for some reason the queue is unexpectedly empty, handle it
                    log3("Queue unexpectedly empty when trying to get message.", "serveCommander", self)
                except Exception as e:
                    # Catch any other issues while processing the message
                    traceback_info = traceback.extract_tb(e.__traceback__)
                    # Extract the file name and line number from the last entry in the traceback
                    log3("Error processing commander msg:" + traceback.format_exc() + " " + str(e), "serveCommander", self)


            await asyncio.sleep(2)
            # log3("watching Commanders...", "serveCommander", self)

    def todoAlreadyExists(self, msg):
        exists = False
        today = datetime.now()
        # Format the date as yyyymmdd
        yyyymmdd = today.strftime("%Y%m%d")
        sf_name = "todos" + yyyymmdd + ".json"
        todays_todo_file = os.path.join(self.my_ecb_data_homepath + "/runlogs", sf_name)

        if os.path.exists(todays_todo_file):
            exists = True

        return exists

    def saveTodaysTodos(self, msg):
        today = datetime.now()
        # Format the date as yyyymmdd
        yyyymmdd = today.strftime("%Y%m%d")
        sf_name = "todos" + yyyymmdd + ".json"
        todays_todo_file = os.path.join(self.my_ecb_data_homepath + "/runlogs", sf_name)
        # print("msg:", msg)
        if msg['todos']:
            with open(todays_todo_file, "w") as tdf:
                json.dump(msg, tdf, indent=4)
                tdf.close()



    # '{"cmd":"reqStatusUpdate", "missions":"all"}'
    # content format varies according to type.
    def processCommanderMsgs(self, msgString):
        try:
            if len(msgString) > 256:
                log3("received from commander: " + msgString[:255] + "...", "serveCommander", self)
            else:
                log3("received from commander: "+msgString, "serveCommander", self)
            if "!connection!" in msgString:
                msg = {"cmd": "connection"}
                msg_parts = msgString.split("!")
                self.commanderIP = msg_parts[0]
                self.commanderName = msg_parts[2]

            elif "!net loss" in msgString:
                msg = {"cmd": "net loss"}
            else:
                msg_parts = msgString.split("!")
                msg_data = "".join(msg_parts[2:])
                msg = json.loads(msg_data)
            # first, check ip and make sure this from a know vehicle.
            if msg["cmd"] == "reqStatusUpdate":
                if msg["missions"] != "":
                    if msg["missions"] == "all":
                        mids = [m.getMid() for m in self.missions]
                    else:
                        mid_chars = msg["missions"].aplit(",")
                        mids = [int(mc) for mc in mid_chars]

                    # capture all the status of all the missions specified and send back the commander...
                    self.sendCommanderMissionsStatMsg(mids)

            elif msg["cmd"] == "reqSendFile":
                # update vehicle status display.
                log3("received a file: "+msg["file_name"], "serveCommander", self)
                file_name = self.ads_profile_dir + msg["file_name"]
                file_type = msg["file_type"]
                file_contents = msg["file_contents"].encode('latin1')  # Encode string to binary data
                with open(file_name, 'wb') as file:
                    file.write(file_contents)

                # first check if the missions are completed or being run or not, if so nothing could be done.
                # otherwise, simply update the mission status to be "Cancelled"
            elif msg["cmd"] == "reqCancelMissions":
                # update vehicle status display.
                self.showMsg(msg["content"])
                # first check if the missions are completed or being run or not, if so nothing could be done.
                # otherwise, simply update the mission status to be "Cancelled"
            elif msg["cmd"] == "reqSetSchedule":
                # schedule work now..... append to array data structure and set up the pointer to the 1st task.
                # the actual running of the tasks will be taken care of by the schduler.

                if not self.todoAlreadyExists(msg):
                    if msg:
                        self.saveTodaysTodos(msg)
                        self.setupScheduledTodos(msg)
                else:
                    log3("commander sent todos exists in the queue. "+json.dumps(self.todays_work["tbd"]), "serveCommander", self)

            elif msg["cmd"] == "reqSetReactiveWorks":
                # schedule work now..... append to array data structure and set up the pointer to the 1st task.
                # the actual running of the tasks will be taken care of by the schduler.
                localworks = msg["todos"]
                self.addBotsMissionsSkillsFromCommander(msg["bots"], msg["missions"], msg["skills"])

                log3("received reactive work request:"+json.dumps(localworks), "serveCommander", self)
                # send work into work Queue which is the self.todays_work["tbd"] data structure.

                self.reactive_work["tbd"].append({"name": "automation", "works": localworks, "status": "yet to start", "current widx": 0, "vname": self.machine_name+":"+self.os_short, "completed": [], "aborted": []})
                log3("after assigned work, "+str(len(self.todays_work["tbd"]))+" todos exists in the queue. "+json.dumps(self.todays_work["tbd"]), "serveCommander", self)

                platform_os = self.platform            # win, mac or linux
                vname = self.machine_name + ":" + self.os_short
                self.todays_scheduled_task_groups[vname] = localworks
                self.unassigned_scheduled_task_groups[vname] = localworks

                # generate ADS loadable batch profiles ((vTasks, vehicle, commander):)
                batched_tasks, ads_profiles = formADSProfileBatchesFor1Vehicle(localworks, self, self)

                # clean up the reports on this vehicle....
                self.todaysReports = []
                self.DONE_WITH_TODAY = False

            elif msg["cmd"] == "reqCancelAllMissions":
                # update vehicle status display.
                log3(json.dumps(msg["content"]), "serveCommander", self)
                self.sendRPAMessage(msg_data)
            elif msg["cmd"] == "reqHaltMissions":
                # update vehicle status display.
                log3(json.dumps(msg["content"]), "serveCommander", self)
                self.sendRPAMessage(msg_data)
                # simply change the mission's status to be "Halted" again, this will make task runner to run this mission
            elif msg["cmd"] == "reqResumeMissions":
                # update vehicle status display.
                log3(json.dumps(msg["content"]), "serveCommander", self)
                self.sendRPAMessage(msg_data)
                # simply change the mission's status to be "Scheduled" again, this will make task runner to run this mission
            elif msg["cmd"] == "reqAddMissions":
                # update vehicle status display.
                log3(json.dumps(msg["content"]), "serveCommander", self)
                # this is for manual generated missions, simply added to the todo list.
            elif msg["cmd"] == "reqSyncFingerPrintProfiles":
                # update vehicle status display.
                # print("profile syncing request received.....")
                # log3(json.dumps(msg["content"]), "serveCommander", self)
                # first gather all finger prints and update them to the latest
                localFingerPrintProfiles = self.gatherFingerPrints()
                self.batchSendFingerPrintProfilesToCommander(localFingerPrintProfiles)

            elif msg["cmd"] == "botsADSProfilesBatchUpdate":
                log3("received commander botsADSProfilesBatchUpdate message")
                # message format {type: chat, msg: msg} msg will be in format of timestamp>from>to>text
                outdated = self.receiveBotsADSProfilesBatchUpdateMessage(msg)
                print("any outdated remote:", outdated)

            elif msg["cmd"] == "ping":
                # respond to ping with pong
                self_info = {"name": platform.node(), "os": platform.system(), "machine": platform.machine()}
                resp = {"ip": self.ip, "type":"pong", "content": self_info}
                # send to commander
                log3("sending "+json.dumps(resp)+ " to commanderIP - " + self.commanderIP, "serveCommander", self)
                print(self.commanderXport)
                msg = json.dumps(resp)
                msg_with_delimiter = msg + "!ENDMSG!"
                if self.commanderXport and not self.commanderXport.is_closing():
                    self.commanderXport.write(msg_with_delimiter.encode('utf8'))
                # asyncio.get_running_loop().call_soon(lambda: print("PONG MSG SENT2COMMANDER..."))

                log3("pong sent!", "serveCommander", self)

            elif msg["cmd"] == "chat":
                # update vehicle status display.
                log3(json.dumps(msg), "serveCommander", self)
                # this message is a chat to a bot/bot group, so forward it to the bot(s)
                # first, find out the bot's queue(which is kind of a temp mailbox for the bot and drop it there)
                self.receiveBotChatMessage(msg["message"])

        except Exception as e:
            # Catch any other issues while processing the message
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorwanProcessCommanderMsgs:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorwanProcessCommanderMsgs traceback information not available:" + str(e)
            log3(f"{ex_stat}", "serveCommander", self)

    def sendCommanderMissionsStatMsg(self, mids):
        statusJson = self.genMissionStatusReport(mids, False)
        msg = "{\"ip\": \"" + self.ip + "\", \"type\":\"status\", \"content\":\"" + json.dumps(statusJson).replace('"', '\\"') + "\"}"

        # Append the delimiter
        msg_with_delimiter = msg + "!ENDMSG!"
        # send to commander
        if self.commanderXport and not self.commanderXport.is_closing():
            self.commanderXport.write(msg_with_delimiter.encode('utf8'))
        # asyncio.get_running_loop().call_soon(lambda: print("MSTAT MSG SENT2COMMANDER..."))

    def sendRPAMessage(self, msg_data):
        asyncio.create_task(self.gui_rpa_msg_queue.put(msg_data))


    # a run report is just an array of the following object:
    # MissionStatus {
    #     mid: ID!
    #     bid: ID!
    #     blevels: String!
    #     status: String!
    #     }
    # 1 report is for 1 TBD workgroup.
    def genRunReport(self, run_type, last_start, last_end, current_mid, current_bid, run_status):
        statReport = None
        tzi = 0
        #only generate report when all done.
        if run_type == "scheduled":
            works = self.todays_work["tbd"][0]["works"]
        else:
            works = self.reactive_work["tbd"][0]["works"]

        if current_bid < 0:
            current_bid = 0

        # self.showMsg("GEN REPORT FOR WORKS:"+json.dumps(works))
        if not self.host_role == "Commander Only":
            current_mission = next((m for m in self.missions if m.getMid() == current_mid), None)
            if current_mission:
                mission_report = current_mission.genSummeryJson()
            else:
                mission_report = {}

            log3("mission_report:"+json.dumps(mission_report), "genRunReport", self)

            if self.host_role != "Platoon":
                # add generated report to report list....
                log3("commander gen run report....."+str(len(self.todaysReport)) + str(len(works)), "genRunReport", self)
                self.todaysReport.append(mission_report)
                # once all of today's task created a report, put the collection of reports into todaysPlatoonReports.
                # on commander machine, todaysPlatoonReports contains a collection of reports from each host machine
                if len(self.todaysReport) == len(works):
                    log3("time to pack today's non-platoon report", "genRunReport", self)
                    rpt = {"ip": self.ip, "type": "report", "content": self.todaysReport}
                    self.todaysPlatoonReports.append(rpt)
                    self.todaysReport = []
            else:
                # self.todaysPlatoonReports.append(str.encode(json.dumps(rpt)))
                # self.showMsg("platoon?? gen run report....."+json.dumps(self.todaysReport))
                self.todaysReport.append(mission_report)
                # once all of today's task created a report, put the collection of reports into todaysPlatoonReports.
                # on platoon machine, todaysPlatoonReports contains a collection of individual task reports on this machine.
                if len(self.todaysReport) == len(works):
                    log3("time to pack today's platoon report", "genRunReport", self)
                    # rpt = {"ip": self.ip, "type": "report", "content": self.todaysReport}
                    # self.todaysPlatoonReports.append(rpt)
                    # self.todaysReport.append(rpt)

        log3(f"GEN REPORT FOR WORKS...[{len(self.todaysReport)}] {json.dumps(self.todaysReport[-1])}", "genRunReport", self)
        return self.todaysReport

    def updateMissionsStatsFromReports(self, all_reports):
        for rpt in all_reports:
            found = next((x for x in self.missions if x.getMid() == rpt["mid"]), None)
            if found:
                found.setStatus(rpt["status"])
                found.setActualStartTime(rpt["starttime"])
                found.setActualEndTime(rpt["endtime"])

    # this function if for all SCHEDULED work done today, now
    # 1) send report to the network,
    # 2) save report to local logs,
    # 3) clear today's work data structures.
    #
    async def doneWithToday(self):
        global commanderXport
        # call reportStatus API to send today's report to API
        log3("Done with today!", "doneWithToday", self)

        if not self.DONE_WITH_TODAY:
            self.DONE_WITH_TODAY = True
            self.rpa_work_assigned_for_today = False

            if not self.host_role == "Platoon":
                # if self.host_role == "Commander":
                #     self.showMsg("commander generate today's report")
                #     rpt = {"ip": self.ip, "type": "report", "content": self.todaysReports}
                #     self.todaysPlatoonReports.append(rpt)

                if len(self.todaysPlatoonReports) > 0:
                    # flatten the report data structure...
                    allTodoReports = [item for pr in self.todaysPlatoonReports for item in pr["content"]]
                    log3("ALLTODOREPORTS:"+json.dumps(allTodoReports), "doneWithToday", self)
                    # missionReports = [item for pr in allTodoReports for item in pr]
                else:
                    allTodoReports = []

                self.updateMissionsStatsFromReports(allTodoReports)

                log3("TO be sent to cloud side::"+json.dumps(allTodoReports), "doneWithToday", self)
                # if this is a commmander, then send report to cloud
                # send_completion_status_to_cloud(self.session, allTodoReports, self.get_auth_token(), self.getWanApiEndpoint())
                eodReportMsg = {
                    "type": "TEAM_REPORT",
                    "bid": "",
                    "report": allTodoReports
                }
                self.gui_manager_msg_queue.put(eodReportMsg)
            else:
                # if this is a platoon, send report to commander today's report is just an list mission status....
                if len(self.todaysReport) > 0:
                    rpt = {"ip": self.ip, "type": "report", "content": self.todaysReport}
                    # Append the delimiter
                    rpt_with_delimiter = json.dumps(rpt) + "!ENDMSG!"
                    log3("Sending report to Commander::"+json.dumps(rpt), "doneWithToday", self)
                    # self.commanderXport.write(str.encode(rpt_with_delimiter))
                    if self.commanderXport and not self.commanderXport.is_closing():
                        self.commanderXport.write(rpt_with_delimiter.encode('utf-8'))
                    # asyncio.get_running_loop().call_soon(lambda: print("DONE MSG SENT2..."))


                # also send updated bot ADS profiles to the commander for backup purose.
                # for bot_profile in self.todays_bot_profiles:
                #     self.send_ads_profile_to_commander(self.commanderXport, "txt", bot_profile)
                localFingerPrintProfiles = self.gatherFingerPrints()
                self.batchSendFingerPrintProfilesToCommander(localFingerPrintProfiles)

            # 2) log reports on local drive.
            self.saveDailyRunReport(self.todaysPlatoonReports)

            # 3) clear data structure, set up for tomorrow morning, this is the case only if this is a commander
            if not self.host_role == "Platoon":
                self.todays_work = {"tbd": [
                    {"name": "fetch schedule", "works": self.gen_default_fetch(), "status": "yet to start",
                     "current widx": 0, "completed": [], "aborted": []}]}
                self.mission_service.update_missions_by_id(self.missions)

            self.todays_completed = []
            self.todaysReports = []                     # per vehicle/host
            self.todaysPlatoonReports = []

    async def sendFingerPrintProfilesToCommander(self, profiles):
        for bot_profile in profiles:
            await self.send_ads_profile_to_commander(self.commanderXport, "txt", bot_profile)

    def batchSendFingerPrintProfilesToCommander(self, profiles):
        self.batch_send_ads_profiles_to_commander(self.commanderXport, "txt", profiles)


    def obtainTZ(self):
        local_time = time.localtime()  # returns a `time.struct_time`
        tzname_local = local_time.tm_zone
        if "East" in tzname_local or "EST" in tzname_local:
            tz = "eastern"
        elif "Pacific" in tzname_local or "PST" in tzname_local:
            tz = "pacific"
        elif "Central" in tzname_local or "CST" in tzname_local:
            tz = "central"
        elif "Mountain" in tzname_local or "MST" in tzname_local:
            tz = "mountain"
        elif "Alaska" in tzname_local or "AST" in tzname_local:
            tz = "alaska"
        elif "Hawaii" in tzname_local or "HST" in tzname_local:
            tz = "hawaii"
        else:
            tz = "eastern"

        return tz

    def getTZ(self):
        return self.tz

    def gen_default_fetch(self):
        FETCH_ROUTINE = [{
                    "mid": 0,
                    "bid": 0,
                    "name": "fetch schedules",
                    "cuspas": "",
                    "todos": None,
                    "start_time": START_TIME,
                    "end_time": "",
                    "stat": "nys"
                }]

        return FETCH_ROUTINE

    def createTrialRunMission(self):
        trMission = EBMISSION(self)
        trMission.setMid(20231225)
        trMission.pubAttributes.setType("user", "Sell")
        trMission.pubAttributes.setBot(0)
        trMission.setCusPAS("win,chrome,amz")
        self.missions.append(trMission)

        return trMission

    def getTrialRunMission(self):
        return self.trMission

    def addSkillToTrialRunMission(self, skid):
        found = False
        for m in self.missions:
            if m.getMid() == 20231225:
                found = True
                break
        if found:
            m.setSkills([skid])

    def getTrialRunMission(self):
        found = False
        for m in self.missions:
            if m.getMid() == 20231225:
                found = True
                break
        if found:
            return m
        else:
            return None

    # Search functions removed - UI components no longer exist


    # build up a dictionary of bot - to be visited site list required by today's mission.
    # this list will be used to filter out cookies of unrelated site, otherwise the
    # naturally saved cookie file by ADS will be too large to fit into an xlsx cell.
    # and the ADS profile import only access xlsx file format.
    def build_cookie_site_lists(self, added=[]):
        today = datetime.today()
        formatted_today = today.strftime('%Y-%m-%d')
        # first, filter out today's missions by createon parameter.

        if added:
            log3("for ADDED only", "build_cookie_site_lists", self)
            targetMissions = added
        else:
            targetMissions = self.missions
            self.bot_cookie_site_lists = {}

        for m in targetMissions:
            log3("mission" + str(m.getMid()) + " created ON:" + m.getBD().split(" ")[0] + " today:" + formatted_today, "build_cookie_site_lists", self)

        missions_today = list(filter(lambda m: formatted_today == m.getBD().split(" ")[0], targetMissions))
        # first ,clear today's bot cookie site list dictionary

        for mission in missions_today:
            bots = [b for b in self.bots if b.getBid() == mission.getBid()]
            if len(bots) > 0:
                bot = bots[0]
                if bot.getEmail() == "":
                    log3("Error: Mission("+str(mission.getMid())+") Bot("+str(bot.getBid())+") running ADS without an Account!!!!!", "build_cookie_site_lists", self)
                else:
                    if bot.getEmail():
                        user_prefix = bot.getEmail().split("@")[0]
                        mail_site_words = bot.getEmail().split("@")[1].split(".")
                        mail_site = mail_site_words[len(mail_site_words) - 2]
                        bot_mission_ads_profile = user_prefix+"_m"+str(mission.getMid()) + ".txt"

                        self.bot_cookie_site_lists[bot_mission_ads_profile] = [mail_site]
                        if mail_site == "gmail":
                            self.bot_cookie_site_lists[bot_mission_ads_profile].append("google")

                        if mission.getSite() == "amz":
                            self.bot_cookie_site_lists[bot_mission_ads_profile].append("amazon")
                        if mission.getSite() == "ebay":
                            self.bot_cookie_site_lists[bot_mission_ads_profile].append("ebay")
                        elif mission.getSite() == "ali":
                            self.bot_cookie_site_lists[bot_mission_ads_profile].append("aliexpress")
                        else:
                            self.bot_cookie_site_lists[bot_mission_ads_profile].append(mission.getSite().lower())

        log3("just build cookie site list:"+json.dumps(self.bot_cookie_site_lists), "build_cookie_site_lists", self)

    def setADSBatchSize(self, batch_size):
        self.config_manager.ads_settings.batch_size = batch_size
        self.config_manager.ads_settings.save()

    def getADSBatchSize(self):
        return self.config_manager.ads_settings.batch_size

    def getADSBatchMethod(self):
        return self.config_manager.ads_settings.batch_method

    def getADSSettings(self):
        return self.config_manager.ads_settings.data

    def saveADSSettings(self, settings):
        try:
            for key, value in settings["fp_browser_settings"].items():
                setattr(self.config_manager.ads_settings, key, value)
            self.config_manager.ads_settings.save()
        except Exception as e:
            logger.error(f"Error saving ADS settings: {e}")

    def getIP(self):
        return self.ip

    def getHostName(self):
        return self.machine_name

    def getCookieSiteLists(self):
        return self.bot_cookie_site_lists

    def getADSProfileDir(self):
        return self.ads_profile_dir


    def send_chat_to_local_bot(self, chat_msg):
        # """ Directly enqueue a message to the asyncio task when the button is clicked. """
        asyncio.create_task(self.gui_chat_msg_queue.put(chat_msg))

    # the message will be in the format of botid:send time stamp in yyyy:mm:dd hh:mm:ss format:msg in html format
    # from network the message will have chatmsg: prepend to the message.
    def update_chat_gui(self, rcvd_msg):
        # Chat GUI removed - no longer updating chat display
        pass

    # this is the interface to the chatting bots, taking message from the running bots and display them on GUI
    async def connectChat(self, chat_msg_queue):
        running = True
        while running:
            if not chat_msg_queue.empty():
                message = await chat_msg_queue.get()
                self.showMsg(f"Rx Chat message from bot: {message}")
                # Chat GUI removed - no longer updating chat display
                if self.host_role != "Staff Officer":
                    response = self.think_about_a_reponse(message)
                    self.c_send_chat(response)
                chat_msg_queue.task_done()

            logger.trace("chat Task ticking....")
            await asyncio.sleep(1)

    def getBV(self, bot):
        if bot.getVehicle():
            return bot.getVehicle()
        else:
            return ""

    def getBotsOnThisVehicle(self):
        thisBots = [b for b in self.bots if self.machine_name in self.getBV(b) ]
        return thisBots

    def getBidsOnThisVehicle(self):
        thisBots = self.getBotsOnThisVehicle()
        thisBids = [b.getBid() for b in thisBots]
        thisBidsString = json.dumps(thisBids)
        self.showMsg("bids on this vehicle:"+thisBidsString)
        return thisBidsString

    def prepFullVehicleReportData(self):
        print("prepFullVehicleReportData...")
        report = []
        try:
            for v in self.vehicles:
                if v.getStatus() == "":
                    vstat = "offline"
                else:
                    vstat = v.getStatus()

                if self.machine_name not in v.getName():
                    vinfo = {
                        "vid": v.getVid(),
                        "vname": v.getName(),
                        "owner": self.user,
                        "status": vstat,
                        "lastseen": v.getLastUpdateTime().strftime("%Y-%m-%d %H:%M:%S.%f")[:19],
                        "functions": v.getFunctions(),
                        "bids": ",".join(str(v.getBotIds())),
                        "hardware": v.getArch(),
                        "software": v.getOS(),
                        "ip": v.getIP(),
                        "created_at": ""
                    }
                else:
                    vinfo = {
                        "vid": 0,
                        "vname": self.machine_name+":"+self.os_short,
                        "owner": self.user,
                        "status": self.working_state,
                        "lastseen": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:19],
                        "functions": self.functions,
                        "bids": self.getBidsOnThisVehicle(),
                        "hardware": self.processor,
                        "software": self.platform,
                        "ip": self.ip,
                        "created_at": ""
                    }
                report.append(vinfo)
            print("vnames:", [v["vname"] for v in report])
            if (self.machine_name+":"+self.os_short) not in [v["vname"] for v in report]:
                if "Only" not in self.host_role and "Staff" not in self.host_role:
                    # add myself as a vehicle resource too.
                    vinfo = {
                        "vid": 0,
                        "vname": self.machine_name+":"+self.os_short,
                        "owner": self.user,
                        "status": self.working_state,
                        "lastseen": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:19],
                        "functions": self.functions,
                        "bids": self.getBidsOnThisVehicle(),
                        "hardware": self.processor,
                        "software": self.platform,
                        "ip": self.ip,
                        "created_at": ""
                    }

                    report.append(vinfo)
                    print("report:", report)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorPrepFullVReport:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorPrepFullVReport traceback information not available:" + str(e)
            print(ex_stat)

        return report


    def prepVehicleReportData(self, v):
        report = []

        if v:
            vinfo = {
                "vid": v.getVid(),
                "vname": v.getName(),
                "owner": self.user,
                "status": v.getStatus(),
                "lastseen": v.getLastUpdateTime().strftime("%Y-%m-%d %H:%M:%S.%f")[:19],
                "functions": v.getFunctions(),
                "bids": json.dumps(v.getBotIds()),
                "hardware": v.getArch(),
                "software": v.getOS(),
                "ip": v.getIP(),
                "created_at": ""
            }
        else:
            vinfo = {
                "vid": 0,
                "vname": self.machine_name+":"+self.os_short,
                "owner": self.user,
                "status": self.working_state,
                "lastseen": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:19],
                "functions": self.functions,
                "bids": self.getBidsOnThisVehicle(),
                "hardware": self.processor,
                "software": self.platform,
                "ip": self.ip,
                "created_at": ""
            }
        report.append(vinfo)

        return report


    # this is the interface to the chatting bots, taking message from the running bots and display them on GUI
    async def runRPAMonitor(self, monitor_msg_queue):
        running = True
        ticks = 0
        while running:
            ticks = ticks + 1
            if ticks > 255:
                ticks = 0

            #ping cloud every 8 second to see whether there is any monitor/control internet. use amazon's sqs
            if ticks % 8 == 0:
                self.showMsg(f"Access Internet Here with Websocket...")

            if ticks % 180 == 0:
                self.showMsg(f"report vehicle status")

                # update vehicles status to local disk, this is done either on platoon or commander
                self.saveVehiclesJsonFile()

                if "Commander" in self.host_role:
                    self.showMsg(f"sending vehicle heartbeat to cloud....")
                    hbInfo = self.stateCapture()
                    # update vehicle info to the chat channel (don't we need to update this to cloud lambda too?)
                    await self.wan_send_heartbeat(hbInfo)

                    # send vehicle status to cloud DB
                    vehicle_report = self.prepFullVehicleReportData()
                    resp = send_report_vehicles_to_cloud(self.session, self.get_auth_token(),
                                                         vehicle_report, self.getWanApiEndpoint())

            if not monitor_msg_queue.empty():
                message = await monitor_msg_queue.get()
                self.showMsg(f"RPA Monitor message: {message}")
                if type(message) != str:
                    if hasattr(self, 'vehicle_monitor'):
                        # Note: VehicleMonitorManager is now a data handler, not a GUI window
                        # You may need to implement a new method to handle log messages
                        self.vehicle_monitor.add_log_message(json.dumps(message))

                monitor_msg_queue.task_done()

            logger.trace("running monitoring Task....", ticks)
            await asyncio.sleep(1)
        print("RPA monitor ended!!!")


    def update_monitor_gui(self, in_message):
        try:
            print("raw rpa monitor incoming msg:", in_message)
            # self.showMsg(f"RPA Monitor:"+in_message)
            if in_message["type"] == "request mission" and self.getIP() not in in_message["sender"]:
                print("request mission:", in_message)
                new_works = json.loads(in_message["contents"])
                print("CONFIG:", new_works['added_missions'][0]['config'])

                # downloaded files if any so that we don't have to do this later on....
                # and set up mission input parameters.
                self.prepareMissionRunAsServer(new_works)
                self.handleCloudScheduledWorks(new_works)

            elif in_message["type"] == "request queued":
                print("processing enqueue notification")
                # a request received on the cloud queue side. here what we will do:
                # enqueue an item on local mirror (we call it virtual cloud queue)
                requester_info = json.loads(in_message["contents"])

                print("requester info:", requester_info)

                asyncio.create_task(self.virtual_cloud_task_queue.put(requester_info))

                print("done local enqueue....")

                #then whenever a task group is finished either local or from remote. in that handler.
                # we will probe virtual cloud queue whethere there is something to work on.
                # if not empty, we will dequeue something from the cloud, once received work, we will deque local
                # and dispatch the work into scheduler.
            elif in_message["type"] == "report results":
                ext_run_results = json.loads(in_message["contents"].replace("\\", "\\\\"))
                handleExtLabelGenResults(self.session, self.get_auth_token(), self.getWanApiEndpoint(), ext_run_results)
            else:
                print("Unknown message type!!!")

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "Errorupdate_monitor_gui:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "Errorupdate_monitor_gui traceback information not available:" + str(e)
            print(ex_stat)


    def downloadForFullfillGenECBLabels(self, orders, worklink):
        try:
            for bi, batch in enumerate(orders):
                if batch['file']:
                    print("batch....", batch)
                    print("about to download....", batch['file'])

                    local_file = download_file(self.session, self.my_ecb_data_homepath, batch['dir'] + "/" + batch['file'],
                                               "", self.get_auth_token(),
                                               self.getWanApiEndpoint(), "general")
                    batch['dir'] = os.path.dirname(local_file)
                    orders[bi]['dir'] = os.path.dirname(local_file)
                    worklink['dir'] = os.path.dirname(local_file)


                    print("local file....", local_file)
                    print("local dir:", os.path.dirname(local_file))

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorDownloadForFullfillGenECBLabels:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorDownloadForFullfillGenECBLabels traceback information not available:" + str(e)
            print(ex_stat)


    # do any download if needed by the missions ONLY IF the mission will be run on this computer.
    # otherwise, don't do download and leave this to whichever computer that will run this mission.
    def prepareMissionRunAsServer(self, new_works):
        try:
            if new_works['added_missions'][0]['type'] == "sellFullfill_genECBLabels":
                first_v = next(iter(new_works['task_groups']))
                if self.machine_name in first_v:
                    self.downloadForFullfillGenECBLabels(new_works['added_missions'][0]['config'][1], new_works['task_groups'][first_v]['eastern'][0]['other_works'][0]['config'][1][0])

                print("updated new work:", new_works)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorPrepareMissionRunAsServer:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorPrepareMissionRunAsServer traceback information not available:" + str(e)
            print(ex_stat)

    # note recipient could be a group ID.
    def sendBotChatMessage(self, sender, recipient, text):
        # first find out where the recipient is at (which vehicle) and then, send the message to it.
        if isinstance(recipient, list):
            recipients = recipient
        else:
            recipients = [recipient]
        dtnow = datetime.now()
        date_word = dtnow.isoformat()
        allbids = [b.getBid() for b in self.bots]
        found = None
        for vidx, v in enumerate(self.vehicles):
            if len(recipients) > 0:
                receivers = set(recipients)
                vbots = set(v.getBotIds())

                # Find the intersection
                intersection = receivers.intersection(vbots)

                # Convert intersection back to a list (optional)
                bids_on_this_vehicle = list(intersection)
                if len(bids_on_this_vehicle) > 0:
                    bids_on_this_vehicle_string = ",".join(bids_on_this_vehicle)

                    #now send the message to bots on this vehicle.
                    full_txt = date_word + ">" + str(sender) + ">" + bids_on_this_vehicle_string + ">" + text
                    cmd = {"cmd": "chat", "message": full_txt.decode('latin1')}
                    cmd_str = json.dumps(cmd)
                    if v.getFieldLink()["transport"] and not v.getFieldLink()["transport"].is_closing():
                        v.getFieldLink()["transport"].write(cmd_str.encode('utf8'))
                    # v.getFieldLink()["transport"].get_loop().call_soon(lambda: print("CHAT MSG SENT2..."))

                    # Remove the intersection from the recipients.
                    receivers.difference_update(intersection)

                    # get updated recipients
                    recipients = list(receivers)

        # if there are still recipients not sent, that means these are local bots,
        #self.send_chat_to_local_bot(text)

        if len(recipient) > 0:
            # recipient here could be comma seperated recipient ids.
            receivers = set(recipients)
            vbots = set(allbids)

            # Find the intersection
            intersection = receivers.intersection(vbots)

            # Convert intersection back to a list (optional)
            bids_on_this_vehicle = list(intersection)

            # Chat GUI removed - no longer adding active chat history
            pass

            if len(bids_on_this_vehicle) > 0:
                # now send the message to local bots on this vehicle.
                unfound_bid_string = ",".join(bids_on_this_vehicle)
                self.showMsg(f"Error: Bot[{unfound_bid_string}] not found")


    # note recipient could be a group ID.
    def receiveBotChatMessage(self, msg_text):
        msg_parts = msg_text.split(">")
        sender = msg_parts[1]
        receiver = msg_parts[2]
        if "," in receiver:
            receivers = [int(rs.strip()) for rs in receiver.split(",")]
        else:
            receivers = [int(receiver.strip())]


        # deliver the message for the other bots. - allowed for inter-bot communication.
        if len(receivers) > 0:
            # now route message to everybody.
            # Chat GUI removed - no longer adding network chat history
            pass

    # note recipient could be a group ID.
    def receiveBotLogMessage(self, msg_text):
        msg_json = json.loads(msg_text)

        sender = msg_json["sender"]

        #logger will only be sent to the boss.
        receivers = [0]

        # deliver the message for the other bots. - allowed for inter-bot communication.
        # Chat GUI removed - no longer adding network chat history
        pass


    async def send_file_to_platoon(self, platoon_link, file_type, file_name_full_path):
        if os.path.exists(file_name_full_path) and platoon_link:
            self.showMsg(f"Sending File [{file_name_full_path}] to platoon: "+platoon_link["ip"])
            with open(file_name_full_path, 'rb') as fileTBSent:
                binary_data = fileTBSent.read()
                encoded_data = base64.b64encode(binary_data).decode('utf-8')

                # Embed in JSON
                json_data = json.dumps({"cmd": "reqSendFile", "file_name": file_name_full_path, "file_type": file_type, "file_contents": encoded_data})
                length_prefix = len(json_data.encode('utf-8')).to_bytes(4, byteorder='big')
                # Send data
                self.showMsg(f"About to send file json with "+str(len(json_data.encode('utf-8')))+ " BYTES!")
                if platoon_link["transport"] and not platoon_link["transport"].is_closing():
                    platoon_link["transport"].write(length_prefix+json_data.encode('utf-8'))
                    # await platoon_link["transport"].drain()
                    asyncio.get_running_loop().call_soon(lambda: print("FILE MSG SENT2PLATOON..."))
                # await xport.drain()

                fileTBSent.close()
        else:
            if not os.path.exists(file_name_full_path):
                self.showMsg(f"ErrorSendFileToPlatoon: File [{file_name_full_path}] not found")
            else:
                self.showMsg(f"ErrorSendFileToPlatoon: TCP link doesn't exist")


    def send_json_to_platoon(self, platoon_link, json_data):
        if json_data and platoon_link:
            json_string = json.dumps(json_data)
            if len(json_string) < 128:
                log3(f"Sending JSON Data to platoon " + platoon_link["ip"] + "::" + json_string, "sendLAN",
                     self)
            else:
                log3(f"Sending JSON Data to platoon " + platoon_link["ip"] + ":: ..." + json_string[-127:], "sendLAN",
                     self)
            encoded_json_string = json_string.encode('utf-8')
            length_prefix = len(encoded_json_string).to_bytes(4, byteorder='big')
            # Send data
            if platoon_link["transport"] and not platoon_link["transport"].is_closing():
                platoon_link["transport"].write(length_prefix+encoded_json_string)
                # await platoon_link["transport"].drain()
        else:
            if json_data == None:
                log3(f"ErrorSendJsonToPlatoon: JSON empty", "sendLAN", self)
            else:
                log3(f"ErrorSendJsonToPlatoon: TCP link doesn't exist", "sendLAN", self)


    async def send_json_to_commander(self, commander_link, json_data):
        if json_data and commander_link:
            json_string = json.dumps(json_data)
            if len(json_string) < 128:
                log3(f"Sending JSON Data to commander ::" + json.dumps(json_data), "sendLAN", self)
            else:
                log3(f"Sending JSON Data to commander " + commander_link["ip"] + ":: ..." + json_string[-127:], "sendLAN",
                     self)
            encoded_json_string = json_string.encode('utf-8')
            length_prefix = len(encoded_json_string).to_bytes(4, byteorder='big')
            # Send data
            if commander_link and not commander_link.is_closing():
                commander_link.write(length_prefix+encoded_json_string)
                # await commander_link.drain()
                asyncio.get_running_loop().call_soon(lambda: print("JSON MSG SENT2COMMANDER..."))

        else:
            if json_data == None:
                log3(f"ErrorSendJsonToCommander: JSON empty", "sendLAN", self)
            else:
                log3(f"ErrorSendJsonToCommander: TCP link doesn't exist", "sendLAN", self)

    async def send_ads_profile_to_commander(self, commander_link, file_type, file_name_full_path):
        if os.path.exists(file_name_full_path) and commander_link:
            log3(f"Sending File [{file_name_full_path}] to commander: " + self.commanderIP, "sendLAN", self)
            with open(file_name_full_path, 'rb') as fileTBSent:
                binary_data = fileTBSent.read()
                encoded_data = base64.b64encode(binary_data).decode('utf-8')

                # Embed in JSON
                json_data = json.dumps({"type": "botsADSProfilesUpdate", "file_name": file_name_full_path, "file_type": file_type,
                                        "file_contents": encoded_data})
                length_prefix = len(json_data.encode('utf-8')).to_bytes(4, byteorder='big')
                # Send data
                if commander_link and not commander_link.is_closing():
                    commander_link.write(length_prefix + json_data.encode('utf-8'))
                    await commander_link.drain()
                # asyncio.get_running_loop().call_soon(lambda: print("FILE SENT2COMMANDER..."))

                # await xport.drain()

                fileTBSent.close()
        else:
            if not os.path.exists(file_name_full_path):
                self.showMsg(f"ErrorSendFileToCommander: File [{file_name_full_path}] not found")
            else:
                self.showMsg(f" : TCP link doesn't exist")

    def batch_send_ads_profiles_to_commander(self, commander_link, file_type, file_paths):
        try:
            if not commander_link:
                log3("ErrorSendFilesToCommander: TCP link doesn't exist", "sendLAN", self)
                return

            profiles = []
            for file_name_full_path in file_paths:
                if os.path.exists(file_name_full_path):
                    log3(f"Sending File [{file_name_full_path}] to commander: {self.commanderIP}", "sendLAN", self)
                    with open(file_name_full_path, 'rb') as fileTBSent:
                        binary_data = fileTBSent.read()
                        encoded_data = base64.b64encode(binary_data).decode('utf-8')

                        # Embed in JSON
                        file_timestamp = os.path.getmtime(file_name_full_path)

                        profiles.append({
                            "file_name": file_name_full_path,
                            "file_type": file_type,
                            "timestamp": file_timestamp,  # Include file timestamp
                            "file_contents": encoded_data
                        })

                else:
                    self.showMsg(f"ErrorSendFileToCommander: File [{file_name_full_path}] not found")

            # Send data
            json_data = json.dumps({
                "type": "botsADSProfilesBatchUpdate",
                "ip": self.ip,
                "profiles": profiles
            })
            length_prefix = len(json_data.encode('utf-8')).to_bytes(4, byteorder='big')
            if len(json_data) < 128:
                print("About to send botsADSProfilesBatchUpdate to commander: "+json_data)
            else:
                print("About to send botsADSProfilesBatchUpdate to commander: ..." + json_data[-127:])

            if commander_link and not commander_link.is_closing():
                commander_link.write(length_prefix + json_data.encode('utf-8'))
                asyncio.get_running_loop().call_soon(lambda: print("ADS FILES SENT2COMMANDER..."))
        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorSendingBatchProfilesToCommander:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorSendingBatchProfilesToCommander traceback information not available:" + str(e)


    def batch_send_ads_profiles_to_platoon(self, platoon_link, file_type, file_paths):
        try:
            if not platoon_link:
                log3("ErrorSendFilesToCommander: TCP link doesn't exist", "sendLAN", self)
                return

            print("# files", len(file_paths))
            profiles = []
            for file_name_full_path in file_paths:
                print("checking", file_name_full_path)
                if os.path.exists(file_name_full_path):
                    print("exists!")
                    # log3(f"Sending File [{file_name_full_path}] to commander: {self.commanderIP}", "gatherFingerPrints", self)
                    # print(f"Sending File [{file_name_full_path}] to commander: {self.commanderIP}")
                    with open(file_name_full_path, 'rb') as fileTBSent:
                        binary_data = fileTBSent.read()
                        encoded_data = base64.b64encode(binary_data).decode('utf-8')

                        # Embed in JSON
                        file_timestamp = os.path.getmtime(file_name_full_path)

                        profiles.append({
                            "file_name": file_name_full_path,
                            "file_type": file_type,
                            "timestamp": file_timestamp,  # Include file timestamp
                            "file_contents": encoded_data
                        })

                else:
                    log3(f"Warning: ADS Profile [{file_name_full_path}] not found", "sendLAN", self)
                    print(f"Warning: ADS Profile [{file_name_full_path}] not found")

            # Send data
            print("profiles ready")
            json_data = json.dumps({
                "cmd": "botsADSProfilesBatchUpdate",
                "ip": self.ip,
                "profiles": profiles
            })

            if len(json_data) < 128:
                print("About to send botsADSProfilesBatchUpdate to platoon: " + json_data)
            else:
                print("About to send botsADSProfilesBatchUpdate to platoon: ..." + json_data[-127:])


            length_prefix = len(json_data.encode('utf-8')).to_bytes(4, byteorder='big')
            if platoon_link["transport"] and not platoon_link["transport"].is_closing():
                platoon_link["transport"].write(length_prefix + json_data.encode('utf-8'))
            # asyncio.get_running_loop().call_soon(lambda: print("ADS FILES SENT2PLATOON..."))
        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorSendingBatchProfilesToCommander:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorSendingBatchProfilesToCommander traceback information not available:" + str(e)



    def send_mission_result_files_to_commander(self, commander_link, mid, file_type, file_name_full_paths):
        try:
            validFiles = [fn for fn in file_name_full_paths if os.path.exists(fn)]

            nFiles = len(validFiles)
            for fidx, file_name_full_path in enumerate(validFiles):
                if os.path.exists(file_name_full_path) and commander_link:
                    self.showMsg(f"Sending File [{file_name_full_path}] to commander: " + self.commanderIP)
                    with open(file_name_full_path, 'rb') as fileTBSent:
                        binary_data = fileTBSent.read()
                        encoded_data = base64.b64encode(binary_data).decode('utf-8')

                        # Embed in JSON
                        json_data = json.dumps({"type": "missionResultFile", "mid": mid, "nFiles": nFiles, "fidx": fidx, "file_name": file_name_full_path, "file_type": file_type,
                                                "file_contents": encoded_data})
                        length_prefix = len(json_data.encode('utf-8')).to_bytes(4, byteorder='big')
                        # Send data
                        if commander_link and not commander_link.is_closing():
                            commander_link.write(length_prefix + json_data.encode('utf-8'))
                        # asyncio.get_running_loop().call_soon(lambda: print("RESULT FILES SENT2COMMANDER..."))
                        # await xport.drain()

                        fileTBSent.close()
                else:
                    if not os.path.exists(file_name_full_path):
                        self.showMsg(f"ErrorSendMissionResultsFilesToCommander: File [{file_name_full_path}] not found")
                    else:
                        self.showMsg(f"ErrorSendMissionResultsFilesToCommander: TCP link doesn't exist")

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorSendMissionResultsFilesToCommander:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorSendMissionResultsFilesToCommander: traceback information not available:" + str(e)
            log3(ex_stat)


    def set_wan_connected(self, wan_stat):
        self.wan_connected = wan_stat

    def set_websocket(self, ws):
        self.websocket = ws

    def get_wan_connected(self):
        return self.wan_connected

    def get_websocket(self):
        return self.websocket

    def get_wan_msg_queue(self):
        return self.wan_chat_msg_queue

    def set_wan_msg_subscribed(self, ss):
        self.wan_msg_subscribed = ss

    def get_wan_msg_subscribed(self):
        return self.wan_msg_subscribed

    def set_staff_officer_online(self, ol):
        self.staff_officer_on_line = ol

    def get_staff_officer_online(self):
        return self.staff_officer_on_line

    # this is an empty task
    async def wait_forever(self):
        await asyncio.Event().wait()  # This will wait indefinitely

    async def wan_ping(self):
        if self.host_role == "Staff Officer":
            commander_chat_id = self.user.replace("@", "_").replace(".", "_") + "_Commander"
            ping_msg = {
                "chatID": commander_chat_id,
                "sender": self.chat_id,
                "receiver": commander_chat_id,
                "type": "ping",
                "contents": json.dumps({"msg": "hello?"}).replace('"', '\\"'),
                "parameters": json.dumps({})
            }

            self.wan_sub_task = asyncio.create_task(wanSendMessage(ping_msg, self))

    async def wan_self_ping(self):
        if self.host_role == "Staff Officer":
            self_chat_id = self.user.replace("@", "_").replace(".", "_") + "_StaffOfficer"
        else:
            self_chat_id = self.user.replace("@", "_").replace(".", "_") + "_Commander"
        print("Self:", self_chat_id)
        ping_msg = {
            "chatID": self_chat_id,
            "sender": self.chat_id,
            "receiver": self_chat_id,
            "type": "loopback",
            "contents": json.dumps({"msg": "hello?"}).replace('"', '\\"'),
            "parameters": json.dumps({})
        }

        self.wan_sub_task = asyncio.create_task(wanSendMessage(ping_msg, self))


    async def wan_pong(self):
        if "Commander" in self.host_role:
            sa_chat_id = self.user.replace("@", "_").replace(".", "_") + "_StaffOfficer"
            pong_msg = {
                # "chatID": sa_chat_id,
                "chatID": self.chat_id,
                "sender": "Commander",
                "receiver": sa_chat_id,
                "type": "pong",
                "contents": json.dumps({"type": "cmd", "cmd": "pong"}).replace('"', '\\"'),
                "parameters": json.dumps({})
            }
            self.wan_sub_task = asyncio.create_task(wanSendMessage8(pong_msg, self))

    def wan_send_log(self, logmsg):
        if self.host_role != "Staff Officer":
            so_chat_id = self.user.replace("@", "_").replace(".", "_") + "_StaffOfficer"
            contents = {"msg": logmsg}
            parameters = {}
            req_msg = {
                "chatID": so_chat_id,
                "sender": "Commander",
                "receiver": self.user,
                "type": "logs",
                "contents": logmsg.replace('"', '\\"'),
                # "contents": json.dumps({"msg": logmsg}).replace('"', '\\"'),
                "parameters": json.dumps(parameters)
            }
            wanSendMessage(req_msg, self)

    async def wan_send_log8(self, logmsg):
        if self.host_role != "Staff Officer":
            so_chat_id = self.user.replace("@", "_").replace(".", "_") + "_StaffOfficer"
            req_msg = {
                "chatID": so_chat_id,
                "sender": "Commander",
                "receiver": self.user,
                "type": "logs",
                "contents": json.dumps({"msg": logmsg}).replace('"', '\\"'),
                "parameters": json.dumps({})
            }
            self.wan_sub_task = asyncio.create_task(wanSendMessage8(req_msg, self))


    async def wan_request_log(self):
        if self.host_role == "Staff Officer":
            commander_chat_id = self.user.replace("@", "_").replace(".", "_") + "_Commander"
            req_msg = {
                "chatID": self.chat_id,
                "sender": "",
                "receiver": commander_chat_id,
                "type": "request",
                "contents": json.dumps({"type": "cmd", "cmd": "start log", "settings": ["all"]}).replace('"', '\\"'),
                "parameters": ""
            }
            self.wan_sub_task = asyncio.create_task(wanSendMessage8(req_msg, self))

    def wan_stop_log(self):
        if "Commander" in self.host_role:
            sa_chat_id = self.user.replace("@", "_").replace(".", "_") + "_StaffOfficer"
            log_msg = {
                "chatID": sa_chat_id,
                "sender": self.chat_id,
                "receiver": sa_chat_id,
                "type": "command",
                "contents": json.dumps({"type": "cmd", "cmd": "stop log", "settings": ["all"]}).replace('"', '\\"'),
                "parameters": ""
            }
            wanSendMessage(log_msg, self)

    async def wan_stop_log(self):
        if self.host_role == "Staff Officer":
            commander_chat_id = self.user.replace("@", "_").replace(".", "_") + "_Commander"
            req_msg = {
                "chatID": commander_chat_id,
                "sender": "",
                "receiver": commander_chat_id,
                "type": "command",
                "contents": json.dumps({"type": "cmd", "cmd": "stop log", "settings": ["all"]}).replace('"', '\\"'),
                "parameters": ""
            }
            self.wan_sub_task = asyncio.create_task(wanSendMessage8(req_msg, self))

    def wan_rpa_ctrl(self, cmd):
        if self.host_role == "Staff Officer":
            commander_chat_id = self.user.replace("@", "_").replace(".", "_") + "_Commander"
            req_msg = {
                "chatID": self.chat_id,
                "sender": "",
                "receiver": commander_chat_id,
                "type": "command",
                "contents": json.dumps({"cmd": cmd, "settings": ["all"]}).replace('"', '\\"'),
                "parameters": ""
            }
            self.wan_sub_task = asyncio.create_task(wanSendMessage(req_msg, self))

    async def wan_send_heartbeat(self, heartbeatInfo):
        if "Commander" in self.host_role:
            sa_chat_id = self.user.replace("@", "_").replace(".", "_") + "_StaffOfficer"
            req_msg = {
                "chatID": sa_chat_id,
                "sender": self.user.replace("@", "_").replace(".", "_") + "_Commander",
                "receiver": sa_chat_id,
                "type": "heartbeat",
                "contents": json.dumps(heartbeatInfo).replace('"', '\\"'),
                "parameters": json.dumps({}),
            }
            # self.wan_sub_task = asyncio.create_task(wanSendMessage8(req_msg, self.get_auth_token(), self.websocket))
            await wanSendMessage8(req_msg, self)


    def send_heartbeat(self):
        if "Commander" in self.host_role:
            print("sending wan heartbeat")
            asyncio.ensure_future(self.wan_send_heartbeat())


    async def wan_chat_test(self):
        """Non-blocking WAN chat test - converted to async to avoid UI freeze"""
        if self.host_role == "Staff Officer":
            asyncio.ensure_future(self.wan_ping())
            await asyncio.sleep(1.0)  # Non-blocking wait
            asyncio.ensure_future(self.wan_self_ping())
        elif self.host_role != "Platoon":
            asyncio.ensure_future(self.wan_pong())
            # asyncio.ensure_future(self.wan_c_send_chat("got it!!!"))
            # await asyncio.sleep(1.0)
            # asyncio.ensure_future(self.wan_self_ping())
            # self.think_about_a_reponse("[abc]'hello?'")


    async def wan_sa_send_chat(self, msg):
        try:
            if self.host_role == "Staff Officer":
                commander_chat_id = self.user.split("@")[0] + "_Commander"
                req_msg = {
                    "chatID": commander_chat_id,
                    "sender": self.chat_id,
                    "receiver": commander_chat_id,
                    "type": "chat",
                    "contents": msg,
                    "parameters": json.dumps({})
                }

                self.wan_sub_task = asyncio.create_task(wanSendMessage8(req_msg, self))

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorWanSaSendChat:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorWanSaSendChat: traceback information not available:" + str(e)
            log3(ex_stat)


    def sa_send_chat(self, msg):
        asyncio.ensure_future(self.wan_sa_send_chat(msg))


    async def wan_c_send_chat(self, msg):
        if "Commander" in self.host_role:
            sa_chat_id = self.user.split("@")[0] + "_StaffOfficer"
            req_msg = {
                "chatID": sa_chat_id,
                "sender": self.chat_id,
                "receiver": sa_chat_id,
                "type": "chat",
                "contents": msg,
                "parameters": json.dumps({})
            }

            self.wan_sub_task = asyncio.create_task(wanSendMessage(req_msg, self))


    def c_send_chat(self, msg):
        asyncio.ensure_future(self.wan_c_send_chat(msg))

        current_time = datetime.now(timezone.utc)
        # Convert to the required AWSDateTime format
        aws_datetime_string = current_time.isoformat()


    def think_about_a_reponse(self, thread):
        logger.info("Thinking about response.")
        current_time = datetime.now(timezone.utc)
        aws_datetime_string = current_time.isoformat()

        session = self.session
        token = self.get_auth_token()
        qs = [{
            "msgID": "1",
            "user": self.user,
            "timeStamp": aws_datetime_string,
            "products": "",
            "goals": "",
            "options": "",
            "background": thread,
            "msg": "provide answer"
        }]
        resp = send_query_chat_request_to_cloud(session, token, qs, self.getWanApiEndpoint())

        logger.info("THINK RESP:", resp)

    # from ip find vehicle, and update its status, and
    def updateVehicleStatusToRunningIdle(self, ip):
        found_vehicles = [v for v in self.vehicles if v.getIP() == ip]
        if found_vehicles:
            found_vehicle = found_vehicles[0]
            found_vehicle.setStatus("running_idle")       # this vehicle is ready to take more work if needed.
            vehicle_report = self.prepVehicleReportData(found_vehicle)
            log3("vehicle status report"+json.dumps(vehicle_report))
            if not self.general_settings.is_test_mode():
                resp = send_report_vehicles_to_cloud(self.session, self.get_auth_token(),
                                                 vehicle_report, self.getWanApiEndpoint())
            self.saveVehiclesJsonFile()

    def updateVehicles(self, vehicles):
        vjs=[self.prepVehicleReportData(v) for v in vehicles]

        resp = send_update_vehicles_request_to_cloud(self.session, vjs, self.get_auth_token(), self.getWanApiEndpoint())
        # for now simply update json file, can put in local db if needed in future.... sc-01/06/2025
        self.saveVehiclesJsonFile()


    # capture current state and send in heartbeat signal to the cloud
    def stateCapture(self):
        current_time = datetime.now()
        for v in self.vehicles:
            current_time = datetime.now()
            if (current_time - v.getLastUpdateTime()).total_seconds() > 480:    # 8 minutes no contact is considered "offline"
                v.setStatus("offline")


        stateInfo = {"vehiclesInfo": [{"vehicles_status": v.getStatus(), "vname": v.getName()} for v in self.vehicles]}

        # we'll capture these info:
        # all vehicle status running_idle/running_working/offline
        # all the mission running status.??? may be this is not the good place for that, we'd have to ping vehicles
        # plus, don't they update periodically already?


        return stateInfo

    def vRunnable(self, vehicle):
        print("vname", vehicle.getName(), self.machine_name, self.host_role)
        runnable = True
        if self.machine_name in vehicle.getName() and self.host_role == "Commander Only":
            runnable = False
        return runnable

    # check whether there is vehicle for hire, if so, check any contract work in the queue
    # if so grab it.
    async def checkCloudWorkQueue(self):
        try:
            taskGroups = {}
            # some debugging here
            # print("N vehicles:", len(self.vehicles))
            # if len(self.vehicles) > 0:
            #     for v in self.vehicles:
            #         print("vname:", v.getName(), "status:", v.getStatus(), )
            # check whether there is any thing in the local mirror: virutal cloud task queue
            if not self.virtual_cloud_task_queue.empty():
                print("something on queue...")
                item = await self.virtual_cloud_task_queue.get()

                # in case there is anything, go ahead and dequeue the cloud side.
                print("all vehicles:", [v.getName() for v in self.vehicles])
                idle_vehicles = [{"vname": v.getName()} for v in self.vehicles if v.getStatus() == "running_idle" and self.vRunnable(v)]
                print("running idel vehicles:", idle_vehicles)
                resp = send_dequeue_tasks_to_cloud(self.session, self.get_auth_token(), idle_vehicles, self.getWanApiEndpoint())
                print("RESP:", resp)
                if "body" in resp:
                    cloudQSize = resp['body']['remainingQSize']
                    taskGroups = resp['body']['task_groups']
                    print("cloudQSize:", cloudQSize)
                    print("newTaskGroups:", taskGroups)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorCheckCloudWorkQueue:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorCheckCloudWorkQueue: traceback information not available:" + str(e)
            log3(ex_stat)
            taskGroups = {"task_groups": None, "added_missions": []}

        return taskGroups

    # if there is actual work, 1) deque from virutal cloud queue, 2) put it into local unassigned work list.
    # and the rest will be taken care of by the work dispatcher...
    # works organized as following....
    # { win: {computer1: {"estern": ..... "central":...} , computer2: ...} , mac:, linux:...}
    def arrangeContractWorks(self, contractWorks):
        if "added_missions" in contractWorks:
            if contractWorks["added_missions"] and contractWorks["task_groups"]:
                # first, download the files.

                log3("ARRANGE external contract work....")
                self.prepareMissionRunAsServer(contractWorks)

                print("updated contract works....", contractWorks)
                # first flatten timezone.
                newlyAddedMissions = self.addNewlyAddedMissions(contractWorks)

                print("newlyAddedMissions config:", newlyAddedMissions[0].getConfig())

                newTaskGroups = self.reGroupByBotVehicles(contractWorks["task_groups"])
                self.unassigned_reactive_task_groups = self.todays_scheduled_task_groups
                for vname in contractWorks["task_groups"]:
                    if vname in self.unassigned_reactive_task_groups:
                        if self.unassigned_reactive_task_groups[vname]:
                            self.unassigned_reactive_task_groups[vname] = self.merge_dicts(self.unassigned_reactive_task_groups[vname], newTaskGroups[vname])
                        else:
                            self.unassigned_reactive_task_groups[vname] = newTaskGroups[vname]
                    else:
                        self.unassigned_reactive_task_groups[vname] = newTaskGroups[vname]

                print("unassigned_reactive_task_groups after adding contract:", self.unassigned_reactive_task_groups)
                self.build_cookie_site_lists(newlyAddedMissions)

    def merge_dicts(self, dict1, dict2):
        merged_dict = {}
        for key in dict1.keys():
            merged_dict[key] = dict1[key] + dict2.get(key, [])
        return merged_dict

    # upon clicking here, it would simulate receiving a websocket message(cmd) and send this
    # message to the relavant queue which will trigger a mission run. (For unit testing purpose)
    def simWanRequest(self):
        contents_data = {
            "task_groups": {
                "DESKTOP-DLLV0:win":
                {
                    "eastern": [],
                    "central": [],
                    "mountain": [],
                    "pacific": [
                        {
                            "bid": 73,
                            "cuspas": "win,ads,ebay",
                            "tz": "pacific",
                            "bw_works": [],
                            "other_works": [{
                                "name": "sellFullfill_routine",
                                "mid": 697,
                                "cuspas": "win,ads,ebay",
                                "config": {
                                    "estRunTime": 2,
                                    "searches": []
                                },
                                "start_time": 30
                            }]
                        }
                    ],
                    "alaska": [],
                    "hawaii": []
                }
            },
            "added_missions": [
                {
                    "mid": 697,
                    "ticket": 0,
                    "owner": "songc@yahoo.com",
                    "botid": 73,
                    "status": "ASSIGNED",
                    "createon": "2024-03-31 05:44:15",
                    "esd": "2024-03-16 05:44:15",
                    "ecd": "2024-03-16 05:44:15",
                    "asd": "2124-03-16 05:44:15",
                    "abd": "2124-03-16 05:44:15",
                    "aad": "2124-03-16 05:44:15",
                    "afd": "2124-03-16 05:44:15",
                    "acd": "2124-03-16 05:44:15",
                    "esttime": 30,
                    "runtime": 2,
                    "trepeat": 3,
                    "cuspas": "win,ads,ebay",
                    "category": "",
                    "phrase": "",
                    "pseudoStore": "",
                    "pseudoBrand": "",
                    "pseudoASIN": "",
                    "type": "sellFullfill_routine",
                    "as_server": True,
                    "config": [
                        "sale",
                        [
                            {
                                "file": "",
                                "dir": ""
                            }
                        ]
                    ],
                    "skills": "87",
                    "delDate": "2124-03-16 05:44:15"
                }
            ]
        }
        sim_contents = json.dumps(contents_data)

        in_message = {
            "type": "request mission",
            "sender": "",
            "id": 0,
            "contents": sim_contents
        }

        asyncio.ensure_future((self.gui_monitor_msg_queue.put(in_message)))

    # check default directory and see whether there is any file dated within the past 24 hrs
    # if so return that file name.
    def checkNewBotsFiles(self):
        bfiles = []

        if bool(self.general_settings.new_bots_file_path):
            bfiles = self.get_new_bot_files(self.general_settings.new_bots_file_path)

        return bfiles

    def get_new_bot_files(self, base_dir="new_bot"):
        # Calculate the timestamp for yesterday at 12 AM
        yesterday_12am = datetime.combine(datetime.now() - timedelta(days=1), datetime.min.time())

        # Convert the timestamp to a Unix timestamp (seconds since epoch)
        timestamp_cutoff = yesterday_12am.timestamp()

        # Get all .xlsx files in the base directory
        bot_files = glob.glob(os.path.join(base_dir, "new_bots_*.xlsx"))

        # Filter files modified after yesterday's 12 AM
        new_bot_files = [file for file in bot_files if os.path.getmtime(file) > timestamp_cutoff]

        last_bots_file = self.general_settings.last_bots_file
        if not last_bots_file:
            latest_file = ""
            latest_time = 0
            last_time = 0
        else:
            latest_file = last_bots_file
            latest_time = self.general_settings.last_bots_file_time
            last_time = latest_time

        not_yet_touched_files = []
        # Print the new bot files (optional)
        if new_bot_files:
            print(f"Found {len(new_bot_files)} new bot files modified after {yesterday_12am}:")
            for file_path in new_bot_files:
                # Get the modification time for each file
                file_mtime = os.path.getmtime(file_path)
                # Update if this file is more recent
                if file_mtime > last_time:
                    not_yet_touched_files.append(file_path)

                if file_mtime > latest_time:
                    latest_time = file_mtime
                    latest_file = file_path

        else:
            print("No new bot files found since yesterday at 12 AM.")

        self.general_settings.last_bots_file = latest_file
        self.general_settings.last_bots_file_time = latest_time
        self.general_settings.save()
        return not_yet_touched_files



    def checkNewMissionsFiles(self):
        mfiles = []

        if bool(self.general_settings.new_orders_dir):
            log3("new_orders_dir:" + self.general_settings.new_orders_dir)
            mfiles = self.get_yesterday_orders_files(self.general_settings.new_orders_dir)
            log3("New order files since yesterday" + json.dumps(mfiles))

        return mfiles

    # make sure the networked dir is escapped correctly: "\\\\HP-ECBOT\\shared"
    def get_yesterday_orders_files(self, base_dir="new orders"):
        # Get yesterday's date
        yesterday = datetime.now() - timedelta(days=1)
        year = yesterday.strftime("%Y")
        month = yesterday.strftime("m%m")  # 'm01' format
        day = yesterday.strftime("d%d")  # 'd01' format

        # Build the path for yesterday's directory
        yesterday_dir = os.path.join(base_dir, year, month, day)

        # Check if the directory exists
        if not os.path.isdir(yesterday_dir):
            print(f"Directory {yesterday_dir} does not exist.")
            return []

        # Find all .xlsx files in yesterday's directory
        order_files = glob.glob(os.path.join(yesterday_dir, "Order*.xlsx"))
        last_order_file = self.general_settings.last_order_file
        if not last_order_file:
            latest_file = ""
            latest_time = 0
            last_time = 0
        else:
            latest_file = last_order_file
            latest_time = self.general_settings.last_order_file_time
            last_time = latest_time

        not_yet_touched_files = []
        # Print found files (optional)
        if order_files:
            print(f"Found {len(order_files)} order files for {yesterday.strftime('%Y-%m-%d')}:")
            for file_path in order_files:
                # Get the modification time for each file
                file_mtime = os.path.getmtime(file_path)

                if file_mtime > last_time:
                    not_yet_touched_files.append(file_path)

                # Update if this file is more recent
                if file_mtime > latest_time:
                    latest_time = file_mtime
                    latest_file = file_path
        else:
            print(f"No order files found for {yesterday.strftime('%Y-%m-%d')}.")

        self.general_settings.last_order_file = latest_file
        self.general_settings.last_order_file_time = latest_time
        self.general_settings.save()
        return not_yet_touched_files

    # assume one sheet only in the xlsx file. at this moment no support for multi-sheet.
    def convert_orders_xlsx_to_json(self, file_path):
        header_to_db_column = {
            "stores": "stores",
            "brand": "brand",
            "execution time": "execution_time",
            "quantity": "quantity",
            "asin": "asin",
            "search term": "phrase",
            "title": "title",
            "page number": "page_number",
            "price": "price",
            "variations": "variations",
            "follow seller": "follow_seller",
            "follow price": "follow_price",
            "product image": "img",
            "fb type": "feedback_type",
            "fb title": "feedback_title",
            "fb contents": "feedback_contents",
            "notes": "order_notes",
            "email": "customer"
        }

        # Load the Excel file
        log3("working on new order xlsx file:"+file_path)
        import pandas as pd
        df = pd.read_excel(file_path, header=2, dtype=str)  # Start reading from the 3rd row

        df.rename(columns=header_to_db_column, inplace=True)

        # Drop any completely empty columns, if there are any
        df.dropna(how="all", axis=1, inplace=True)
        df.dropna(how="all", axis=0, inplace=True)

        # Convert DataFrame to a list of dictionaries (JSON format)
        orders_json = df.to_dict(orient="records")
        return orders_json

    def generate_key_from_string(self, password: str) -> bytes:
        """Generate a 32-byte key from the given string (password) using a key derivation function."""
        password = password.encode()  # Convert string to bytes
        salt = b'some_salt_value'  # You can generate this securely if you want, here it's fixed for simplicity
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        key = base64.urlsafe_b64encode(kdf.derive(password))
        return key

    def encrypt_string(self, key: bytes, plaintext: str) -> str:
        """Encrypt the given plaintext using the derived key."""
        fernet = Fernet(key)
        encrypted = fernet.encrypt(plaintext.encode())  # Encrypt the plaintext
        return encrypted.decode()  # Return as a string

    def decrypt_string(self, key: bytes, encrypted_text: str) -> str:
        """Decrypt the given encrypted text using the derived key."""
        fernet = Fernet(key)
        decrypted = fernet.decrypt(encrypted_text.encode())  # Decrypt the text
        return decrypted.decode()  # Return as a string



    def generateShortHash(self, text: str, length: int = 32) -> str:
        """Generate a fixed-length hash for the input text."""
        hash_obj = hashlib.sha256(text.encode())
        hashed_bytes = hash_obj.digest()
        # Encode in base64 to make it URL-safe and take the desired length
        hashed = base64.urlsafe_b64encode(hashed_bytes).decode()[:length]
        # log3("hashed:"+hashed)
        return hashed

    def createNewMissionsFromOrdersXlsx(self):
        newMisionsFiles = self.checkNewMissionsFiles()
        if newMisionsFiles:
            last_order_info = self.get_last_order_file_info()
            log3("last_order_file:"+last_order_info["file"]+"..."+str(last_order_info["time"]))
            self.createMissionsFromFilesOrJsData(newMisionsFiles)

    def createNewBotsFromBotsXlsx(self):
        newBotsFiles = self.checkNewBotsFiles()
        log3("newBotsFiles:"+json.dumps(newBotsFiles))
        if newBotsFiles:
            firstNewBid, addedBots = self.createBotsFromFilesOrJsData(newBotsFiles)

    def isPlatoon(self):
        return (self.machine_role == "Platoon")


    def findManagerOfThisVehicle(self):
        # for bot in self.bots:
        #     print("bot:", bot.getRoles(), bot.getVehicle(), self.machine_name)
        foundBots = [x for x in self.bots if "manage" in x.getRoles().lower() and self.machine_name in x.getVehicle()]
        return foundBots

    def findManagerMissionsOfThisVehicle(self):
        managerBots = self.findManagerOfThisVehicle()
        logger.trace("#manager::", len(managerBots))
        managerBids = [x.getBid() for x in managerBots]
        logger.trace("#managerBids::", managerBids)
        managerMissions = [x for x in self.missions if x.getBid() in managerBids and ("completed" not in x.getStatus().lower())]
        return managerBots, managerMissions

    def getDailyFailedBots(self):
        failed = [b for b in self.bots if b.getStatus().lower() == "failed"]
        return failed

    def isValidAddr(self, addr):
        val = True

        if "Any,Any" in addr or not addr.split("\n")[0].strip():
            val = False

        return val

    def screenBuyerBotCandidates(self, acctRows, all_bots):
        # note the acctRows is in format of following....
        # just look at the ip, vccard, bot assignment
        allBotEmails = [b.getEmail() for b in all_bots]
        botsNeedsUpdate = [b for b in all_bots if (not b.getEmail()) or (not b.getBackEm()) or (not self.isValidAddr(b.getAddr())) or (not self.isValidAddr(b.getShippingAddr())) or (not b.getVehicle()) or (not b.getOrg())]

        print('allBotEmails:', allBotEmails)
        qualified = [row for row in acctRows if row["email"] and row["vcard_num"] and row["proxy_host"] and (not row["bot"]) and (row["email"] not in allBotEmails)]
        rowsNeedsUpdate = [row for row in acctRows if row["email"] and row["proxy_host"] and (not row["bot"]) and (row["email"] in allBotEmails)]
        vehiclesNeedsUpdate = []
        # for rows missing bot id, fill it in.
        for row in rowsNeedsUpdate:
            foundBot = next((x for x in self.bots if x.getEmail() == row["email"]), None)
            if foundBot:
                row["bot"] = foundBot.getBid()

        # update in data structure
        print("row bot:", [r["bot"] for r in acctRows])
        for bot in botsNeedsUpdate:
            print("ids:", bot.getBid())
            row = next((r for i, r in enumerate(acctRows) if bot.getBid() == r["bot"]), None)
            if row:
                print("found row....", row["addr_street_line1"])
                if row["email"]:
                    bot.setEmail(row["email"])
                    bot.setEPW(row["email_pw"])

                if row["backup_email"]:
                    bot.setBackEmail(row["backup_email"])
                    bot.setEBPW(row["backup_email_pw"])

                if row["addr_street_line1"]:
                    print("set address....")
                    bot.setAddr(row["addr_street_line1"], row["addr_street_line2"], row["addr_city"], row["addr_state"], row["addr_zip"])
                    print("after set addr:", bot.getAddr())

                if not self.isValidAddr(bot.getShippingAddr()) and row["addr_street_line1"]:
                    bot.setShippingAddr(row["addr_street_line1"], row["addr_street_line2"], row["addr_city"], row["addr_state"], row["addr_zip"])

            self.assignBotVehicle(bot, vehiclesNeedsUpdate)

            if not bot.getOrg():
                bot.setOrg("{}")


        print("bot ids for ones need update:", [b.getBid() for b in botsNeedsUpdate])
        return qualified, rowsNeedsUpdate, botsNeedsUpdate, vehiclesNeedsUpdate

    def assignBotVehicle(self, bot, vehiclesNeedsUpdate):
        if not bot.getVehicle():
            bv = self.genBotVehicle(bot)
            if bv:
                bot.setVehicle(bv.getName())
                print("setting vehicle:", bot.getVehicle())
                bv.addBot(bot.getBid())
                if bv not in vehiclesNeedsUpdate:
                    vehiclesNeedsUpdate.append(bv)
            else:
                log3("vehicle not found for a bot")

    # turn acct into bots/agents
    def hireBuyerBotCandidates(self, acctRows):
        newBotsJs = []
        for row in acctRows:
            # format conversion and some.
            newBotJS = {
                "pubProfile": {
                    "bid":0,
                    "pseudo_nick_name":row["first_name"],
                    "pseudo_name": self.genPseudoName(row["first_name"],row["last_name"]),
                    "location": self.genBotLoc(row["addr_state"]),
                    "pubbirthday": self.genBotPubBirthday(),
                    "gender": self.getBotGender(),
                    "interests":"Any,Any,Any,Any,Any",
                    "roles": "amz:buyer",
                    "org":"{}",
                    "levels": "amz:green:buyer",
                    "levelStart": "",
                    "vehicle": self.genBotVehicle("buyer").getName(),
                    "status": "Active"
                },
                "privateProfile":{
                    "first_name": row["first_name"],
                    "last_name": row["last_name"],
                    "email": row["email"],
                    "email_pw": row["email_pw"],
                    "phone": "",
                    "backup_email": row["backup_email"],
                    "acct_pw": row["backup_email_pw"],
                    "backup_email_site": "",
                    "birthday": "2000-02-02",
                    "addrl1": row["addr_street_line1"],
                    "addrl2": row["addr_street_line2"],
                    "addrcity": row["addr_city"],
                    "addrstate": row["addr_state"],
                    "addrzip": row["addr_zip"],
                    "shipaddrl1": row["addr_street_line1"],
                    "shipaddrl2": row["addr_street_line2"],
                    "shipaddrcity": row["addr_city"],
                    "shipaddrstate": row["addr_state"],
                    "shipaddrzip": row["addr_zip"],
                    "adsProfile": ""
                },
                "settings": {
                    "platform":"win",
                    "os":"win",
                    "browser":"ads",
                    "machine":""
                }
            }
            newBotsJs.append(newBotJS)

        print("newBotsJs:", newBotsJs)
        firstNewBid, addedBots = self.createBotsFromFilesOrJsData(newBotsJs)
        print("firstNewBid:", firstNewBid)
        if firstNewBid:
            newBid = firstNewBid
            for row in acctRows:
                row["bot"] = newBid
                newBid = newBid + 1


    def genPseudoName(self, fn, ln):
        pfn = fn
        pln = ln

        return pfn+" "+pln


    def genBotLoc(self,state):
        LARGEST_CITY = { "CA": "Los Angeles", "NY": "New York", "IL": "Chicago", "D.C.": "Washington", "WA": "Seattle", "TX": "Dallas"}
        # for simplicity, just use largest city of that state.
        loc = LARGEST_CITY[state]+","+state
        print("gen loc:", loc)
        return loc

    def genBotPubBirthday(self):
        # Randomly pick
        yyyy = random.randint(1995, 2005)
        mm = random.randint(1, 12)
        dd = random.randint(1, 28)
        # Format with leading zeros
        pbd = f"{yyyy}-{str(mm).zfill(2)}-{str(dd).zfill(2)}"
        print("pbd:", pbd)
        return pbd

    def getBotGender(self):
        # randomely pick
        gends = ["F", "M"]
        random_number = random.randint(0, 1)
        gend = gends[random_number]
        print("gend", gend)
        return gend

    def botFunctionMatchVehicle(self, b, v):
        fit = False
        if isinstance(b, str):
            roles = b.lower()
        else:
            roles = b.getRoles().lower()

        vfws = v.getFunctions().split(",")
        vfs = [r.strip().lower() for r in vfws]
        for vf in vfs:
            if vf in roles:
                fit = True

        return fit


    def genBotVehicle(self, bot):
        # fill the least filled vehicle first.
        fitV = ""
        functionMatchedV = [v for v in self.vehicles if self.botFunctionMatchVehicle(bot, v)]
        sortedV = sorted(functionMatchedV, key=lambda v: len(v.getBotIds()), reverse=False)
        print("sorted vehicles:", [(v.getName(), len(v.getBotIds())) for v in sortedV])
        fitV = ""
        for v in sortedV:
            if not v.getBotsOverCapStatus():
                fitV = v
                break

        return fitV

    # this function sends request to all on-line platoons and request they send back
    # all the latest finger print profiles of the troop members on that team.
    # we will stores them onto the local dir, if there is existing ones, compare the time stamp of incoming file and existing file,
    # if the incoming file has a later time stamp, then overwrite the existing one.
    async def syncFingerPrintRequest(self):
        """Sync fingerprint request - converted to async to avoid UI freeze"""
        try:
            self.botFingerPrintsReady = False
            if self.machine_role == "Commander":
                log3("syncing finger prints...")

                reqMsg = {"cmd": "reqSyncFingerPrintProfiles", "content": "now"}

                # send over scheduled tasks to platton.
                self.expected_vehicle_responses = {}
                for vehicle in self.vehicles:
                    print("vehicle:", vehicle.getName(), vehicle.getStatus())
                    if vehicle.getFieldLink() and "running" in vehicle.getStatus():
                        self.showMsg(get_printable_datetime() + "SENDING [" + vehicle.getName() + "]PLATOON[" + vehicle.getFieldLink()[
                            "ip"] + "]: " + json.dumps(reqMsg))

                        self.send_json_to_platoon(vehicle.getFieldLink(), reqMsg)
                        self.expected_vehicle_responses[vehicle.getName()] = None

                #now wait for the response to all come back. for each v, give it 10 seconds.
                VTIMEOUT = 2
                sync_time_out =  len(self.expected_vehicle_responses.keys())*VTIMEOUT
                sync_time_out = VTIMEOUT
                print("waiting for ", sync_time_out, "seconds...")
                await asyncio.sleep(8.0)  # Non-blocking wait for responses


        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorSyncFingerPrintRequest:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorSyncFingerPrintRequest: traceback information not available:" + str(e)
            log3(ex_stat)


    async def syncFingerPrintOnConnectedVehicle(self, vehicle):
        """Sync fingerprint on connected vehicle - converted to async to avoid UI freeze"""
        try:
            self.botFingerPrintsReady = False
            if self.machine_role == "Commander":
                log3("syncing finger prints", "gatherFingerPrints", self)

                reqMsg = {"cmd": "reqSyncFingerPrintProfiles", "content": "now"}

                # send over scheduled tasks to platton.
                self.expected_vehicle_responses = {}

                log3(f"vehicle: {vehicle.getName()} {vehicle.getStatus()}", "gatherFingerPrints", self)
                if vehicle.getFieldLink() and "running" in vehicle.getStatus():
                    self.showMsg(get_printable_datetime() + "SENDING [" + vehicle.getName() + "]PLATOON[" + vehicle.getFieldLink()[
                        "ip"] + "]: " + json.dumps(reqMsg))

                    self.send_json_to_platoon(vehicle.getFieldLink(), reqMsg)
                    self.expected_vehicle_responses[vehicle.getName()] = None

                #now wait for the response to all come back. for each v, give it 10 seconds.
                VTIMEOUT = 12
                sync_time_out = VTIMEOUT
                print("waiting for ", sync_time_out, "seconds....")
                # Non-blocking wait with progress indication
                for i in range(sync_time_out):
                    await asyncio.sleep(1.0)
                    print("tick...", sync_time_out - i - 1)


        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorSyncFingerPrintOnConnectedVehicle:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorSyncFingerPrintOnConnectedVehicle: traceback information not available:" + str(e)
            log3(ex_stat)


    # this function updates latest finger prints on this vehicle.
    # 1) go to ads dir, look for all xlsx, gather unique emails from username column
    # 2) then string part before "@" will be the user name to use.
    #    in the finger prints directory, there could be three types of files:
    #    i) individual user's text version of finger print profile named {username}.txt for example, JohnSmith.txt, (may or may not exist)
    #    ii) text version of batched finger print profiles which starts with "profiles", for example profiles*.txt file, this file could contains multiple individual user's finger print profile
    #    iii) xlsx version of batched finger print profiles which starts with "profiles", for example profiles*.txt file, this file could contains multiple individual user's finger print profile (may or may not exist)
    # 3) a individual's profile could exist in all three type of files.
    # 4) is it easier to just call batch to singles?
    # 5) finally make a backup copy of all updated profiles for record keeping. the backup
    #    dir should be self.ads_profile_dir/backup_datestring for example backup_20250110
    #    and at the same time, get rid of backup folders dated more than 2 weeks older than
    #    current date.
    def gatherFingerPrints(self):
        try:
            updated_profiles = []
            duplicated_profiles = []
            if self.machine_role == "Platoon":
                log3("gathering finger prints....", "gatherFingerPrints", self)
                print("gaterhing fp profiles", self.ads_profile_dir)

                # Define the directory containing profiles*.txt and individual profiles

                # Get all profiles*.txt files, sorted by timestamp (latest first)
                batch_files = sorted(
                    [
                        os.path.join(self.ads_profile_dir, f)
                        for f in os.listdir(self.ads_profile_dir)
                        if f.startswith("profiles") and f.endswith(".txt")
                    ],
                    key=os.path.getmtime,
                    reverse=True,
                )
                log3("time sorted batch_files:"+json.dumps(batch_files), "gatherFingerPrints", self)

                # Track already updated usernames
                updated_usernames = set()

                # Process each batch file
                for batch_file in batch_files:

                    # Extract usernames from the batch file
                    with open(batch_file, "r") as bf:
                        batch_content = bf.readlines()

                    usernames = set(
                        line.split("=")[1].strip().split("@")[0]  # Extract username before "@"
                        for line in batch_content
                        if line.startswith("username=")
                    )
                    # print("usernames in this batch file:", batch_file, usernames)
                    # Exclude already updated usernames when processing this batch
                    remaining_usernames = usernames - updated_usernames
                    log3(f"remaining_usernames: {remaining_usernames}", "gatherFingerPrints", self)

                    if remaining_usernames:
                        updateIndividualProfileFromBatchSavedTxt(self, batch_file,
                                                                      excludeUsernames=list(updated_usernames))
                        # obtain batch file's time stamp
                        batch_file_timestamp = os.path.getmtime(batch_file)

                        # Add updated profiles to the list
                        for username in remaining_usernames:
                            individual_profile_path = os.path.join(self.ads_profile_dir, f"{username}.txt")
                            updated_profiles.append(individual_profile_path)

                            # Set the timestamp of the individual profile to match the batch profile's timestamp
                            if os.path.exists(individual_profile_path):
                                os.utime(individual_profile_path, (batch_file_timestamp, batch_file_timestamp))
                    else:
                        duplicated_profiles.append(batch_file)

                    # Add processed usernames to the updated list
                    updated_usernames.update(usernames)

                log3(f"Updating usernames: {len(updated_usernames)} {updated_usernames}", "gatherFingerPrints", self)
                # Point #5: Save Backups and Delete Old Backup Directories
                # Create backup directory with a date suffix
                today_date = datetime.now().strftime("%Y%m%d")
                backup_dir = os.path.join(self.ads_profile_dir, f"backup_{today_date}")
                os.makedirs(backup_dir, exist_ok=True)
                # print(f"backup_dir: {backup_dir}")

                # Backup all updated profiles
                for profile in updated_profiles:
                    if os.path.exists(profile):
                        shutil.copy2(profile, backup_dir)  # Preserve file metadata during copy

                log3(f"Backup created at: {backup_dir}", "gatherFingerPrints", self)

                # Delete old backups (older than 2 weeks)
                two_weeks_ago = datetime.now() - timedelta(weeks=2)
                for folder in os.listdir(self.ads_profile_dir):
                    if folder.startswith("backup_"):
                        folder_path = os.path.join(self.ads_profile_dir, folder)
                        # Parse the date suffix from the folder name
                        try:
                            folder_date = datetime.strptime(folder.split("_")[1], "%Y%m%d")
                            if folder_date < two_weeks_ago:
                                shutil.rmtree(folder_path)
                                log3(f"Deleted old backup folder: {folder_path}", "gatherFingerPrints", self)
                        except (IndexError, ValueError):
                            log3(f"Skipped invalid backup folder: {folder_path}", "gatherFingerPrints", self)

                # **Remove Old Duplicated Profiles**
                for duplicate_file in duplicated_profiles:
                    duplicate_file_date = os.path.basename(duplicate_file).split("_")[1:4]  # Extract yyyy, mm, dd
                    try:
                        duplicate_file_date = datetime.strptime("_".join(duplicate_file_date),
                                                                         "%Y_%m_%d")
                        if duplicate_file_date < two_weeks_ago:
                            os.remove(duplicate_file)
                            log3(f"Deleted old duplicate profile: {duplicate_file}", "gatherFingerPrints", self)
                    except (IndexError, ValueError):
                        log3(f"Skipped invalid duplicate file: {duplicate_file}", "gatherFingerPrints", self)

            return updated_profiles

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorGatherFingerPrints:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorGatherFingerPrints: traceback information not available:" + str(e)
            log3(ex_stat)
            return []

    async def vehicleSetupTeam(self, vehicle):
        vname = vehicle.getName()
        print("send all ads profiles to the vehicle", vname)

        # find all bots on this vehicle,
        vbs = [b for b in self.bots if b.getVehicle() == vname]
        allvbids = [b.getBid() for b in vbs]
        print([(b.getBid(), b.getVehicle(), b.getEmail()) for b in self.bots])
        print("all bids on vehicle:", len(vbs), allvbids)
        # gather all ads profiles fo the bots, and send over the files.
        bot_profile_paths = [os.path.join(self.ads_profile_dir, b.getEmail().split("@")[0]+".txt") for b in vbs]
        # bot_profile_paths = [b.getADSProfile() for b in vbs]
        print("all vb profiles:", bot_profile_paths)
        print("all fl names:", [fl["name"] for fl in fieldLinks])
        #
        vlink = next((fl for i, fl in enumerate(fieldLinks) if fl["name"] in vname), None)

        # if not self.tcpServer == None:
        if vlink:
            print("sending all bot profiles to platoon...")
            self.batch_send_ads_profiles_to_platoon(vlink, "text", bot_profile_paths)
        else:
            print("WARNING: vehicle not connected!")


    async def setUpBotADSProfilesOnVehicles(self):
        print("send all ads profiles to all running vehicle")
        for v in self.vehicles:
            if "running" in v.getStatus():
                await self.vehicleSetupTeam(v)

    # from task group extract the vehicle related work, and get bots, missions, skills
    # all ready and send over to the vehicle to get the work started.
    async def vehicleSetupWorkSchedule(self, vehicle, p_task_groups, scheduled=True):
        try:
            vname = vehicle.getName()
            if vehicle and "running" in vehicle.getStatus():
                log3("working on remote task group vehicle : " + vname, "assignWork", self)
                # flatten tasks and regroup them based on sites, and divide them into batches
                # all_works = [work for tg in p_task_groups for work in tg.get("works", [])]
                batched_tasks, ads_profiles = formADSProfileBatchesFor1Vehicle(p_task_groups, vehicle, self)
                # print("add buy search", batched_tasks)
                # self.add_buy_searchs(batched_tasks)

                print("ads_profiles:", ads_profiles)
                # send fingerprint browser profiles to platoon/vehicle
                # for profile in ads_profiles:
                #     self.send_file_to_platoon(vehicle.getFieldLink(), "ads profile", profile)
                await self.vehicleSetupTeam(vehicle)

                # now need to fetch this task associated bots, mission, skills
                # get all bots IDs involved. get all mission IDs involved.
                tg_botids, tg_mids, tg_skids = self.getAllBotidsMidsSkidsFromTaskGroup(p_task_groups)
                vehicle.setBotIds(tg_botids)
                vehicle.setMids(tg_botids)

                log3("tg_skids:" + json.dumps(tg_skids), "assignWork", self)
                # put togehter all bots, missions, needed skills infommation in one batch and put onto the vehicle to
                # execute
                # resource_string = self.formBotsMissionsSkillsString(tg_botids, tg_mids, tg_skids)
                resource_bots, resource_missions, resource_skills = self.formBotsMissionsSkillsJsonData(tg_botids, tg_mids,
                                                                                                        tg_skids)
                if scheduled:
                    workCmd = "reqSetSchedule"
                else:
                    workCmd = "reqSetReactiveWorks"
                schedule = {"cmd": workCmd, "todos": batched_tasks, "bots": resource_bots,
                            "missions": resource_missions, "skills": resource_skills}

                # send over scheduled tasks to platton.
                if vehicle.getFieldLink():
                    log3(get_printable_datetime() + "SENDING [" + vname + "]PLATOON[" + vehicle.getFieldLink()[
                        "ip"] + "] SCHEDULE::: " + json.dumps(schedule), "assignWork", self)

                    self.send_json_to_platoon(vehicle.getFieldLink(), schedule)

                    # send over skills to platoon
                    await self.empower_platoon_with_skills(vehicle.getFieldLink(), tg_skids)

                else:
                    log3(get_printable_datetime() + "scheduled vehicle " + vname + " is not FOUND on LAN.", "assignWork", self)
            else:
                log3("WARNING: scheduled vehicle not found on network at the moment: " + vname, "assignWork", self)

        except Exception as e:
            # Get the traceback information
            traceback_info = traceback.extract_tb(e.__traceback__)
            # Extract the file name and line number from the last entry in the traceback
            if traceback_info:
                ex_stat = "ErrorVehicleSetupWorkSchedule:" + traceback.format_exc() + " " + str(e)
            else:
                ex_stat = "ErrorVehicleSetupWorkSchedule: traceback information not available:" + str(e)
            log3(ex_stat, "assignWork", self)


    def vehiclePing(self, vehicle):
        self.sendToVehicleByVip(vehicle.getIP())

    def vehicleShowMonitor(self, vehicle):
        if hasattr(self, 'vehicle_monitor'):
            # Note: VehicleMonitorManager is now a data handler, not a GUI window
            # You may need to implement a new GUI or use existing vehicle display methods
            self.vehicle_monitor.start_monitoring(vehicle)
        else:
            self.vehicle_monitor = VehicleMonitorManager(self, vehicle)
            # Note: VehicleMonitorManager is now a data handler, not a GUI window
            # You may need to implement a new GUI or use existing vehicle display methods
            self.vehicle_monitor.start_monitoring(vehicle)

    def genFeedbacks(self, mids):
        #assumption: all mids corresponds to the same product, there is only 1 product invovled here
        fbs = {}
        dtnow = datetime.now()
        date_word = dtnow.isoformat()
        foundM = next((x for x in self.missions if x.getMid() == mids[0]), None)

        if foundM:
            product = foundM.getTitle()
            qs = [{
                "msgID": "1",
                "user": "000",
                "timeStamp": date_word,
                "products": "",
                "goals": "",
                "options": json.dumps({}).replace('"', '\\"'),
                "background": "assume you are happy customers, and would like to write good reviews for both the seller and the product",
                "msg": f"please help generate {len(mids)} review for both seller and the product ({product}), each review should have a title and a review body, \n" +
                       f"the titles sh ould satisfy the following criteria: 1) concise, no more than 6 words long, 2) non repeating 3) title contents should match the body contents.\n"+
                       f"the body of the seller reivew should satisfy the following criteria: 1) concise, no more 2 sentences long, each sentence should have no more than 8 words, 2) non repeating 3) for seller usually, the good review is about a combinartion of good products, fast deliver, prompt communication, fiendly support etc.\n" +
                       f"the body of the product reivew  the following criteria: 1) no more 5 sentences long, best to be less than 3 sentences long, each sentence should contain no more than 12 words. 2) non repeating 3) try comment on product's quality, price, or particular attributes or good for certain occassions (for example, a gift to a closed friend or relative). 4) avoid exaggerating wording, it's perferrable to sound cliche \n" +
                       f"the return reponse should be json structure data, it should a list of json dicts with seller_fb_title, sell_fb_body, product_fb_title, product_fb_body as the key, and the corresponding text as the value.\n"
            }]
            response = send_query_chat_request_to_cloud(self.session, self.get_auth_token(), qs, self.getWanApiEndpoint())
            for midx, mid in enumerate(mids):
                fbs[str(mid)] = response[midx]

        return fbs

    def getRPAReports(self, start_date, end_date):
        log_settings = self.config_manager.get_log_settings()
        base_dir = log_settings.get("", "")
        """
        Get report files from start_date (non-inclusive) to end_date (inclusive),
        clean up the files, and return a dictionary with report data.

        :param base_dir: The base directory containing year-named folders.
        :param start_date: Start date in YYYYMMDD format (non-inclusive).
        :param end_date: End date in YYYYMMDD format (inclusive).
        :return: A dictionary with keys as report dates (YYYYMMDD) and values as JSON data.
        """
        reports = {}

        # Convert start_date and end_date to datetime objects
        start_date = datetime.strptime(start_date, "%Y%m%d")
        end_date = datetime.strptime(end_date, "%Y%m%d")

        # Ensure the base directory exists
        if not os.path.exists(base_dir):
            print(f"Base directory does not exist: {base_dir}")
            return reports

        # Iterate over year-named folders
        for year_folder in sorted(os.listdir(base_dir)):
            year_path = os.path.join(base_dir, year_folder)

            # Skip non-directory or invalid year folders
            if not os.path.isdir(year_path) or not year_folder.isdigit() or len(year_folder) != 4:
                continue

            # Iterate over files in the year folder
            for file_name in sorted(os.listdir(year_path)):
                # Check if the file matches the report naming convention
                if re.match(r"reports\d{8}", file_name):
                    report_date_str = file_name[7:15]
                    report_date = datetime.strptime(report_date_str, "%Y%m%d")

                    # Include only files within the date range
                    if start_date < report_date <= end_date:
                        file_path = os.path.join(year_path, file_name)
                        try:
                            # Read and clean the file content
                            with open(file_path, "r") as f:
                                cleaned_content = []
                                for line in f:
                                    if not re.match(r"^========+", line):
                                        cleaned_content.append(line.strip())
                                cleaned_json = "\n".join(cleaned_content)

                            # Load cleaned content as JSON and add to reports
                            report_data = json.loads(cleaned_json)
                            reports[report_date_str] = report_data
                        except Exception as e:
                            print(f"Error processing file {file_path}: {e}")

        return reports

    def dailyHousekeeping(self):
        # send a message to manager task to trigger the daily housekeeping task
        in_message = {"type": "ALL_WORK_DONE"}
        print("sending manager msg:", in_message)
        asyncio.ensure_future((self.gui_manager_msg_queue.put(in_message)))


    def dailyTeamPrep(self):
        # send a message to manager task to trigger the daily team prep task
        in_message = {"type": "SCHEDULE_READY"}
        print("sending manager msg:", in_message)
        asyncio.ensure_future((self.gui_manager_msg_queue.put(in_message)))


    def runExternalHook(self, hook, params):
        global symTab
        symTab["hook_flag"] = False
        symTab["hook_result"] = None
        symTab["hook_params"] = params
        hook_path = self.my_ecb_data_homepath + '/my_skills/hooks'
        hook_file = hook+".py"
        stepjson = {
            "type": "External Hook",
            "file_name_type": "direct",
            "file_path": hook_path,
            "file_name": hook_file,
            "params": "hook_params",  # Optional dictionary of parameters for the external script
            "result": "hook_result",
            "flag": "hook_flag"
        }
        i, runStat = processExternalHook(stepjson, 1)
        # print("hook result:", symTab["hook_result"])
        return runStat


print(TimeUtil.formatted_now_with_ms() + " load MainGui all finished...")